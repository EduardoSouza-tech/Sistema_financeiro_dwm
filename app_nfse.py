"""
Aplicação Flask standalone para Consulta de NFS-e (Nota Fiscal de Serviço Eletrônica)
Microserviço separado para operações com NFS-e via SOAP

Provedores suportados:
- GINFES (500+ municípios)
- ISS.NET (200+ municípios)
- BETHA (1000+ municípios)
- e-ISS (150+ municípios)
- WebISS (50+ municípios)
- SimplISS (300+ municípios)
"""
import os
import sys
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
import logging

# Importar módulos do sistema NFS-e
from nfse_database import NFSeDatabase
from nfse_service import NFSeService, descobrir_provedor, testar_conexao
from nfse_functions import (
    salvar_pdf_nfse, salvar_xml_nfse,
    configurar_municipio, buscar_nfse_por_periodo,
    consultar_nfse_por_numero
)

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Criar app Flask
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# CORS para permitir chamadas do frontend principal
CORS(app, resources={
    r"/api/*": {
        "origins": os.environ.get('FRONTEND_URL', '*'),
        "methods": ["GET", "POST", "PUT", "DELETE"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Configuração do banco
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    logger.error("DATABASE_URL não configurada!")
    sys.exit(1)

# Inicializar módulos NFS-e
try:
    nfse_db = NFSeDatabase(DATABASE_URL)
    nfse_service = NFSeService()
    logger.info("✅ Módulos NFS-e inicializados com sucesso")
except Exception as e:
    logger.error(f"❌ Erro ao inicializar módulos NFS-e: {e}")
    sys.exit(1)


def get_db_connection():
    """Retorna conexão com banco de dados"""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        logger.error(f"Erro ao conectar ao banco: {e}")
        raise


def get_empresa_id_from_token(token):
    """Extrai empresa_id do token JWT (simplificado)"""
    # TODO: Implementar validação JWT completa
    empresa_id = request.headers.get('X-Empresa-ID', '1')
    return int(empresa_id)


def require_auth(f):
    """Decorator simples de autenticação"""
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token não fornecido'}), 401
        
        # TODO: Validar token JWT
        return f(*args, **kwargs)
    
    decorated_function.__name__ = f.__name__
    return decorated_function


# ==================== ROTAS DA API ====================

@app.route('/health')
def health_check():
    """Health check para Railway"""
    try:
        conn = get_db_connection()
        conn.close()
        return jsonify({
            'status': 'healthy',
            'service': 'nfse-consulta',
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 503


@app.route('/')
def index():
    """Página inicial"""
    return render_template('nfse_dashboard.html')


@app.route('/api/nfse/config', methods=['GET'])
@require_auth
def listar_configuracoes():
    """Lista configurações de municípios"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        configs = nfse_db.obter_configuracoes_municipio(empresa_id)
        
        return jsonify({
            'success': True,
            'configuracoes': configs
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar configurações: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/config', methods=['POST'])
@require_auth
def criar_configuracao():
    """Cadastra nova configuração de município"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        # Validações
        campos_obrigatorios = ['codigo_municipio', 'cnpj_prestador', 'inscricao_municipal', 'provedor']
        for campo in campos_obrigatorios:
            if not data.get(campo):
                return jsonify({'error': f'Campo {campo} é obrigatório'}), 400
        
        # Criar configuração
        resultado = configurar_municipio(
            empresa_id=empresa_id,
            codigo_municipio=data['codigo_municipio'],
            nome_municipio=data.get('nome_municipio', ''),
            cnpj_prestador=data['cnpj_prestador'],
            inscricao_municipal=data['inscricao_municipal'],
            provedor=data['provedor'],
            url_webservice=data.get('url_webservice'),
            usuario=data.get('usuario'),
            senha=data.get('senha')
        )
        
        if resultado.get('success'):
            return jsonify(resultado), 201
        else:
            return jsonify(resultado), 400
            
    except Exception as e:
        logger.error(f"Erro ao criar configuração: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/config/<int:config_id>', methods=['PUT'])
@require_auth
def atualizar_configuracao(config_id):
    """Atualiza configuração de município"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        resultado = nfse_db.atualizar_configuracao_municipio(
            config_id=config_id,
            empresa_id=empresa_id,
            dados=data
        )
        
        if resultado:
            return jsonify({'success': True, 'message': 'Configuração atualizada'})
        else:
            return jsonify({'error': 'Configuração não encontrada'}), 404
            
    except Exception as e:
        logger.error(f"Erro ao atualizar configuração: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/config/<int:config_id>', methods=['DELETE'])
@require_auth
def deletar_configuracao(config_id):
    """Desativa configuração de município"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        resultado = nfse_db.desativar_configuracao_municipio(config_id, empresa_id)
        
        if resultado:
            return jsonify({'success': True, 'message': 'Configuração desativada'})
        else:
            return jsonify({'error': 'Configuração não encontrada'}), 404
            
    except Exception as e:
        logger.error(f"Erro ao deletar configuração: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/buscar', methods=['POST'])
@require_auth
def buscar_nfse():
    """Busca NFS-e em um período"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        config_id = data.get('config_id')
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        if not all([config_id, data_inicio, data_fim]):
            return jsonify({'error': 'Parâmetros incompletos'}), 400
        
        # Buscar NFS-e
        resultado = buscar_nfse_por_periodo(
            empresa_id=empresa_id,
            config_id=config_id,
            data_inicio=data_inicio,
            data_fim=data_fim
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao buscar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/consultar', methods=['POST'])
@require_auth
def consultar_nfse():
    """Consulta NFS-e por número"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        config_id = data.get('config_id')
        numero_nfse = data.get('numero_nfse')
        
        if not config_id or not numero_nfse:
            return jsonify({'error': 'config_id e numero_nfse são obrigatórios'}), 400
        
        # Consultar NFS-e
        resultado = consultar_nfse_por_numero(
            empresa_id=empresa_id,
            config_id=config_id,
            numero_nfse=numero_nfse
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao consultar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse', methods=['GET'])
@require_auth
def listar_nfse():
    """Lista NFS-e com paginação"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        # Parâmetros de paginação e filtros
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))
        config_id = request.args.get('config_id')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        offset = (pagina - 1) * por_pagina
        
        # Buscar NFS-e do banco
        resultado = nfse_db.listar_nfse(
            empresa_id=empresa_id,
            config_id=config_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limite=por_pagina,
            offset=offset
        )
        
        return jsonify({
            'success': True,
            'nfse': resultado.get('nfse', []),
            'paginacao': resultado.get('paginacao', {})
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/<int:nfse_id>', methods=['GET'])
@require_auth
def detalhes_nfse(nfse_id):
    """Retorna detalhes de uma NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        nfse = nfse_db.obter_nfse_por_id(nfse_id, empresa_id)
        
        if not nfse:
            return jsonify({'error': 'NFS-e não encontrada'}), 404
        
        return jsonify({
            'success': True,
            'nfse': nfse
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/<int:nfse_id>', methods=['DELETE'])
@require_auth
def deletar_nfse(nfse_id):
    """Deleta uma NFS-e do banco"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        resultado = nfse_db.deletar_nfse(nfse_id, empresa_id)
        
        if resultado:
            return jsonify({'success': True, 'message': 'NFS-e deletada'})
        else:
            return jsonify({'error': 'NFS-e não encontrada'}), 404
            
    except Exception as e:
        logger.error(f"Erro ao deletar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/<int:nfse_id>/pdf', methods=['GET'])
@require_auth
def download_pdf_nfse(nfse_id):
    """Download do PDF (DANFSe) da NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        nfse = nfse_db.obter_nfse_por_id(nfse_id, empresa_id)
        
        if not nfse or not nfse.get('caminho_pdf'):
            return jsonify({'error': 'PDF não encontrado'}), 404
        
        caminho_pdf = nfse['caminho_pdf']
        
        if not os.path.exists(caminho_pdf):
            return jsonify({'error': 'Arquivo PDF não encontrado no storage'}), 404
        
        return send_file(
            caminho_pdf,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"NFS-e_{nfse['numero_nfse']}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Erro ao baixar PDF: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/<int:nfse_id>/xml', methods=['GET'])
@require_auth
def download_xml_nfse(nfse_id):
    """Download do XML da NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        nfse = nfse_db.obter_nfse_por_id(nfse_id, empresa_id)
        
        if not nfse or not nfse.get('caminho_xml'):
            return jsonify({'error': 'XML não encontrado'}), 404
        
        caminho_xml = nfse['caminho_xml']
        
        if not os.path.exists(caminho_xml):
            return jsonify({'error': 'Arquivo XML não encontrado no storage'}), 404
        
        return send_file(
            caminho_xml,
            mimetype='application/xml',
            as_attachment=True,
            download_name=f"NFS-e_{nfse['numero_nfse']}.xml"
        )
        
    except Exception as e:
        logger.error(f"Erro ao baixar XML: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/estatisticas', methods=['GET'])
@require_auth
def estatisticas_nfse():
    """Retorna estatísticas das NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        stats = nfse_db.obter_estatisticas(empresa_id)
        
        return jsonify({
            'success': True,
            'estatisticas': stats
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar estatísticas: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/resumo-mensal', methods=['POST'])
@require_auth
def resumo_mensal_nfse():
    """Resumo mensal de NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        ano = data.get('ano', datetime.now().year)
        mes = data.get('mes', datetime.now().month)
        
        resumo = nfse_db.obter_resumo_mensal(empresa_id, ano, mes)
        
        return jsonify({
            'success': True,
            'resumo': resumo
        })
        
    except Exception as e:
        logger.error(f"Erro ao gerar resumo mensal: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/export/excel', methods=['POST'])
@require_auth
def exportar_excel_nfse():
    """Exporta NFS-e para Excel"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        config_id = data.get('config_id')
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        # Buscar dados
        resultado = nfse_db.listar_nfse(
            empresa_id=empresa_id,
            config_id=config_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limite=10000
        )
        
        nfse_list = resultado.get('nfse', [])
        
        if not nfse_list:
            return jsonify({'error': 'Nenhuma NFS-e encontrada'}), 404
        
        # Gerar Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "NFS-e"
        
        # Cabeçalhos
        headers = [
            'Número', 'Código Verificação', 'Data Emissão', 'Prestador (CNPJ)',
            'Prestador (Nome)', 'Tomador (CNPJ)', 'Tomador (Nome)',
            'Valor Serviços', 'Valor ISS', 'Status', 'Município'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for row, nfse in enumerate(nfse_list, 2):
            ws.cell(row=row, column=1, value=nfse.get('numero_nfse'))
            ws.cell(row=row, column=2, value=nfse.get('codigo_verificacao'))
            ws.cell(row=row, column=3, value=nfse.get('data_emissao', '').strftime('%d/%m/%Y') if nfse.get('data_emissao') else '')
            ws.cell(row=row, column=4, value=nfse.get('cnpj_prestador'))
            ws.cell(row=row, column=5, value=nfse.get('razao_social_prestador'))
            ws.cell(row=row, column=6, value=nfse.get('cnpj_tomador'))
            ws.cell(row=row, column=7, value=nfse.get('razao_social_tomador'))
            ws.cell(row=row, column=8, value=nfse.get('valor_servicos'))
            ws.cell(row=row, column=9, value=nfse.get('valor_iss'))
            ws.cell(row=row, column=10, value=nfse.get('status'))
            ws.cell(row=row, column=11, value=nfse.get('nome_municipio'))
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"nfse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/certificado/upload', methods=['POST'])
@require_auth
def upload_certificado():
    """Upload de certificado digital A1 (.pfx)"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        if 'arquivo' not in request.files:
            return jsonify({'error': 'Arquivo não enviado'}), 400
        
        arquivo = request.files['arquivo']
        senha = request.form.get('senha')
        
        if not senha:
            return jsonify({'error': 'Senha não fornecida'}), 400
        
        # Salvar certificado
        resultado = nfse_db.salvar_certificado(
            empresa_id=empresa_id,
            arquivo_pfx=arquivo.read(),
            senha=senha
        )
        
        if resultado.get('success'):
            return jsonify(resultado), 201
        else:
            return jsonify(resultado), 400
            
    except Exception as e:
        logger.error(f"Erro ao fazer upload de certificado: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/certificado', methods=['GET'])
@require_auth
def listar_certificados():
    """Lista certificados digitais"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        certificados = nfse_db.listar_certificados(empresa_id)
        
        return jsonify({
            'success': True,
            'certificados': certificados
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar certificados: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/provedores', methods=['GET'])
def listar_provedores():
    """Lista provedores NFS-e disponíveis"""
    from nfse_service import PROVEDORES_NFSE
    
    provedores = [
        {
            'codigo': codigo,
            'nome': info['nome'],
            'padrao_abrasf': info['padrao_abrasf'],
            'municipios_estimados': info.get('municipios_estimados', 'N/A')
        }
        for codigo, info in PROVEDORES_NFSE.items()
    ]
    
    return jsonify({
        'success': True,
        'provedores': provedores
    })


@app.route('/api/nfse/testar-conexao', methods=['POST'])
@require_auth
def testar_conexao_municipio():
    """Testa conexão com webservice do município"""
    try:
        data = request.json
        
        codigo_municipio = data.get('codigo_municipio')
        provedor = data.get('provedor')
        url_webservice = data.get('url_webservice')
        
        if not all([codigo_municipio, provedor]):
            return jsonify({'error': 'Parâmetros incompletos'}), 400
        
        resultado = testar_conexao(
            codigo_municipio=codigo_municipio,
            provedor=provedor,
            url_webservice=url_webservice
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao testar conexão: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== INICIALIZAÇÃO ====================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5002))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    logger.info(f"🚀 Iniciando serviço NFS-e na porta {port}")
    logger.info(f"📊 Ambiente: {'PRODUÇÃO' if not debug else 'DESENVOLVIMENTO'}")
    
    app.run(
        host='0.0.0.0',
        port=port,
        debug=debug
    )
