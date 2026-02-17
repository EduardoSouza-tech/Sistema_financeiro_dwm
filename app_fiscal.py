"""
Aplicação Flask standalone para Documentos Fiscais
Microserviço unificado: NF-e, CT-e e NFS-e

Funcionalidades:
- NF-e/CT-e: Busca automática via SEFAZ (SOAP/REST)
- NFS-e: Consulta via provedores municipais (SOAP)

Provedores NFS-e: GINFES, ISS.NET, BETHA, e-ISS, WebISS, SimplISS
"""
import os
import sys
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
import logging

# Importar módulos NF-e/CT-e
from relatorios.nfe.nfe_api import NFeBuscaAPI
from relatorios.nfe.nfe_busca import NFeBusca
from relatorios.nfe.nfe_processor import NFeProcessor
from relatorios.nfe.nfe_storage import NFeStorage

# Importar módulos NFS-e
from nfse_database import NFSeDatabase
from nfse_service import NFSeService, descobrir_provedor, testar_conexao
from nfse_functions import (
    salvar_pdf_nfse, salvar_xml_nfse,
    configurar_municipio, buscar_nfse_por_periodo,
    consultar_nfse_por_numero
)

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Criar app Flask
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# CORS para permitir chamadas do frontend principal
CORS(app, resources={
    r"/api/*": {
        "origins": os.environ.get('FRONTEND_URL', '*'),
        "methods": ["GET", "POST", "PUT", "DELETE"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Configuração do banco
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    logger.error("DATABASE_URL não configurada!")
    sys.exit(1)

# Inicializar módulos
try:
    # NF-e/CT-e
    nfe_storage = NFeStorage()
    nfe_processor = NFeProcessor()
    nfe_busca = NFeBusca()
    nfe_api = NFeBuscaAPI()
    
    # NFS-e
    nfse_db = NFSeDatabase(DATABASE_URL)
    nfse_service = NFSeService()
    
    logger.info("✅ Módulos fiscais inicializados com sucesso")
except Exception as e:
    logger.error(f"❌ Erro ao inicializar módulos: {e}")
    sys.exit(1)


def get_db_connection():
    """Retorna conexão com banco de dados"""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        logger.error(f"Erro ao conectar ao banco: {e}")
        raise


def get_empresa_id_from_token(token):
    """Extrai empresa_id do token JWT"""
    empresa_id = request.headers.get('X-Empresa-ID', '1')
    return int(empresa_id)


def require_auth(f):
    """Decorator de autenticação"""
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token não fornecido'}), 401
        return f(*args, **kwargs)
    
    decorated_function.__name__ = f.__name__
    return decorated_function


# ==================== ROTAS GERAIS ====================

@app.route('/health')
def health_check():
    """Health check para Railway"""
    try:
        conn = get_db_connection()
        conn.close()
        return jsonify({
            'status': 'healthy',
            'service': 'documentos-fiscais',
            'modules': ['nfe', 'cte', 'nfse'],
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 503


@app.route('/')
def index():
    """Página inicial"""
    return render_template('fiscal_dashboard.html')


# ==================== ROTAS NF-e/CT-e ====================

@app.route('/api/nfe/certificados', methods=['GET'])
@require_auth
def listar_certificados_nfe():
    """Lista certificados digitais NF-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT 
                id, empresa_id, tipo_certificado, cnpj, 
                razao_social, valido_ate, ativo, criado_em,
                (valido_ate > NOW()) as valido
            FROM certificados_digitais
            WHERE empresa_id = %s
            ORDER BY ativo DESC, valido_ate DESC
        """, (empresa_id,))
        
        certificados = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'certificados': certificados
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar certificados NF-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfe/certificados/novo', methods=['POST'])
@require_auth
def criar_certificado_nfe():
    """Cadastra certificado digital NF-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        if 'arquivo_pfx' not in request.files:
            return jsonify({'error': 'Arquivo PFX não enviado'}), 400
        
        arquivo = request.files['arquivo_pfx']
        senha = request.form.get('senha')
        tipo = request.form.get('tipo', 'A1')
        
        if not senha:
            return jsonify({'error': 'Senha não fornecida'}), 400
        
        resultado = nfe_api.cadastrar_certificado(
            empresa_id=empresa_id,
            arquivo_pfx=arquivo.read(),
            senha=senha,
            tipo_certificado=tipo
        )
        
        if resultado.get('success'):
            return jsonify(resultado), 201
        else:
            return jsonify(resultado), 400
            
    except Exception as e:
        logger.error(f"Erro ao criar certificado NF-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfe/buscar-documentos', methods=['POST'])
@require_auth
def buscar_nfe():
    """Busca automática NF-e/CT-e na SEFAZ"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        resultado = nfe_api.buscar_documentos_automatico(
            empresa_id=empresa_id,
            certificado_id=data.get('certificado_id'),
            tipo_ambiente=data.get('ambiente', 'homologacao'),
            data_inicio=data.get('data_inicio'),
            data_fim=data.get('data_fim')
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao buscar NF-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfe/consultar-chave', methods=['POST'])
@require_auth
def consultar_nfe_por_chave():
    """Consulta NF-e por chave de acesso"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        resultado = nfe_api.consultar_por_chave(
            empresa_id=empresa_id,
            certificado_id=data.get('certificado_id'),
            chave_acesso=data.get('chave_acesso'),
            tipo_ambiente=data.get('ambiente', 'homologacao')
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao consultar NF-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfe/documentos', methods=['GET'])
@require_auth
def listar_nfe():
    """Lista NF-e/CT-e com paginação"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))
        tipo_doc = request.args.get('tipo')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        offset = (pagina - 1) * por_pagina
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        where_clauses = ['empresa_id = %s']
        params = [empresa_id]
        
        if tipo_doc:
            where_clauses.append('tipo_documento = %s')
            params.append(tipo_doc)
        
        if data_inicio:
            where_clauses.append('data_emissao >= %s')
            params.append(data_inicio)
        
        if data_fim:
            where_clauses.append('data_emissao <= %s')
            params.append(data_fim)
        
        where_sql = ' AND '.join(where_clauses)
        
        cursor.execute(f"""
            SELECT COUNT(*) as total
            FROM documentos_fiscais_log
            WHERE {where_sql}
        """, params)
        total = cursor.fetchone()['total']
        
        cursor.execute(f"""
            SELECT *
            FROM documentos_fiscais_log
            WHERE {where_sql}
            ORDER BY data_emissao DESC, criado_em DESC
            LIMIT %s OFFSET %s
        """, params + [por_pagina, offset])
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'documentos': documentos,
            'paginacao': {
                'pagina': pagina,
                'por_pagina': por_pagina,
                'total': total,
                'total_paginas': (total + por_pagina - 1) // por_pagina
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar NF-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfe/documento/<int:doc_id>/xml', methods=['GET'])
@require_auth
def download_xml_nfe(doc_id):
    """Download XML NF-e/CT-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT chave_acesso, numero, tipo_documento
            FROM documentos_fiscais_log
            WHERE id = %s AND empresa_id = %s
        """, (doc_id, empresa_id))
        
        documento = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not documento:
            return jsonify({'error': 'Documento não encontrado'}), 404
        
        tipo_doc = 'nfe' if documento['tipo_documento'] == 'NF-e' else 'cte'
        xml_path = nfe_storage.obter_caminho_xml(
            empresa_id,
            tipo_doc,
            documento['chave_acesso']
        )
        
        if not os.path.exists(xml_path):
            return jsonify({'error': 'XML não encontrado'}), 404
        
        return send_file(
            xml_path,
            mimetype='application/xml',
            as_attachment=True,
            download_name=f"{documento['tipo_documento']}_{documento['numero']}.xml"
        )
        
    except Exception as e:
        logger.error(f"Erro ao baixar XML NF-e: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== ROTAS NFS-e ====================

@app.route('/api/nfse/config', methods=['GET'])
@require_auth
def listar_config_nfse():
    """Lista configurações municípios NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        configs = nfse_db.obter_configuracoes_municipio(empresa_id)
        
        return jsonify({
            'success': True,
            'configuracoes': configs
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar configurações NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/config', methods=['POST'])
@require_auth
def criar_config_nfse():
    """Cadastra município NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        resultado = configurar_municipio(
            empresa_id=empresa_id,
            codigo_municipio=data['codigo_municipio'],
            nome_municipio=data.get('nome_municipio', ''),
            cnpj_prestador=data['cnpj_prestador'],
            inscricao_municipal=data['inscricao_municipal'],
            provedor=data['provedor'],
            url_webservice=data.get('url_webservice'),
            usuario=data.get('usuario'),
            senha=data.get('senha')
        )
        
        if resultado.get('success'):
            return jsonify(resultado), 201
        else:
            return jsonify(resultado), 400
            
    except Exception as e:
        logger.error(f"Erro ao criar configuração NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/buscar', methods=['POST'])
@require_auth
def buscar_nfse():
    """Busca NFS-e por período"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        resultado = buscar_nfse_por_periodo(
            empresa_id=empresa_id,
            config_id=data.get('config_id'),
            data_inicio=data.get('data_inicio'),
            data_fim=data.get('data_fim')
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao buscar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse', methods=['GET'])
@require_auth
def listar_nfse():
    """Lista NFS-e com paginação"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))
        config_id = request.args.get('config_id')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        offset = (pagina - 1) * por_pagina
        
        resultado = nfse_db.listar_nfse(
            empresa_id=empresa_id,
            config_id=config_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limite=por_pagina,
            offset=offset
        )
        
        return jsonify({
            'success': True,
            'nfse': resultado.get('nfse', []),
            'paginacao': resultado.get('paginacao', {})
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/<int:nfse_id>/pdf', methods=['GET'])
@require_auth
def download_pdf_nfse(nfse_id):
    """Download PDF NFS-e"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        nfse = nfse_db.obter_nfse_por_id(nfse_id, empresa_id)
        
        if not nfse or not nfse.get('caminho_pdf'):
            return jsonify({'error': 'PDF não encontrado'}), 404
        
        return send_file(
            nfse['caminho_pdf'],
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"NFS-e_{nfse['numero_nfse']}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Erro ao baixar PDF NFS-e: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nfse/provedores', methods=['GET'])
def listar_provedores_nfse():
    """Lista provedores NFS-e"""
    from nfse_service import PROVEDORES_NFSE
    
    provedores = [
        {
            'codigo': codigo,
            'nome': info['nome'],
            'padrao_abrasf': info['padrao_abrasf']
        }
        for codigo, info in PROVEDORES_NFSE.items()
    ]
    
    return jsonify({
        'success': True,
        'provedores': provedores
    })


# ==================== ROTAS DE EXPORTAÇÃO ====================

@app.route('/api/fiscal/exportar-excel', methods=['POST'])
@require_auth
def exportar_excel():
    """Exporta documentos fiscais para Excel"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        tipo = data.get('tipo')  # 'nfe' ou 'nfse'
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import io
        
        wb = Workbook()
        ws = wb.active
        
        if tipo == 'nfe':
            ws.title = "NF-e e CT-e"
            
            # Buscar NF-e/CT-e
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("""
                SELECT 
                    tipo_documento, numero, serie, chave_acesso,
                    data_emissao, cnpj_emitente, nome_emitente,
                    cnpj_destinatario, nome_destinatario,
                    valor_total, status_sefaz
                FROM documentos_fiscais_log
                WHERE empresa_id = %s
                    AND data_emissao BETWEEN %s AND %s
                ORDER BY data_emissao DESC
            """, (empresa_id, data_inicio, data_fim))
            
            documentos = cursor.fetchall()
            cursor.close()
            conn.close()
            
            # Cabeçalhos
            headers = ['Tipo', 'Número', 'Série', 'Chave', 'Data', 'CNPJ Emit.', 
                      'Nome Emit.', 'CNPJ Dest.', 'Nome Dest.', 'Valor', 'Status']
            
        else:  # nfse
            ws.title = "NFS-e"
            
            resultado = nfse_db.listar_nfse(
                empresa_id=empresa_id,
                data_inicio=data_inicio,
                data_fim=data_fim,
                limite=10000
            )
            
            documentos = resultado.get('nfse', [])
            
            # Cabeçalhos
            headers = ['Número', 'Código Verif.', 'Data', 'CNPJ Prest.', 
                      'Nome Prest.', 'CNPJ Tom.', 'Nome Tom.', 'Valor', 'ISS', 'Município']
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Escrever dados
        for row, doc in enumerate(documentos, 2):
            if tipo == 'nfe':
                ws.cell(row=row, column=1, value=doc.get('tipo_documento'))
                ws.cell(row=row, column=2, value=doc.get('numero'))
                ws.cell(row=row, column=3, value=doc.get('serie'))
                ws.cell(row=row, column=4, value=doc.get('chave_acesso'))
                ws.cell(row=row, column=5, value=doc.get('data_emissao', '').strftime('%d/%m/%Y') if doc.get('data_emissao') else '')
                ws.cell(row=row, column=6, value=doc.get('cnpj_emitente'))
                ws.cell(row=row, column=7, value=doc.get('nome_emitente'))
                ws.cell(row=row, column=8, value=doc.get('cnpj_destinatario'))
                ws.cell(row=row, column=9, value=doc.get('nome_destinatario'))
                ws.cell(row=row, column=10, value=doc.get('valor_total'))
                ws.cell(row=row, column=11, value=doc.get('status_sefaz'))
            else:
                ws.cell(row=row, column=1, value=doc.get('numero_nfse'))
                ws.cell(row=row, column=2, value=doc.get('codigo_verificacao'))
                ws.cell(row=row, column=3, value=doc.get('data_emissao', '').strftime('%d/%m/%Y') if doc.get('data_emissao') else '')
                ws.cell(row=row, column=4, value=doc.get('cnpj_prestador'))
                ws.cell(row=row, column=5, value=doc.get('razao_social_prestador'))
                ws.cell(row=row, column=6, value=doc.get('cnpj_tomador'))
                ws.cell(row=row, column=7, value=doc.get('razao_social_tomador'))
                ws.cell(row=row, column=8, value=doc.get('valor_servicos'))
                ws.cell(row=row, column=9, value=doc.get('valor_iss'))
                ws.cell(row=row, column=10, value=doc.get('nome_municipio'))
        
        # Salvar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"documentos_fiscais_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/fiscal/estatisticas', methods=['GET'])
@require_auth
def estatisticas_gerais():
    """Estatísticas gerais de documentos fiscais"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Stats NF-e/CT-e
        cursor.execute("""
            SELECT 
                COUNT(*) as total_docs,
                COUNT(CASE WHEN tipo_documento = 'NF-e' THEN 1 END) as total_nfes,
                COUNT(CASE WHEN tipo_documento = 'CT-e' THEN 1 END) as total_ctes,
                COALESCE(SUM(valor_total), 0) as valor_total_nfe
            FROM documentos_fiscais_log
            WHERE empresa_id = %s
        """, (empresa_id,))
        
        stats_nfe = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Stats NFS-e
        stats_nfse = nfse_db.obter_estatisticas(empresa_id)
        
        return jsonify({
            'success': True,
            'nfe_cte': stats_nfe,
            'nfse': stats_nfse
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar estatísticas: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== INICIALIZAÇÃO ====================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    logger.info(f"🚀 Iniciando serviço Documentos Fiscais na porta {port}")
    logger.info(f"📊 Módulos: NF-e, CT-e, NFS-e")
    logger.info(f"📊 Ambiente: {'PRODUÇÃO' if not debug else 'DESENVOLVIMENTO'}")
    
    app.run(
        host='0.0.0.0',
        port=port,
        debug=debug
    )
