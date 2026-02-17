"""
Aplicação Flask standalone para Busca Automática de NF-e/CT-e via SEFAZ
Microserviço separado para operações fiscais
"""
import os
import sys
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
import logging

# Importar módulos do sistema NF-e
from relatorios.nfe.nfe_api import NFeBuscaAPI
from relatorios.nfe.nfe_busca import NFeBusca
from relatorios.nfe.nfe_processor import NFeProcessor
from relatorios.nfe.nfe_storage import NFeStorage

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Criar app Flask
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# CORS para permitir chamadas do frontend principal
CORS(app, resources={
    r"/api/*": {
        "origins": os.environ.get('FRONTEND_URL', '*'),
        "methods": ["GET", "POST", "PUT", "DELETE"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Configuração do banco
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    logger.error("DATABASE_URL não configurada!")
    sys.exit(1)

# Inicializar módulos NF-e
try:
    nfe_storage = NFeStorage()
    nfe_processor = NFeProcessor()
    nfe_busca = NFeBusca()
    nfe_api = NFeBuscaAPI()
    logger.info("✅ Módulos NF-e inicializados com sucesso")
except Exception as e:
    logger.error(f"❌ Erro ao inicializar módulos NF-e: {e}")
    sys.exit(1)


def get_db_connection():
    """Retorna conexão com banco de dados"""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        logger.error(f"Erro ao conectar ao banco: {e}")
        raise


def get_empresa_id_from_token(token):
    """Extrai empresa_id do token JWT (simplificado)"""
    # TODO: Implementar validação JWT completa
    # Por enquanto, retorna empresa_id do header ou default
    empresa_id = request.headers.get('X-Empresa-ID', '1')
    return int(empresa_id)


def require_auth(f):
    """Decorator simples de autenticação"""
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token não fornecido'}), 401
        
        # TODO: Validar token JWT
        # Por enquanto, aceita qualquer token
        return f(*args, **kwargs)
    
    decorated_function.__name__ = f.__name__
    return decorated_function


# ==================== ROTAS DA API ====================

@app.route('/health')
def health_check():
    """Health check para Railway"""
    try:
        conn = get_db_connection()
        conn.close()
        return jsonify({
            'status': 'healthy',
            'service': 'busca-nfe-cte',
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 503


@app.route('/')
def index():
    """Página inicial"""
    return render_template('nfe_dashboard.html')


@app.route('/api/certificados', methods=['GET'])
@require_auth
def listar_certificados():
    """Lista certificados digitais da empresa"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT 
                id, empresa_id, tipo_certificado, cnpj, 
                razao_social, valido_ate, ativo, criado_em,
                (valido_ate > NOW()) as valido
            FROM certificados_digitais
            WHERE empresa_id = %s
            ORDER BY ativo DESC, valido_ate DESC
        """, (empresa_id,))
        
        certificados = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'certificados': certificados
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar certificados: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/certificados/novo', methods=['POST'])
@require_auth
def criar_certificado():
    """Cadastra novo certificado digital"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        # Validar arquivo PFX
        if 'arquivo_pfx' not in request.files:
            return jsonify({'error': 'Arquivo PFX não enviado'}), 400
        
        arquivo = request.files['arquivo_pfx']
        senha = request.form.get('senha')
        tipo = request.form.get('tipo', 'A1')
        
        if not senha:
            return jsonify({'error': 'Senha não fornecida'}), 400
        
        # Processar certificado
        resultado = nfe_api.cadastrar_certificado(
            empresa_id=empresa_id,
            arquivo_pfx=arquivo.read(),
            senha=senha,
            tipo_certificado=tipo
        )
        
        if resultado.get('success'):
            return jsonify(resultado), 201
        else:
            return jsonify(resultado), 400
            
    except Exception as e:
        logger.error(f"Erro ao criar certificado: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/certificados/<int:certificado_id>/desativar', methods=['POST'])
@require_auth
def desativar_certificado(certificado_id):
    """Desativa certificado digital"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE certificados_digitais
            SET ativo = false, atualizado_em = NOW()
            WHERE id = %s AND empresa_id = %s
        """, (certificado_id, empresa_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Certificado desativado'})
        
    except Exception as e:
        logger.error(f"Erro ao desativar certificado: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/buscar-documentos', methods=['POST'])
@require_auth
def buscar_documentos():
    """Busca automática de documentos na SEFAZ"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        certificado_id = data.get('certificado_id')
        tipo_ambiente = data.get('ambiente', 'homologacao')
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        if not certificado_id:
            return jsonify({'error': 'Certificado não especificado'}), 400
        
        # Executar busca
        resultado = nfe_api.buscar_documentos_automatico(
            empresa_id=empresa_id,
            certificado_id=certificado_id,
            tipo_ambiente=tipo_ambiente,
            data_inicio=data_inicio,
            data_fim=data_fim
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao buscar documentos: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/consultar-chave', methods=['POST'])
@require_auth
def consultar_por_chave():
    """Consulta documento por chave de acesso"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        chave_acesso = data.get('chave_acesso')
        certificado_id = data.get('certificado_id')
        tipo_ambiente = data.get('ambiente', 'homologacao')
        
        if not chave_acesso or not certificado_id:
            return jsonify({'error': 'Chave de acesso e certificado são obrigatórios'}), 400
        
        # Consultar chave
        resultado = nfe_api.consultar_por_chave(
            empresa_id=empresa_id,
            certificado_id=certificado_id,
            chave_acesso=chave_acesso,
            tipo_ambiente=tipo_ambiente
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao consultar chave: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/documentos', methods=['GET'])
@require_auth
def listar_documentos():
    """Lista documentos fiscais com paginação"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        # Parâmetros de paginação e filtros
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 50))
        tipo_doc = request.args.get('tipo')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        offset = (pagina - 1) * por_pagina
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Query base
        where_clauses = ['empresa_id = %s']
        params = [empresa_id]
        
        if tipo_doc:
            where_clauses.append('tipo_documento = %s')
            params.append(tipo_doc)
        
        if data_inicio:
            where_clauses.append('data_emissao >= %s')
            params.append(data_inicio)
        
        if data_fim:
            where_clauses.append('data_emissao <= %s')
            params.append(data_fim)
        
        where_sql = ' AND '.join(where_clauses)
        
        # Contar total
        cursor.execute(f"""
            SELECT COUNT(*) as total
            FROM documentos_fiscais_log
            WHERE {where_sql}
        """, params)
        total = cursor.fetchone()['total']
        
        # Buscar documentos
        cursor.execute(f"""
            SELECT *
            FROM documentos_fiscais_log
            WHERE {where_sql}
            ORDER BY data_emissao DESC, criado_em DESC
            LIMIT %s OFFSET %s
        """, params + [por_pagina, offset])
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'documentos': documentos,
            'paginacao': {
                'pagina': pagina,
                'por_pagina': por_pagina,
                'total': total,
                'total_paginas': (total + por_pagina - 1) // por_pagina
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar documentos: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/documento/<int:doc_id>', methods=['GET'])
@require_auth
def detalhes_documento(doc_id):
    """Retorna detalhes de um documento fiscal"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT *
            FROM documentos_fiscais_log
            WHERE id = %s AND empresa_id = %s
        """, (doc_id, empresa_id))
        
        documento = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not documento:
            return jsonify({'error': 'Documento não encontrado'}), 404
        
        return jsonify({
            'success': True,
            'documento': documento
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar documento: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/documento/<int:doc_id>/xml', methods=['GET'])
@require_auth
def download_xml(doc_id):
    """Download do XML do documento"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT chave_acesso, numero, tipo_documento
            FROM documentos_fiscais_log
            WHERE id = %s AND empresa_id = %s
        """, (doc_id, empresa_id))
        
        documento = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not documento:
            return jsonify({'error': 'Documento não encontrado'}), 404
        
        # Buscar XML do storage
        tipo_doc = 'nfe' if documento['tipo_documento'] == 'NF-e' else 'cte'
        xml_path = nfe_storage.obter_caminho_xml(
            empresa_id,
            tipo_doc,
            documento['chave_acesso']
        )
        
        if not os.path.exists(xml_path):
            return jsonify({'error': 'XML não encontrado no storage'}), 404
        
        return send_file(
            xml_path,
            mimetype='application/xml',
            as_attachment=True,
            download_name=f"{documento['tipo_documento']}_{documento['numero']}.xml"
        )
        
    except Exception as e:
        logger.error(f"Erro ao baixar XML: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/estatisticas', methods=['GET'])
@require_auth
def estatisticas():
    """Retorna estatísticas dos documentos"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT 
                COUNT(*) as total_documentos,
                COUNT(CASE WHEN tipo_documento = 'NF-e' THEN 1 END) as total_nfes,
                COUNT(CASE WHEN tipo_documento = 'CT-e' THEN 1 END) as total_ctes,
                COALESCE(SUM(valor_total), 0) as valor_total,
                COUNT(DISTINCT DATE(data_emissao)) as dias_com_documentos
            FROM documentos_fiscais_log
            WHERE empresa_id = %s
        """, (empresa_id,))
        
        stats = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'estatisticas': stats
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar estatísticas: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nsu-status', methods=['GET'])
@require_auth
def nsu_status():
    """Retorna status dos NSUs pendentes"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT 
                c.id,
                c.cnpj,
                c.razao_social,
                c.ultimo_nsu,
                c.proximo_nsu,
                COUNT(d.id) as documentos_processados,
                MAX(d.criado_em) as ultima_busca
            FROM certificados_digitais c
            LEFT JOIN documentos_fiscais_log d ON d.certificado_id = c.id
            WHERE c.empresa_id = %s AND c.ativo = true
            GROUP BY c.id
            ORDER BY c.razao_social
        """, (empresa_id,))
        
        status = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'status': status
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar status NSU: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/exportar-excel', methods=['POST'])
@require_auth
def exportar_excel():
    """Exporta documentos para Excel"""
    try:
        empresa_id = get_empresa_id_from_token(request.headers.get('Authorization'))
        data = request.json
        
        tipo_doc = data.get('tipo')
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        # Buscar documentos
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        where_clauses = ['empresa_id = %s']
        params = [empresa_id]
        
        if tipo_doc:
            where_clauses.append('tipo_documento = %s')
            params.append(tipo_doc)
        
        if data_inicio:
            where_clauses.append('data_emissao >= %s')
            params.append(data_inicio)
        
        if data_fim:
            where_clauses.append('data_emissao <= %s')
            params.append(data_fim)
        
        where_sql = ' AND '.join(where_clauses)
        
        cursor.execute(f"""
            SELECT 
                tipo_documento, numero, serie, chave_acesso,
                data_emissao, cnpj_emitente, nome_emitente,
                cnpj_destinatario, nome_destinatario,
                valor_total, status_sefaz, criado_em
            FROM documentos_fiscais_log
            WHERE {where_sql}
            ORDER BY data_emissao DESC
        """, params)
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not documentos:
            return jsonify({'error': 'Nenhum documento encontrado'}), 404
        
        # Gerar Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Documentos Fiscais"
        
        # Cabeçalhos
        headers = [
            'Tipo', 'Número', 'Série', 'Chave de Acesso', 'Data Emissão',
            'CNPJ Emitente', 'Nome Emitente', 'CNPJ Destinatário',
            'Nome Destinatário', 'Valor Total', 'Status SEFAZ', 'Data Importação'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for row, doc in enumerate(documentos, 2):
            ws.cell(row=row, column=1, value=doc['tipo_documento'])
            ws.cell(row=row, column=2, value=doc['numero'])
            ws.cell(row=row, column=3, value=doc['serie'])
            ws.cell(row=row, column=4, value=doc['chave_acesso'])
            ws.cell(row=row, column=5, value=doc['data_emissao'].strftime('%d/%m/%Y') if doc['data_emissao'] else '')
            ws.cell(row=row, column=6, value=doc['cnpj_emitente'])
            ws.cell(row=row, column=7, value=doc['nome_emitente'])
            ws.cell(row=row, column=8, value=doc['cnpj_destinatario'])
            ws.cell(row=row, column=9, value=doc['nome_destinatario'])
            ws.cell(row=row, column=10, value=doc['valor_total'])
            ws.cell(row=row, column=11, value=doc['status_sefaz'])
            ws.cell(row=row, column=12, value=doc['criado_em'].strftime('%d/%m/%Y %H:%M') if doc['criado_em'] else '')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['I'].width = 40
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"documentos_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== INICIALIZAÇÃO ====================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    logger.info(f"🚀 Iniciando serviço Busca NF-e/CT-e na porta {port}")
    logger.info(f"📊 Ambiente: {'PRODUÇÃO' if not debug else 'DESENVOLVIMENTO'}")
    
    app.run(
        host='0.0.0.0',
        port=port,
        debug=debug
    )
