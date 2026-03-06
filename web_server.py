"""
Servidor Web para o Sistema Financeiro
Otimizado para PostgreSQL com pool de conex�es

Deploy: 2026-02-15 16:40 - Fix campos associacao e numero_documento no GET
"""
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, session, redirect, url_for
from flask_cors import CORS
from functools import wraps
import os
import sys

# ============================================================================
# LOGGING E MONITORAMENTO
# ============================================================================
from logger_config import setup_logging, get_logger, log_request, log_error
from sentry_config import init_sentry, set_user_context, clear_user_context, add_breadcrumb, capture_exception

# ============================================================================
# CSRF PROTECTION
# ============================================================================
from csrf_config import init_csrf, csrf, inject_csrf_token, register_csrf_error_handlers

# ============================================================================
# MOBILE DETECTION
# ============================================================================
from mobile_config import is_mobile_device, get_device_info

# Configurar logging estruturado
logger = setup_logging(
    app_name='sistema_financeiro',
    log_level=os.getenv('LOG_LEVEL', 'INFO'),
    enable_json=bool(os.getenv('RAILWAY_ENVIRONMENT'))  # JSON em produ��o
)

# Inicializar Sentry em produ��o
SENTRY_ENABLED = init_sentry(
    environment='production' if os.getenv('RAILWAY_ENVIRONMENT') else 'development',
    traces_sample_rate=0.1  # 10% das transa��es
)

logger.info("="*80)
logger.info("Sistema de logging e monitoramento inicializado")
logger.info(f"Sentry: {'? Ativo' if SENTRY_ENABLED else '??  Desabilitado'}")
logger.info("="*80)

# Importa��o opcional do flask-limiter (para compatibilidade durante deploy)
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    LIMITER_AVAILABLE = True
    print("? Flask-Limiter carregado")
except ImportError:
    LIMITER_AVAILABLE = False
    print("?? Flask-Limiter n�o dispon�vel - Rate limiting desabilitado")

# ============================================================================
# IMPORTA��ES DO BANCO DE DADOS - APENAS POSTGRESQL
# ============================================================================
# FOR�A REIMPORT DO M�DULO database_postgresql
# Remove do cache para garantir que m�todos novos sejam carregados
if 'database_postgresql' in sys.modules:
    print("?? For�ando reimport de database_postgresql...")
    del sys.modules['database_postgresql']
    
try:
    import database_postgresql as database
    import database_postgresql as auth_db
    from database_postgresql import DatabaseManager, get_db_connection
    from database_postgresql import pagar_lancamento as db_pagar_lancamento
    from database_postgresql import cancelar_lancamento as db_cancelar_lancamento
    from database_postgresql import obter_lancamento as db_obter_lancamento
    from database_postgresql import atualizar_cliente, atualizar_fornecedor
    print("? M�dulo PostgreSQL carregado com sucesso")
except Exception as e:
    print(f"? ERRO CR�TICO: N�o foi poss�vel carregar o m�dulo PostgreSQL")
    print(f"   Erro: {e}")
    print(f"   Certifique-se de que DATABASE_URL est� configurado")
    raise

from auth_middleware import require_auth, require_admin, require_permission, get_usuario_logado, filtrar_por_cliente, aplicar_filtro_cliente
from database_postgresql import ContaBancaria, Lancamento, Categoria, TipoLancamento, StatusLancamento
from decimal import Decimal
from datetime import datetime, date, timedelta
import json
import os
import secrets
import time
import psycopg2
import psycopg2.extras

# ============================================================================
# VALIDA��O DE DOCUMENTOS
# ============================================================================
# IMPORTS COMENTADOS - movidos para dentro das fun��es espec�ficas
# from cpf_validator import CPFValidator
# from cpf_corrector import CPFCorrector

# ============================================================================
# UTILIT�RIOS COMPARTILHADOS (FASE 4)
# ============================================================================
from app.utils import (
    parse_date,
    format_date_br,
    format_date_iso,
    get_current_date_br,
    get_current_date_filename,
    format_currency,
    parse_currency
)

app = Flask(__name__, static_folder='static', template_folder='templates')

# ============================================================================
# AUTO-EXECUTAR MIGRATION DE EVENTOS (STARTUP)
# ============================================================================
def auto_execute_migrations():
    """Executa migrations automaticamente no startup - DESABILITADA TEMPORARIAMENTE"""
    logger.info("?? auto_execute_migrations desabilitada (causa timeout no deploy)")
    return  # DESABILITADO - causava timeout no Railway
    
    # C�digo comentado abaixo para refer�ncia futura
    try:
        logger.info("="*80)
        logger.info("?? AUTO-EXECUTANDO MIGRATIONS DE EVENTOS")
        logger.info("="*80)
        
        # Verificar se tabelas j� existem
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*) 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name IN ('funcoes_evento', 'evento_funcionarios')
        """)
        
        _row = cursor.fetchone()
        count = _row['count'] if isinstance(_row, dict) else _row[0]
        
        if count == 2:
            logger.info("? Tabelas j� existem. Verificando colunas adicionais...")
            
            # Adicionar colunas de hor�rio se n�o existirem
            try:
                cursor.execute("""
                    ALTER TABLE evento_funcionarios 
                    ADD COLUMN IF NOT EXISTS hora_inicio TIME
                """)
                conn.commit()
                logger.info("? Coluna hora_inicio adicionada/verificada em evento_funcionarios")
            except Exception as e:
                logger.warning(f"?? Erro ao adicionar coluna hora_inicio: {e}")
                conn.rollback()
            
            try:
                cursor.execute("""
                    ALTER TABLE evento_funcionarios 
                    ADD COLUMN IF NOT EXISTS hora_fim TIME
                """)
                conn.commit()
                logger.info("? Coluna hora_fim adicionada/verificada em evento_funcionarios")
            except Exception as e:
                logger.warning(f"?? Erro ao adicionar coluna hora_fim: {e}")
                conn.rollback()
            
            cursor.close()
            return
        
        logger.info(f"?? Encontradas {count}/2 tabelas. Executando migration...")
        
        # Ler e executar SQL
        sql_file = os.path.join(os.path.dirname(__file__), 'migration_evento_funcionarios.sql')
        
        if not os.path.exists(sql_file):
            logger.error(f"? Arquivo SQL n�o encontrado: {sql_file}")
            return
        
        with open(sql_file, 'r', encoding='utf-8') as f:
            sql_content = f.read()
        
        logger.info("?? Executando SQL...")
        cursor.execute(sql_content)
        conn.commit()
        logger.info("? SQL executado e commitado")
        
        # Verificar cria��o
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name IN ('funcoes_evento', 'evento_funcionarios')
            ORDER BY table_name
        """)
        
        tables = cursor.fetchall()
        logger.info(f"? {len(tables)} TABELAS CRIADAS")
        
        # Contar fun��es
        cursor.execute("SELECT COUNT(*) as total FROM funcoes_evento")
        result = cursor.fetchone()
        count_funcoes = result['total'] if isinstance(result, dict) else result[0]
        logger.info(f"? {count_funcoes} FUN��ES INSERIDAS")
        
        # Adicionar colunas de hor�rio se n�o existirem
        try:
            cursor.execute("""
                ALTER TABLE evento_funcionarios 
                ADD COLUMN IF NOT EXISTS hora_inicio TIME
            """)
            conn.commit()
            logger.info("? Coluna hora_inicio adicionada/verificada em evento_funcionarios")
        except Exception as e:
            logger.warning(f"?? Erro ao adicionar coluna hora_inicio: {e}")
            conn.rollback()
        
        try:
            cursor.execute("""
                ALTER TABLE evento_funcionarios 
                ADD COLUMN IF NOT EXISTS hora_fim TIME
            """)
            conn.commit()
            logger.info("? Coluna hora_fim adicionada/verificada em evento_funcionarios")
        except Exception as e:
            logger.warning(f"?? Erro ao adicionar coluna hora_fim: {e}")
            conn.rollback()
        
        # Colunas para controle de horas nas sess�es (finalizar_sessao)
        for col_def in [
            ("horas_trabalhadas", "DECIMAL(10,2)"),
            ("finalizada_em",     "TIMESTAMP"),
            ("finalizada_por",    "INTEGER"),
        ]:
            col_name, col_type = col_def
            try:
                cursor.execute(f"ALTER TABLE sessoes ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                conn.commit()
                logger.info(f"? Coluna {col_name} adicionada/verificada em sessoes")
            except Exception as e:
                logger.warning(f"?? Erro ao adicionar coluna {col_name} em sessoes: {e}")
                conn.rollback()

        cursor.close()
        
        logger.info("="*80)
        logger.info("? MIGRATION CONCLU�DA!")
        logger.info("="*80)
        
    except Exception as e:
        logger.error(f"? Erro na auto-migration: {e}")
        import traceback
        traceback.print_exc()

# Detectar ambiente de produ��o
IS_PRODUCTION = bool(os.getenv('RAILWAY_ENVIRONMENT'))

# Build timestamp para cache busting (atualizado a cada restart)
BUILD_TIMESTAMP = str(int(time.time()))

# Configurar secret key para sess�es
app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = IS_PRODUCTION  # True em produ��o com HTTPS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Configurar CORS - Em produ��o usa a origem da requisi��o, em dev permite tudo
if IS_PRODUCTION:
    ALLOWED_ORIGINS = ['https://sistema-financeiro-dwm-production.up.railway.app']
else:
    ALLOWED_ORIGINS = ['*']  # Permitir tudo apenas em desenvolvimento local

CORS(app, 
     resources={r"/api/*": {
         "origins": ALLOWED_ORIGINS,
         "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"]
     }}, 
     supports_credentials=True)

# ============================================================================
# AUTO-RENOVA��O DE SESS�O (KEEP-ALIVE)
# ============================================================================
@app.before_request
def renovar_sessao():
    """
    Renova a sess�o automaticamente a cada requisi��o para evitar timeout
    durante uso ativo do sistema. A sess�o � marcada como modificada para
    for�ar o Flask a atualizar o cookie de sess�o.
    
    IMPORTANTE: Verifica 'session_token' que � a chave usada pelo sistema
    de autentica��o (n�o 'user_id' nem 'usuario_id').
    """
    # Verificar se h� token de sess�o ativo (chave correta do sistema)
    if 'session_token' in session:
        session.modified = True  # For�a renova��o do cookie de sess�o
        # O Flask automaticamente atualiza o timestamp da sess�o
        # logger.debug desabilitado para evitar 500+ logs/sec em produ��o
        # logger.debug(f"?? [SESS�O] Renovada automaticamente para token: {session.get('session_token', '')[:20]}...")

# ============================================================================
# INICIALIZAR CSRF PROTECTION
# ============================================================================
csrf_instance = init_csrf(app)
register_csrf_error_handlers(app)

# NOTA: Isen��es CSRF s�o aplicadas via decorador @csrf_instance.exempt
# diretamente nas view functions (n�o na lista de rotas)
# Ver exemplos: /api/auth/login, /api/admin/import/upload

# Injetar CSRF token em todos os templates
@app.context_processor
def inject_csrf():
    return inject_csrf_token()

logger.info("? CSRF Protection configurado")

# ============================================================================
# REGISTRAR BLUEPRINTS (ARQUITETURA MODULAR)
# ============================================================================
from app.routes import register_blueprints
register_blueprints(app)
logger.info("? Blueprints registrados")

# Configurar Rate Limiting (apenas se dispon�vel)
if LIMITER_AVAILABLE:
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=["200 per day", "50 per hour"],
        storage_uri="memory://"
    )
    print("? Rate Limiting ativado")
else:
    # Criar um decorador dummy se limiter n�o estiver dispon�vel
    class DummyLimiter:
        def limit(self, *args, **kwargs):
            def decorator(f):
                return f
            return decorator
        
        def exempt(self, f):
            """Decorador exempt dummy - retorna fun��o sem modifica��o"""
            return f
    limiter = DummyLimiter()
    print("?? Rate Limiting desabilitado (flask-limiter n�o instalado)")

# ============================================================================
# MANIPULADORES DE ERRO GLOBAIS
# ============================================================================

@app.before_request
def log_request_info():
    """Log de todas as requisi��es para debug - DESABILITADO para reduzir polui��o"""
    # Logs comentados - descomentar apenas para debug profundo
    # if request.path.startswith('/api/'):
    #     print(f"\n{'??'*40}")
    #     print(f"?? REQUISI��O: {request.method} {request.path}")
    #     print(f"   Session token: {'Presente' if session.get('session_token') else 'AUSENTE'}")
    #     print(f"   Cookies: {list(request.cookies.keys())}")
    #     print(f"   Headers Authorization: {request.headers.get('Authorization', 'N�o presente')}")
    #     print(f"   CSRF Token no header: {request.headers.get('X-CSRFToken', 'AUSENTE')}")
        
    # Gerar CSRF token automaticamente se n�o existir na sess�o
    from flask_wtf.csrf import generate_csrf
    if '_csrf_token' not in session and request.path.startswith('/api/'):
        generate_csrf()
        # print(f"   ?? CSRF Token gerado automaticamente: {token[:20]}...")
    # else:
    #     print(f"   ?? CSRF Token j� existe na sess�o")
    # print(f"{'??'*40}")

@app.after_request
def add_no_cache_headers(response):
    """For�a navegador a NUNCA cachear HTML, CSS e JS"""
    # Para arquivos est�ticos (JS, CSS), desabilita cache agressivamente
    if request.path.startswith('/static/') or request.path.endswith(('.html', '.js', '.css')):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
        response.headers['X-Content-Type-Options'] = 'nosniff'
    return response

@app.before_request
def log_request_info():
    """Log de todas as requisi��es HTTP para auditoria e detec��o mobile"""
    # Pular verifica��es para rotas de API mobile (j� autenticadas via JWT)
    if request.path.startswith('/api/mobile/'):
        return None
    
    # Obter usu�rio se autenticado
    user_id = session.get('usuario_id')
    proprietario_id = session.get('proprietario_id')
    
    # Log estruturado
    log_request(request, user_id=user_id, proprietario_id=proprietario_id)
    
    # Breadcrumb para Sentry
    if SENTRY_ENABLED:
        add_breadcrumb(
            f"{request.method} {request.path}",
            category='http',
            data={
                'url': request.url,
                'method': request.method,
                'ip': request.remote_addr
            }
        )

@app.errorhandler(404)
def handle_404_error(e):
    """Captura erros 404 e loga detalhes"""
    logger.warning(
        f"404 - Rota n�o encontrada: {request.method} {request.path}",
        extra={'ip': request.remote_addr}
    )
    return jsonify({'error': 'Rota n�o encontrada', 'path': request.path}), 404

@app.errorhandler(500)
def handle_500_error(e):
    """Captura erros 500 e loga detalhes"""
    error_context = {
        'path': request.path,
        'method': request.method,
        'ip': request.remote_addr,
        'user_id': session.get('usuario_id')
    }
    
    # Log local
    logger.error(f"500 - Erro interno: {str(e)}", extra=error_context, exc_info=True)
    
    # Enviar para Sentry
    if SENTRY_ENABLED:
        capture_exception(e, context=error_context)
    
    return jsonify({'error': 'Erro interno do servidor', 'details': str(e)}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    """Captura TODAS as exce��es n�o tratadas"""
    error_context = {
        'path': request.path,
        'method': request.method,
        'ip': request.remote_addr,
        'user_id': session.get('usuario_id'),
        'proprietario_id': session.get('proprietario_id')
    }
    
    # Log local cr�tico
    logger.critical(
        f"Exce��o n�o tratada: {type(e).__name__} - {str(e)}",
        extra=error_context,
        exc_info=True
    )
    
    # Enviar para Sentry com alta prioridade
    if SENTRY_ENABLED:
        capture_exception(e, context=error_context, level='fatal')
    print("="*80)
    print(f"Rota: {request.path}")
    print(f"M�todo: {request.method}")
    print(f"Tipo: {type(e).__name__}")
    print(f"Mensagem: {str(e)}")
    import traceback
    traceback.print_exc()
    print("="*80 + "\n")
    return jsonify({'error': 'Erro interno', 'type': type(e).__name__, 'message': str(e)}), 500

# ============================================================================
# CONFIGURA��O E INICIALIZA��O DO SISTEMA
# ============================================================================

# Flag para controlar execu��o de migrations no startup
# ATEN��O: Desabilitado pois causava timeout no Railway (deploy > 10 min)
EXECUTAR_MIGRATIONS_STARTUP = False

print("\n" + "="*70)
print("?? SISTEMA FINANCEIRO - INICIALIZA��O")
print("="*70)
print(f"?? Banco de Dados: PostgreSQL (Pool de Conex�es)")
print(f"?? DATABASE_URL: {'? Configurado' if os.getenv('DATABASE_URL') else '? N�o configurado'}")
print(f"?? Ambiente: {'Produ��o (Railway)' if os.getenv('RAILWAY_ENVIRONMENT') else 'Desenvolvimento'}")
print(f"?? Migrations no Startup: {'? Ativado' if EXECUTAR_MIGRATIONS_STARTUP else '? Desabilitado'}")
print("="*70 + "\n")

# Inicializar banco de dados com pool de conex�es
try:
    print("?? Inicializando DatabaseManager com pool de conex�es...")
    db = DatabaseManager()
    print("DatabaseManager inicializado com sucesso!")
    print(f"   Pool de conexoes: 2-20 conexoes simultaneas")
    
    # Executar migra��es necess�rias (controlado por flag EXECUTAR_MIGRATIONS_STARTUP)
    if EXECUTAR_MIGRATIONS_STARTUP:
        try:
            print("\n?? Executando migra��o Usu�rio Multi-Empresa...")
            from migration_usuario_multi_empresa import executar_migracao as migrar_usuario_multi_empresa
            if migrar_usuario_multi_empresa(db):
                print("? Sistema Usu�rio Multi-Empresa configurado com sucesso!\n")
            else:
                print("?? Migra��o Usu�rio Multi-Empresa falhou (pode j� estar aplicada)\n")
        except Exception as e:
            print(f"?? Aviso: N�o foi poss�vel executar migra��o usu�rio multi-empresa: {e}")
        
        try:
            print("\n?? Executando migra��o Tipo Saldo Inicial...")
            from migration_tipo_saldo_inicial import executar_migracao as migrar_tipo_saldo
            if migrar_tipo_saldo(db):
                print("? Coluna tipo_saldo_inicial adicionada com sucesso!\n")
            else:
                print("?? Migra��o tipo_saldo_inicial falhou (pode j� estar aplicada)\n")
        except Exception as e:
            print(f"?? Aviso: N�o foi poss�vel executar migra��o tipo_saldo_inicial: {e}")
        
        # ?? AUTO-EXECUTAR MIGRATIONS DE EVENTOS (ap�s db estar pronto)
        try:
            print("\n?? Executando migra��o de Eventos...")
            auto_execute_migrations()
            print("? Migration de eventos verificada!\n")
        except Exception as e:
            print(f"?? Aviso: N�o foi poss�vel executar auto-migration de eventos: {e}")
        
        try:
            print("\n?? Executando migra��o Data de In�cio...")
            from migration_data_inicio import executar_migracao as migrar_data_inicio
            if migrar_data_inicio(db):
                print("? Coluna data_inicio adicionada com sucesso!\n")
            else:
                print("?? Migra��o data_inicio falhou (pode j� estar aplicada)\n")
        except Exception as e:
            print(f"?? Aviso: N�o foi poss�vel executar migra��o data_inicio: {e}")
    else:
        print("?? Migrations de startup desabilitadas (EXECUTAR_MIGRATIONS_STARTUP=False)")
    
    # ?? MIGRATION CR�TICA: Sempre executar (independente de flag)
    try:
        print("\n?? Verificando coluna usa_integracao_folha...")
        from migration_add_usa_integracao_folha import executar_migration
        executar_migration()
    except Exception as e:
        print(f"?? Aviso: {e}")

    # ?? Criar tabela ofx_filtros_memo (Ajuste de OFX)
    try:
        print("\n?? Verificando tabela ofx_filtros_memo...")
        with db.get_connection() as _conn:
            _cur = _conn.cursor()
            _cur.execute("""
                CREATE TABLE IF NOT EXISTS ofx_filtros_memo (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER NOT NULL,
                    conta_bancaria VARCHAR(500) NOT NULL,
                    memo_filtro TEXT NOT NULL,
                    criado_em TIMESTAMP DEFAULT NOW(),
                    UNIQUE (empresa_id, conta_bancaria, memo_filtro)
                )
            """)
            _conn.commit()
            _cur.close()
        print("? Tabela ofx_filtros_memo verificada/criada!")
    except Exception as e:
        print(f"?? Aviso ao criar ofx_filtros_memo: {e}")

    # ?? Criar tabelas M�dulo Fiscal Federal
    try:
        print("\n?? Verificando tabelas do M�dulo Fiscal Federal...")
        with db.get_connection() as _fc:
            _fcur = _fc.cursor()
            _fcur.execute("""
                CREATE TABLE IF NOT EXISTS logs_fiscais (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    tipo_operacao VARCHAR(100),
                    endpoint VARCHAR(300),
                    request JSONB,
                    response JSONB,
                    status_http INTEGER,
                    protocolo VARCHAR(200),
                    data TIMESTAMP DEFAULT NOW()
                );
                CREATE INDEX IF NOT EXISTS idx_logs_fiscais_empresa
                    ON logs_fiscais(empresa_id, data DESC);

                CREATE TABLE IF NOT EXISTS fiscal_cnpj_historico (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    cnpj VARCHAR(14),
                    dados JSONB,
                    consultado_em TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS fiscal_certidoes (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    cnpj VARCHAR(14),
                    tipo VARCHAR(50) DEFAULT 'CND_FEDERAL',
                    numero VARCHAR(100),
                    data_emissao DATE,
                    data_vencimento DATE,
                    pdf_base64 TEXT,
                    status VARCHAR(50) DEFAULT 'emitida',
                    protocolo VARCHAR(200),
                    criado_em TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS fiscal_dctfweb (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    cnpj VARCHAR(14),
                    competencia VARCHAR(6),
                    situacao VARCHAR(100),
                    valor_total NUMERIC(15,2) DEFAULT 0,
                    dados JSONB,
                    consultado_em TIMESTAMP DEFAULT NOW(),
                    UNIQUE(empresa_id, cnpj, competencia)
                );

                CREATE TABLE IF NOT EXISTS fiscal_reinf (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    cnpj VARCHAR(14),
                    competencia VARCHAR(6),
                    evento VARCHAR(20),
                    recibo VARCHAR(200),
                    status VARCHAR(100),
                    dados JSONB,
                    consultado_em TIMESTAMP DEFAULT NOW(),
                    UNIQUE(empresa_id, cnpj, competencia, evento)
                );

                CREATE TABLE IF NOT EXISTS fiscal_darf (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    cnpj VARCHAR(14),
                    codigo_receita VARCHAR(20),
                    competencia VARCHAR(6),
                    valor NUMERIC(15,2),
                    data_vencimento DATE,
                    status VARCHAR(50) DEFAULT 'emitido',
                    pdf_base64 TEXT,
                    protocolo VARCHAR(200),
                    lancamento_id INTEGER,
                    criado_em TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS fiscal_fila (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    tipo VARCHAR(100),
                    parametros JSONB,
                    status VARCHAR(50) DEFAULT 'pendente',
                    tentativas INTEGER DEFAULT 0,
                    resultado JSONB,
                    criado_em TIMESTAMP DEFAULT NOW(),
                    processado_em TIMESTAMP
                );
            """)
            _fc.commit()
            _fcur.close()
        print("? Tabelas do M�dulo Fiscal Federal verificadas/criadas!")
    except Exception as e:
        print(f"?? Aviso ao criar tabelas fiscais: {e}")

    # -- Tabelas EFD-Reinf ----------------------------------------------------
    try:
        with db.get_connection() as _rc:
            _rcur = _rc.cursor()
            _rcur.execute("""
                CREATE TABLE IF NOT EXISTS reinf_eventos (
                    id TEXT PRIMARY KEY,
                    empresa_id INTEGER,
                    competencia VARCHAR(7),
                    evento VARCHAR(10),
                    identificador_evento VARCHAR(100),
                    status VARCHAR(30) DEFAULT 'pendente',
                    protocolo TEXT,
                    recibo TEXT,
                    xml_enviado TEXT,
                    xml_retorno TEXT,
                    erro TEXT,
                    enviado_em TIMESTAMP,
                    versao INTEGER DEFAULT 1,
                    created_at TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS reinf_dados (
                    id TEXT PRIMARY KEY,
                    evento_id TEXT REFERENCES reinf_eventos(id) ON DELETE CASCADE,
                    payload JSONB,
                    created_at TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS reinf_totalizadores (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    competencia VARCHAR(7),
                    total_base NUMERIC(15,2) DEFAULT 0,
                    total_inss NUMERIC(15,2) DEFAULT 0,
                    total_ir NUMERIC(15,2) DEFAULT 0,
                    total_csll NUMERIC(15,2) DEFAULT 0,
                    total_pis NUMERIC(15,2) DEFAULT 0,
                    total_cofins NUMERIC(15,2) DEFAULT 0,
                    status_competencia VARCHAR(20) DEFAULT 'aberta',
                    created_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(empresa_id, competencia)
                );
                CREATE TABLE IF NOT EXISTS reinf_motor_sugestoes (
                    id SERIAL PRIMARY KEY,
                    empresa_id INTEGER,
                    competencia VARCHAR(7),
                    evento_sugerido VARCHAR(10),
                    motivo TEXT,
                    origem VARCHAR(50),
                    aceito BOOLEAN,
                    created_at TIMESTAMP DEFAULT NOW()
                );
            """)
            _rc.commit()
            _rcur.close()
        print("? Tabelas EFD-Reinf verificadas/criadas!")
    except Exception as e:
        print(f"?? Aviso ao criar tabelas REINF: {e}")

    # -- Migração: adicionar colunas novas em nfse_baixadas -------------------
    try:
        with db.get_connection() as _nm:
            _ncur = _nm.cursor()
            _ncur.execute("""
                ALTER TABLE nfse_baixadas
                    ADD COLUMN IF NOT EXISTS tp_ret_issqn VARCHAR(5) DEFAULT NULL,
                    ADD COLUMN IF NOT EXISTS data_pagamento DATE DEFAULT NULL,
                    ADD COLUMN IF NOT EXISTS situacao_recebimento VARCHAR(20) DEFAULT NULL;
            """)
            _nm.commit()
            _ncur.close()
        print("✅ Colunas tp_ret_issqn, data_pagamento, situacao_recebimento verificadas em nfse_baixadas!")
    except Exception as e:
        print(f"⚠️ Aviso ao migrar nfse_baixadas: {e}")

    print("? DatabaseManager pronto!")
    print("="*70 + "\n")
        
except Exception as e:
    print(f"? ERRO CR�TICO ao inicializar DatabaseManager: {e}")
    import traceback
    traceback.print_exc()
    raise

# ============================================================================
# ROTAS DE AUTENTICA��O
# ============================================================================

@app.route('/api/auth/login', methods=['POST'])
@csrf_instance.exempt
@limiter.limit("5 per minute")  # M�ximo 5 tentativas por minuto
def login():
    """Endpoint de login com prote��o contra brute force"""
    try:
        print(f"\n{'='*80}")
        print(f"?? [LOGIN] Iniciando processo de login...")
        print(f"{'='*80}")
        
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        print(f"?? Dados recebidos:")
        print(f"   - username: {username}")
        print(f"   - password: {'***' if password else 'VAZIO'}")
        
        if not username or not password:
            print(f"? Username ou senha vazios")
            return jsonify({
                'success': False,
                'error': 'Username e senha s�o obrigat�rios'
            }), 400
        
        # Verificar se conta est� bloqueada
        print(f"?? Verificando se conta est� bloqueada...")
        from auth_functions import verificar_conta_bloqueada
        if verificar_conta_bloqueada(username, db):
            print(f"?? Conta bloqueada!")
            return jsonify({
                'success': False,
                'error': 'Conta temporariamente bloqueada por excesso de tentativas. Tente novamente em 15 minutos.'
            }), 429
        print(f"? Conta n�o bloqueada")
        
        # Autenticar usu�rio
        print(f"?? Chamando auth_db.autenticar_usuario('{username}', '***')...")
        usuario = auth_db.autenticar_usuario(username, password)
        print(f"?? Resultado autentica��o: {usuario if usuario else 'FALHOU'}")
        
        if not usuario:
            print(f"? Autentica��o falhou!")
            # Registrar tentativa falha
            auth_db.registrar_log_acesso(
                usuario_id=None,
                acao='login_failed',
                descricao=f'Tentativa de login falhou para username: {username}',
                ip_address=request.remote_addr,
                sucesso=False
            )
            print(f"{'='*80}\n")
            return jsonify({
                'success': False,
                'error': 'Usu�rio ou senha inv�lidos'
            }), 401
        
        print(f"? Usu�rio autenticado:")
        print(f"   - id: {usuario.get('id')}")
        print(f"   - username: {usuario.get('username')}")
        print(f"   - tipo: {usuario.get('tipo')}")
        
        # Criar sess�o
        print(f"?? Criando sess�o...")
        token = auth_db.criar_sessao(
            usuario['id'],
            request.remote_addr,
            request.headers.get('User-Agent', '')
        )
        print(f"? Sess�o criada: {token[:20]}...")
        
        # Guardar token e user_id na sess�o do Flask
        session['session_token'] = token
        session['user_id'] = usuario['id']  # ? Necess�rio para rotas que usam session.get('user_id')
        session.permanent = True
        
        # Registrar login bem-sucedido
        auth_db.registrar_log_acesso(
            usuario_id=usuario['id'],
            acao='login',
            descricao='Login realizado com sucesso',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        # ============================================================
        # MULTI-EMPRESA: Carregar empresas do usu�rio
        # ============================================================
        empresas_disponiveis = []
        empresa_selecionada = None
        
        if usuario['tipo'] == 'admin':
            # Super admin tem acesso a todas as empresas
            empresas_disponiveis = database.listar_empresas({})
            # N�o selecionar empresa automaticamente para super admin
        else:
            # Carregar empresas que o usu�rio tem acesso
            from auth_functions import listar_empresas_usuario, obter_empresa_padrao
            empresas_disponiveis = listar_empresas_usuario(usuario['id'], auth_db)
            
            if empresas_disponiveis:
                # Buscar empresa padr�o
                empresa_padrao_id = obter_empresa_padrao(usuario['id'], auth_db)
                
                if empresa_padrao_id:
                    empresa_selecionada = next((e for e in empresas_disponiveis if e.get('empresa_id') == empresa_padrao_id), None)
                else:
                    # Se n�o tem padr�o, selecionar a primeira
                    empresa_selecionada = empresas_disponiveis[0]
                
                if empresa_selecionada:
                    session['empresa_id'] = empresa_selecionada.get('empresa_id')
                    print(f"? Empresa selecionada no login: {empresa_selecionada.get('razao_social')}")
        
        # Obter permiss�es do usu�rio
        if usuario['tipo'] == 'admin':
            permissoes = ['*']  # Super admin tem todas as permiss�es
        elif empresa_selecionada:
            from auth_functions import obter_permissoes_usuario_empresa
            permissoes = obter_permissoes_usuario_empresa(usuario['id'], empresa_selecionada.get('empresa_id'), auth_db)
        else:
            permissoes = auth_db.obter_permissoes_usuario(usuario['id'])
        
        # Preparar resposta
        response_data = {
            'success': True,
            'message': 'Login realizado com sucesso',
            'usuario': {
                'id': usuario['id'],
                'username': usuario['username'],
                'nome_completo': usuario['nome_completo'],
                'tipo': usuario['tipo'],
                'email': usuario['email']
            },
            'permissoes': permissoes,
            'empresas_disponiveis': [{
                'id': e.get('empresa_id'),
                'razao_social': e.get('razao_social'),
                'is_padrao': e.get('is_empresa_padrao', False)
            } for e in empresas_disponiveis] if empresas_disponiveis else []
        }
        
        # Adicionar empresa selecionada se houver
        if empresa_selecionada:
            response_data['empresa_selecionada'] = {
                'id': empresa_selecionada.get('empresa_id'),
                'razao_social': empresa_selecionada.get('razao_social')
            }
        
        # Se usu�rio tem m�ltiplas empresas, indicar que precisa escolher
        if len(empresas_disponiveis) > 1 and usuario['tipo'] != 'admin':
            response_data['require_empresa_selection'] = True
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"? Erro no login: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Erro ao processar login'
        }), 500


@app.route('/api/auth/logout', methods=['POST'])
@require_auth
def logout():
    """Endpoint de logout"""
    try:
        token = session.get('session_token')
        
        if token:
            auth_db.invalidar_sessao(token)
            
            # Registrar logout
            usuario = request.usuario
            auth_db.registrar_log_acesso(
                usuario_id=usuario['id'],
                acao='logout',
                descricao='Logout realizado',
                ip_address=request.remote_addr,
                sucesso=True
            )
        
        session.clear()
        
        return jsonify({
            'success': True,
            'message': 'Logout realizado com sucesso'
        })
        
    except Exception as e:
        print(f"? Erro no logout: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro ao processar logout'
        }), 500


@app.route('/api/auth/verify', methods=['GET'])
def verify_session():
    """Verifica se a sess�o est� v�lida"""
    try:
        usuario = get_usuario_logado()
        
        if not usuario:
            return jsonify({
                'success': False,
                'authenticated': False
            })
        
        # ============================================================
        # MULTI-EMPRESA: Carregar empresa atual e empresas dispon�veis
        # ============================================================
        empresa_atual = None
        empresas_disponiveis = []
        
        if usuario['tipo'] == 'admin':
            # Super admin
            permissoes = ['*']
            empresas_disponiveis = database.listar_empresas({})
            empresa_id = session.get('empresa_id')
            if empresa_id:
                empresa_atual = database.obter_empresa(empresa_id)
        else:
            # Usu�rio normal
            from auth_functions import listar_empresas_usuario
            empresas_disponiveis = listar_empresas_usuario(usuario['id'], auth_db)
            
            empresa_id = session.get('empresa_id')
            
            if empresa_id:
                # Carregar permiss�es espec�ficas da empresa
                from auth_functions import obter_permissoes_usuario_empresa
                permissoes = obter_permissoes_usuario_empresa(usuario['id'], empresa_id, auth_db)
                
                # Buscar dados da empresa atual
                empresa_atual = next((e for e in empresas_disponiveis if e.get('empresa_id') == empresa_id), None)
            else:
                # Sem empresa selecionada
                permissoes = auth_db.obter_permissoes_usuario(usuario['id'])
        
        response = {
            'success': True,
            'authenticated': True,
            'usuario': {
                'id': usuario['id'],
                'username': usuario['username'],
                'nome_completo': usuario['nome_completo'],
                'tipo': usuario['tipo'],
                'email': usuario['email'],
                'permissoes': permissoes
            },
            'permissoes': permissoes,
            'empresas_disponiveis': [{
                'id': e.get('empresa_id') if usuario['tipo'] != 'admin' else e.get('id'),
                'razao_social': e.get('razao_social'),
                'is_padrao': e.get('is_empresa_padrao', False)
            } for e in empresas_disponiveis] if empresas_disponiveis else []
        }
        
        # Adicionar empresa atual se houver
        if empresa_atual:
            response['empresa_atual'] = {
                'id': empresa_atual.get('empresa_id') if usuario['tipo'] != 'admin' else empresa_atual.get('id'),
                'razao_social': empresa_atual.get('razao_social')
            }
        
        return jsonify(response)
        
    except Exception as e:
        print(f"\n? ERRO ao verificar sess�o:")
        print(f"   Tipo: {type(e).__name__}")
        print(f"   Mensagem: {e}")
        import traceback
        traceback.print_exc()
        print(f"{'='*80}\n")
        return jsonify({
            'success': False,
            'error': 'Erro ao verificar sess�o'
        }), 500


@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def change_password():
    """Alterar senha do usu�rio logado"""
    try:
        data = request.json
        senha_atual = data.get('senha_atual')
        senha_nova = data.get('senha_nova')
        
        if not senha_atual or not senha_nova:
            return jsonify({
                'success': False,
                'error': 'Senha atual e nova senha s�o obrigat�rias'
            }), 400
        
        # Validar for�a da nova senha
        from auth_functions import validar_senha_forte
        valida, mensagem = validar_senha_forte(senha_nova)
        if not valida:
            return jsonify({
                'success': False,
                'error': f'Nova senha fraca: {mensagem}'
            }), 400
        
        usuario = request.usuario
        
        # Verificar senha atual
        usuario_verificado = auth_db.autenticar_usuario(
            usuario['username'],
            senha_atual
        )
        
        if not usuario_verificado:
            return jsonify({
                'success': False,
                'error': 'Senha atual incorreta'
            }), 401
        
        # Atualizar senha
        auth_db.atualizar_usuario(
            usuario['id'],
            {'password': senha_nova}
        )
        
        # Registrar altera��o
        auth_db.registrar_log_acesso(
            usuario_id=usuario['id'],
            acao='change_password',
            descricao='Senha alterada com sucesso',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        return jsonify({
            'success': True,
            'message': 'Senha alterada com sucesso'
        })
        
    except Exception as e:
        print(f"? Erro ao alterar senha: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro ao alterar senha'
        }), 500


# ===================================================================
# ROTAS DE GEST�O MULTI-EMPRESA (Usu�rio com Acesso a M�ltiplas Empresas)
# ===================================================================

@app.route('/api/auth/minhas-empresas', methods=['GET'])
@require_auth
def minhas_empresas():
    """Lista todas as empresas que o usu�rio tem acesso"""
    try:
        usuario = request.usuario
        
        # Super admin tem acesso a todas as empresas
        if usuario['tipo'] == 'admin':
            empresas = database.listar_empresas({})
            return jsonify({
                'success': True,
                'empresas': [{
                    'id': e['id'],
                    'razao_social': e['razao_social'],
                    'cnpj': e.get('cnpj'),
                    'papel': 'admin',
                    'is_padrao': False,
                    'permissoes': ['*']  # Todas as permiss�es
                } for e in empresas]
            })
        
        # Usu�rios normais: buscar empresas vinculadas
        from auth_functions import listar_empresas_usuario
        empresas = listar_empresas_usuario(usuario['id'], auth_db)
        
        if not empresas:
            return jsonify({
                'success': True,
                'empresas': [],
                'message': 'Usu�rio n�o est� vinculado a nenhuma empresa'
            })
        
        return jsonify({
            'success': True,
            'empresas': empresas
        })
        
    except Exception as e:
        print(f"? Erro ao listar empresas do usu�rio: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auth/switch-empresa', methods=['POST'])
@require_auth
def switch_empresa():
    """Troca a empresa atual do usu�rio na sess�o"""
    print(f"\n{'='*80}")
    print(f"?? [SWITCH-EMPRESA] Requisi��o recebida")
    try:
        data = request.json
        print(f"?? Dados recebidos: {data}")
        empresa_id = data.get('empresa_id')
        print(f"?? Empresa ID: {empresa_id}")
        
        if not empresa_id:
            print(f"? empresa_id n�o fornecido")
            return jsonify({
                'success': False,
                'error': 'empresa_id � obrigat�rio'
            }), 400
        
        usuario = request.usuario
        print(f"?? Usu�rio: {usuario['username']} (tipo: {usuario['tipo']})")
        
        # Super admin pode acessar qualquer empresa
        if usuario['tipo'] != 'admin':
            # Validar se usu�rio tem acesso � empresa
            from auth_functions import tem_acesso_empresa
            print(f"?? Validando acesso do usu�rio � empresa...")
            if not tem_acesso_empresa(usuario['id'], empresa_id, auth_db):
                print(f"? Acesso negado")
                return jsonify({
                    'success': False,
                    'error': 'Acesso negado a esta empresa'
                }), 403
            print(f"? Acesso validado")
        else:
            print(f"?? Admin - acesso total")
        
        # Buscar dados da empresa
        empresa = database.obter_empresa(empresa_id)
        if not empresa:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o encontrada'
            }), 404
        
        # Atualizar sess�o com nova empresa
        session['empresa_id'] = empresa_id
        session.modified = True
        
        # Registrar troca de empresa
        auth_db.registrar_log_acesso(
            usuario_id=usuario['id'],
            acao='switch_empresa',
            descricao=f'Trocou para empresa: {empresa["razao_social"]}',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        # Carregar permiss�es da nova empresa
        print(f"?? Carregando permiss�es...")
        if usuario['tipo'] != 'admin':
            from auth_functions import obter_permissoes_usuario_empresa
            permissoes = obter_permissoes_usuario_empresa(usuario['id'], empresa_id, auth_db)
        else:
            permissoes = ['*']  # Super admin tem todas as permiss�es
        print(f"?? Permiss�es carregadas: {len(permissoes)}")
        
        print(f"? Troca de empresa conclu�da com sucesso")
        print(f"{'='*80}\n")
        return jsonify({
            'success': True,
            'message': 'Empresa alterada com sucesso',
            'empresa': {
                'id': empresa['id'],
                'razao_social': empresa['razao_social'],
                'cnpj': empresa.get('cnpj')
            },
            'permissoes': permissoes
        })
        
    except Exception as e:
        print(f"? ERRO em switch-empresa: {e}")
        print(f"? Tipo do erro: {type(e)}")
        import traceback
        traceback.print_exc()
        print(f"{'='*80}\n")
        return jsonify({
            'success': False,
            'error': 'Erro ao trocar empresa'
        }), 500


@app.route('/api/auth/empresa-padrao', methods=['PUT'])
@require_auth
def definir_empresa_padrao():
    """Define a empresa padr�o do usu�rio (selecionada automaticamente no login)"""
    try:
        data = request.json
        empresa_id = data.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'empresa_id � obrigat�rio'
            }), 400
        
        usuario = request.usuario
        
        # Super admin n�o precisa de empresa padr�o
        if usuario['tipo'] == 'admin':
            return jsonify({
                'success': False,
                'error': 'Super admin n�o precisa de empresa padr�o'
            }), 400
        
        # Validar acesso � empresa
        from auth_functions import tem_acesso_empresa, atualizar_usuario_empresa
        if not tem_acesso_empresa(usuario['id'], empresa_id, auth_db):
            return jsonify({
                'success': False,
                'error': 'Acesso negado a esta empresa'
            }), 403
        
        # Atualizar empresa padr�o
        sucesso = atualizar_usuario_empresa(
            usuario['id'], 
            empresa_id,
            is_padrao=True,
            db=auth_db
        )
        
        if not sucesso:
            return jsonify({
                'success': False,
                'error': 'Erro ao definir empresa padr�o'
            }), 500
        
        return jsonify({
            'success': True,
            'message': 'Empresa padr�o definida com sucesso'
        })
        
    except Exception as e:
        print(f"? Erro ao definir empresa padr�o: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/usuario-empresas', methods=['POST'])
@require_admin
def vincular_usuario_empresa_admin():
    """Vincula um usu�rio a uma empresa (apenas admin)"""
    try:
        data = request.json
        usuario_id = data.get('usuario_id')
        empresa_id = data.get('empresa_id')
        papel = data.get('papel', 'usuario')
        permissoes = data.get('permissoes', [])
        is_padrao = data.get('is_padrao', False)
        
        if not usuario_id or not empresa_id:
            return jsonify({
                'success': False,
                'error': 'usuario_id e empresa_id s�o obrigat�rios'
            }), 400
        
        if papel not in ['admin_empresa', 'usuario', 'visualizador']:
            return jsonify({
                'success': False,
                'error': 'Papel inv�lido. Use: admin_empresa, usuario ou visualizador'
            }), 400
        
        admin = request.usuario
        
        # Vincular usu�rio � empresa
        from auth_functions import vincular_usuario_empresa
        vinculo_id = vincular_usuario_empresa(
            usuario_id=usuario_id,
            empresa_id=empresa_id,
            papel=papel,
            permissoes=permissoes,
            is_padrao=is_padrao,
            criado_por=admin['id'],
            db=auth_db
        )
        
        # Registrar a��o
        auth_db.registrar_log_acesso(
            usuario_id=admin['id'],
            acao='vincular_usuario_empresa',
            descricao=f'Vinculou usu�rio {usuario_id} � empresa {empresa_id}',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        return jsonify({
            'success': True,
            'message': 'Usu�rio vinculado � empresa com sucesso',
            'id': vinculo_id
        }), 201
        
    except Exception as e:
        logger.error(f"Erro ao vincular usu�rio � empresa: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/usuario-empresas/<int:usuario_id>/<int:empresa_id>', methods=['PUT'])
@require_admin
def atualizar_usuario_empresa_admin(usuario_id: int, empresa_id: int):
    """Atualiza o v�nculo de um usu�rio com uma empresa (apenas admin)"""
    try:
        data = request.json
        papel = data.get('papel')
        permissoes = data.get('permissoes')
        is_padrao = data.get('is_padrao')
        
        from auth_functions import atualizar_usuario_empresa
        sucesso = atualizar_usuario_empresa(
            usuario_id=usuario_id,
            empresa_id=empresa_id,
            papel=papel,
            permissoes=permissoes,
            is_padrao=is_padrao,
            db=auth_db
        )
        
        if not sucesso:
            return jsonify({
                'success': False,
                'error': 'Erro ao atualizar v�nculo'
            }), 500
        
        # Registrar a��o
        admin = request.usuario
        auth_db.registrar_log_acesso(
            usuario_id=admin['id'],
            acao='atualizar_usuario_empresa',
            descricao=f'Atualizou v�nculo do usu�rio {usuario_id} com empresa {empresa_id}',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        return jsonify({
            'success': True,
            'message': 'V�nculo atualizado com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao atualizar v�nculo: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/usuario-empresas/<int:usuario_id>/<int:empresa_id>', methods=['DELETE'])
@require_admin
def remover_usuario_empresa_admin(usuario_id: int, empresa_id: int):
    """Remove o v�nculo de um usu�rio com uma empresa (apenas admin)"""
    try:
        from auth_functions import remover_usuario_empresa
        sucesso = remover_usuario_empresa(usuario_id, empresa_id, auth_db)
        
        if not sucesso:
            return jsonify({
                'success': False,
                'error': 'Erro ao remover v�nculo'
            }), 500
        
        # Registrar a��o
        admin = request.usuario
        auth_db.registrar_log_acesso(
            usuario_id=admin['id'],
            acao='remover_usuario_empresa',
            descricao=f'Removeu v�nculo do usu�rio {usuario_id} com empresa {empresa_id}',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        return jsonify({
            'success': True,
            'message': 'V�nculo removido com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao remover v�nculo: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/usuarios/<int:usuario_id>/empresas', methods=['GET'])
@require_admin
def listar_empresas_do_usuario_admin(usuario_id: int):
    """Lista todas as empresas que um usu�rio tem acesso (apenas admin)"""
    try:
        from auth_functions import listar_empresas_usuario
        empresas = listar_empresas_usuario(usuario_id, auth_db)
        
        return jsonify({
            'success': True,
            'empresas': empresas
        })
        
    except Exception as e:
        print(f"? Erro ao listar empresas do usu�rio: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== FIM DAS ROTAS MULTI-EMPRESA =====

# ===== ROTAS DE GERENCIAMENTO DE USU�RIOS (APENAS ADMIN) =====

@app.route('/api/usuarios', methods=['GET', 'POST'])
@require_admin
def gerenciar_usuarios():
    """Listar ou criar usu�rios"""
    print(f"\n?? [gerenciar_usuarios] FUN��O CHAMADA - M�todo: {request.method}")
    if request.method == 'GET':
        try:
            print(f"\n{'='*80}")
            print(f"?? GET /api/usuarios - Listando usu�rios...")
            print(f"{'='*80}")
            
            # Verificar se usu�rio est� autenticado
            usuario = getattr(request, 'usuario', None)
            if not usuario:
                print(f"   ? Usu�rio n�o autenticado")
                return jsonify({'success': False, 'error': 'N�o autenticado'}), 401
            
            print(f"   ? Usu�rio autenticado: {usuario.get('username')} (tipo: {usuario.get('tipo')})")
            
            # Listar usu�rios
            usuarios = auth_db.listar_usuarios()
            print(f"   ?? Tipo retornado: {type(usuarios)}")
            
            # Garantir que � uma lista
            if not isinstance(usuarios, list):
                print(f"   ?? N�o � lista! Convertendo...")
                if usuarios is None:
                    usuarios = []
                else:
                    usuarios = [usuarios] if isinstance(usuarios, dict) else []
            
            # Converter datas para string (JSON serializable)
            usuarios_serializaveis = []
            for user in usuarios:
                user_dict = dict(user) if not isinstance(user, dict) else user
                
                # Converter datetime para string
                if 'created_at' in user_dict and user_dict['created_at']:
                    user_dict['created_at'] = str(user_dict['created_at'])
                if 'ultima_sessao' in user_dict and user_dict['ultima_sessao']:
                    user_dict['ultima_sessao'] = str(user_dict['ultima_sessao'])
                
                usuarios_serializaveis.append(user_dict)
            
            print(f"   ? Retornando {len(usuarios_serializaveis)} usu�rios")
            print(f"{'='*80}\n")
            
            return jsonify({'success': True, 'usuarios': usuarios_serializaveis})
            
        except Exception as e:
            print(f"? Erro ao listar usu�rios: {e}")
            import traceback
            traceback.print_exc()
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    else:  # POST
        try:
            data = request.json
            admin = request.usuario
            data['created_by'] = admin['id']
            
            print(f"?? Dados recebidos do frontend: {data}")
            print(f"   - empresas_ids: {data.get('empresas_ids')}")
            print(f"   - empresa_id_padrao: {data.get('empresa_id_padrao')}")
            print(f"   - tipo: {data.get('tipo')}")
            
            # Validar empresas
            empresas_ids = data.get('empresas_ids', [])
            if not empresas_ids or len(empresas_ids) == 0:
                return jsonify({
                    'success': False,
                    'error': 'Selecione ao menos uma empresa'
                }), 400
            
            # Validar for�a da senha
            from auth_functions import validar_senha_forte
            if 'password' in data:
                valida, mensagem = validar_senha_forte(data['password'])
                if not valida:
                    return jsonify({
                        'success': False,
                        'error': f'Senha fraca: {mensagem}'
                    }), 400
            
            # ?? MULTI-EMPRESA: Usar primeira empresa para cria��o (compatibilidade)
            data['empresa_id'] = empresas_ids[0]
            
            print(f"?? Dados para criar_usuario: {data}")
            usuario_id = auth_db.criar_usuario(data)
            print(f"? Usu�rio criado com ID: {usuario_id}")
            
            # ?? MULTI-EMPRESA: Criar v�nculos na tabela usuario_empresas
            from auth_functions import vincular_usuario_empresa
            empresa_id_padrao = data.get('empresa_id_padrao')
            
            for empresa_id in empresas_ids:
                is_padrao = (empresa_id == empresa_id_padrao)
                
                print(f"?? Vinculando usu�rio {usuario_id} � empresa {empresa_id} (padr�o: {is_padrao})")
                
                vincular_usuario_empresa(
                    usuario_id=usuario_id,
                    empresa_id=empresa_id,
                    papel='usuario',  # Papel padr�o
                    permissoes=data.get('permissoes', []),
                    is_padrao=is_padrao,
                    criado_por=admin['id'],
                    db=auth_db
                )
            
            # Conceder permiss�es globais se fornecidas (legado)
            if 'permissoes' in data:
                print(f"?? Concedendo {len(data['permissoes'])} permiss�es")
                auth_db.sincronizar_permissoes_usuario(
                    usuario_id,
                    data['permissoes'],
                    admin['id']
                )
            
            # Registrar cria��o
            auth_db.registrar_log_acesso(
                usuario_id=admin['id'],
                acao='create_user',
                descricao=f'Usu�rio criado: {data["username"]} com {len(empresas_ids)} empresa(s)',
                ip_address=request.remote_addr,
                sucesso=True
            )
            
            return jsonify({
                'success': True,
                'message': 'Usu�rio criado com sucesso',
                'id': usuario_id
            }), 201
            
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)}), 400
        except Exception as e:
            print(f"? Erro ao criar usu�rio: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/usuarios/<int:usuario_id>', methods=['GET', 'PUT', 'DELETE'])
@require_admin
def gerenciar_usuario_especifico(usuario_id):
    """Obter, atualizar ou deletar usu�rio espec�fico"""
    print(f"\n?? [gerenciar_usuario_especifico] FUN��O CHAMADA - ID: {usuario_id}, M�todo: {request.method}")
    if request.method == 'GET':
        try:
            print(f"\n{'='*80}")
            print(f"   ?? GET /api/usuarios/{usuario_id}")
            print(f"{'='*80}")
            print(f"   ?? Buscando usu�rio ID {usuario_id}...")
            
            usuario = auth_db.obter_usuario(usuario_id)
            print(f"   ?? Tipo do resultado: {type(usuario)}")
            print(f"   ?? Resultado: {usuario if usuario else 'N�O ENCONTRADO'}")
            
            if not usuario:
                print(f"   ? Usu�rio {usuario_id} n�o encontrado")
                return jsonify({'success': False, 'error': 'Usu�rio n�o encontrado'}), 404
            
            print(f"   ?? Convertendo para dict...")
            # Converter para dict se necess�rio
            usuario_dict = dict(usuario) if not isinstance(usuario, dict) else usuario.copy()
            print(f"   ? Dict criado. Keys: {list(usuario_dict.keys())}")
            
            print(f"   ?? Serializando campos datetime...")
            # Converter datetime para string (JSON serializable)
            datetime_fields = ['created_at', 'ultima_sessao', 'updated_at', 'ultimo_acesso']
            for field in datetime_fields:
                if field in usuario_dict and usuario_dict[field]:
                    try:
                        print(f"      - {field}: {type(usuario_dict[field])} ? str")
                        usuario_dict[field] = str(usuario_dict[field])
                    except Exception as e:
                        print(f"      ?? Erro ao serializar {field}: {e}")
                        usuario_dict[field] = None
            
            # Garantir que empresa_id � int ou None
            if 'empresa_id' in usuario_dict and usuario_dict['empresa_id']:
                try:
                    usuario_dict['empresa_id'] = int(usuario_dict['empresa_id'])
                except:
                    usuario_dict['empresa_id'] = None
            
            print(f"   ?? Obtendo permiss�es...")
            # Incluir permiss�es
            permissoes = auth_db.obter_permissoes_usuario(usuario_id)
            print(f"   ?? Permiss�es: {permissoes}")
            usuario_dict['permissoes'] = permissoes
            
            print(f"   ?? Serializando para JSON...")
            result = jsonify(usuario_dict)
            print(f"   ? JSON criado com sucesso")
            print(f"{'='*80}\n")
            return result
            
        except Exception as e:
            print(f"\n{'='*80}")
            print(f"? ERRO ao obter usu�rio {usuario_id}")
            print(f"? Tipo do erro: {type(e).__name__}")
            print(f"? Mensagem: {e}")
            print(f"? Stacktrace:")
            import traceback
            traceback.print_exc()
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.json
            admin = request.usuario
            
            print(f"\n{'='*80}")
            print(f"?? PUT /api/usuarios/{usuario_id} - INICIANDO")
            print(f"{'='*80}")
            print(f"?? DADOS RECEBIDOS DO FRONTEND:")
            print(f"   - Tipo de data: {type(data)}")
            print(f"   - Keys presentes: {list(data.keys()) if data else 'NENHUMA'}")
            print(f"   - JSON completo: {json.dumps(data, indent=2, default=str)}")
            print(f"\n?? CAMPOS ESPEC�FICOS:")
            print(f"   - username: {data.get('username')} (tipo: {type(data.get('username'))})")
            print(f"   - nome_completo: {data.get('nome_completo')} (tipo: {type(data.get('nome_completo'))})")
            print(f"   - email: {data.get('email')} (tipo: {type(data.get('email'))})")
            print(f"   - telefone: {data.get('telefone')} (tipo: {type(data.get('telefone'))})")
            print(f"   - tipo: {data.get('tipo')} (tipo: {type(data.get('tipo'))})")
            print(f"   - ativo: {data.get('ativo')} (tipo: {type(data.get('ativo'))})")
            print(f"   - empresa_id: {data.get('empresa_id')} (tipo: {type(data.get('empresa_id'))})")
            print(f"   - empresas_ids: {data.get('empresas_ids')} (tipo: {type(data.get('empresas_ids'))})")
            print(f"   - empresa_id_padrao: {data.get('empresa_id_padrao')} (tipo: {type(data.get('empresa_id_padrao'))})")
            print(f"   - permissoes: {data.get('permissoes')} (tipo: {type(data.get('permissoes'))})")
            print(f"   - password presente: {'Sim' if 'password' in data else 'N�o'}")
            
            # Validar for�a da senha se estiver sendo alterada
            if 'password' in data and data['password']:
                print(f"\n?? Validando senha...")
                from auth_functions import validar_senha_forte
                valida, mensagem = validar_senha_forte(data['password'])
                if not valida:
                    print(f"? Senha fraca: {mensagem}")
                    print(f"{'='*80}\n")
                    return jsonify({
                        'success': False,
                        'error': f'Senha fraca: {mensagem}'
                    }), 400
                print(f"? Senha v�lida")
            
            print(f"\n?? Chamando auth_db.atualizar_usuario({usuario_id}, data)...")
            print(f"   Fun��o: {auth_db.atualizar_usuario}")
            # Atualizar dados do usu�rio
            success = auth_db.atualizar_usuario(usuario_id, data)
            print(f"   Resultado: {success} (tipo: {type(success)})")
            
            if not success:
                print(f"? Usu�rio {usuario_id} n�o encontrado")
                return jsonify({'success': False, 'error': 'Usu�rio n�o encontrado'}), 404
            
            print(f"? Dados do usu�rio atualizados")
            
            # ?? MULTI-EMPRESA: Atualizar v�nculos se empresas_ids fornecido
            if 'empresas_ids' in data:
                print(f"?? Atualizando v�nculos multi-empresa...")
                from auth_functions import (
                    vincular_usuario_empresa,
                    remover_usuario_empresa,
                    listar_empresas_usuario
                )
                
                empresas_ids = data['empresas_ids']
                empresa_id_padrao = data.get('empresa_id_padrao')
                
                print(f"   - Empresas selecionadas: {empresas_ids}")
                print(f"   - Empresa padr�o: {empresa_id_padrao}")
                
                # Obter empresas atuais
                print(f"   ?? Obtendo empresas atuais...")
                empresas_atuais = listar_empresas_usuario(usuario_id, auth_db)
                empresas_atuais_ids = [e['empresa_id'] for e in empresas_atuais]
                print(f"   - Empresas atuais: {empresas_atuais_ids}")
                
                # Remover v�nculos que n�o est�o mais selecionados
                for empresa_id_atual in empresas_atuais_ids:
                    if empresa_id_atual not in empresas_ids:
                        print(f"??? Removendo v�nculo com empresa {empresa_id_atual}")
                        remover_usuario_empresa(usuario_id, empresa_id_atual, auth_db)
                
                # Adicionar novos v�nculos
                for empresa_id in empresas_ids:
                    if empresa_id not in empresas_atuais_ids:
                        is_padrao = (empresa_id == empresa_id_padrao)
                        permissoes_para_empresa = data.get('permissoes', [])
                        print(f"? Adicionando v�nculo com empresa {empresa_id} (padr�o: {is_padrao})")
                        print(f"   ?? Permiss�es a serem salvas: {permissoes_para_empresa}")
                        
                        vincular_usuario_empresa(
                            usuario_id=usuario_id,
                            empresa_id=empresa_id,
                            papel='usuario',
                            permissoes=permissoes_para_empresa,
                            is_padrao=is_padrao,
                            criado_por=admin['id'],
                            db=auth_db
                        )
                    else:
                        # Atualizar empresa padr�o se necess�rio
                        from auth_functions import atualizar_usuario_empresa
                        is_padrao = (empresa_id == empresa_id_padrao)
                        permissoes_para_empresa = data.get('permissoes', [])
                        
                        # Obter v�nculo atual
                        vinculo_atual = next((e for e in empresas_atuais if e['empresa_id'] == empresa_id), None)
                        
                        print(f"?? Atualizando v�nculo com empresa {empresa_id} (padr�o: {is_padrao})")
                        print(f"   ?? Permiss�es a serem salvas: {permissoes_para_empresa}")
                        
                        atualizar_usuario_empresa(
                            usuario_id=usuario_id,
                            empresa_id=empresa_id,
                            papel=vinculo_atual.get('papel', 'usuario') if vinculo_atual else 'usuario',
                            permissoes=permissoes_para_empresa,
                            is_padrao=is_padrao,
                            db=auth_db
                        )
            
            # Atualizar permiss�es globais se fornecidas (legado)
            if 'permissoes' in data:
                print(f"?? Atualizando permiss�es globais...")
                print(f"   - Permiss�es: {data['permissoes']}")
                auth_db.sincronizar_permissoes_usuario(
                    usuario_id,
                    data['permissoes'],
                    admin['id']
                )
                print(f"   ? Permiss�es atualizadas")
            
            # Registrar atualiza��o
            auth_db.registrar_log_acesso(
                usuario_id=admin['id'],
                acao='update_user',
                descricao=f'Usu�rio atualizado: ID {usuario_id}',
                ip_address=request.remote_addr,
                sucesso=True
            )
            
            print(f"? Usu�rio {usuario_id} atualizado com sucesso!")
            print(f"{'='*80}\n")
            
            return jsonify({
                'success': True,
                'message': 'Usu�rio atualizado com sucesso'
            })
            
        except Exception as e:
            print(f"\n{'='*80}")
            print(f"? ERRO ao atualizar usu�rio {usuario_id}")
            print(f"? Tipo do erro: {type(e).__name__}")
            print(f"? Mensagem: {e}")
            print(f"? Stacktrace:")
            import traceback
            traceback.print_exc()
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    else:  # DELETE
        try:
            admin = request.usuario
            success = auth_db.deletar_usuario(usuario_id)
            
            if not success:
                return jsonify({'success': False, 'error': 'Usu�rio n�o encontrado'}), 404
            
            # Registrar exclus�o
            auth_db.registrar_log_acesso(
                usuario_id=admin['id'],
                acao='delete_user',
                descricao=f'Usu�rio deletado: ID {usuario_id}',
                ip_address=request.remote_addr,
                sucesso=True
            )
            
            return jsonify({
                'success': True,
                'message': 'Usu�rio deletado com sucesso'
            })
            
        except Exception as e:
            print(f"? Erro ao deletar usu�rio: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/usuarios/<int:usuario_id>/permissoes', methods=['PUT'])
@require_admin
def atualizar_permissoes_usuario(usuario_id):
    """Atualizar apenas as permiss�es de um usu�rio"""
    print(f"\n?? [atualizar_permissoes_usuario] FUN��O CHAMADA - ID: {usuario_id}")
    try:
        data = request.json
        permissoes = data.get('permissoes', [])
        
        print(f"?? Permiss�es recebidas: {permissoes}")
        
        # Verificar se usu�rio existe
        usuario = auth_db.obter_usuario(usuario_id)
        if not usuario:
            print(f"? Usu�rio {usuario_id} n�o encontrado")
            return jsonify({'success': False, 'error': 'Usu�rio n�o encontrado'}), 404
        
        # Atualizar permiss�es
        print(f"?? Atualizando permiss�es...")
        success = auth_db.atualizar_permissoes_usuario(usuario_id, permissoes)
        
        if success:
            print(f"? Permiss�es atualizadas com sucesso!")
            return jsonify({
                'success': True,
                'message': 'Permiss�es atualizadas com sucesso'
            })
        else:
            print(f"? Falha ao atualizar permiss�es")
            return jsonify({'success': False, 'error': 'Falha ao atualizar permiss�es'}), 500
            
    except Exception as e:
        print(f"? Erro ao atualizar permiss�es: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/permissoes', methods=['GET'])
@require_admin
def listar_permissoes():
    """Listar todas as permiss�es dispon�veis"""
    print(f"\n?? [listar_permissoes] FUN��O CHAMADA")
    try:
        categoria = request.args.get('categoria')
        permissoes = auth_db.listar_permissoes(categoria)
        return jsonify(permissoes)
    except Exception as e:
        print(f"? Erro ao listar permiss�es: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# === ROTAS DE CONTAS BANC�RIAS ===

@app.route('/api/contas', methods=['GET'])
@require_permission('contas_view')
@aplicar_filtro_cliente
def listar_contas():
    """Lista todas as contas banc�rias com saldo real e filtro de multi-tenancy"""
    try:
        # ?? CORRE��O: Usar empresa_id da sess�o ao inv�s de proprietario_id
        from flask import session
        empresa_id = session.get('empresa_id')
        
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
        
        # Preparar resposta - CALCULAR saldo real com base nos lan�amentos pagos OU extrato banc�rio
        contas_com_saldo = []
        for c in contas:
            # ?? PRIORIDADE 1: Buscar saldo do extrato banc�rio (fonte de verdade)
            saldo_real = None
            
            try:
                with get_db_connection(empresa_id=empresa_id) as conn:
                    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                    
                    # Verificar se existem transações de extrato para esta conta
                    cursor.execute("""
                        SELECT COUNT(*) AS total
                        FROM transacoes_extrato
                        WHERE empresa_id = %s AND conta_bancaria = %s
                    """, (empresa_id, c.nome))
                    tem_extrato = cursor.fetchone()['total'] > 0

                    if tem_extrato:
                        # ✅ CÁLCULO CORRETO: saldo_inicial + SUM(valor de todas as transações)
                        # Mesma estratégia do extrato_functions.py — imune a saldo armazenado errado
                        cursor.execute("""
                            SELECT
                                cb.saldo_inicial,
                                cb.data_inicio,
                                COALESCE(SUM(
                                    CASE
                                        WHEN te.data >= COALESCE(cb.data_inicio, '1900-01-01')
                                        THEN te.valor
                                        ELSE 0
                                    END
                                ), 0) AS soma_transacoes
                            FROM contas_bancarias cb
                            LEFT JOIN transacoes_extrato te
                                   ON te.conta_bancaria = cb.nome
                                  AND te.empresa_id     = cb.empresa_id
                            WHERE cb.empresa_id = %s AND cb.nome = %s
                            GROUP BY cb.saldo_inicial, cb.data_inicio
                        """, (empresa_id, c.nome))
                        row = cursor.fetchone()
                        if row:
                            saldo_real = float(row['saldo_inicial'] or 0) + float(row['soma_transacoes'] or 0)
                            print(f"✅ Conta {c.nome}: saldo_inicial={row['saldo_inicial']} + soma={row['soma_transacoes']:.2f} = R$ {saldo_real:.2f}")
                        else:
                            saldo_real = float(c.saldo_inicial)
                    else:
                        # 📋 FALLBACK: Calcular com base nos lançamentos manuais (sem extrato OFX)
                        print(f"📋 Conta {c.nome}: Sem extrato, calculando com lançamentos...")
                        cursor.execute("""
                            SELECT
                                COALESCE(SUM(CASE WHEN tipo = 'receita' THEN valor ELSE 0 END), 0) AS total_receitas,
                                COALESCE(SUM(CASE WHEN tipo = 'despesa' THEN valor ELSE 0 END), 0) AS total_despesas
                            FROM lancamentos
                            WHERE empresa_id = %s AND conta_bancaria = %s AND status = 'pago'
                        """, (empresa_id, c.nome))
                        result = cursor.fetchone()
                        saldo_real = float(c.saldo_inicial) + float(result['total_receitas']) - float(result['total_despesas'])
                        print(f"📋 Conta {c.nome}: Saldo calculado = R$ {saldo_real:.2f}")
                    
                    cursor.close()
                    
            except Exception as e:
                print(f"?? Erro ao calcular saldo real da conta {c.nome}: {e}")
                import traceback
                traceback.print_exc()
                # Em caso de erro, usar saldo_inicial
                saldo_real = float(c.saldo_inicial)
            
            contas_com_saldo.append({
                'nome': c.nome,
                'banco': c.banco,
                'agencia': c.agencia,
                'conta': c.conta,
                'saldo_inicial': float(c.saldo_inicial),
                'saldo': saldo_real,  # Saldo real calculado
                'saldo_real': saldo_real,  # Alias para compatibilidade
                'ativa': c.ativa if hasattr(c, 'ativa') else True
            })
        
        return jsonify({
            'success': True,
            'data': contas_com_saldo,
            'total': len(contas_com_saldo),
            'message': 'Nenhuma conta cadastrada' if len(contas_com_saldo) == 0 else None
        })
    except Exception as e:
        print(f"? Erro em /api/contas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/contas', methods=['POST'])
@require_permission('contas_create')
@aplicar_filtro_cliente
def adicionar_conta():
    """Adiciona uma nova conta banc�ria"""
    try:
        from flask import session
        
        # ?? Obter empresa_id da sess�o (OBRIGAT�RIO)
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        data = request.json
        
        # Validar campos obrigat�rios
        if not data.get('nome'):
            return jsonify({'success': False, 'error': 'Nome da conta � obrigat�rio'}), 400
        if not data.get('banco'):
            return jsonify({'success': False, 'error': 'Banco � obrigat�rio'}), 400
        
        # ?? proprietario_id = ID do USU�RIO logado (se aplic�vel), n�o empresa_id!
        usuario = get_usuario_logado()
        proprietario_id = usuario.get('id') if usuario.get('tipo') == 'cliente' else None
        
        print(f"\n?? [POST /api/contas] Adicionando conta:")
        print(f"   - empresa_id: {empresa_id}")
        print(f"   - proprietario_id (usuario): {proprietario_id}")
        print(f"   - nome: {data.get('nome')}")
        print(f"   - banco: {data.get('banco')}")
        
        # Verificar contas existentes da mesma empresa antes de adicionar
        contas_existentes = db.listar_contas_por_empresa(empresa_id=empresa_id)
        
        # Verificar se j� existe conta com mesmo nome na mesma empresa
        for c in contas_existentes:
            if c.nome == data['nome']:
                print(f"   ? CONFLITO: Conta '{data['nome']}' j� existe na empresa {empresa_id}!")
                return jsonify({'success': False, 'error': f'J� existe uma conta cadastrada com: Banco: {data["banco"]}, Ag�ncia: {data["agencia"]}, Conta: {data["conta"]}'}), 400
        
        conta = ContaBancaria(
            nome=data['nome'],  # type: ignore
            banco=data['banco'],  # type: ignore
            agencia=data['agencia'],  # type: ignore
            conta=data['conta'],  # type: ignore
            saldo_inicial=float(data.get('saldo_inicial', 0) if data else 0),  # type: ignore
            tipo_saldo_inicial=data.get('tipo_saldo_inicial', 'credor'),  # type: ignore
            data_inicio=data.get('data_inicio')  # type: ignore
        )
        
        conta_id = db.adicionar_conta(conta, proprietario_id=proprietario_id, empresa_id=empresa_id)
        print(f"   ? Conta criada com ID: {conta_id}")
        return jsonify({'success': True, 'id': conta_id})
    except Exception as e:
        print(f"   ? Erro ao criar conta: {str(e)}")
        import traceback
        traceback.print_exc()
        error_msg = str(e)
        if 'UNIQUE constraint' in error_msg:
            error_msg = 'J� existe uma conta com este nome'
        elif 'foreign key constraint' in error_msg.lower():
            error_msg = 'Erro ao vincular conta: proprietario_id inv�lido'
        return jsonify({'success': False, 'error': error_msg}), 400


@app.route('/api/contas/<path:nome>', methods=['GET', 'PUT', 'DELETE', 'OPTIONS'])  # type: ignore
@require_permission('contas_view')
def modificar_conta(nome):
    """Busca, atualiza ou remove uma conta banc�ria"""
    
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    # Decode do nome que vem URL-encoded
    from urllib.parse import unquote
    nome = unquote(nome)
    
    # Responder ao preflight OPTIONS
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    
    if request.method == 'GET':
        try:
            contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
            for conta in contas:
                if conta.nome == nome:
                    return jsonify({
                        'nome': conta.nome,
                        'banco': conta.banco,
                        'agencia': conta.agencia,
                        'conta': conta.conta,
                        'saldo_inicial': float(conta.saldo_inicial),
                        'tipo_saldo_inicial': conta.tipo_saldo_inicial,
                        'data_inicio': conta.data_inicio.isoformat() if hasattr(conta.data_inicio, 'isoformat') else str(conta.data_inicio)
                    })
            return jsonify({'success': False, 'error': 'Conta n�o encontrada'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400
    
    elif request.method == 'PUT':
        try:
            data = request.json
            
            print(f"\n{'='*80}")
            print(f"?? PUT /api/contas/{nome}")
            print(f"{'='*80}")
            print(f"?? Nome da conta a atualizar (par�metro URL): {nome}")
            print(f"?? Dados recebidos: {data}")
            print(f"   - Nome novo: {data.get('nome')}")
            print(f"   - Banco: {data.get('banco')}")
            print(f"   - Ag�ncia: {data.get('agencia')}")
            print(f"   - Conta: {data.get('conta')}")
            print(f"   - Saldo inicial: {data.get('saldo_inicial')}")
            print(f"   - Data in�cio: {data.get('data_inicio')}")
            print(f"   - Tipo saldo: {data.get('tipo_saldo_inicial')}")
            
            conta = ContaBancaria(
                nome=data['nome'],  # type: ignore
                banco=data['banco'],  # type: ignore
                agencia=data['agencia'],  # type: ignore
                conta=data['conta'],  # type: ignore
                saldo_inicial=float(data.get('saldo_inicial', 0) if data else 0),  # type: ignore
                tipo_saldo_inicial=data.get('tipo_saldo_inicial', 'credor'),  # type: ignore
                data_inicio=data.get('data_inicio')  # type: ignore
            )
            
            print(f"? Objeto ContaBancaria criado:")
            print(f"   - Nome: {conta.nome}")
            print(f"?? Chamando db.atualizar_conta(nome_original='{nome}', conta={conta.nome})")
            
            success = db.atualizar_conta(nome, conta)
            
            print(f"?? Resultado: success={success}")
            print(f"{'='*80}\n")
            
            return jsonify({'success': success})
        except Exception as e:
            import traceback
            traceback.print_exc()
            error_msg = str(e)
            if 'UNIQUE constraint' in error_msg:
                error_msg = 'J� existe uma conta com este nome'
            return jsonify({'success': False, 'error': error_msg}), 400
    
    elif request.method == 'DELETE':
        try:
            print(f"\n{'='*80}")
            print(f"??? DELETE /api/contas/{nome}")
            print(f"{'='*80}")
            
            # Verificar se h� lan�amentos vinculados
            lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
            lancamentos_conta = [l for l in lancamentos if l.conta_bancaria == nome]
            
            print(f"?? Lan�amentos vinculados � conta: {len(lancamentos_conta)}")
            
            if lancamentos_conta:
                print(f"? Exclus�o bloqueada: conta possui {len(lancamentos_conta)} lan�amento(s)")
                print(f"{'='*80}\n")
                return jsonify({
                    'success': False, 
                    'error': f'N�o � poss�vel excluir esta conta. Ela possui {len(lancamentos_conta)} lan�amento(s) vinculado(s). Use "Inativar" em vez de excluir.'
                }), 400
            
            # Verificar se h� transa��es de extrato vinculadas
            import psycopg2.extras
            
            conn = db.get_connection()
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Contar transa��es de extrato vinculadas � conta
            cursor.execute("""
                SELECT COUNT(*) as total 
                FROM transacoes_extrato 
                WHERE conta_bancaria LIKE %s
            """, (f'%{nome}%',))
            
            result = cursor.fetchone()
            total_extratos = result['total'] if result else 0
            
            cursor.close()
            conn.close()
            
            print(f"?? Transa��es de extrato vinculadas: {total_extratos}")
            
            if total_extratos > 0:
                print(f"? Exclus�o bloqueada: conta possui {total_extratos} transa��o(�es) de extrato")
                print(f"{'='*80}\n")
                return jsonify({
                    'success': False,
                    'error': f'N�o � poss�vel excluir esta conta. Ela possui {total_extratos} transa��o(�es) de extrato importada(s). Use "Inativar" em vez de excluir.'
                }), 400
            
            # Se n�o h� movimenta��es, pode excluir
            print(f"? Nenhuma movimenta��o encontrada. Excluindo conta...")
            success = db.excluir_conta(nome)
            print(f"?? Resultado: success={success}")
            print(f"{'='*80}\n")
            
            return jsonify({'success': success})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/contas/<path:nome>/toggle-ativo', methods=['POST'])
@require_permission('contas_edit')
def toggle_ativo_conta(nome):
    """Ativa ou inativa uma conta banc�ria"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        from urllib.parse import unquote
        nome = unquote(nome)
        
        print(f"\n{'='*80}")
        print(f"?? POST /api/contas/{nome}/toggle-ativo")
        print(f"{'='*80}")
        
        # Buscar conta atual
        contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
        conta_atual = None
        for c in contas:
            if c.nome == nome:
                conta_atual = c
                break
        
        if not conta_atual:
            print(f"? Conta n�o encontrada")
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': 'Conta n�o encontrada'}), 404
        
        # Inverter status
        novo_status = not conta_atual.ativa
        print(f"?? Status atual: {conta_atual.ativa}")
        print(f"?? Novo status: {novo_status}")
        
        # Atualizar conta com novo status
        conta_atual.ativa = novo_status
        success = db.atualizar_conta(nome, conta_atual)
        
        acao = "ativada" if novo_status else "inativada"
        print(f"? Conta {acao} com sucesso")
        print(f"{'='*80}\n")
        
        return jsonify({
            'success': success,
            'ativa': novo_status,
            'message': f'Conta {acao} com sucesso'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/transferencias', methods=['POST'])
@require_permission('lancamentos_create')
def criar_transferencia():
    """Cria uma transfer�ncia entre contas banc�rias"""
    try:
        data = request.json
        empresa_id = data.get('empresa_id') if data else None
        
        # Validar dados
        if not data or not data.get('conta_origem') or not data.get('conta_destino'):
            return jsonify({'success': False, 'error': 'Contas de origem e destino s�o obrigat�rias'}), 400
        
        if data['conta_origem'] == data['conta_destino']:
            return jsonify({'success': False, 'error': 'Conta de origem e destino n�o podem ser iguais'}), 400
        
        valor = float(data.get('valor', 0))
        if valor <= 0:
            return jsonify({'success': False, 'error': 'Valor deve ser maior que zero'}), 400
        
        # Buscar contas
        conta_origem = db.buscar_conta(data['conta_origem'])
        conta_destino = db.buscar_conta(data['conta_destino'])
        
        if not conta_origem:
            return jsonify({'success': False, 'error': 'Conta de origem n�o encontrada'}), 404
        if not conta_destino:
            return jsonify({'success': False, 'error': 'Conta de destino n�o encontrada'}), 404
        
        # Validar se as contas est�o ativas
        if hasattr(conta_origem, 'ativa') and not conta_origem.ativa:
            print(f"? Tentativa de criar transfer�ncia com conta origem inativa: {conta_origem.nome}")
            return jsonify({
                'success': False,
                'error': f'N�o � poss�vel criar transfer�ncia. A conta de origem "{conta_origem.nome}" est� inativa. Reative a conta antes de criar transfer�ncias.'
            }), 400
        
        if hasattr(conta_destino, 'ativa') and not conta_destino.ativa:
            print(f"? Tentativa de criar transfer�ncia com conta destino inativa: {conta_destino.nome}")
            return jsonify({
                'success': False,
                'error': f'N�o � poss�vel criar transfer�ncia. A conta de destino "{conta_destino.nome}" est� inativa. Reative a conta antes de criar transfer�ncias.'
            }), 400
        
        if not conta_origem:
            return jsonify({'success': False, 'error': 'Conta de origem n�o encontrada'}), 404
        if not conta_destino:
            return jsonify({'success': False, 'error': 'Conta de destino n�o encontrada'}), 404
        
        # Criar data da transfer�ncia
        data_transferencia = datetime.fromisoformat(data['data']) if data.get('data') else datetime.now()
        
        # Criar lan�amento de transfer�ncia
        lancamento = Lancamento(
            descricao=f"Transfer�ncia: {conta_origem.nome} ? {conta_destino.nome}",
            valor=valor,
            tipo=TipoLancamento.TRANSFERENCIA,
            categoria="Transfer�ncia Interna",
            data_vencimento=data_transferencia,
            data_pagamento=data_transferencia,
            conta_bancaria=data['conta_origem'],
            pessoa="",
            observacoes=f"Destino: {conta_destino.nome}. {data.get('observacoes', '')}",
            num_documento="",
            subcategoria=data['conta_destino']  # Usar subcategoria para armazenar conta destino
        )
        
        lancamento.status = StatusLancamento.PAGO
        lancamento_id = db.adicionar_lancamento(lancamento, empresa_id=empresa_id)
        
        return jsonify({'success': True, 'id': lancamento_id})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# === ROTAS DE CATEGORIAS ===

@app.route('/api/categorias', methods=['GET'])
@require_permission('categorias_view')
def listar_categorias():
    """Lista todas as categorias"""
    try:
        # Filtrar por empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        
        # Listar categorias da empresa
        categorias = db.listar_categorias(empresa_id=empresa_id)
        for i, c in enumerate(categorias):
            print(f'   [{i+1}] {c.nome} (tipo: {c.tipo.value}, empresa_id: {getattr(c, "empresa_id", "N/A")})')
        
        resultado = [{
            'id': c.id,  # ? Adicionar ID da categoria
            'nome': c.nome,
            'tipo': c.tipo.value,
            'subcategorias': c.subcategorias,
            'empresa_id': getattr(c, 'empresa_id', None)
        } for c in categorias]
        
        print(f'   ? Retornando {len(resultado)} categorias')
        print('='*80 + '\n')
        return jsonify({
            'success': True,
            'data': resultado,
            'total': len(resultado),
            'message': 'Nenhuma categoria cadastrada. Adicione categorias para organizar suas transa��es.' if len(resultado) == 0 else None
        })
    except Exception as e:
        print(f'   ? Erro ao listar categorias: {str(e)}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categorias', methods=['POST'])
@require_permission('categorias_create')
def adicionar_categoria():
    """Adiciona uma nova categoria"""
    try:
        print('\n' + '='*80)
        print('?? POST /api/categorias - NOVA CATEGORIA')
        print(f'   ?? Headers: {dict(request.headers)}')
        print(f'   ?? CSRF Token no header: {request.headers.get("X-CSRFToken", "AUSENTE")}')
        print(f'   ?? Empresa na sess�o: {session.get("empresa_id")}')
        print(f'   ?? Usu�rio na sess�o: {session.get("usuario_id")}')
        
        data = request.json
        print(f'   ?? Dados recebidos: {data}')
        
        # Extrair empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            print('   ? ERRO: Empresa n�o identificada na sess�o!')
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
        
        # Converter tipo para min�scula para compatibilidade com o enum
        tipo_str = data['tipo'].lower() if data and data.get('tipo') else 'receita'  # type: ignore
        
        # Normalizar nome: uppercase e trim
        nome_normalizado = data['nome'].strip().upper() if data and data.get('nome') else ''  # type: ignore
        
        print(f'   ?? Nome normalizado: {nome_normalizado}')
        print(f'   ?? Tipo: {tipo_str}')
        print(f'   ?? Empresa ID: {empresa_id}')
        
        categoria = Categoria(
            nome=nome_normalizado,  # type: ignore
            tipo=TipoLancamento(tipo_str),  # type: ignore
            subcategorias=data.get('subcategorias', []) if data else [],  # type: ignore
            empresa_id=empresa_id  # type: ignore
        )
        categoria_id = db.adicionar_categoria(categoria)
        
        print(f'   ? Categoria criada com ID: {categoria_id}')
        print('='*80 + '\n')
        
        return jsonify({'success': True, 'id': categoria_id})
    except Exception as e:
        import traceback
        print('   ? ERRO ao adicionar categoria:')
        traceback.print_exc()
        print('='*80 + '\n')
        
        error_msg = str(e)
        if 'UNIQUE constraint' in error_msg:
            error_msg = 'J� existe uma categoria com este nome'
        return jsonify({'success': False, 'error': error_msg}), 400


@app.route('/api/categorias/<path:nome>', methods=['PUT', 'DELETE'])  # type: ignore
@require_permission('categorias_edit')
def modificar_categoria(nome):
    """Atualiza ou remove uma categoria"""
    if request.method == 'PUT':
        try:
            data = request.json
            
            print('\n' + '='*80)
            print('?? PUT /api/categorias - ATUALIZAR CATEGORIA')
            print(f'   ?? Nome original (URL): {nome}')
            print(f'   ?? CSRF Token no header: {request.headers.get("X-CSRFToken", "AUSENTE")}')
            print(f'   ?? Dados recebidos: {data}')
            print(f'   ?? Empresa na sess�o: {session.get("empresa_id")}')
            print(f'   ?? Usu�rio na sess�o: {session.get("usuario_id")}')
            
            # Extrair empresa_id do request ou sess�o
            empresa_id = data.get('empresa_id') if data else None
            if not empresa_id:
                empresa_id = session.get('empresa_id')
            
            print(f'   ?? empresa_id a ser usado: {empresa_id}')
            
            # Converter tipo para min�scula para compatibilidade com o enum
            tipo_str = data['tipo'].lower() if data and data.get('tipo') else 'receita'  # type: ignore
            
            # Normalizar nome: uppercase e trim
            nome_normalizado = data['nome'].strip().upper() if data and data.get('nome') else ''  # type: ignore
            
            # Se o nome mudou, precisamos atualizar com atualizar_nome_categoria primeiro
            nome_original_normalizado = nome.strip().upper()
            
            print(f'   ?? Nome original normalizado: {nome_original_normalizado}')
            print(f'   ?? Nome novo normalizado: {nome_normalizado}')
            print(f'   ?? Nome mudou? {nome_normalizado != nome_original_normalizado}')
            
            # Criar objeto categoria com os novos dados
            categoria = Categoria(
                nome=nome_normalizado,  # type: ignore
                tipo=TipoLancamento(tipo_str),  # type: ignore
                subcategorias=data.get('subcategorias', []) if data else [],  # type: ignore
                empresa_id=empresa_id  # type: ignore
            )
            
            print(f'   ?? Atualizando categoria: {categoria.nome} (tipo: {categoria.tipo.value}, empresa: {categoria.empresa_id})')
            print(f'   ?? Usando nome_original para localizar: {nome_original_normalizado}')
            
            # Passar nome_original para a fun��o UPDATE usar no WHERE
            success = db.atualizar_categoria(categoria, nome_original=nome_original_normalizado)
            
            print(f'   {"?" if success else "?"} Resultado: {success}')
            print('='*80 + '\n')
            
            return jsonify({'success': success})
        except Exception as e:
            import traceback
            print('   ? ERRO ao atualizar categoria:')
            traceback.print_exc()
            print('='*80 + '\n')
            
            error_msg = str(e)
            if 'UNIQUE constraint' in error_msg:
                error_msg = 'J� existe uma categoria com este nome'
            return jsonify({'success': False, 'error': error_msg}), 400
    
    elif request.method == 'DELETE':
        try:
            print('\n' + '='*80)
            print('??? DELETE /api/categorias - EXCLUIR CATEGORIA')
            print(f'   ?? Nome (URL): {nome}')
            print(f'   ?? CSRF Token no header: {request.headers.get("X-CSRFToken", "AUSENTE")}')
            print(f'   ?? Usu�rio: {session.get("usuario_id")}')
            print(f'   ?? Empresa: {session.get("empresa_id")}')
            
            # Normalizar nome
            nome_normalizado = nome.strip().upper()
            print(f'   ?? Nome normalizado: {nome_normalizado}')
            
            success = db.excluir_categoria(nome)
            
            print(f'   {"?" if success else "?"} Resultado: {success}')
            print('='*80 + '\n')
            
            return jsonify({'success': success})
        except Exception as e:
            print('   ? ERRO ao excluir categoria:')
            print(f'   Mensagem: {str(e)}')
            import traceback
            traceback.print_exc()
            print('='*80 + '\n')
            
            return jsonify({'success': False, 'error': str(e)}), 400


# === IMPORTA��O DE CATEGORIAS ENTRE EMPRESAS ===

@app.route('/api/categorias/empresas-disponiveis', methods=['GET'])
@require_permission('categorias_view')
def listar_empresas_com_categorias():
    """Lista empresas do usu�rio com suas categorias para importa��o"""
    try:
        usuario = get_usuario_logado()
        empresa_atual_id = session.get('empresa_id')
        
        print(f"\n?? [IMPORTAR CATEGORIAS] Buscando empresas dispon�veis")
        print(f"   ?? Usu�rio: {usuario.get('nome')}")
        print(f"   ?? Empresa atual: {empresa_atual_id}")
        
        # Buscar empresas do usu�rio
        from auth_functions import listar_empresas_usuario
        empresas = listar_empresas_usuario(usuario.get('id'), auth_db)
        print(f"   ?? Total de empresas do usu�rio: {len(empresas)}")
        
        empresas_com_categorias = []
        for empresa in empresas:
            empresa_id = empresa.get('empresa_id')
            razao_social = empresa.get('razao_social')
            
            print(f"\n   ?? Analisando empresa: {razao_social} (ID: {empresa_id})")
            
            # N�o listar a empresa atual
            if empresa_id == empresa_atual_id:
                print(f"      ?? Pulando (� a empresa atual)")
                continue
            
            # Buscar categorias desta empresa
            categorias = db.listar_categorias(empresa_id=empresa_id)
            print(f"      ?? Categorias encontradas: {len(categorias)}")
            
            if categorias:  # S� incluir empresas que t�m categorias
                categorias_list = []
                for cat in categorias:
                    # Verificar se � objeto ou dicion�rio
                    if hasattr(cat, 'nome'):
                        cat_dict = {
                            'nome': cat.nome,
                            'tipo': cat.tipo.value if hasattr(cat.tipo, 'value') else cat.tipo,
                            'subcategorias': cat.subcategorias if hasattr(cat, 'subcategorias') else []
                        }
                    else:
                        cat_dict = {
                            'nome': cat.get('nome', 'Sem nome'),
                            'tipo': cat.get('tipo', 'despesa'),
                            'subcategorias': cat.get('subcategorias', [])
                        }
                    categorias_list.append(cat_dict)
                
                empresas_com_categorias.append({
                    'empresa_id': empresa_id,
                    'razao_social': razao_social,
                    'total_categorias': len(categorias),
                    'categorias': categorias_list
                })
                print(f"      ? Empresa inclu�da com {len(categorias)} categoria(s)")
        
        print(f"\n? Total de empresas dispon�veis para importa��o: {len(empresas_com_categorias)}")
        
        return jsonify({
            'success': True,
            'data': empresas_com_categorias
        })
        
    except Exception as e:
        print(f"? Erro ao listar empresas com categorias: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categorias/importar-de-empresa', methods=['POST'])
@require_permission('categorias_create')
def importar_categorias_de_empresa():
    """Importa categorias de outra empresa do usu�rio"""
    print("\n" + "="*80)
    print("?? IMPORTAR CATEGORIAS - IN�CIO")
    print("="*80)
    
    try:
        data = request.json
        empresa_origem_id = data.get('empresa_origem_id')
        categorias_ids = data.get('categorias')  # Lista de nomes de categorias para importar
        
        print(f"?? Request data: {data}")
        print(f"?? Empresa origem: {empresa_origem_id}")
        print(f"?? Categorias espec�ficas: {categorias_ids}")
        
        if not empresa_origem_id:
            return jsonify({'success': False, 'error': 'empresa_origem_id � obrigat�rio'}), 400
        
        usuario = get_usuario_logado()
        empresa_destino_id = session.get('empresa_id')
        
        print(f"?? Usu�rio: {usuario.get('nome')}")
        print(f"?? Empresa destino: {empresa_destino_id}")
        
        if not empresa_destino_id:
            return jsonify({'success': False, 'error': 'Empresa destino n�o identificada'}), 400
        
        # Verificar se usu�rio tem acesso � empresa origem
        from auth_functions import listar_empresas_usuario
        empresas_usuario = listar_empresas_usuario(usuario.get('id'), auth_db)
        tem_acesso = any(e.get('empresa_id') == empresa_origem_id for e in empresas_usuario)
        
        print(f"? Tem acesso � empresa origem? {tem_acesso}")
        
        if not tem_acesso:
            return jsonify({'success': False, 'error': 'Sem permiss�o para acessar empresa origem'}), 403
        
        # Buscar categorias da empresa origem
        categorias_origem = db.listar_categorias(empresa_id=empresa_origem_id)
        print(f"?? Categorias da origem: {len(categorias_origem)}")
        for cat in categorias_origem:
            print(f"   - {cat.nome} ({cat.tipo.value if hasattr(cat.tipo, 'value') else cat.tipo})")
        
        # Filtrar categorias selecionadas (se especificado)
        if categorias_ids:
            categorias_origem = [c for c in categorias_origem if c.nome in categorias_ids]
            print(f"?? Ap�s filtro: {len(categorias_origem)} categorias")
        
        # Buscar categorias j� existentes na empresa destino
        categorias_destino = db.listar_categorias(empresa_id=empresa_destino_id)
        nomes_existentes = {c.nome.upper() for c in categorias_destino}
        print(f"?? Categorias no destino: {len(categorias_destino)} ({nomes_existentes})")
        
        importadas = 0
        duplicadas = 0
        erros = []
        
        print(f"\n?? Iniciando loop de importa��o...")
        for cat_origem in categorias_origem:
            try:
                print(f"\n   ?? Processando: {cat_origem.nome}")
                
                # Verificar se j� existe (case insensitive)
                if cat_origem.nome.upper() in nomes_existentes:
                    print(f"      ?? Duplicada")
                    duplicadas += 1
                    continue
                
                print(f"      ? Nova categoria - criando...")
                
                # Criar nova categoria na empresa destino
                nova_categoria = Categoria(
                    nome=cat_origem.nome,
                    tipo=cat_origem.tipo,
                    descricao=getattr(cat_origem, 'descricao', ''),
                    subcategorias=getattr(cat_origem, 'subcategorias', []),
                    cor=getattr(cat_origem, 'cor', '#000000'),
                    icone=getattr(cat_origem, 'icone', 'folder'),
                    empresa_id=empresa_destino_id
                )
                
                print(f"      ?? Objeto Categoria criado: nome={nova_categoria.nome}, tipo={nova_categoria.tipo}, empresa_id={nova_categoria.empresa_id}")
                
                categoria_id = db.adicionar_categoria(nova_categoria)
                print(f"      ? Categoria adicionada com ID: {categoria_id}")
                importadas += 1
                
            except Exception as e:
                print(f"      ? ERRO ao processar {cat_origem.nome}: {e}")
                import traceback
                traceback.print_exc()
                erros.append(f"{cat_origem.nome}: {str(e)}")
        
        print(f"\n?? RESULTADO:")
        print(f"   ? Importadas: {importadas}")
        print(f"   ?? Duplicadas: {duplicadas}")
        print(f"   ? Erros: {len(erros)}")
        if erros:
            for erro in erros:
                print(f"      - {erro}")
        
        print("="*80)
        print("?? IMPORTAR CATEGORIAS - FIM")
        print("="*80 + "\n")
        
        return jsonify({
            'success': True,
            'importadas': importadas,
            'duplicadas': duplicadas,
            'erros': erros,
            'message': f'{importadas} categoria(s) importada(s) com sucesso'
        })
        
    except Exception as e:
        print(f"? ERRO FATAL ao importar categorias: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# === ROTAS DE CLIENTES ===

@app.route('/api/clientes', methods=['GET'])
@require_permission('clientes_view')
@aplicar_filtro_cliente
def listar_clientes():
    """Lista clientes ativos ou inativos com filtro de multi-tenancy"""
    ativos = request.args.get('ativos', 'true').lower() == 'true'
    
    # ? CORRE��O: Usar filtro do decorator (empresa_id do usu�rio)
    # O decorator @aplicar_filtro_cliente seta request.filtro_cliente_id = empresa_id
    # As fun��es de DB agora filtram por empresa_id (n�o mais proprietario_id)
    filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
    
    usuario = get_usuario_logado()
    clientes = db.listar_clientes(ativos=ativos, filtro_cliente_id=filtro_cliente_id)
    
    # Adicionar cliente_id para cada cliente (usando nome como identificador)
    for cliente in clientes:
        cliente['cliente_id'] = cliente.get('nome')
    
    return jsonify({
        'success': True,
        'data': clientes,
        'total': len(clientes),
        'message': 'Nenhum cliente cadastrado' if len(clientes) == 0 else None
    })


@app.route('/api/clientes', methods=['POST'])
@require_permission('clientes_create')
@aplicar_filtro_cliente
def adicionar_cliente():
    """Adiciona um novo cliente"""
    try:
        from flask import session
        from app.utils.validators import validate_cpf, validate_cnpj, validate_email
        
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        data = request.json
        
        # Validar campos obrigat�rios
        if not data.get('nome'):
            return jsonify({'success': False, 'error': 'Nome do cliente � obrigat�rio'}), 400
        
        # ?? Validar CPF/CNPJ se fornecido
        if data.get('cpf_cnpj'):
            cpf_cnpj = data['cpf_cnpj'].strip()
            # Remover formata��o para detectar se � CPF (11) ou CNPJ (14)
            import re
            numeros = re.sub(r'[^0-9]', '', cpf_cnpj)
            
            if len(numeros) == 11:
                is_valid, error_msg = validate_cpf(cpf_cnpj)
                if not is_valid:
                    return jsonify({'success': False, 'error': f'CPF inv�lido: {error_msg}'}), 400
            elif len(numeros) == 14:
                is_valid, error_msg = validate_cnpj(cpf_cnpj)
                if not is_valid:
                    return jsonify({'success': False, 'error': f'CNPJ inv�lido: {error_msg}'}), 400
            elif numeros:  # Se tem algum n�mero mas n�o � 11 nem 14
                return jsonify({'success': False, 'error': 'CPF deve ter 11 d�gitos ou CNPJ deve ter 14 d�gitos'}), 400
        
        # ?? Validar email se fornecido
        if data.get('email'):
            is_valid, error_msg = validate_email(data['email'])
            if not is_valid:
                return jsonify({'success': False, 'error': f'Email inv�lido: {error_msg}'}), 400
        
        # ?? Garantir que empresa_id est� nos dados
        data['empresa_id'] = empresa_id
        
        # ?? Obter proprietario_id do usu�rio logado (ID na tabela usuarios)
        usuario = get_usuario_logado()
        proprietario_id = None
        if usuario and usuario.get('tipo') == 'cliente':
            proprietario_id = usuario.get('id')  # ID do usu�rio, N�O empresa_id
        
        print(f"\n?? [POST /api/clientes] Adicionando cliente:")
        print(f"   - empresa_id: {empresa_id}")
        print(f"   - usuario.id: {usuario.get('id') if usuario else None}")
        print(f"   - usuario.tipo: {usuario.get('tipo') if usuario else None}")
        print(f"   - proprietario_id: {proprietario_id}")
        print(f"   - nome: {data.get('nome')}")
        
        cliente_id = db.adicionar_cliente(data, proprietario_id=proprietario_id)  # type: ignore
        print(f"   ? Cliente criado com ID: {cliente_id}")
        return jsonify({'success': True, 'id': cliente_id})
    except ValueError as e:
        # Erro de valida��o (ex: CPF/CNPJ duplicado)
        error_msg = str(e)
        print(f"   ?? Valida��o: {error_msg}")
        return jsonify({'success': False, 'error': error_msg}), 400
    except Exception as e:
        print(f"   ? Erro ao criar cliente: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clientes/<path:nome>', methods=['GET'])
@require_permission('clientes_view')
@aplicar_filtro_cliente
def obter_cliente(nome):
    """Busca um cliente espec�fico pelo nome"""
    try:
        # Decode do nome que vem URL-encoded
        from urllib.parse import unquote
        nome = unquote(nome)
        
        filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
        
        print(f"\n=== Buscando cliente ===")
        print(f"Nome: {nome}")
        print(f"Filtro cliente ID (empresa_id): {filtro_cliente_id}")
        
        cliente = db.obter_cliente_por_nome(nome)
        
        if not cliente:
            return jsonify({'success': False, 'error': 'Cliente n�o encontrado'}), 404
        
        # ? CORRE��O: Validar por empresa_id (n�o mais proprietario_id)
        # filtro_cliente_id cont�m o empresa_id do usu�rio logado
        if filtro_cliente_id is not None:
            cliente_empresa_id = cliente.get('empresa_id')
            if cliente_empresa_id != filtro_cliente_id:
                print(f"? Acesso negado: cliente.empresa_id={cliente_empresa_id}, filtro={filtro_cliente_id}")
                return jsonify({'success': False, 'error': 'Cliente n�o encontrado ou sem permiss�o'}), 403
        
        print(f"? Cliente encontrado: {cliente.get('nome')}")
        print(f"   - empresa_id: {cliente.get('empresa_id')}")
        print(f"   - cpf_cnpj: {cliente.get('cpf_cnpj')}")
        return jsonify(cliente)
    except Exception as e:
        print(f"? ERRO ao buscar cliente: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clientes/<path:nome>', methods=['PUT', 'DELETE'])  # type: ignore
@require_permission('clientes_edit')
@aplicar_filtro_cliente
def modificar_cliente(nome):
    """Atualiza ou remove um cliente com valida��o de empresa"""
    # Decode do nome que vem URL-encoded
    from urllib.parse import unquote
    nome = unquote(nome)
    
    filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
    
    if request.method == 'PUT':
        try:
            from app.utils.validators import validate_cpf, validate_cnpj, validate_email
            import re
            
            data = request.json
            print(f"\n=== Atualizando cliente ===")
            print(f"URL recebida: {request.url}")
            print(f"Nome da URL (raw): '{nome}'")
            print(f"Dados recebidos: {data}")
            
            # ?? Validar CPF/CNPJ se fornecido
            if data.get('cpf_cnpj'):
                cpf_cnpj = data['cpf_cnpj'].strip()
                numeros = re.sub(r'[^0-9]', '', cpf_cnpj)
                
                if len(numeros) == 11:
                    is_valid, error_msg = validate_cpf(cpf_cnpj)
                    if not is_valid:
                        return jsonify({'success': False, 'error': f'CPF inv�lido: {error_msg}'}), 400
                elif len(numeros) == 14:
                    is_valid, error_msg = validate_cnpj(cpf_cnpj)
                    if not is_valid:
                        return jsonify({'success': False, 'error': f'CNPJ inv�lido: {error_msg}'}), 400
                elif numeros:
                    return jsonify({'success': False, 'error': 'CPF deve ter 11 d�gitos ou CNPJ deve ter 14 d�gitos'}), 400
            
            # ?? Validar email se fornecido
            if data.get('email'):
                is_valid, error_msg = validate_email(data['email'])
                if not is_valid:
                    return jsonify({'success': False, 'error': f'Email inv�lido: {error_msg}'}), 400
            
            # Validar propriedade antes de atualizar (se n�o for admin)
            if filtro_cliente_id is not None:
                cliente_atual = db.obter_cliente_por_nome(nome)
                if not cliente_atual or cliente_atual.get('empresa_id') != filtro_cliente_id:
                    return jsonify({'success': False, 'error': 'Cliente n�o encontrado ou sem permiss�o'}), 403
            
            success = atualizar_cliente(nome, data)
            print(f"Cliente atualizado: {success}")
            return jsonify({'success': success})
        except Exception as e:
            print(f"ERRO ao atualizar cliente: {str(e)}")
            import traceback
            traceback.print_exc()
            error_msg = str(e)
            if 'UNIQUE constraint' in error_msg:
                error_msg = 'J� existe um cliente com este nome'
            return jsonify({'success': False, 'error': error_msg}), 400
    
    elif request.method == 'DELETE':
        try:
            # Validar propriedade antes de deletar (se n�o for admin)
            if filtro_cliente_id is not None:
                cliente_atual = db.obter_cliente_por_nome(nome)
                if not cliente_atual or cliente_atual.get('empresa_id') != filtro_cliente_id:
                    return jsonify({'success': False, 'error': 'Cliente n�o encontrado ou sem permiss�o'}), 403
            
            success, mensagem = db.excluir_cliente(nome)
            if success:
                return jsonify({'success': True, 'message': mensagem})
            else:
                return jsonify({'success': False, 'error': mensagem}), 400
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


# === ROTAS DE FORNECEDORES ===

@app.route('/api/fornecedores', methods=['GET'])
@require_permission('fornecedores_view')
@aplicar_filtro_cliente
def listar_fornecedores():
    """Lista fornecedores ativos ou inativos com filtro de multi-tenancy"""
    ativos = request.args.get('ativos', 'true').lower() == 'true'
    
    # ? CORRE��O: Usar filtro do decorator (empresa_id do usu�rio)
    # O decorator @aplicar_filtro_cliente seta request.filtro_cliente_id = empresa_id
    # As fun��es de DB agora filtram por empresa_id (n�o mais proprietario_id)
    filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
    
    fornecedores = db.listar_fornecedores(ativos=ativos, filtro_cliente_id=filtro_cliente_id)
    
    return jsonify({
        'success': True,
        'data': fornecedores,
        'total': len(fornecedores),
        'message': 'Nenhum fornecedor cadastrado' if len(fornecedores) == 0 else None
    })


@app.route('/api/fornecedores', methods=['POST'])
@require_permission('fornecedores_create')
@aplicar_filtro_cliente
def adicionar_fornecedor():
    """Adiciona um novo fornecedor"""
    try:
        from app.utils.validators import validate_cpf, validate_cnpj, validate_email
        import re
        
        # ?? VALIDA��O DE SEGURAN�A
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.json
        
        # ?? Validar campo obrigat�rio
        if not data.get('nome'):
            return jsonify({'success': False, 'error': 'Nome do fornecedor � obrigat�rio'}), 400
        
        # ?? Validar CPF/CNPJ se fornecido
        if data.get('cpf_cnpj'):
            cpf_cnpj = data['cpf_cnpj'].strip()
            numeros = re.sub(r'[^0-9]', '', cpf_cnpj)
            
            if len(numeros) == 11:
                is_valid, error_msg = validate_cpf(cpf_cnpj)
                if not is_valid:
                    return jsonify({'success': False, 'error': f'CPF inv�lido: {error_msg}'}), 400
            elif len(numeros) == 14:
                is_valid, error_msg = validate_cnpj(cpf_cnpj)
                if not is_valid:
                    return jsonify({'success': False, 'error': f'CNPJ inv�lido: {error_msg}'}), 400
            elif numeros:
                return jsonify({'success': False, 'error': 'CPF deve ter 11 d�gitos ou CNPJ deve ter 14 d�gitos'}), 400
        
        # ?? Validar email se fornecido
        if data.get('email'):
            is_valid, error_msg = validate_email(data['email'])
            if not is_valid:
                return jsonify({'success': False, 'error': f'Email inv�lido: {error_msg}'}), 400
        
        # ?? Adicionar empresa_id aos dados
        data['empresa_id'] = empresa_id
        
        # ?? Obter proprietario_id do usu�rio logado (ID na tabela usuarios)
        usuario = get_usuario_logado()
        proprietario_id = None
        if usuario and usuario.get('tipo') == 'cliente':
            proprietario_id = usuario.get('id')  # ID do usu�rio, N�O empresa_id
        
        print(f"\n?? [POST /api/fornecedores]")
        print(f"   - empresa_id: {empresa_id}")
        print(f"   - usuario.id: {usuario.get('id') if usuario else None}")
        print(f"   - usuario.tipo: {usuario.get('tipo') if usuario else None}")
        print(f"   - proprietario_id: {proprietario_id}")
        print(f"   - nome: {data.get('nome')}")
        
        fornecedor_id = db.adicionar_fornecedor(data, proprietario_id=proprietario_id)  # type: ignore
        print(f"   ? Fornecedor criado com ID: {fornecedor_id}")
        return jsonify({'success': True, 'id': fornecedor_id})
    except ValueError as e:
        # Erro de valida��o (ex: CPF/CNPJ duplicado)
        error_msg = str(e)
        print(f"   ?? Valida��o: {error_msg}")
        return jsonify({'success': False, 'error': error_msg}), 400
    except Exception as e:
        print(f"   ? Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fornecedores/<path:nome>', methods=['GET'])
@require_permission('fornecedores_view')
@aplicar_filtro_cliente
def obter_fornecedor(nome):
    """Obt�m dados de um fornecedor espec�fico"""
    try:
        # Decode do nome que vem URL-encoded
        from urllib.parse import unquote
        nome = unquote(nome)
        
        filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
        
        print(f"\n=== Buscando fornecedor ===")
        print(f"Nome: {nome}")
        print(f"Filtro cliente ID (empresa_id): {filtro_cliente_id}")
        
        # Buscar fornecedor
        fornecedor = db.obter_fornecedor_por_nome(nome)
        
        if not fornecedor:
            return jsonify({'error': 'Fornecedor n�o encontrado'}), 404
        
        # ? CORRE��O: Validar por empresa_id (n�o mais proprietario_id)
        # filtro_cliente_id cont�m o empresa_id do usu�rio logado
        if filtro_cliente_id is not None:
            fornecedor_empresa_id = fornecedor.get('empresa_id')
            if fornecedor_empresa_id != filtro_cliente_id:
                print(f"? Acesso negado: fornecedor.empresa_id={fornecedor_empresa_id}, filtro={filtro_cliente_id}")
                return jsonify({'error': 'Sem permiss�o para visualizar este fornecedor'}), 403
        
        print(f"? Fornecedor encontrado: {fornecedor.get('nome')}")
        print(f"   - empresa_id: {fornecedor.get('empresa_id')}")
        print(f"   - cpf_cnpj: {fornecedor.get('cpf_cnpj')}")
        
        # Retornar dados completos do fornecedor
        return jsonify(fornecedor)
        
    except Exception as e:
        print(f"? ERRO ao obter fornecedor {nome}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/fornecedores/<path:nome>', methods=['PUT', 'DELETE'])  # type: ignore
@require_permission('fornecedores_edit')
@aplicar_filtro_cliente
def modificar_fornecedor(nome):
    """Atualiza ou remove um fornecedor com valida��o de empresa"""
    # Decode do nome que vem URL-encoded
    from urllib.parse import unquote
    nome = unquote(nome)
    
    filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
    
    if request.method == 'PUT':
        try:
            from app.utils.validators import validate_cpf, validate_cnpj, validate_email
            import re
            
            data = request.json
            
            # ?? Validar CPF/CNPJ se fornecido
            if data.get('cpf_cnpj'):
                cpf_cnpj = data['cpf_cnpj'].strip()
                numeros = re.sub(r'[^0-9]', '', cpf_cnpj)
                
                if len(numeros) == 11:
                    is_valid, error_msg = validate_cpf(cpf_cnpj)
                    if not is_valid:
                        return jsonify({'success': False, 'error': f'CPF inv�lido: {error_msg}'}), 400
                elif len(numeros) == 14:
                    is_valid, error_msg = validate_cnpj(cpf_cnpj)
                    if not is_valid:
                        return jsonify({'success': False, 'error': f'CNPJ inv�lido: {error_msg}'}), 400
                elif numeros:
                    return jsonify({'success': False, 'error': 'CPF deve ter 11 d�gitos ou CNPJ deve ter 14 d�gitos'}), 400
            
            # ?? Validar email se fornecido
            if data.get('email'):
                is_valid, error_msg = validate_email(data['email'])
                if not is_valid:
                    return jsonify({'success': False, 'error': f'Email inv�lido: {error_msg}'}), 400
            
            # Validar propriedade antes de atualizar (se n�o for admin)
            if filtro_cliente_id is not None:
                fornecedor_atual = db.obter_fornecedor_por_nome(nome)
                if not fornecedor_atual or fornecedor_atual.get('empresa_id') != filtro_cliente_id:
                    return jsonify({'success': False, 'error': 'Fornecedor n�o encontrado ou sem permiss�o'}), 403
            
            success = atualizar_fornecedor(nome, data)
            return jsonify({'success': success})
        except Exception as e:
            import traceback
            traceback.print_exc()
            error_msg = str(e)
            if 'UNIQUE constraint' in error_msg:
                error_msg = 'J� existe um fornecedor com este nome'
            return jsonify({'success': False, 'error': error_msg}), 400
    
    elif request.method == 'DELETE':
        try:
            # Validar propriedade antes de deletar (se n�o for admin)
            if filtro_cliente_id is not None:
                fornecedor_atual = db.obter_fornecedor_por_nome(nome)
                if not fornecedor_atual or fornecedor_atual.get('empresa_id') != filtro_cliente_id:
                    return jsonify({'success': False, 'error': 'Fornecedor n�o encontrado ou sem permiss�o'}), 403
            
            success, mensagem = db.excluir_fornecedor(nome)
            if success:
                return jsonify({'success': True, 'message': mensagem})
            else:
                return jsonify({'success': False, 'error': mensagem}), 400
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/clientes/<path:nome>/inativar', methods=['POST'])
@require_permission('clientes_edit')
def inativar_cliente(nome):
    """Inativa um cliente com motivo"""
    try:
        # Decode do nome que vem URL-encoded
        from urllib.parse import unquote
        nome = unquote(nome)
        
        data = request.json or {}
        motivo = data.get('motivo', 'Inativado pelo usu�rio')
        
        success, mensagem = db.inativar_cliente(nome, motivo)
        return jsonify({'success': success, 'message': mensagem})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/clientes/<path:nome>/reativar', methods=['POST'])
@require_permission('clientes_edit')
def reativar_cliente(nome):
    """Reativa um cliente"""
    try:
        # Decode do nome que vem URL-encoded
        from urllib.parse import unquote
        nome = unquote(nome)
        
        success = db.reativar_cliente(nome)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/fornecedores/<path:nome>/inativar', methods=['POST'])
@require_permission('fornecedores_edit')
def inativar_fornecedor(nome):
    """Inativa um fornecedor com motivo"""
    try:
        # Decode do nome que vem URL-encoded
        from urllib.parse import unquote
        nome = unquote(nome)
        
        data = request.json or {}
        motivo = data.get('motivo', 'Inativado pelo usu�rio')
        
        success, mensagem = db.inativar_fornecedor(nome, motivo)
        return jsonify({'success': success, 'message': mensagem})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/fornecedores/<path:nome>/reativar', methods=['POST'])
@require_permission('fornecedores_edit')
def reativar_fornecedor(nome):
    """Reativa um fornecedor"""
    try:
        # Decode do nome que vem URL-encoded
        from urllib.parse import unquote
        nome = unquote(nome)
        
        success = db.reativar_fornecedor(nome)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# === ROTAS DE LAN�AMENTOS ===

@app.route('/api/lancamentos', methods=['GET'])
@require_permission('lancamentos_view')
@aplicar_filtro_cliente
def listar_lancamentos():
    """Lista todos os lan�amentos com filtro de multi-tenancy e pagina��o"""
    try:
        print("\n" + "="*80)
        print("?? ROTA /api/lancamentos chamada")
        
        # Obter empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        
        # Par�metros de filtro
        tipo_filtro = request.args.get('tipo')
        status_filtro = request.args.get('status')
        categoria_filtro = request.args.get('categoria')
        subcategoria_filtro = request.args.get('subcategoria')
        conta_bancaria_filtro = request.args.get('conta_bancaria')
        data_inicio_filtro = request.args.get('data_inicio')
        data_fim_filtro = request.args.get('data_fim')
        ano_filtro = request.args.get('ano')
        mes_filtro = request.args.get('mes')
        search_filtro = request.args.get('search')
        cliente_filtro = request.args.get('cliente')
        fornecedor_filtro = request.args.get('fornecedor')
        filtro_cliente_id = getattr(request, 'filtro_cliente_id', None)
        
        # Par�metros de pagina��o
        page = request.args.get('page', type=int)
        per_page = request.args.get('per_page', default=300, type=int)
        per_page = min(per_page, 300)  # M�ximo de 300 registros por p�gina
        
        print(f"?? Par�metros recebidos:")
        print(f"   - empresa_id: {empresa_id}")
        print(f"   - tipo_filtro: {tipo_filtro}")
        print(f"   - status_filtro: {status_filtro}")
        print(f"   - categoria_filtro: {categoria_filtro}")
        print(f"   - conta_bancaria_filtro: {conta_bancaria_filtro}")
        print(f"   - data_inicio_filtro: {data_inicio_filtro}")
        print(f"   - data_fim_filtro: {data_fim_filtro}")
        print(f"   - ano_filtro: {ano_filtro}")
        print(f"   - mes_filtro: {mes_filtro}")
        print(f"   - search_filtro: {search_filtro}")
        print(f"   - cliente_filtro: {cliente_filtro}")
        print(f"   - fornecedor_filtro: {fornecedor_filtro}")
        print(f"   - filtro_cliente_id: {filtro_cliente_id}")
        print(f"   - page: {page}")
        print(f"   - per_page: {per_page}")
        
        # Criar dicion�rio de filtros
        filtros = {}
        if tipo_filtro:
            filtros['tipo'] = tipo_filtro.upper()
        if status_filtro:
            filtros['status'] = status_filtro.upper()
        if categoria_filtro:
            filtros['categoria'] = categoria_filtro
        if subcategoria_filtro:
            filtros['subcategoria'] = subcategoria_filtro
        if conta_bancaria_filtro:
            filtros['conta_bancaria'] = conta_bancaria_filtro
        if data_inicio_filtro:
            filtros['data_inicio'] = data_inicio_filtro
        if data_fim_filtro:
            filtros['data_fim'] = data_fim_filtro
        
        # Filtros especiais: ano e m�s (converter para data_inicio/data_fim)
        if ano_filtro:
            from datetime import date
            ano = int(ano_filtro)
            if mes_filtro:
                mes = int(mes_filtro)
                # Filtrar por m�s espec�fico
                import calendar
                ultimo_dia = calendar.monthrange(ano, mes)[1]
                filtros['data_inicio'] = date(ano, mes, 1).isoformat()
                filtros['data_fim'] = date(ano, mes, ultimo_dia).isoformat()
                print(f"   ??? Filtro ano+m�s: {filtros['data_inicio']} at� {filtros['data_fim']}")
            else:
                # Filtrar por ano inteiro
                filtros['data_inicio'] = date(ano, 1, 1).isoformat()
                filtros['data_fim'] = date(ano, 12, 31).isoformat()
                print(f"   ??? Filtro ano: {filtros['data_inicio']} at� {filtros['data_fim']}")
        
        # Filtro de busca textual (search) - ser� aplicado ap�s consulta
        # Filtro de cliente/fornecedor - ser� aplicado ap�s consulta
        
        print(f"?? Filtros montados: {filtros}")
        
        # Chamar m�todo com todos os par�metros
        print(f"?? Chamando database.listar_lancamentos()...")
        lancamentos = database.listar_lancamentos(
            empresa_id=empresa_id,
            filtros=filtros,
            filtro_cliente_id=filtro_cliente_id,
            page=page,
            per_page=per_page
        )
        
        print(f"? Retornaram {len(lancamentos)} lan�amentos")
        
        # VERIFICAR DUPLICATAS
        if lancamentos:
            ids = [l.id for l in lancamentos if hasattr(l, 'id')]
            ids_unicos = set(ids)
            if len(ids) != len(ids_unicos):
                duplicados = [id for id in ids if ids.count(id) > 1]
                print(f"?? ATEN��O: QUERY RETORNOU IDs DUPLICADOS!")
                print(f"   Total IDs: {len(ids)}, �nicos: {len(ids_unicos)}")
                print(f"   IDs duplicados: {set(duplicados)}")
            else:
                print(f"? Todos os IDs s�o �nicos ({len(ids_unicos)} registros)")
        
        # Converter para lista de dicts
        lancamentos_list = []
        for idx, l in enumerate(lancamentos):
            try:
                item = {
                    'id': l.id if hasattr(l, 'id') else None,
                    'tipo': l.tipo.value if hasattr(l.tipo, 'value') else str(l.tipo),
                    'descricao': l.descricao,
                    'valor': float(l.valor),
                    'data_vencimento': l.data_vencimento.isoformat() if l.data_vencimento else None,
                    'data_pagamento': l.data_pagamento.isoformat() if l.data_pagamento else None,
                    'status': l.status.value if hasattr(l.status, 'value') else str(l.status),
                    'categoria': l.categoria,
                    'subcategoria': l.subcategoria,
                    'conta_bancaria': l.conta_bancaria,
                    'pessoa': l.pessoa,
                    'observacoes': l.observacoes,
                    'num_documento': getattr(l, 'num_documento', ''),
                    'associacao': getattr(l, 'associacao', ''),
                    'numero_documento': getattr(l, 'numero_documento', ''),
                    'recorrente': getattr(l, 'recorrente', False),
                    'frequencia_recorrencia': getattr(l, 'frequencia_recorrencia', ''),
                    'cliente_id': getattr(l, 'pessoa', None)
                }
                lancamentos_list.append(item)
            except Exception as e:
                print(f"?? Erro ao converter lan�amento {idx} (ID: {getattr(l, 'id', '?')}): {e}")
                continue
        
        # Aplicar filtros adicionais (search, cliente, fornecedor) em mem�ria
        if search_filtro:
            search_lower = search_filtro.lower()
            lancamentos_list = [
                l for l in lancamentos_list 
                if (search_lower in (l.get('descricao') or '').lower() or
                    search_lower in (l.get('pessoa') or '').lower() or
                    search_lower in (l.get('observacoes') or '').lower())
            ]
            print(f"?? Ap�s filtro search: {len(lancamentos_list)} registros")
        
        if cliente_filtro:
            lancamentos_list = [l for l in lancamentos_list if l.get('pessoa') == cliente_filtro]
            print(f"?? Ap�s filtro cliente: {len(lancamentos_list)} registros")
        
        if fornecedor_filtro:
            lancamentos_list = [l for l in lancamentos_list if l.get('pessoa') == fornecedor_filtro]
            print(f"?? Ap�s filtro fornecedor: {len(lancamentos_list)} registros")
        
        print(f"?? Retornando {len(lancamentos_list)} lan�amentos no JSON")
        print("="*80 + "\n")
        
        return jsonify({
            'success': True,
            'data': lancamentos_list,
            'total': len(lancamentos_list),
            'message': 'Nenhum lan�amento encontrado' if len(lancamentos_list) == 0 else None
        })
    except Exception as e:
        print(f"? ERRO CR�TICO em listar_lancamentos: {e}")
        import traceback
        traceback.print_exc()
        print("="*80 + "\n")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos', methods=['POST'])
@require_permission('lancamentos_create')
@aplicar_filtro_cliente
def adicionar_lancamento():
    """Adiciona um novo lan�amento (com suporte a parcelamento)"""
    try:
        # ?? VALIDA��O DE SEGURAN�A
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.json
        
        # ?? Obter proprietario_id do usu�rio logado (ID na tabela usuarios)
        usuario = get_usuario_logado()
        proprietario_id = None
        if usuario and usuario.get('tipo') == 'cliente':
            proprietario_id = usuario.get('id')  # ID do usu�rio, N�O empresa_id
        
        print(f"\n?? [POST /api/lancamentos]")
        print(f"   - empresa_id: {empresa_id}")
        print(f"   - usuario.id: {usuario.get('id') if usuario else None}")
        print(f"   - usuario.tipo: {usuario.get('tipo') if usuario else None}")
        print(f"   - proprietario_id: {proprietario_id}")
        
        # Validar se a conta banc�ria est� ativa
        if data and data.get('conta_bancaria'):
            conta_nome = data['conta_bancaria']
            try:
                # Buscar conta diretamente do banco de dados
                conn = db.get_connection()
                cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                cursor.execute(
                    "SELECT nome, ativa FROM contas_bancarias WHERE nome = %s AND empresa_id = %s",
                    (conta_nome, empresa_id)
                )
                conta = cursor.fetchone()
                cursor.close()
                from database_postgresql import return_to_pool
                return_to_pool(conn)
                
                if conta and 'ativa' in conta and not conta['ativa']:
                    print(f"? Tentativa de criar lan�amento em conta inativa: {conta_nome}")
                    return jsonify({
                        'success': False,
                        'error': f'N�o � poss�vel criar lan�amento. A conta banc�ria "{conta_nome}" est� inativa. Reative a conta antes de criar novos lan�amentos.'
                    }), 400
            except Exception as e:
                print(f"?? Erro ao validar conta banc�ria: {e}")
                # Continuar mesmo se a valida��o falhar (n�o bloquear cria��o)
        
        parcelas = int(data.get('parcelas', 1)) if data else 1
        
        if parcelas > 1:
            # Criar m�ltiplos lan�amentos para parcelas
            from dateutil.relativedelta import relativedelta  # type: ignore
            data_base = datetime.fromisoformat(data['data_vencimento']) if data and data.get('data_vencimento') else datetime.now()
            valor_parcela = float(data['valor']) / parcelas if data else 0.0
            tipo_str = data['tipo'].lower() if data and data.get('tipo') else 'receita'
            
            lancamentos_ids = []
            for i in range(parcelas):
                data_venc = data_base + relativedelta(months=i)
                descricao_parcela = f"{data['descricao']} ({i+1}/{parcelas})" if data else f"Parcela {i+1}/{parcelas}"
                
                lancamento = Lancamento(
                    descricao=descricao_parcela,
                    valor=valor_parcela,
                    tipo=TipoLancamento(tipo_str),
                    categoria=data.get('categoria', '') if data else '',
                    data_vencimento=data_venc,
                    data_pagamento=None,
                    conta_bancaria=data['conta_bancaria'] if data else '',
                    pessoa=data.get('pessoa', '') if data else '',
                    observacoes=data.get('observacoes', '') if data else '',
                    subcategoria=data.get('subcategoria', '') if data else '',
                    associacao=data.get('associacao', '') if data else ''
                )
                
                if data and data.get('status'):
                    lancamento.status = StatusLancamento(data['status'])
                
                lancamento_id = db.adicionar_lancamento(lancamento, proprietario_id=proprietario_id, empresa_id=empresa_id)
                lancamentos_ids.append(lancamento_id)
            
            print(f"Lan�amentos parcelados adicionados! IDs: {lancamentos_ids}")
            return jsonify({'success': True, 'ids': lancamentos_ids})
        else:
            # Lan�amento �nico (sem parcelamento)
            data_venc = datetime.fromisoformat(data['data_vencimento']) if data and data.get('data_vencimento') else datetime.now()
            data_pag = datetime.fromisoformat(data['data_pagamento']) if data and data.get('data_pagamento') else None
            tipo_str = data['tipo'].lower() if data and data.get('tipo') else 'receita'
            
            lancamento = Lancamento(
                descricao=data['descricao'] if data else '',
                valor=float(data['valor']) if data else 0.0,
                tipo=TipoLancamento(tipo_str),
                categoria=data.get('categoria', '') if data else '',
                data_vencimento=data_venc,
                data_pagamento=data_pag,
                conta_bancaria=data['conta_bancaria'] if data else '',
                pessoa=data.get('pessoa', '') if data else '',
                observacoes=data.get('observacoes', '') if data else '',
                subcategoria=data.get('subcategoria', '') if data else '',
                associacao=data.get('associacao', '') if data else ''
            )
            
            if data and data.get('status'):
                lancamento.status = StatusLancamento(data['status'])
            
            lancamento_id = db.adicionar_lancamento(lancamento, proprietario_id=proprietario_id, empresa_id=empresa_id)
            return jsonify({'success': True, 'id': lancamento_id})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/lancamentos/corrigir-tipos-conciliacao', methods=['POST'])
@require_permission('lancamentos_edit')
def corrigir_tipos_lancamentos_conciliacao():
    """
    Corrige automaticamente o tipo (receita/despesa) de lançamentos conciliados
    que divergem do tipo registrado na transação do extrato.

    Regra:
        extrato DÉBITO  → lançamento deve ser 'despesa'
        extrato CRÉDITO → lançamento deve ser 'receita'

    Retorna:
        { corrigidos: int }  — número de lançamentos atualizados
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa não identificada'}), 403

        import psycopg2.extras
        with database.get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE lancamentos l
                SET tipo = CASE
                    WHEN te.tipo ILIKE '%DEB%' THEN 'despesa'
                    WHEN te.tipo ILIKE '%CRE%' THEN 'receita'
                    WHEN te.tipo ILIKE '%CRÉ%' THEN 'receita'
                END
                FROM conciliacoes c
                JOIN transacoes_extrato te
                     ON te.id = c.transacao_extrato_id
                    AND te.empresa_id = c.empresa_id
                WHERE c.lancamento_id   = l.id
                  AND c.empresa_id       = l.empresa_id
                  AND l.empresa_id       = %s
                  AND (
                    (te.tipo ILIKE '%%DEB%%' AND l.tipo = 'receita')
                    OR
                    (te.tipo ILIKE '%%CRE%%' AND l.tipo = 'despesa')
                    OR
                    (te.tipo ILIKE '%%CRÉ%%' AND l.tipo = 'despesa')
                  )
            """, (empresa_id,))
            corrigidos = cursor.rowcount
            conn.commit()
            cursor.close()

        if corrigidos > 0:
            logger.info(f"🔧 corrigir_tipos_conciliacao: {corrigidos} lançamento(s) corrigido(s) para empresa {empresa_id}")

        return jsonify({'success': True, 'corrigidos': corrigidos}), 200

    except Exception as e:
        logger.error(f"Erro ao corrigir tipos: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos/<int:lancamento_id>', methods=['GET'])
@aplicar_filtro_cliente
@require_permission('lancamentos_view')
def obter_lancamento_route(lancamento_id):
    """Retorna os dados de um lan�amento espec�fico"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        print(f"\n{'='*80}")
        print(f"?? GET /api/lancamentos/{lancamento_id}")
        print(f"{'='*80}")
        
        lancamento = db_obter_lancamento(empresa_id=empresa_id, lancamento_id=lancamento_id)
        print(f"Resultado db_obter_lancamento: {lancamento}")
        print(f"Tipo: {type(lancamento)}")
        
        if lancamento:
            # Converter Lancamento para dict
            lancamento_dict = {
                'id': lancamento.id,
                'tipo': lancamento.tipo.value if hasattr(lancamento.tipo, 'value') else str(lancamento.tipo),
                'descricao': lancamento.descricao,
                'valor': float(lancamento.valor),
                'data_vencimento': lancamento.data_vencimento.isoformat() if lancamento.data_vencimento else None,
                'data_pagamento': lancamento.data_pagamento.isoformat() if lancamento.data_pagamento else None,
                'categoria': lancamento.categoria,
                'subcategoria': lancamento.subcategoria,
                'conta_bancaria': lancamento.conta_bancaria,
                'cliente_fornecedor': lancamento.cliente_fornecedor,
                'pessoa': lancamento.pessoa,
                'status': lancamento.status.value if hasattr(lancamento.status, 'value') else str(lancamento.status),
                'observacoes': lancamento.observacoes,
                'anexo': lancamento.anexo,
                'recorrente': lancamento.recorrente,
                'frequencia_recorrencia': lancamento.frequencia_recorrencia,
                'dia_vencimento': lancamento.dia_vencimento,
                'juros': float(getattr(lancamento, 'juros', 0)),
                'desconto': float(getattr(lancamento, 'desconto', 0)),
                'associacao': getattr(lancamento, 'associacao', ''),
                'numero_documento': getattr(lancamento, 'numero_documento', '')
            }
            print(f"? Lan�amento convertido para dict: {lancamento_dict}")
            print(f"{'='*80}\n")
            return jsonify(lancamento_dict), 200
        else:
            print(f"? Lan�amento n�o encontrado")
            print(f"{'='*80}\n")
            return jsonify({'error': 'Lan�amento n�o encontrado'}), 404
    except Exception as e:
        print(f"? ERRO ao obter lan�amento:")
        print(f"   Tipo: {type(e).__name__}")
        print(f"   Mensagem: {str(e)}")
        import traceback
        print(f"   Traceback:\n{traceback.format_exc()}")
        print(f"{'='*80}\n")
        return jsonify({'error': str(e)}), 500


@app.route('/api/lancamentos/<int:lancamento_id>', methods=['PUT', 'DELETE', 'OPTIONS'])
@require_permission('lancamentos_edit')
def gerenciar_lancamento(lancamento_id):
    """Atualiza ou remove um lan�amento"""
    if request.method == 'OPTIONS':
        return jsonify({'success': True}), 200
    
    if request.method == 'PUT':
        try:
            print(f"\n{'='*80}")
            print(f"?? PUT /api/lancamentos/{lancamento_id}")
            print(f"{'='*80}")
            
            data = request.get_json()
            print(f"?? Dados recebidos: {data}")
            print(f"?? CAMPO CR�TICO - associacao recebido: '{data.get('associacao', '')}' (tipo: {type(data.get('associacao', ''))})")
            
            # ?? VALIDA��O DE SEGURAN�A
            empresa_id = session.get('empresa_id')
            if not empresa_id:
                return jsonify({'erro': 'Empresa n�o selecionada'}), 403
            
            # Verificar se lan�amento existe
            lancamento_atual = db_obter_lancamento(empresa_id=empresa_id, lancamento_id=lancamento_id)
            if not lancamento_atual:
                print("? Lan�amento n�o encontrado")
                return jsonify({'success': False, 'error': 'Lan�amento n�o encontrado'}), 404
            
            # Preservar dados de pagamento se j� foi pago
            status_atual = lancamento_atual.status.value if hasattr(lancamento_atual.status, 'value') else str(lancamento_atual.status)
            data_pgto_atual = lancamento_atual.data_pagamento
            conta_bancaria_atual = lancamento_atual.conta_bancaria
            juros_atual = getattr(lancamento_atual, 'juros', 0)
            desconto_atual = getattr(lancamento_atual, 'desconto', 0)
            
            print(f"?? Preservando dados de pagamento:")
            print(f"   - Status: {status_atual}")
            print(f"   - Data pagamento: {data_pgto_atual}")
            print(f"   - Conta: {conta_bancaria_atual}")
            
            # Criar objeto Lancamento atualizado
            from database_postgresql import TipoLancamento, StatusLancamento
            from decimal import Decimal
            
            tipo_enum = TipoLancamento(data['tipo'].lower())
            status_enum = StatusLancamento(status_atual.lower()) if status_atual else StatusLancamento.PENDENTE
            
            lancamento_atualizado = Lancamento(
                id=lancamento_id,
                tipo=tipo_enum,
                descricao=data.get('descricao', ''),
                valor=Decimal(str(data['valor'])),
                data_vencimento=datetime.fromisoformat(data['data_vencimento']),
                data_pagamento=data_pgto_atual,
                categoria=data.get('categoria', ''),
                subcategoria=data.get('subcategoria', ''),
                conta_bancaria=conta_bancaria_atual,
                cliente_fornecedor=data.get('cliente_fornecedor', ''),
                pessoa=data.get('pessoa', ''),
                status=status_enum,
                observacoes=data.get('observacoes', ''),
                anexo=data.get('anexo', ''),
                recorrente=data.get('recorrente', False),
                frequencia_recorrencia=data.get('frequencia_recorrencia', ''),
                dia_vencimento=data.get('dia_vencimento', 0),
                juros=juros_atual,
                desconto=desconto_atual,
                associacao=data.get('associacao', '')
            )
            
            print(f"? Objeto Lancamento criado:")
            print(f"   - ID: {lancamento_atualizado.id}")
            print(f"   - associacao: '{lancamento_atualizado.associacao}' (tipo: {type(lancamento_atualizado.associacao)})")
            
            # Atualizar no banco COM empresa_id para RLS
            success = db.atualizar_lancamento(lancamento_atualizado, empresa_id=empresa_id)
            
            print(f"? Resultado: {success}")
            print(f"{'='*80}\n")
            
            if success:
                return jsonify({'success': True, 'id': lancamento_id})
            else:
                return jsonify({'success': False, 'error': 'Falha ao atualizar'}), 400
            
        except Exception as e:
            print(f"? ERRO ao atualizar lan�amento:")
            print(f"   Tipo: {type(e).__name__}")
            print(f"   Mensagem: {str(e)}")
            import traceback
            print(f"   Traceback:\n{traceback.format_exc()}")
            print(f"{'='*80}\n")
            return jsonify({'success': False, 'error': str(e)}), 400
    
    # DELETE
    try:
        print(f"\n=== Excluindo lan�amento ID: {lancamento_id} ===")
        
        # ?? VALIDA��O DE SEGURAN�A
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'erro': 'Empresa n�o selecionada'}), 403
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            # 1. Encontrar transacoes_extrato vinculadas via conciliacoes
            cursor.execute(
                "SELECT transacao_extrato_id FROM conciliacoes WHERE lancamento_id = %s AND empresa_id = %s",
                (lancamento_id, empresa_id)
            )
            transacao_ids = [row['transacao_extrato_id'] for row in cursor.fetchall()]
            print(f"  Transações de extrato vinculadas: {transacao_ids}")

            # 1.5. Registrar no historico ANTES de deletar (lancamento ainda existe)
            _garantir_tabela_historico_conciliacoes(conn)
            for _tid in transacao_ids:
                _inserir_historico_conciliacao(conn, empresa_id, 'desconciliado', _tid, lancamento_id)

            # 2. Remover registros de conciliação
            cursor.execute(
                "DELETE FROM conciliacoes WHERE lancamento_id = %s AND empresa_id = %s",
                (lancamento_id, empresa_id)
            )
            print(f"  Conciliações removidas: {cursor.rowcount}")

            # 3. Desmarcar transacoes_extrato como conciliadas
            if transacao_ids:
                cursor.execute(
                    "UPDATE transacoes_extrato SET conciliado = FALSE WHERE id = ANY(%s) AND empresa_id = %s",
                    (transacao_ids, empresa_id)
                )
                print(f"  Extratos desconciliados: {cursor.rowcount}")

            # 4. Excluir o lançamento
            cursor.execute(
                "DELETE FROM lancamentos WHERE id = %s AND empresa_id = %s",
                (lancamento_id, empresa_id)
            )
            success = cursor.rowcount > 0
            cursor.close()

        print(f"Resultado da exclusão: {success}")

        if not success:
            print("AVISO: Nenhum registro foi excluído (ID não encontrado?)")
            return jsonify({'success': False, 'error': 'Lançamento não encontrado'}), 404

        print("Lançamento excluído com sucesso!")
        return jsonify({'success': True})
    except Exception as e:
        print(f"ERRO ao excluir lançamento: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/lancamentos/bulk-delete', methods=['POST'])
@limiter.exempt
@require_permission('lancamentos_delete')
def bulk_delete_lancamentos():
    """
    Exclui múltiplos lançamentos de uma vez (sem rate-limit).
    Limpa conciliacoes e transacoes_extrato.conciliado em cascata.

    Body JSON:
        ids: lista de IDs (int)
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa não selecionada'}), 403

        dados = request.get_json(force=True) or {}
        ids = [int(i) for i in (dados.get('ids') or []) if str(i).isdigit()]

        if not ids:
            return jsonify({'success': False, 'error': 'Nenhum ID fornecido'}), 400
        if len(ids) > 500:
            return jsonify({'success': False, 'error': 'Máximo 500 IDs por requisição'}), 400

        print(f"\n=== Bulk delete: {len(ids)} lançamentos ===")

        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()

            # 1. Encontrar transacoes_extrato vinculadas
            cursor.execute(
                "SELECT transacao_extrato_id FROM conciliacoes "
                "WHERE lancamento_id = ANY(%s) AND empresa_id = %s",
                (ids, empresa_id)
            )
            transacao_ids = [row['transacao_extrato_id'] for row in cursor.fetchall()]

            # 2. Registrar no historico ANTES de deletar
            _garantir_tabela_historico_conciliacoes(conn)
            for _lid in ids:
                cursor.execute(
                    "SELECT transacao_extrato_id FROM conciliacoes "
                    "WHERE lancamento_id = %s AND empresa_id = %s",
                    (_lid, empresa_id)
                )
                for _row in cursor.fetchall():
                    _inserir_historico_conciliacao(
                        conn, empresa_id, 'desconciliado', _row['transacao_extrato_id'], _lid
                    )

            # 3. Remover conciliacoes
            cursor.execute(
                "DELETE FROM conciliacoes WHERE lancamento_id = ANY(%s) AND empresa_id = %s",
                (ids, empresa_id)
            )
            print(f"  Conciliações removidas: {cursor.rowcount}")

            # 4. Desmarcar transacoes_extrato
            if transacao_ids:
                cursor.execute(
                    "UPDATE transacoes_extrato SET conciliado = FALSE "
                    "WHERE id = ANY(%s) AND empresa_id = %s",
                    (transacao_ids, empresa_id)
                )
                print(f"  Extratos desconciliados: {cursor.rowcount}")

            # 5. Excluir os lançamentos
            cursor.execute(
                "DELETE FROM lancamentos WHERE id = ANY(%s) AND empresa_id = %s",
                (ids, empresa_id)
            )
            deleted = cursor.rowcount
            cursor.close()

        print(f"  Lançamentos excluídos: {deleted}")
        return jsonify({'success': True, 'deleted': deleted})

    except Exception as e:
        print(f"ERRO bulk-delete: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400


# ============================================================================
# ROTAS DE EXTRATO BANCARIO (IMPORTACAO OFX)
# ============================================================================

# ---- Ajuste de OFX: CRUD de filtros de MEMO --------------------------------

@app.route('/api/ofx-filtros', methods=['GET'])
@require_permission('lancamentos_view')
def listar_ofx_filtros():
    """Lista filtros de MEMO configurados para a empresa atual"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute(
                "SELECT id, conta_bancaria, memo_filtro, criado_em FROM ofx_filtros_memo "
                "WHERE empresa_id = %s ORDER BY conta_bancaria, memo_filtro",
                (empresa_id,)
            )
            filtros = cursor.fetchall()
            cursor.close()
        return jsonify({'success': True, 'filtros': [dict(f) for f in filtros]})
    except Exception as e:
        logger.error(f"Erro ao listar ofx_filtros: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ofx-filtros', methods=['POST'])
@require_permission('lancamentos_edit')
def criar_ofx_filtro():
    """Cria um novo filtro de MEMO para a empresa atual"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        data = request.json or {}
        conta_bancaria = (data.get('conta_bancaria') or '').strip()
        memo_filtro = (data.get('memo_filtro') or '').strip()
        if not conta_bancaria:
            return jsonify({'success': False, 'error': 'Conta banc�ria obrigat�ria'}), 400
        if not memo_filtro:
            return jsonify({'success': False, 'error': 'Texto do MEMO obrigat�rio'}), 400
        with db.get_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute(
                "INSERT INTO ofx_filtros_memo (empresa_id, conta_bancaria, memo_filtro) "
                "VALUES (%s, %s, %s) ON CONFLICT DO NOTHING RETURNING id",
                (empresa_id, conta_bancaria, memo_filtro)
            )
            row = cursor.fetchone()
            conn.commit()
            cursor.close()
        if row:
            return jsonify({'success': True, 'id': row['id'], 'message': 'Filtro criado com sucesso'})
        else:
            return jsonify({'success': False, 'error': 'Filtro j� existe para esta conta/MEMO'}), 409
    except Exception as e:
        logger.error(f"Erro ao criar ofx_filtro: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ofx-filtros/<int:filtro_id>', methods=['DELETE'])
@require_permission('lancamentos_edit')
def deletar_ofx_filtro(filtro_id):
    """Deleta um filtro de MEMO"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "DELETE FROM ofx_filtros_memo WHERE id = %s AND empresa_id = %s",
                (filtro_id, empresa_id)
            )
            deleted = cursor.rowcount
            conn.commit()
            cursor.close()
        if deleted:
            return jsonify({'success': True, 'message': 'Filtro removido com sucesso'})
        else:
            return jsonify({'success': False, 'error': 'Filtro n�o encontrado'}), 404
    except Exception as e:
        logger.error(f"Erro ao deletar ofx_filtro: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ---- Fim Ajuste de OFX -------------------------------------------------------

import extrato_functions

@app.route('/api/extratos/upload', methods=['POST'])
@require_permission('lancamentos_edit')
def upload_extrato_ofx():
    """Upload e processamento de arquivo OFX"""
    try:
        print(f"\n{'='*60}")
        print(f"?? UPLOAD DE EXTRATO OFX INICIADO")
        print(f"{'='*60}")
        
        # Log dos arquivos recebidos
        print(f"?? Arquivos em request.files: {list(request.files.keys())}")
        print(f"?? Dados em request.form: {dict(request.form)}")
        
        if 'file' not in request.files:
            print(f"? Erro: Nenhum arquivo enviado")
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        conta_bancaria = request.form.get('conta_bancaria')
        
        print(f"?? Arquivo: {file.filename}")
        print(f"?? Conta banc�ria: {conta_bancaria}")
        
        if not conta_bancaria:
            print(f"? Erro: Conta banc�ria n�o informada")
            return jsonify({'success': False, 'error': 'Conta bancaria e obrigatoria'}), 400
        
        if file.filename == '':
            print(f"? Erro: Nome do arquivo vazio")
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.lower().endswith('.ofx'):
            print(f"? Erro: Extens�o inv�lida: {file.filename}")
            return jsonify({'success': False, 'error': 'Apenas arquivos .ofx sao permitidos'}), 400
        
        # Buscar informa��es da conta banc�ria cadastrada
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o (empresa selecionada)
        empresa_id = session.get('empresa_id')
        
        if not empresa_id:
            print(f"? Erro: Empresa n�o identificada na sess�o")
            return jsonify({'success': False, 'error': 'Empresa n�o identificada. Fa�a login novamente.'}), 403
        
        print(f"?? EMPRESA ATUAL (sess�o): {empresa_id}")
        print(f"?? Transa��es ser�o salvas APENAS para empresa: {empresa_id}")
        
        # ?? Buscar APENAS contas da empresa atual (isolamento multi-tenant)
        from database_postgresql import DatabaseManager
        db_manager = DatabaseManager()
        
        try:
            contas_cadastradas = db_manager.listar_contas_por_empresa(empresa_id=empresa_id)
            print(f"?? Total de contas da empresa {empresa_id}: {len(contas_cadastradas)}")
            print(f"?? Nomes das contas: {[c.nome for c in contas_cadastradas]}")
        except Exception as e:
            print(f"? Erro ao buscar contas da empresa {empresa_id}: {e}")
            return jsonify({'success': False, 'error': f'Erro ao buscar contas banc�rias: {str(e)}'}), 500
        
        conta_info = next((c for c in contas_cadastradas if c.nome == conta_bancaria), None)
        
        if not conta_info:
            print(f"? Erro: Conta '{conta_bancaria}' n�o encontrada na lista")
            return jsonify({'success': False, 'error': f'Conta banc�ria "{conta_bancaria}" n�o encontrada'}), 400
        
        print(f"? Conta encontrada: {conta_info.nome}")
        
        # Validar se a conta est� ativa
        if hasattr(conta_info, 'ativa') and not conta_info.ativa:
            print(f"? Tentativa de importar extrato para conta inativa: {conta_bancaria}")
            return jsonify({
                'success': False,
                'error': f'N�o � poss�vel importar extrato. A conta banc�ria "{conta_bancaria}" est� inativa. Reative a conta antes de importar extratos.'
            }), 400
        
        print(f"? Conta est� ativa, prosseguindo com o upload...")
        
        # Parse OFX � ler bytes brutos e recodificar para evitar erros de charset
        # Arquivos OFX de bancos brasileiros costumam usar CP1252/ISO-8859-1 com
        # bytes invalidos (ex: 0x8d que nao existe no charmap/cp1252).
        # Estrategia: latin-1 aceita TODOS os bytes 0x00-0xFF garantidamente,
        # depois recodifica para UTF-8 limpo e normaliza o header CHARSET do OFX.
        try:
            import ofxparse
            import io
            import re as _re

            raw_bytes = file.read()

            # latin-1 (ISO-8859-1) mapeia 1:1 bytes ? unicode para 0x00-0xFF, nunca falha
            decoded_content = raw_bytes.decode('latin-1', errors='replace')

            # Normalizar header SGML do OFX para evitar que ofxparse tente
            # redecodificar com charset errado (ex: CHARSET:1252 ? CHARSET:0 / UTF-8)
            # O header fica antes da tag <OFX> e tem linhas KEY:VALUE
            decoded_content = _re.sub(
                r'(?im)^(CHARSET\s*:\s*).*$', r'\g<1>0', decoded_content
            )
            decoded_content = _re.sub(
                r'(?im)^(ENCODING\s*:\s*).*$', r'\g<1>UTF-8', decoded_content
            )
            # Remover caracteres de controle invis�veis que quebram o parser
            # (manter \t, \n, \r que s�o v�lidos em XML/SGML)
            decoded_content = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', decoded_content)

            # Recodificar para UTF-8 limpo e passar ao ofxparse via BytesIO
            clean_bytes = decoded_content.encode('utf-8', errors='replace')
            ofx = ofxparse.OfxParser.parse(io.BytesIO(clean_bytes))
        except Exception as e:
            return jsonify({'success': False, 'error': f'Erro ao processar OFX: {str(e)}'}), 400
        
        # ?? VALIDA��O: Verificar se j� existe importa��o no per�odo (ANTES de processar)
        try:
            for account in ofx.accounts:
                # Extrair per�odo do OFX (ignorar timezone para evitar bugs)
                start_date_ofx = account.statement.start_date
                end_date_ofx = account.statement.end_date
                
                # Converter para date (ignorar timezone)
                if hasattr(start_date_ofx, 'year'):
                    periodo_inicio = date(start_date_ofx.year, start_date_ofx.month, start_date_ofx.day)
                elif hasattr(start_date_ofx, 'date'):
                    periodo_inicio = start_date_ofx.date()
                else:
                    periodo_inicio = start_date_ofx
                
                if hasattr(end_date_ofx, 'year'):
                    periodo_fim = date(end_date_ofx.year, end_date_ofx.month, end_date_ofx.day)
                elif hasattr(end_date_ofx, 'date'):
                    periodo_fim = end_date_ofx.date()
                else:
                    periodo_fim = end_date_ofx
                
                print(f"\n?? VALIDANDO PER�ODO: {periodo_inicio} at� {periodo_fim}")
                
                # Consultar per�odos j� importados para esta conta/empresa
                with db.get_connection() as conn:
                    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                    
                    # ?? DIAGN�STICO: Contar transa��es �rf�s (sem importacao_id)
                    cursor.execute("""
                        SELECT COUNT(*) as total_orfas
                        FROM transacoes_extrato
                        WHERE empresa_id = %s 
                        AND conta_bancaria = %s 
                        AND (importacao_id IS NULL OR importacao_id = '')
                    """, (empresa_id, conta_bancaria))
                    
                    resultado_orfas = cursor.fetchone()
                    total_orfas = resultado_orfas['total_orfas'] if resultado_orfas else 0
                    
                    if total_orfas > 0:
                        print(f"?? ATEN��O: {total_orfas} transa��es �rf�s detectadas (sem importacao_id)")
                        logger.warning(f"Conta {conta_bancaria}: {total_orfas} transa��es sem importacao_id")
                    
                    # Buscar apenas importa��es v�lidas (com importacao_id preenchido)
                    cursor.execute("""
                        SELECT 
                            importacao_id,
                            MIN(data) as inicio,
                            MAX(data) as fim,
                            COUNT(*) as qtd_transacoes
                        FROM transacoes_extrato
                        WHERE empresa_id = %s 
                        AND conta_bancaria = %s
                        AND importacao_id IS NOT NULL 
                        AND importacao_id != ''
                        GROUP BY importacao_id
                        HAVING COUNT(*) > 0
                        ORDER BY MAX(data) DESC
                    """, (empresa_id, conta_bancaria))
                    
                    periodos_existentes = cursor.fetchall()
                    
                    print(f"?? Total de importa��es encontradas: {len(periodos_existentes)}")
                    for i, p in enumerate(periodos_existentes, 1):
                        print(f"   [{i}] ID: {p['importacao_id'][:8]}... | {p['inicio']} a {p['fim']} ({p['qtd_transacoes']} transa��es)")
                    
                    cursor.close()
                
                # Verificar sobreposi��o com cada per�odo existente
                for periodo_existente in periodos_existentes:
                    inicio_existente = periodo_existente['inicio']
                    fim_existente = periodo_existente['fim']
                    importacao_id_existente = periodo_existente['importacao_id']
                    
                    # L�gica de sobreposi��o: novo_inicio <= existente_fim AND novo_fim >= existente_inicio
                    if periodo_inicio <= fim_existente and periodo_fim >= inicio_existente:
                        print(f"? SOBREPOSI��O DETECTADA!")
                        print(f"   Per�odo tentando importar: {periodo_inicio} at� {periodo_fim}")
                        print(f"   Per�odo j� existente (ID {importacao_id_existente[:8]}...): {inicio_existente} at� {fim_existente}")
                        
                        # Mensagem detalhada para o usu�rio
                        erro_msg = f'? J� existe uma importa��o no per�odo de {inicio_existente.strftime("%d/%m/%Y")} at� {fim_existente.strftime("%d/%m/%Y")}'
                        
                        if total_orfas > 0:
                            erro_msg = f'?? ATEN��O: {total_orfas} transa��o(�es) �rf�(s) detectada(s) sem ID de importa��o! Exclua manualmente na tela de Extrato Banc�rio antes de reimportar.'
                        
                        return jsonify({
                            'success': False,
                            'error': erro_msg,
                            'details': {
                                'periodo_tentado': {
                                    'inicio': periodo_inicio.strftime('%d/%m/%Y'),
                                    'fim': periodo_fim.strftime('%d/%m/%Y')
                                },
                                'periodo_existente': {
                                    'importacao_id': importacao_id_existente,
                                    'inicio': inicio_existente.strftime('%d/%m/%Y'),
                                    'fim': fim_existente.strftime('%d/%m/%Y'),
                                    'transacoes': periodo_existente['qtd_transacoes']
                                },
                                'transacoes_orfas': total_orfas,
                                'mensagem': 'Use o bot�o "Deletar Extrato" na tela de Extrato Banc�rio (filtrar por per�odo e clicar em "Deletar Extrato").',
                                'solucao': f'DELETE FROM transacoes_extrato WHERE importacao_id = \'{importacao_id_existente}\' AND empresa_id = {empresa_id};'
                            }
                        }), 409  # 409 Conflict
                
                print(f"? Per�odo v�lido, sem sobreposi��o com importa��es existentes")
        
        except Exception as e:
            logger.error(f"Erro na valida��o de per�odo: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': f'Erro ao validar per�odo: {str(e)}'}), 500
        
        # Extrair transacoes
        transacoes = []

        # --- Carregar filtros de MEMO (Ajuste de OFX) para esta empresa + conta ---
        _memos_ignorar = set()
        try:
            with db.get_connection() as _fc:
                _fcur = _fc.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                _fcur.execute(
                    "SELECT memo_filtro FROM ofx_filtros_memo "
                    "WHERE empresa_id = %s AND conta_bancaria = %s",
                    (empresa_id, conta_bancaria)
                )
                for _fr in _fcur.fetchall():
                    _memos_ignorar.add(_fr['memo_filtro'].strip().upper())
                _fcur.close()
            if _memos_ignorar:
                logger.info(f"?? Ajuste OFX: {len(_memos_ignorar)} MEMO(s) ser�o ignorados: {_memos_ignorar}")
        except Exception as _fe:
            logger.warning(f"?? N�o foi poss�vel carregar filtros de MEMO: {_fe}")
        # --------------------------------------------------------------------------

        for account in ofx.accounts:
            # Obter saldo final do OFX
            saldo_final = float(account.statement.balance) if hasattr(account.statement, 'balance') else None
            
            print(f"\n{'='*60}")
            print(f"?? AN�LISE DO ARQUIVO OFX")
            print(f"{'='*60}")
            print(f"?? Conta: {account.number if hasattr(account, 'number') else 'N/A'}")
            print(f"?? Per�odo: {account.statement.start_date} a {account.statement.end_date}")
            print(f"?? Saldo Final (OFX): R$ {saldo_final:,.2f}" if saldo_final else "?? Saldo Final: N�O INFORMADO")
            print(f"?? Total de transa��es: {len(account.statement.transactions)}")
            
            # Ordenar transa��es por data (mais antiga primeiro)
            transactions_list = sorted(account.statement.transactions, key=lambda t: t.date)
            
            # Aplicar filtros de MEMO (Ajuste de OFX) � remover transa��es ignoradas
            if _memos_ignorar:
                _antes = len(transactions_list)
                transactions_list = [
                    t for t in transactions_list
                    if (t.memo or '').strip().upper() not in _memos_ignorar
                    and (t.payee or '').strip().upper() not in _memos_ignorar
                ]
                _ignoradas = _antes - len(transactions_list)
                if _ignoradas:
                    logger.info(f"?? Ajuste OFX: {_ignoradas} transa��o(�es) ignorada(s) por filtro de MEMO")
                    print(f"?? Ajuste OFX: {_ignoradas} transa��o(�es) ignorada(s) por filtro de MEMO")
            
            # PRIMEIRO: processar transa��es para corrigir sinais
            transacoes_processadas = []
            for trans in transactions_list:
                valor_ofx = float(trans.amount)
                trans_type = getattr(trans, 'type', None)
                
                # Determinar tipo e corrigir sinal
                # Tipos OFX padrão de DÉBITO (saída de dinheiro)
                _TIPOS_DEBITO = {'DEBIT', 'DEB', 'PAYMENT', 'ATM', 'DIRECTDEBIT', 'REPEATPMT',
                                 'FEE', 'SRVCHG', 'CHECK', 'POS', 'DÉBITO', 'DEBITO'}
                # Tipos OFX padrão de CRÉDITO (entrada de dinheiro)
                _TIPOS_CREDITO = {'CREDIT', 'DIRECTDEP', 'DEP', 'INT', 'DIV', 'XFER',
                                  'HOLD', 'OTHER', 'CASH', 'CRÉDITO', 'CREDITO'}
                if trans_type:
                    _tt = trans_type.upper()
                    if _tt in _TIPOS_DEBITO:
                        tipo = 'debito'
                        valor_correto = -abs(valor_ofx)  # DÉBITO sempre negativo
                    elif _tt in _TIPOS_CREDITO:
                        tipo = 'credito'
                        valor_correto = abs(valor_ofx)   # CRÉDITO sempre positivo
                    else:
                        # Tipo desconhecido: usar sinal do valor como desempate
                        if valor_ofx < 0:
                            tipo = 'debito'
                            valor_correto = valor_ofx
                        else:
                            tipo = 'credito'
                            valor_correto = valor_ofx
                        print(f"⚠️  OFX: TRNTYPE desconhecido '{trans_type}' — usando sinal do valor ({valor_ofx:.2f} → {tipo})")
                else:
                    # Sem trans_type: usar sinal do valor
                    if valor_ofx < 0:
                        tipo = 'debito'
                        valor_correto = valor_ofx
                    else:
                        tipo = 'credito'
                        valor_correto = valor_ofx
                
                transacoes_processadas.append({
                    'trans': trans,
                    'valor_ofx': valor_ofx,
                    'valor_correto': valor_correto,
                    'tipo': tipo
                })
            
            # Calcular saldo inicial baseado no saldo final e soma correta das transa��es
            # OU usar saldo_inicial da conta se data_inicio for anterior �s transa��es
            if saldo_final is not None:
                soma_transacoes = sum(t['valor_correto'] for t in transacoes_processadas)
                saldo_inicial_calculado_ofx = saldo_final - soma_transacoes
                
                # ?? FIX: Extrair data da primeira transa��o ignorando timezone
                if hasattr(transactions_list[0].date, 'year'):
                    data_primeira_transacao = date(transactions_list[0].date.year, 
                                                   transactions_list[0].date.month, 
                                                   transactions_list[0].date.day)
                elif hasattr(transactions_list[0].date, 'date'):
                    data_primeira_transacao = transactions_list[0].date.date()
                else:
                    data_primeira_transacao = transactions_list[0].date
                
                usar_saldo_conta = False
                if hasattr(conta_info, 'data_inicio') and conta_info.data_inicio:
                    data_inicio_conta = conta_info.data_inicio.date() if hasattr(conta_info.data_inicio, 'date') else conta_info.data_inicio
                    
                    # Se data_inicio da conta for anterior ou igual � primeira transa��o, usar saldo_inicial da conta
                    if data_inicio_conta <= data_primeira_transacao:
                        usar_saldo_conta = True
                        saldo_atual = float(conta_info.saldo_inicial)
                        print(f"\n? USANDO SALDO INICIAL DA CONTA:")
                        print(f"   Data de in�cio da conta: {data_inicio_conta}")
                        print(f"   Primeira transa��o OFX: {data_primeira_transacao}")
                        print(f"   Saldo inicial da conta: R$ {saldo_atual:,.2f}")
                        print(f"   (Saldo calculado pelo OFX seria: R$ {saldo_inicial_calculado_ofx:,.2f})")
                
                if not usar_saldo_conta:
                    saldo_atual = saldo_inicial_calculado_ofx
                    print(f"\n?? C�LCULOS (Saldo calculado pelo OFX):")
                    print(f"   Soma de todas transa��es (corrigida): R$ {soma_transacoes:+,.2f}")
                    print(f"   Saldo Final (OFX): R$ {saldo_final:,.2f}")
                    print(f"   Saldo Inicial calculado: R$ {saldo_inicial_calculado_ofx:,.2f}")
                    print(f"   F�rmula: {saldo_final:,.2f} - ({soma_transacoes:+,.2f}) = {saldo_inicial_calculado_ofx:,.2f}")
            else:
                print(f"\n?? AVISO: Saldo final n�o informado no OFX")
                # Usar saldo_inicial da conta se dispon�vel
                if hasattr(conta_info, 'saldo_inicial'):
                    saldo_atual = float(conta_info.saldo_inicial)
                    print(f"   Usando saldo inicial da conta: R$ {saldo_atual:,.2f}")
                else:
                    saldo_atual = 0
                    print(f"   Iniciando em R$ 0,00")
            
            print(f"\n?? PROCESSANDO TRANSA��ES (cronol�gica):")
            print(f"{'Data':<12} {'Tipo':<15} {'Valor OFX':>15} {'Valor Correto':>15} {'Saldo Ap�s':>15}")
            print(f"{'-'*72}")
            
            # Processar cada transa��o j� calculada e atualizar saldo
            for t_proc in transacoes_processadas:
                trans = t_proc['trans']
                valor_ofx = t_proc['valor_ofx']
                valor_correto = t_proc['valor_correto']
                tipo = t_proc['tipo']
                
                # Atualizar saldo: saldo += valor (negativo diminui, positivo aumenta)
                saldo_atual += valor_correto
                
                # ?? FIX: Extrair data ignorando timezone (previne bug -1 dia no Railway/UTC)
                # Em vez de trans.date.date() que pode ser afetado por timezone do servidor,
                # usar componentes year/month/day direto do datetime
                if hasattr(trans.date, 'year'):
                    data_transacao = date(trans.date.year, trans.date.month, trans.date.day)
                elif hasattr(trans.date, 'date'):
                    data_transacao = trans.date.date()
                else:
                    data_transacao = trans.date
                
                data_str = str(data_transacao)
                tipo_label = '?? D�BITO' if tipo == 'debito' else '?? CR�DITO'
                print(f"{data_str:<12} {tipo_label:<15} {valor_ofx:>+15.2f} {valor_correto:>+15.2f} {saldo_atual:>15.2f}")
                
                transacoes.append({
                    'data': data_transacao,
                    'descricao': trans.payee or trans.memo or 'Sem descricao',
                    'valor': valor_correto,  # Guardar valor com sinal (negativo para d�bito, positivo para cr�dito)
                    'tipo': tipo.upper(),  # DEBITO ou CREDITO (mai�sculo)
                    'saldo': saldo_atual,  # Saldo ap�s esta transa��o
                    'fitid': trans.id,
                    'memo': trans.memo,
                    'checknum': trans.checknum if hasattr(trans, 'checknum') else None
                })
        
        if not transacoes:
            return jsonify({'success': False, 'error': 'Nenhuma transacao encontrada no arquivo'}), 400
        
        # Salvar no banco (empresa_id j� foi obtido no in�cio da fun��o)
        resultado = extrato_functions.salvar_transacoes_extrato(
            database, 
            empresa_id, 
            conta_bancaria, 
            transacoes
        )
        
        if resultado['success']:
            # Formatar resposta para o frontend
            return jsonify({
                'success': True,
                'message': 'Extrato importado com sucesso',
                'transacoes_importadas': resultado.get('inseridas', 0),
                'transacoes_duplicadas': resultado.get('duplicadas', 0),
                'importacao_id': resultado.get('importacao_id')
            }), 200
        else:
            return jsonify(resultado), 400
        
    except Exception as e:
        logger.info(f"Erro ao processar OFX: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/extratos', methods=['GET'])
@require_permission('lancamentos_view')
def listar_extratos():
    """Lista transacoes do extrato com filtros"""
    try:
        usuario = get_usuario_logado()
        
        # Usar empresa_id da sess�o (empresa selecionada pelo usu�rio)
        empresa_id = session.get('empresa_id') or usuario.get('cliente_id') or usuario.get('empresa_id') or 1
        
        logger.info(f"?? /api/extratos: empresa_id={empresa_id}, usuario={usuario.get('nome', 'N/A')}")
        
        # Validar e sanitizar datas (rejeitar anos absurdos, ex: 0202 vindo de JS Date com ano truncado)
        def _sanitize_date(value):
            if not value:
                return None
            try:
                from datetime import datetime as _dt
                parsed = _dt.strptime(value, '%Y-%m-%d')
                if parsed.year < 2000 or parsed.year > 2100:
                    logger.warning(f"?? Data com ano inv�lido ignorada: {value}")
                    return None
                return value
            except (ValueError, TypeError):
                return None

        filtros = {
            'conta_bancaria': request.args.get('conta'),
            'data_inicio': _sanitize_date(request.args.get('data_inicio')),
            'data_fim': _sanitize_date(request.args.get('data_fim')),
            'conciliado': request.args.get('conciliado')
        }
        
        logger.info(f"?? Filtros recebidos: {filtros}")
        
        # Converter conciliado para boolean
        if filtros['conciliado'] is not None:
            filtros['conciliado'] = filtros['conciliado'].lower() == 'true'
        
        # Fun��o agora retorna dict com 'transacoes' e 'saldo_anterior'
        logger.info(f"?? Chamando extrato_functions.listar_transacoes_extrato...")
        resultado = extrato_functions.listar_transacoes_extrato(
            database,
            empresa_id,
            filtros
        )
        
        # Log do resultado (evitar backslash em f-string)
        if isinstance(resultado, dict):
            qtd_transacoes = len(resultado.get('transacoes', []))
            logger.info(f"?? Resultado tipo: dict com {qtd_transacoes} transa��es")
        else:
            logger.info(f"?? Resultado tipo: {type(resultado)}, conte�do: {resultado}")
        
        # Manter compatibilidade: se retornou lista (c�digo antigo), converter
        if isinstance(resultado, list):
            transacoes = resultado
            saldo_anterior = None
        else:
            transacoes = resultado.get('transacoes', [])
            saldo_anterior = resultado.get('saldo_anterior')
        
        logger.info(f"? Retornando {len(transacoes)} transa��o(�es) para o frontend")
        
        # Retornar no formato esperado pelo frontend
        resposta = {
            'transacoes': transacoes,
            'saldo_anterior': saldo_anterior
        }
        
        return jsonify(resposta), 200
        
    except Exception as e:
        logger.error(f"? ERRO ao listar extratos: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/extratos/<int:transacao_id>/conciliar', methods=['POST'])
@require_permission('lancamentos_edit')
def conciliar_extrato(transacao_id):
    """
    Concilia uma transacao do extrato com um lancamento
    
    Security:
        ?? Validado empresa_id da sess�o
    """
    try:
        # ?? VALIDA��O DE SEGURAN�A OBRIGAT�RIA
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'erro': 'Empresa n�o selecionada'}), 403
        
        dados = request.json
        lancamento_id = dados.get('lancamento_id')
        
        # ?? Passar empresa_id explicitamente
        resultado = extrato_functions.conciliar_transacao(
            database,
            empresa_id,
            transacao_id,
            lancamento_id
        )
        
        return jsonify(resultado), 200 if resultado['success'] else 400
        
    except Exception as e:
        logger.info(f"Erro ao conciliar: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# HISTORICO PERMANENTE DE CONCILIACOES - Funcoes auxiliares
# ============================================================================

def _garantir_tabela_historico_conciliacoes(conn):
    """Cria a tabela historico_conciliacoes se nao existir (idempotente)."""
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS historico_conciliacoes (
            id                   BIGSERIAL PRIMARY KEY,
            empresa_id           INT NOT NULL,
            evento               VARCHAR(20) NOT NULL,
            data_evento          TIMESTAMPTZ DEFAULT NOW(),
            transacao_extrato_id INT,
            lancamento_id        INT,
            data_transacao       DATE,
            conta_bancaria       VARCHAR(255),
            descricao_extrato    TEXT,
            valor                NUMERIC(15,2),
            tipo_extrato         VARCHAR(50),
            descricao_lancamento TEXT,
            categoria            VARCHAR(255),
            subcategoria         VARCHAR(255),
            pessoa               VARCHAR(255),
            observacoes          TEXT,
            memo                 TEXT,
            fitid                VARCHAR(255)
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_hist_conc_empresa_data
            ON historico_conciliacoes (empresa_id, data_evento DESC)
    """)
    cur.close()


def _inserir_historico_conciliacao(conn, empresa_id, evento, transacao_id, lancamento_id=None):
    """Insere um evento no historico de conciliacoes, copiando dados do extrato/lancamento."""
    try:
        _garantir_tabela_historico_conciliacoes(conn)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO historico_conciliacoes (
                empresa_id, evento, transacao_extrato_id, lancamento_id,
                data_transacao, conta_bancaria, descricao_extrato, valor, tipo_extrato,
                descricao_lancamento, categoria, subcategoria, pessoa, observacoes, memo, fitid
            )
            SELECT
                te.empresa_id, %s, te.id, %s,
                te.data, te.conta_bancaria, te.descricao, ABS(COALESCE(te.valor, 0)), te.tipo,
                COALESCE(l.descricao, te.descricao),
                COALESCE(l.categoria, te.categoria),
                COALESCE(l.subcategoria, te.subcategoria),
                COALESCE(l.pessoa, te.pessoa),
                l.observacoes, te.memo, te.fitid
            FROM transacoes_extrato te
            LEFT JOIN lancamentos l ON l.id = %s AND l.empresa_id = te.empresa_id
            WHERE te.id = %s AND te.empresa_id = %s
        """, (evento, lancamento_id, lancamento_id, transacao_id, empresa_id))
        cur.close()
    except Exception as exc:
        logger.warning(f"_inserir_historico_conciliacao falhou: {exc}")


@app.route('/api/extratos/historico-conciliacao', methods=['GET'])
@require_permission('lancamentos_view')
def historico_conciliacao():
    """
    Retorna o historico permanente de conciliacoes (tabela historico_conciliacoes).
    Inclui eventos 'conciliado' e 'desconciliado' para auditoria completa.
    Auto-migra registros existentes da tabela conciliacoes na primeira chamada.

    Query params opcionais:
        conta         - filtrar por conta bancaria
        data_inicio   - YYYY-MM-DD
        data_fim      - YYYY-MM-DD
        evento        - 'conciliado' ou 'desconciliado' (padrao: todos)
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'erro': 'Empresa nao selecionada'}), 403

        conta         = request.args.get('conta', '').strip()
        data_inicio   = request.args.get('data_inicio', '').strip()
        data_fim      = request.args.get('data_fim', '').strip()
        evento_filtro = request.args.get('evento', '').strip()

        filtros = []
        params  = [empresa_id]

        if conta:
            filtros.append("h.conta_bancaria = %s")
            params.append(conta)
        if data_inicio:
            filtros.append("h.data_transacao >= %s")
            params.append(data_inicio)
        if data_fim:
            filtros.append("h.data_transacao <= %s")
            params.append(data_fim)
        if evento_filtro in ('conciliado', 'desconciliado'):
            filtros.append("h.evento = %s")
            params.append(evento_filtro)

        where_extra = ('AND ' + ' AND '.join(filtros)) if filtros else ''

        import psycopg2.extras
        with database.get_db_connection(empresa_id=empresa_id) as conn:
            # Garantir que a tabela existe
            _garantir_tabela_historico_conciliacoes(conn)

            # Auto-migrar conciliacoes existentes que ainda nao estao no historico
            conn.cursor().execute("""
                INSERT INTO historico_conciliacoes (
                    empresa_id, evento, data_evento, transacao_extrato_id, lancamento_id,
                    data_transacao, conta_bancaria, descricao_extrato, valor, tipo_extrato,
                    descricao_lancamento, categoria, subcategoria, pessoa, observacoes, memo, fitid
                )
                SELECT
                    c.empresa_id, 'conciliado',
                    COALESCE(c.data_conciliacao, te.created_at, NOW()),
                    te.id, l.id,
                    te.data, te.conta_bancaria, te.descricao, ABS(COALESCE(te.valor, 0)), te.tipo,
                    COALESCE(l.descricao, te.descricao),
                    COALESCE(l.categoria, te.categoria),
                    COALESCE(l.subcategoria, te.subcategoria),
                    COALESCE(l.pessoa, te.pessoa),
                    l.observacoes, te.memo, te.fitid
                FROM conciliacoes c
                JOIN transacoes_extrato te ON te.id = c.transacao_extrato_id
                                          AND te.empresa_id = c.empresa_id
                LEFT JOIN lancamentos l ON l.id = c.lancamento_id
                                       AND l.empresa_id = c.empresa_id
                WHERE c.empresa_id = %s
                  AND NOT EXISTS (
                      SELECT 1 FROM historico_conciliacoes h2
                      WHERE h2.transacao_extrato_id = te.id
                        AND h2.empresa_id = c.empresa_id
                        AND h2.evento = 'conciliado'
                  )
            """, (empresa_id,))

            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute(f"""
                SELECT
                    h.id                     AS conciliacao_id,
                    h.data_evento            AS data_conciliacao,
                    h.transacao_extrato_id   AS transacao_id,
                    h.data_transacao,
                    h.conta_bancaria,
                    h.descricao_extrato,
                    h.valor,
                    h.tipo_extrato,
                    h.memo,
                    h.fitid,
                    h.lancamento_id,
                    h.descricao_lancamento,
                    h.categoria,
                    h.subcategoria,
                    h.pessoa,
                    h.observacoes,
                    h.evento
                FROM historico_conciliacoes h
                WHERE h.empresa_id = %s
                {where_extra}
                ORDER BY h.data_evento DESC, h.id DESC
                LIMIT 2000
            """, params)
            rows = cursor.fetchall()
            cursor.close()

        from decimal import Decimal
        from datetime import date as _date, datetime as _datetime
        result = []
        for row in rows:
            d = dict(row)
            for k, v in d.items():
                if isinstance(v, _datetime):
                    d[k] = v.isoformat()
                elif isinstance(v, _date):
                    d[k] = v.isoformat()
                elif isinstance(v, Decimal):
                    d[k] = float(v)
            result.append(d)

        return jsonify(result), 200

    except Exception as e:
        logger.error(f"Erro no historico_conciliacao: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'erro': str(e)}), 500


@app.route('/api/extratos/conciliacao/<int:conciliacao_id>', methods=['PATCH'])
@require_permission('lancamentos_edit')
def editar_conciliacao(conciliacao_id):
    """
    Edita os campos de um registro de conciliação (atualiza o lançamento vinculado).

    Body JSON (todos opcionais):
        descricao_lancamento, categoria, subcategoria, pessoa, observacoes

    Security:
        🔒 Validado empresa_id da sessão
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'erro': 'Empresa não selecionada'}), 403

        dados = request.get_json(force=True) or {}
        campos_permitidos = ['descricao', 'categoria', 'subcategoria', 'pessoa', 'observacoes']

        # Mapear descricao_lancamento → descricao no lancamento
        if 'descricao_lancamento' in dados:
            dados['descricao'] = dados.pop('descricao_lancamento')

        sets   = []
        params = []
        for campo in campos_permitidos:
            if campo in dados:
                sets.append(f"{campo} = %s")
                params.append(dados[campo])

        if not sets:
            return jsonify({'erro': 'Nenhum campo para atualizar'}), 400

        import psycopg2.extras
        with database.get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            # Obter lancamento_id vinculado (com segurança multi-tenant)
            cursor.execute("""
                SELECT lancamento_id FROM historico_conciliacoes
                WHERE id = %s AND empresa_id = %s
            """, (conciliacao_id, empresa_id))
            row = cursor.fetchone()
            if not row:
                cursor.close()
                return jsonify({'erro': 'Conciliação não encontrada'}), 404

            lancamento_id = row['lancamento_id']
            params.append(lancamento_id)
            params.append(empresa_id)

            cursor.execute(f"""
                UPDATE lancamentos
                SET {', '.join(sets)}, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, params)

            conn.commit()
            cursor.close()

        return jsonify({'success': True}), 200

    except Exception as e:
        logger.error(f"Erro ao editar conciliação {conciliacao_id}: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/extratos/regerar-conciliacao', methods=['POST'])
@require_permission('lancamentos_create')
def regerar_conciliacao():
    """
    Regenera as conciliações de um período, recriando os lançamentos corretamente.

    Fluxo:
    1. Busca todas as transações conciliadas no período/conta
    2. Exclui os lançamentos e registros de conciliação em bloco (transação atômica)
    3. Recria os lançamentos com tipo correto (derivado de transacoes_extrato.tipo)
       preservando categoria/subcategoria/pessoa/descrição do histórico
    4. Reconcilia novamente

    Body JSON:
        conta       - nome da conta bancária (obrigatório)
        data_inicio - YYYY-MM-DD (obrigatório)
        data_fim    - YYYY-MM-DD (obrigatório)

    Security:
        🔒 Validado empresa_id da sessão
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa não identificada'}), 403

        dados = request.get_json(force=True) or {}
        conta       = dados.get('conta', '').strip()
        data_inicio = dados.get('data_inicio', '').strip()
        data_fim    = dados.get('data_fim', '').strip()

        if not conta:
            return jsonify({'success': False, 'error': 'Conta bancária é obrigatória'}), 400
        if not data_inicio:
            return jsonify({'success': False, 'error': 'Data início é obrigatória'}), 400
        if not data_fim:
            return jsonify({'success': False, 'error': 'Data fim é obrigatória'}), 400

        logger.info(f"🔁 Regerar conciliação: empresa={empresa_id}, conta={conta}, {data_inicio} a {data_fim}")

        import psycopg2.extras
        # Variáveis acessíveis após o bloco with
        transacoes = []
        criados    = 0
        erros      = []

        with database.get_db_connection(empresa_id=empresa_id) as conn:
            conn.autocommit = False
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            try:
                # ── 1. Buscar transações conciliadas no período ─────────────────────────
                cursor.execute("""
                    SELECT
                        te.id                AS transacao_id,
                        te.data,
                        te.descricao,
                        te.valor,
                        te.tipo              AS tipo_extrato,
                        te.conta_bancaria,
                        te.categoria         AS te_categoria,
                        te.subcategoria      AS te_subcategoria,
                        te.pessoa            AS te_pessoa,
                        c.lancamento_id,
                        -- Fallback: usa dados do histórico quando lançamento foi excluído
                        COALESCE(l.descricao,    h.descricao_lancamento) AS l_descricao,
                        COALESCE(l.categoria,    h.categoria)            AS l_categoria,
                        COALESCE(l.subcategoria, h.subcategoria)         AS l_subcategoria,
                        COALESCE(l.pessoa,       h.pessoa)               AS l_pessoa,
                        COALESCE(l.observacoes,  h.observacoes)          AS l_observacoes
                    FROM transacoes_extrato te
                    LEFT JOIN conciliacoes c ON c.transacao_extrato_id = te.id
                                            AND c.empresa_id = te.empresa_id
                    LEFT JOIN lancamentos  l ON l.id = c.lancamento_id
                                            AND l.empresa_id = te.empresa_id
                    -- Recupera dados do histórico mais recente para transações desconciliadas
                    LEFT JOIN LATERAL (
                        SELECT descricao_lancamento, categoria, subcategoria, pessoa, observacoes
                        FROM historico_conciliacoes
                        WHERE transacao_extrato_id = te.id
                          AND empresa_id           = te.empresa_id
                        ORDER BY data_evento DESC
                        LIMIT 1
                    ) h ON TRUE
                    WHERE te.empresa_id     = %s
                      AND te.conta_bancaria = %s
                      AND te.data          >= %s
                      AND te.data          <= %s
                      AND (
                          -- Atualmente conciliada
                          te.conciliado = TRUE
                          -- OU foi conciliada/desconciliada anteriormente (lançamento excluído)
                          OR EXISTS (
                              SELECT 1 FROM historico_conciliacoes hc
                              WHERE hc.transacao_extrato_id = te.id
                                AND hc.empresa_id = te.empresa_id
                          )
                      )
                    ORDER BY te.data ASC, te.id ASC
                """, (empresa_id, conta, data_inicio, data_fim))

                transacoes = list(cursor.fetchall())

                if not transacoes:
                    cursor.close()
                    return jsonify({
                        'success': False,
                        'error': (f'Nenhuma conciliação (ativa ou desconciliada) encontrada no período '
                                  f'{data_inicio} a {data_fim} para a conta "{conta}"')
                    }), 404

                # Garantir tabela historico existe
                _garantir_tabela_historico_conciliacoes(conn)

                ids_transacoes  = [t['transacao_id'] for t in transacoes]
                ids_lancamentos = [t['lancamento_id'] for t in transacoes if t['lancamento_id']]

                logger.info(f"🔁 {len(transacoes)} transações encontradas, "
                            f"{len(ids_lancamentos)} lançamentos a excluir")

                # 1.5: Registrar desconciliacoes no historico ANTES de excluir
                for _t in transacoes:
                    if _t['lancamento_id']:
                        _inserir_historico_conciliacao(
                            conn, empresa_id, 'desconciliado',
                            _t['transacao_id'], _t['lancamento_id']
                        )

                # ── 2. Excluir lançamentos vinculados ───────────────────────────────────
                if ids_lancamentos:
                    cursor.execute(
                        "DELETE FROM lancamentos WHERE id = ANY(%s) AND empresa_id = %s",
                        (ids_lancamentos, empresa_id)
                    )
                    logger.info(f"🗑️  Deletados {cursor.rowcount} lançamento(s)")

                # ── 3. Excluir registros de conciliação ─────────────────────────────────
                cursor.execute(
                    "DELETE FROM conciliacoes WHERE transacao_extrato_id = ANY(%s) AND empresa_id = %s",
                    (ids_transacoes, empresa_id)
                )
                logger.info(f"🗑️  Deletadas {cursor.rowcount} conciliação(ões)")

                # ── 4. Marcar transações como não conciliadas ───────────────────────────
                cursor.execute(
                    "UPDATE transacoes_extrato SET conciliado = FALSE "
                    "WHERE id = ANY(%s) AND empresa_id = %s",
                    (ids_transacoes, empresa_id)
                )

                # ── 5. Recriar lançamentos e conciliações ───────────────────────────────
                for t in transacoes:
                    try:
                        # Tipo correto derivado do extrato (NÃO do lançamento problemático)
                        tipo_extrato_raw = (t['tipo_extrato'] or '').upper()
                        valor_float      = float(t['valor'] or 0)

                        if tipo_extrato_raw in ('DÉBITO', 'DEBITO'):
                            tipo_lancamento = 'despesa'
                        elif tipo_extrato_raw in ('CRÉDITO', 'CREDITO'):
                            tipo_lancamento = 'receita'
                        else:
                            tipo_lancamento = 'receita' if valor_float >= 0 else 'despesa'
                            logger.warning(
                                f"⚠️  Tipo desconhecido '{t['tipo_extrato']}' para transação "
                                f"{t['transacao_id']} — usando sinal do valor → {tipo_lancamento}"
                            )

                        valor_abs    = abs(valor_float)
                        # Usar dados mais completos: lançamento editado > extrato como fallback
                        categoria    = t['l_categoria']    or t['te_categoria']    or 'Conciliação Bancária'
                        subcategoria = t['l_subcategoria'] or t['te_subcategoria']
                        pessoa       = t['l_pessoa']       or t['te_pessoa']
                        descricao_lc = (t['l_descricao']   or t['descricao'] or
                                        'Lançamento recriado via Regerar Conciliação')
                        observacoes  = (t['l_observacoes'] or '').strip()

                        cursor.execute("""
                            INSERT INTO lancamentos (
                                empresa_id, tipo, descricao, valor,
                                data_vencimento, data_pagamento, status,
                                conta_bancaria, categoria, subcategoria, pessoa,
                                observacoes
                            ) VALUES (
                                %s, %s, %s, %s,
                                %s, %s, 'pago',
                                %s, %s, %s, %s,
                                %s
                            )
                            RETURNING id
                        """, (
                            empresa_id,
                            tipo_lancamento,
                            descricao_lc,
                            valor_abs,
                            t['data'],
                            t['data'],
                            t['conta_bancaria'],
                            categoria,
                            subcategoria,
                            pessoa,
                            observacoes
                        ))

                        novo_id = cursor.fetchone()['id']

                        cursor.execute("""
                            INSERT INTO conciliacoes (empresa_id, transacao_extrato_id, lancamento_id)
                            VALUES (%s, %s, %s)
                        """, (empresa_id, t['transacao_id'], novo_id))

                        cursor.execute(
                            "UPDATE transacoes_extrato SET conciliado = TRUE "
                            "WHERE id = %s AND empresa_id = %s",
                            (t['transacao_id'], empresa_id)
                        )

                        criados += 1
                        # Registrar conciliacao no historico
                        _inserir_historico_conciliacao(
                            conn, empresa_id, 'conciliado',
                            t['transacao_id'], novo_id
                        )
                        logger.info(f"✅ Transação {t['transacao_id']} → lançamento #{novo_id} ({tipo_lancamento})")

                    except Exception as item_err:
                        erros.append(f"Transação #{t['transacao_id']}: {str(item_err)}")
                        logger.error(f"❌ Erro ao recriar lançamento para transação "
                                     f"{t['transacao_id']}: {item_err}")

                conn.commit()
                cursor.close()
                logger.info(f"✅ Regerar concluído: {criados} criados, {len(erros)} erros")

            except Exception as inner_err:
                conn.rollback()
                cursor.close()
                raise inner_err

        return jsonify({
            'success': criados > 0,
            'total':   len(transacoes),
            'criados': criados,
            'erros':   erros,
            'message': (f'{criados} lançamento(s) recriado(s) com sucesso'
                        + (f'. {len(erros)} erro(s).' if erros else ''))
        }), 200

    except Exception as e:
        logger.error(f"Erro ao regerar conciliação: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/extratos/<int:transacao_id>/sugestoes', methods=['GET'])
@require_permission('lancamentos_view')
def sugerir_conciliacoes_extrato(transacao_id):
    """Sugere lancamentos para conciliar com uma transacao"""
    try:
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        sugestoes = extrato_functions.sugerir_conciliacoes(
            database,
            empresa_id,
            transacao_id
        )
        
        return jsonify(sugestoes), 200
        
    except Exception as e:
        logger.info(f"Erro ao sugerir conciliacoes: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/extratos/importacao/<importacao_id>', methods=['DELETE'])
@require_permission('lancamentos_delete')
def deletar_importacao_extrato(importacao_id):
    """
    Deleta todas as transacoes de uma importacao
    
    Security:
        ?? Validado empresa_id da sess�o
    """
    try:
        # ?? VALIDA��O DE SEGURAN�A OBRIGAT�RIA
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'erro': 'Empresa n�o selecionada'}), 403
        
        # ?? VERIFICAR SE H� TRANSA��ES CONCILIADAS
        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*) 
            FROM transacoes_extrato 
            WHERE importacao_id = %s AND empresa_id = %s AND conciliado = TRUE
        """, (importacao_id, empresa_id))
        
        _row = cursor.fetchone()
        total_conciliados = _row['count'] if isinstance(_row, dict) else _row[0]
        cursor.close()
        from database_postgresql import return_to_pool
        return_to_pool(conn)
        
        # Se houver concilia��es, avisar o usu�rio
        if total_conciliados > 0:
            logger.warning(f"?? Tentativa de deletar extrato com {total_conciliados} transa��es conciliadas")
            return jsonify({
                'success': False,
                'error': f'?? ATEN��O: Este extrato cont�m {total_conciliados} transa��o(�es) conciliada(s). A exclus�o ir� desfazer todas as concilia��es. Confirme para continuar.',
                'transacoes_conciliadas': total_conciliados,
                'requer_confirmacao': True
            }), 409  # 409 Conflict
        
        # ?? Passar empresa_id explicitamente
        resultado = extrato_functions.deletar_transacoes_extrato(
            database,
            empresa_id,
            importacao_id
        )
        
        return jsonify(resultado), 200 if resultado['success'] else 400
        
    except Exception as e:
        logger.info(f"Erro ao deletar importacao: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/extratos/diagnostico', methods=['GET'])
@require_permission('lancamentos_view')
def diagnostico_extrato():
    """Verifica problemas no extrato (duplicatas, saldo, etc.)"""
    try:
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        conta_bancaria = request.args.get('conta')
        if not conta_bancaria:
            return jsonify({'success': False, 'error': 'Conta banc�ria n�o informada'}), 400
        
        logger.info(f"?? Diagn�stico do extrato - empresa_id: {empresa_id}, conta: {conta_bancaria}")
        
        # Criar inst�ncia local do DatabaseManager
        db_manager = DatabaseManager()
        conn = db_manager.get_connection()
        
        try:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # 1. Total de transa��es
            cursor.execute("""
                SELECT COUNT(*) as total, MIN(data) as data_inicio, MAX(data) as data_fim
                FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s
            """, (empresa_id, conta_bancaria))
            resumo = cursor.fetchone()
            logger.info(f"   ?? Total transa��es: {resumo['total'] if resumo else 0}")
            
            # 2. Saldo atual (�ltima transa��o)
            cursor.execute("""
                SELECT data, saldo, descricao
                FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s
                ORDER BY data DESC, id DESC
                LIMIT 1
            """, (empresa_id, conta_bancaria))
            ultima = cursor.fetchone()
            if ultima:
                logger.info(f"   ?? Saldo atual: R$ {ultima['saldo']}")
            
            # 3. Verificar duplicatas por FITID
            cursor.execute("""
                SELECT fitid, COUNT(*) as qtd,
                       STRING_AGG(CAST(id AS TEXT), ', ') as ids,
                       STRING_AGG(CAST(data AS TEXT), ', ') as datas
                FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s AND fitid IS NOT NULL
                GROUP BY fitid
                HAVING COUNT(*) > 1
                ORDER BY qtd DESC
            """, (empresa_id, conta_bancaria))
            duplicatas_fitid = cursor.fetchall()
            logger.info(f"   ?? Duplicatas FITID: {len(duplicatas_fitid)}")
            
            # 4. Verificar duplicatas por data+valor+descri��o
            cursor.execute("""
                SELECT data, valor, descricao, COUNT(*) as qtd,
                       STRING_AGG(CAST(id AS TEXT), ', ') as ids
                FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s
                GROUP BY data, valor, descricao
                HAVING COUNT(*) > 1
                ORDER BY qtd DESC, data DESC
            """, (empresa_id, conta_bancaria))
            duplicatas_conteudo = cursor.fetchall()
            logger.info(f"   ?? Duplicatas conte�do: {len(duplicatas_conteudo)}")
            
            # 5. Importa��es
            cursor.execute("""
                SELECT importacao_id, COUNT(*) as transacoes, MIN(data) as inicio, MAX(data) as fim
                FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s
                GROUP BY importacao_id
                ORDER BY importacao_id DESC
            """, (empresa_id, conta_bancaria))
            importacoes = cursor.fetchall()
            logger.info(f"   ?? Importa��es: {len(importacoes)}")
            
            # ?? 5.1. Detectar transa��es �rf�s (sem importacao_id)
            cursor.execute("""
                SELECT COUNT(*) as total_orfas,
                       MIN(data) as data_inicio,
                       MAX(data) as data_fim,
                       STRING_AGG(CAST(id AS TEXT), ', ' ORDER BY data DESC LIMIT 10) as ids_exemplo
                FROM transacoes_extrato
                WHERE empresa_id = %s 
                AND conta_bancaria = %s
                AND (importacao_id IS NULL OR importacao_id = '')
            """, (empresa_id, conta_bancaria))
            orfas_info = cursor.fetchone()
            total_orfas = orfas_info['total_orfas'] if orfas_info else 0
            
            if total_orfas > 0:
                logger.warning(f"   ?? TRANSA��ES �RF�S: {total_orfas} (sem importacao_id)")
                logger.warning(f"      Per�odo: {orfas_info['data_inicio']} a {orfas_info['data_fim']}")
            else:
                logger.info(f"   ? Nenhuma transa��o �rf� detectada")
            
            # 6. Saldo da conta cadastrada
            cursor.execute("""
                SELECT saldo_inicial, data_inicio
                FROM contas_bancariasempresa
                WHERE empresa_id = %s AND nome = %s
            """, (empresa_id, conta_bancaria))
            conta = cursor.fetchone()
            
            cursor.close()
            
            resultado = {
                'success': True,
                'conta': conta_bancaria,
                'resumo': {
                    'total_transacoes': int(resumo['total']) if resumo else 0,
                    'periodo': {
                        'inicio': str(resumo['data_inicio']) if resumo and resumo['data_inicio'] else None,
                        'fim': str(resumo['data_fim']) if resumo and resumo['data_fim'] else None
                    }
                },
                'saldo_atual': {
                    'valor': float(ultima['saldo']) if ultima and ultima['saldo'] is not None else 0,
                    'data': str(ultima['data']) if ultima else None,
                    'descricao': ultima['descricao'] if ultima else None
                } if ultima else None,
                'conta_cadastrada': {
                    'saldo_inicial': float(conta['saldo_inicial']) if conta and conta['saldo_inicial'] is not None else 0,
                    'data_inicio': str(conta['data_inicio']) if conta and conta['data_inicio'] else None
                } if conta else None,
                'duplicatas': {
                    'por_fitid': [dict(d) for d in duplicatas_fitid],
                    'por_conteudo': [dict(d) for d in duplicatas_conteudo],
                    'total_fitid': len(duplicatas_fitid),
                    'total_conteudo': len(duplicatas_conteudo)
                },
                'importacoes': [dict(i) for i in importacoes],
                'transacoes_orfas': {
                    'total': total_orfas,
                    'periodo': {
                        'inicio': str(orfas_info['data_inicio']) if orfas_info and orfas_info['data_inicio'] else None,
                        'fim': str(orfas_info['data_fim']) if orfas_info and orfas_info['data_fim'] else None
                    } if total_orfas > 0 else None,
                    'ids_exemplo': orfas_info['ids_exemplo'] if orfas_info else None
                },
                'problemas_detectados': []
            }
            
            # Adicionar problemas detectados
            if total_orfas > 0:
                resultado['problemas_detectados'].append({
                    'tipo': 'transacoes_orfas',
                    'severidade': 'ALTA',
                    'mensagem': f'{total_orfas} transa��o(�es) sem ID de importa��o detectada(s)',
                    'solucao': 'Use o bot�o "Deletar Extrato" filtrando pelo per�odo para remover estas transa��es',
                    'periodo': f"{orfas_info['data_inicio']} at� {orfas_info['data_fim']}"
                })
            
            if len(duplicatas_fitid) > 0:
                resultado['problemas_detectados'].append({
                    'tipo': 'duplicatas_fitid',
                    'severidade': 'M�DIA',
                    'mensagem': f'{len(duplicatas_fitid)} grupo(s) de transa��es duplicadas por FITID',
                    'solucao': 'Execute o script de limpeza de duplicatas'
                })
            
            logger.info(f"? Diagn�stico conclu�do com sucesso")
            return jsonify(resultado), 200
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.error(f"? Erro no diagn�stico: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/extratos/deletar-tudo-conta', methods=['DELETE'])
@require_permission('lancamentos_delete')
def deletar_tudo_extrato_conta():
    """
    Deleta TODAS as transa��es do extrato de uma conta espec�fica
    ?? CUIDADO: A��o irrevers�vel!
    """
    try:
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        conta_bancaria = request.args.get('conta')
        if not conta_bancaria:
            return jsonify({'success': False, 'error': 'Conta banc�ria n�o informada'}), 400
        
        # Verificar se usu�rio confirmou a exclus�o (caso haja concilia��es)
        confirmado = request.args.get('confirmar', 'false').lower() == 'true'
        
        logger.info(f"??? Deletando TODAS transa��es - empresa_id: {empresa_id}, conta: {conta_bancaria}")
        
        # Criar inst�ncia local do DatabaseManager
        db_manager = DatabaseManager()
        conn = db_manager.get_connection()
        
        try:
            conn.autocommit = False
            cursor = conn.cursor()
            
            #  Contar quantas transa��es ser�o deletadas (para log)
            cursor.execute("""
                SELECT COUNT(*) FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s
            """, (empresa_id, conta_bancaria))
            _row_antes = cursor.fetchone()
            total_antes = _row_antes['count'] if isinstance(_row_antes, dict) else _row_antes[0]
            
            # ?? VERIFICAR SE H� TRANSA��ES CONCILIADAS
            cursor.execute("""
                SELECT COUNT(*) 
                FROM transacoes_extrato 
                WHERE empresa_id = %s AND conta_bancaria = %s AND conciliado = TRUE
            """, (empresa_id, conta_bancaria))
            
            _row_conc = cursor.fetchone()
            total_conciliados = _row_conc['count'] if isinstance(_row_conc, dict) else _row_conc[0]
            
            # Se houver concilia��es e n�o foi confirmado, avisar o usu�rio
            if total_conciliados > 0 and not confirmado:
                cursor.close()
                conn.close()
                logger.warning(f"?? Tentativa de deletar extrato com {total_conciliados} transa��es conciliadas")
                return jsonify({
                    'success': False,
                    'error': f'?? ATEN��O: {total_conciliados} de {total_antes} transa��o(�es) est�(�o) conciliada(s). A exclus�o ir� desfazer todas as concilia��es. Confirme para continuar.',
                    'transacoes_conciliadas': total_conciliados,
                    'total_transacoes': total_antes,
                    'requer_confirmacao': True
                }), 409  # 409 Conflict
            
            logger.info(f"   ?? Total de transa��es a deletar: {total_antes} ({total_conciliados} conciliadas)")
            
            # Deletar TODAS as transa��es desta conta/empresa
            cursor.execute("""
                DELETE FROM transacoes_extrato
                WHERE empresa_id = %s AND conta_bancaria = %s
            """, (empresa_id, conta_bancaria))
            
            deletados = cursor.rowcount
            
            conn.commit()
            cursor.close()
            
            logger.info(f"   ? {deletados} transa��es deletadas com sucesso")
            
            mensagem = f'? {deletados} transa��o(�es) deletada(s) com sucesso.'
            if total_conciliados > 0:
                mensagem += f' {total_conciliados} concilia��o(�es) foi(ram) desfeita(s).'
            mensagem += ' Agora voc� pode reimportar o arquivo OFX.'
            
            return jsonify({
                'success': True,
                'deletados': deletados,
                'conciliacoes_desfeitas': total_conciliados,
                'message': mensagem
            }), 200
            
        except Exception as e:
            if conn:
                conn.rollback()
            raise e
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.info(f"Erro ao deletar todas transa� �es: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/extratos/deletar-filtrado', methods=['DELETE'])
@require_permission('lancamentos_delete')
def deletar_extrato_filtrado():
    """Deleta transacoes do extrato baseado em filtros"""
    try:
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        filtros = {
            'conta_bancaria': request.args.get('conta'),
            'data_inicio': request.args.get('data_inicio'),
            'data_fim': request.args.get('data_fim')
        }
        
        # Validar que pelo menos um filtro foi fornecido
        if not any(filtros.values()):
            return jsonify({
                'success': False, 
                'error': 'Pelo menos um filtro deve ser fornecido (conta, data_inicio ou data_fim)'
            }), 400
        
        # Verificar se usu�rio confirmou a exclus�o (caso haja concilia��es)
        confirmado = request.args.get('confirmar', 'false').lower() == 'true'
        
        # Deletar transa��es que correspondem aos filtros
        with db.get_connection() as conn:
            conn.autocommit = False
            cursor = conn.cursor()
            
            # Primeiro verificar quantas transa��es conciliadas ser�o afetadas
            query_count = "SELECT COUNT(*) FROM transacoes_extrato WHERE empresa_id = %s AND conciliado = TRUE"
            params = [empresa_id]
            
            if filtros['conta_bancaria']:
                query_count += " AND conta_bancaria = %s"
                params.append(filtros['conta_bancaria'])
            
            if filtros['data_inicio']:
                query_count += " AND data >= %s"
                params.append(filtros['data_inicio'])
            
            if filtros['data_fim']:
                query_count += " AND data <= %s"
                params.append(filtros['data_fim'])
            
            cursor.execute(query_count, params)
            _row = cursor.fetchone()
            total_conciliados = _row['count'] if isinstance(_row, dict) else _row[0]
            
            # Se houver concilia��es e n�o foi confirmado, avisar o usu�rio
            if total_conciliados > 0 and not confirmado:
                cursor.close()
                logger.warning(f"?? Tentativa de deletar extrato com {total_conciliados} transa��es conciliadas")
                return jsonify({
                    'success': False,
                    'error': f'?? ATEN��O: {total_conciliados} transa��o(�es) conciliada(s) ser�(�o) afetada(s). A exclus�o ir� desfazer todas as concilia��es. Confirme para continuar.',
                    'transacoes_conciliadas': total_conciliados,
                    'requer_confirmacao': True
                }), 409  # 409 Conflict
            
            # Executar a dele��o
            query = "DELETE FROM transacoes_extrato WHERE empresa_id = %s"
            params = [empresa_id]
            
            if filtros['conta_bancaria']:
                query += " AND conta_bancaria = %s"
                params.append(filtros['conta_bancaria'])
            
            if filtros['data_inicio']:
                query += " AND data >= %s"
                params.append(filtros['data_inicio'])
            
            if filtros['data_fim']:
                query += " AND data <= %s"
                params.append(filtros['data_fim'])
            
            cursor.execute(query, params)
            deletados = cursor.rowcount
            
            conn.commit()
            cursor.close()
            
            mensagem = f'{deletados} transa��o(�es) deletada(s) com sucesso'
            if total_conciliados > 0:
                mensagem += f'. {total_conciliados} concilia��o(�es) foi(ram) desfeita(s)'
            
            return jsonify({
                'success': True,
                'deletados': deletados,
                'conciliacoes_desfeitas': total_conciliados,
                'message': mensagem
            }), 200
        
    except Exception as e:
        logger.info(f"Erro ao deletar extratos filtrados: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/auditoria/pagamentos-duplicados', methods=['GET'])
@require_permission('relatorios_view')
def auditoria_pagamentos_duplicados():
    """
    Auditoria de Pagamentos - Detecta pagamentos duplicados
    Crit�rio: Mesma data + Mesmo valor + Mesmo benefici�rio (nome/CPF)
    """
    try:
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        # Filtros opcionais
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        conta_bancaria = request.args.get('conta')
        
        logger.info(f"?? Auditoria de Pagamentos - empresa_id: {empresa_id}")
        
        # Criar inst�ncia local do DatabaseManager
        db_manager = DatabaseManager()
        conn = db_manager.get_connection()
        
        try:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Query para detectar duplicatas no EXTRATO BANC�RIO
            query_extrato = """
                SELECT 
                    data,
                    valor,
                    descricao,
                    conta_bancaria,
                    COUNT(*) as quantidade,
                    STRING_AGG(CAST(id AS TEXT), ', ' ORDER BY id) as ids,
                    'extrato' as origem
                FROM transacoes_extrato
                WHERE empresa_id = %s
                  AND tipo = 'DEBITO'  -- Apenas d�bitos (sa�das/pagamentos)
            """
            params_extrato = [empresa_id]
            
            if data_inicio:
                query_extrato += " AND data >= %s"
                params_extrato.append(data_inicio)
            
            if data_fim:
                query_extrato += " AND data <= %s"
                params_extrato.append(data_fim)
            
            if conta_bancaria:
                query_extrato += " AND conta_bancaria = %s"
                params_extrato.append(conta_bancaria)
            
            query_extrato += """
                GROUP BY data, valor, descricao, conta_bancaria
                HAVING COUNT(*) > 1
                ORDER BY quantidade DESC, data DESC, ABS(valor) DESC
            """
            
            cursor.execute(query_extrato, params_extrato)
            duplicatas_extrato = cursor.fetchall()
            
            logger.info(f"   ?? Duplicatas no extrato: {len(duplicatas_extrato)}")
            
            # Query para detectar duplicatas nos LAN�AMENTOS (Contas a Pagar)
            # Nota: tabela lancamentos usa colunas cliente_fornecedor/pessoa para nomes, n�o IDs
            query_lancamentos = """
                SELECT 
                    data_vencimento as data,
                    valor,
                    COALESCE(cliente_fornecedor, pessoa, 'Sem benefici�rio') as beneficiario,
                    categoria,
                    conta_bancaria,
                    COUNT(*) as quantidade,
                    STRING_AGG(CAST(id AS TEXT), ', ' ORDER BY id) as ids,
                    'lancamentos' as origem
                FROM lancamentos
                WHERE empresa_id = %s
                  AND UPPER(tipo) = 'DESPESA'
                  AND UPPER(status) = 'PAGO'
            """
            params_lancamentos = [empresa_id]
            
            if data_inicio:
                query_lancamentos += " AND data_vencimento >= %s"
                params_lancamentos.append(data_inicio)
            
            if data_fim:
                query_lancamentos += " AND data_vencimento <= %s"
                params_lancamentos.append(data_fim)
            
            if conta_bancaria:
                query_lancamentos += " AND conta_bancaria = %s"
                params_lancamentos.append(conta_bancaria)
            
            query_lancamentos += """
                GROUP BY data_vencimento, valor, COALESCE(cliente_fornecedor, pessoa, 'Sem benefici�rio'), 
                         categoria, conta_bancaria
                HAVING COUNT(*) > 1
                ORDER BY quantidade DESC, data DESC, ABS(valor) DESC
            """
            
            cursor.execute(query_lancamentos, params_lancamentos)
            duplicatas_lancamentos = cursor.fetchall()
            
            logger.info(f"   ?? Duplicatas em lan�amentos: {len(duplicatas_lancamentos)}")
            
            # Calcular valor total duplicado
            total_duplicado_extrato = sum(
                abs(float(d['valor'])) * (int(d['quantidade']) - 1) 
                for d in duplicatas_extrato
            )
            
            total_duplicado_lancamentos = sum(
                abs(float(d['valor'])) * (int(d['quantidade']) - 1) 
                for d in duplicatas_lancamentos
            )
            
            cursor.close()
            
            resultado = {
                'success': True,
                'resumo': {
                    'total_grupos_extrato': len(duplicatas_extrato),
                    'total_grupos_lancamentos': len(duplicatas_lancamentos),
                    'valor_duplicado_extrato': float(total_duplicado_extrato),
                    'valor_duplicado_lancamentos': float(total_duplicado_lancamentos),
                    'total_valor_duplicado': float(total_duplicado_extrato + total_duplicado_lancamentos)
                },
                'duplicatas_extrato': [dict(d) for d in duplicatas_extrato],
                'duplicatas_lancamentos': [dict(d) for d in duplicatas_lancamentos],
                'filtros_aplicados': {
                    'data_inicio': data_inicio,
                    'data_fim': data_fim,
                    'conta_bancaria': conta_bancaria
                }
            }
            
            logger.info(f"? Auditoria conclu�da - Total duplicado: R$ {total_duplicado_extrato + total_duplicado_lancamentos:,.2f}")
            return jsonify(resultado), 200
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.error(f"? Erro na auditoria: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


print("?? Registrando rota: /api/extratos/conciliacao-geral")

@app.route('/api/extratos/conciliacao-geral', methods=['POST'])
@require_permission('lancamentos_create')
def conciliacao_geral_extrato():
    """Concilia��o autom�tica em massa de transa��es do extrato para contas a pagar/receber"""
    # Logs reduzidos para evitar polui��o
    try:
        logger.info("?? CONCILIA��O GERAL INICIADA")
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        logger.info(f"?? Usu�rio: {usuario.get('username')} | Empresa ID: {empresa_id}")
        
        dados = request.json
        transacoes = dados.get('transacoes', [])
        print(f"?? Recebidas {len(transacoes)} transa��o(�es) para conciliar")
        logger.info(f"?? Recebidas {len(transacoes)} transa��o(�es) para conciliar")
        print(f"?? Dados: {dados}")
        logger.info(f"?? Dados recebidos: {dados}")
        
        if not transacoes:
            return jsonify({'success': False, 'error': 'Nenhuma transa��o selecionada'}), 400
        
        # Buscar clientes e fornecedores para matching de CPF/CNPJ
        clientes = db.listar_clientes(ativos=True)
        fornecedores = db.listar_fornecedores(ativos=True)
        
        # Criar dicion�rios de busca r�pida por CPF/CNPJ
        clientes_dict = {}
        for cliente in clientes:
            cpf_cnpj = cliente.get('cpf') or cliente.get('cnpj')
            if cpf_cnpj:
                # Normalizar (remover pontos, tra�os, barras)
                cpf_cnpj_limpo = ''.join(filter(str.isdigit, str(cpf_cnpj)))
                clientes_dict[cpf_cnpj_limpo] = cliente['nome']
        
        fornecedores_dict = {}
        for fornecedor in fornecedores:
            cpf_cnpj = fornecedor.get('cpf') or fornecedor.get('cnpj')
            if cpf_cnpj:
                cpf_cnpj_limpo = ''.join(filter(str.isdigit, str(cpf_cnpj)))
                fornecedores_dict[cpf_cnpj_limpo] = fornecedor['nome']
        
        criados = 0
        erros = []
        
        for item in transacoes:
            try:
                transacao_id = item.get('transacao_id')
                categoria = item.get('categoria')
                subcategoria = item.get('subcategoria', '')
                razao_social = item.get('razao_social', '')
                descricao_personalizada = item.get('descricao', '')
                
                # Buscar transa��o do extrato
                with db.get_connection() as conn:
                    import psycopg2.extras
                    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                    cursor.execute(
                        "SELECT * FROM transacoes_extrato WHERE id = %s AND empresa_id = %s",
                        (transacao_id, empresa_id)
                    )
                    transacao = cursor.fetchone()
                    cursor.close()
                
                if not transacao:
                    erros.append(f"Transa��o {transacao_id} n�o encontrada")
                    continue
                
                # Validar se a conta banc�ria est� ativa
                conta_bancaria = transacao['conta_bancaria']
                print(f"?? Validando conta banc�ria: {conta_bancaria}")
                contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
                print(f"?? Total de contas encontradas: {len(contas)}")
                
                # Debug: listar todas as contas
                for c in contas:
                    print(f"   - Conta cadastrada: '{c.nome}' (ativa={c.ativa if hasattr(c, 'ativa') else 'N/A'})")
                
                conta = next((c for c in contas if c.nome == conta_bancaria), None)
                
                if not conta:
                    erros.append(f"Transa��o {transacao_id}: A conta banc�ria '{conta_bancaria}' n�o est� cadastrada no sistema ou o nome n�o corresponde exatamente. Verifique o cadastro de contas.")
                    print(f"? Concilia��o bloqueada: conta '{conta_bancaria}' n�o encontrada")
                    logger.warning(f"? Tentativa de conciliar com conta n�o cadastrada: {conta_bancaria}")
                    continue
                
                print(f"? Conta encontrada: {conta.nome}")
                print(f"?? Campo ativa existe? {hasattr(conta, 'ativa')}")
                print(f"?? Valor do campo ativa: {conta.ativa if hasattr(conta, 'ativa') else 'N/A'}")
                
                if hasattr(conta, 'ativa') and not conta.ativa:
                    erros.append(f"Transa��o {transacao_id}: A conta banc�ria '{conta_bancaria}' est� inativa. Reative a conta antes de conciliar.")
                    print(f"? Concilia��o bloqueada: conta {conta_bancaria} est� inativa")
                    logger.warning(f"? Tentativa de conciliar com conta inativa: {conta_bancaria}")
                    continue
                
                # Detectar CPF/CNPJ na descri��o (regex simples)
                import re
                descricao = transacao['descricao']
                cpf_cnpj_encontrado = None
                
                # Buscar CPF (11 d�gitos) ou CNPJ (14 d�gitos)
                numeros = ''.join(filter(str.isdigit, descricao))
                if len(numeros) == 11 or len(numeros) == 14:
                    cpf_cnpj_encontrado = numeros
                
                # Tentar matching autom�tico se n�o foi fornecida raz�o social
                if not razao_social and cpf_cnpj_encontrado:
                    if transacao['tipo'].upper() == 'CREDITO':
                        razao_social = clientes_dict.get(cpf_cnpj_encontrado, '')
                    else:
                        razao_social = fornecedores_dict.get(cpf_cnpj_encontrado, '')
                
                # ?? CONCILIAR TRANSA��O E CRIAR LAN�AMENTO AUTOMATICAMENTE
                # Usa a fun��o conciliar_transacao() que agora cria lan�amento com status='PAGO'
                
                print(f"?? Conciliando transa��o {transacao_id} (criando lan�amento)...")
                logger.info(f"?? Conciliando transa��o {transacao_id} - empresa_id: {empresa_id}")
                
                # Importar fun��o de concilia��o
                from extrato_functions import conciliar_transacao
                
                # Conciliar com cria��o autom�tica de lan�amento (lancamento_id='auto')
                resultado = conciliar_transacao(
                    database=db,
                    empresa_id=empresa_id,
                    transacao_id=transacao_id,
                    lancamento_id='auto'  # Cria novo lan�amento automaticamente
                )
                
                if resultado.get('success'):
                    lancamento_id = resultado.get('lancamento_id')
                    
                    # Atualizar campos extras no LANÇAMENTO e na transação do extrato
                    if lancamento_id and (categoria or subcategoria or razao_social or descricao_personalizada):
                        with db.get_db_connection(empresa_id=empresa_id) as conn:
                            cursor_update = conn.cursor()

                            # ✅ FIX: Atualizar o LANÇAMENTO com os dados selecionados pelo usuário
                            # (categoria, subcategoria, pessoa, descrição personalizada)
                            cursor_update.execute("""
                                UPDATE lancamentos
                                SET
                                    categoria    = COALESCE(NULLIF(%s, ''), categoria),
                                    subcategoria = COALESCE(NULLIF(%s, ''), subcategoria),
                                    pessoa       = COALESCE(NULLIF(%s, ''), pessoa),
                                    descricao    = CASE WHEN %s != '' THEN %s ELSE descricao END
                                WHERE id = %s AND empresa_id = %s
                            """, (
                                categoria or '',
                                subcategoria or '',
                                razao_social or '',
                                descricao_personalizada or '',
                                descricao_personalizada or '',
                                lancamento_id,
                                empresa_id
                            ))

                            # Atualizar também a transação do extrato para consistência
                            cursor_update.execute("""
                                UPDATE transacoes_extrato 
                                SET 
                                    categoria = COALESCE(%s, categoria),
                                    subcategoria = COALESCE(%s, subcategoria),
                                    pessoa = COALESCE(%s, pessoa)
                                WHERE id = %s AND empresa_id = %s
                            """, (
                                categoria if categoria else None,
                                subcategoria if subcategoria else None,
                                razao_social if razao_social else None,
                                transacao_id,
                                empresa_id
                            ))
                            conn.commit()
                            cursor_update.close()
                    
                    print(f"? Transa��o {transacao_id} conciliada ? lan�amento #{lancamento_id} criado com status PAGO")
                    logger.info(f"? Transa��o {transacao_id} conciliada ? lan�amento #{lancamento_id}")
                    criados += 1
                else:
                    erro_msg = resultado.get('error', 'Erro desconhecido')
                    erros.append(f"Transa��o {transacao_id}: {erro_msg}")
                    logger.error(f"? Falha ao conciliar transa��o {transacao_id}: {erro_msg}")
                    continue
                
            except Exception as e:
                erro_msg = f"Erro na transa��o {item.get('transacao_id')}: {str(e)}"
                print(f"? {erro_msg}")
                erros.append(erro_msg)
                logger.error(f"Erro ao conciliar transa��o {item.get('transacao_id')}: {e}")
                import traceback
                print(traceback.format_exc())
                traceback.print_exc()
        
        # Determinar status de sucesso
        success = criados > 0
        status_code = 200 if success else 400
        
        if not success and erros:
            # Se nenhuma transa��o foi conciliada e h� erros, retornar erro
            return jsonify({
                'success': False,
                'criados': 0,
                'erros': erros,
                'message': erros[0] if len(erros) == 1 else f'{len(erros)} erro(s) encontrado(s)'
            }), 400
        
        return jsonify({
            'success': success,
            'criados': criados,
            'erros': erros,
            'message': f'{criados} lan�amento(s) criado(s) com sucesso' + (f'. {len(erros)} erro(s).' if erros else '')
        }), status_code
        
    except Exception as e:
        logger.error(f"Erro na concilia��o geral: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/extratos/<int:transacao_id>/desconciliar', methods=['POST'])
@require_permission('lancamentos_delete')
def desconciliar_extrato(transacao_id):
    """Desfaz a concilia��o de uma transa��o do extrato e exclui o lan�amento"""
    try:
        print("\n" + "="*80)
        print(f"?? DESCONCILIA��O INICIADA - Transa��o ID: {transacao_id}")
        
        usuario = get_usuario_logado()
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        conn = db.get_connection()
        import psycopg2.extras
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            # Buscar transa��o do extrato
            cursor.execute(
                "SELECT * FROM transacoes_extrato WHERE id = %s AND empresa_id = %s",
                (transacao_id, empresa_id)
            )
            transacao = cursor.fetchone()
            
            if not transacao:
                cursor.close()
                from database_postgresql import return_to_pool
                return_to_pool(conn)
                return jsonify({'success': False, 'error': 'Transa��o n�o encontrada'}), 404
            
            if not transacao['conciliado']:
                cursor.close()
                from database_postgresql import return_to_pool
                return_to_pool(conn)
                return jsonify({'success': False, 'error': 'Transa��o n�o est� conciliada'}), 400
            
            print(f"?? Transa��o: ID={transacao_id}, Conciliado={transacao['conciliado']}")
            
            # Buscar lancamento_id da tabela conciliacoes
            cursor.execute(
                "SELECT lancamento_id FROM conciliacoes WHERE transacao_extrato_id = %s AND empresa_id = %s",
                (transacao_id, empresa_id)
            )
            conciliacao = cursor.fetchone()
            
            lancamento_id = conciliacao['lancamento_id'] if conciliacao else None

            # Registrar no historico ANTES de excluir o lancamento
            try:
                _garantir_tabela_historico_conciliacoes(conn)
                if lancamento_id:
                    cursor.execute("""
                        INSERT INTO historico_conciliacoes (
                            empresa_id, evento, transacao_extrato_id, lancamento_id,
                            data_transacao, conta_bancaria, descricao_extrato, valor, tipo_extrato,
                            descricao_lancamento, categoria, subcategoria, pessoa, observacoes, memo, fitid
                        )
                        SELECT
                            te.empresa_id, 'desconciliado', te.id, %s,
                            te.data, te.conta_bancaria, te.descricao, ABS(COALESCE(te.valor, 0)), te.tipo,
                            COALESCE(l.descricao, te.descricao),
                            COALESCE(l.categoria, te.categoria),
                            COALESCE(l.subcategoria, te.subcategoria),
                            COALESCE(l.pessoa, te.pessoa),
                            l.observacoes, te.memo, te.fitid
                        FROM transacoes_extrato te
                        LEFT JOIN lancamentos l ON l.id = %s AND l.empresa_id = te.empresa_id
                        WHERE te.id = %s AND te.empresa_id = %s
                    """, (lancamento_id, lancamento_id, transacao_id, empresa_id))
            except Exception as _he:
                print(f"Historico insert warning: {_he}")
            print(f"?? Lan�amento ID: {lancamento_id}")
            
            # Excluir lan�amento se existir
            if lancamento_id:
                print(f"??? Excluindo lan�amento ID={lancamento_id}")
                db.excluir_lancamento(lancamento_id)
                print(f"? Lan�amento {lancamento_id} exclu�do")
            
            # Deletar da tabela conciliacoes
            print(f"??? Removendo da tabela conciliacoes...")
            cursor.execute(
                "DELETE FROM conciliacoes WHERE transacao_extrato_id = %s AND empresa_id = %s",
                (transacao_id, empresa_id)
            )
            affected_conciliacoes = cursor.rowcount
            print(f"?? DELETE conciliacoes: {affected_conciliacoes} linha(s) deletada(s)")
            
            # Atualizar transa��o: desconciliar
            print(f"?? Atualizando flag conciliado -> FALSE")
            cursor.execute(
                "UPDATE transacoes_extrato SET conciliado = FALSE WHERE id = %s AND empresa_id = %s",
                (transacao_id, empresa_id)
            )
            affected_rows = cursor.rowcount
            print(f"?? UPDATE transacoes_extrato: {affected_rows} linha(s) atualizada(s)")
            
            conn.commit()
            print("? COMMIT OK")
            
            cursor.close()
            from database_postgresql import return_to_pool
            return_to_pool(conn)
            
            print(f"? Desconcilia��o conclu�da com sucesso!")
            print("="*80 + "\n")
            
            return jsonify({
                'success': True,
                'message': 'Desconcilia��o realizada com sucesso'
            }), 200
            
        except Exception as e:
            conn.rollback()
            cursor.close()
            from database_postgresql import return_to_pool
            return_to_pool(conn)
            raise
        
    except Exception as e:
        print(f"? Erro na desconcilia��o: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ROTAS DE REGRAS DE AUTO-CONCILIA��O
# ============================================================================

@app.route('/api/regras-conciliacao', methods=['GET'])
@require_permission('regras_conciliacao_view')
def listar_regras_conciliacao():
    """Lista todas as regras de auto-concilia��o da empresa"""
    try:
        print("?? [DEBUG] Iniciando listar_regras_conciliacao")
        
        empresa_id = session.get('empresa_id')
        print(f"?? [DEBUG] empresa_id: {empresa_id}")
        
        if not empresa_id:
            print("? [DEBUG] Empresa n�o selecionada")
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        print(f"?? [DEBUG] Chamando db.listar_regras_conciliacao(empresa_id={empresa_id})")
        regras = db.listar_regras_conciliacao(empresa_id=empresa_id)
        print(f"? [DEBUG] Regras retornadas: {len(regras) if regras else 0}")
        
        return jsonify({
            'success': True,
            'data': regras
        }), 200
        
    except Exception as e:
        print(f"? [DEBUG] ERRO: {e}")
        import traceback
        traceback.print_exc()
        logger.error(f"Erro ao listar regras de concilia��o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/regras-conciliacao', methods=['POST'])
@require_permission('regras_conciliacao_create')
def criar_regra_conciliacao():
    """Cria nova regra de auto-concilia��o"""
    try:
        print("?? [DEBUG] Iniciando criar_regra_conciliacao")
        
        empresa_id = session.get('empresa_id')
        print(f"?? [DEBUG] empresa_id: {empresa_id}")
        
        if not empresa_id:
            print("? [DEBUG] Empresa n�o selecionada")
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        dados = request.json
        print(f"?? [DEBUG] Dados recebidos: {dados}")
        
        # Validar campos obrigat�rios
        if not dados.get('palavra_chave'):
            print("? [DEBUG] Palavra-chave n�o fornecida")
            return jsonify({'success': False, 'error': 'Palavra-chave � obrigat�ria'}), 400
        
        print(f"?? [DEBUG] Chamando db.criar_regra_conciliacao")
        regra = db.criar_regra_conciliacao(
            empresa_id=empresa_id,
            palavra_chave=dados.get('palavra_chave'),
            categoria=dados.get('categoria'),
            subcategoria=dados.get('subcategoria'),
            cliente_padrao=dados.get('cliente_padrao'),
            descricao=dados.get('descricao')
        )
        print(f"? [DEBUG] Regra criada: {regra}")
        
        return jsonify({
            'success': True,
            'message': 'Regra criada com sucesso',
            'data': regra
        }), 201
        
    except ValueError as e:
        # Erro de valida��o (ex: regra duplicada)
        print(f"?? [DEBUG] ERRO DE VALIDA��O: {e}")
        return jsonify({'success': False, 'error': str(e)}), 400
        
    except Exception as e:
        print(f"? [DEBUG] ERRO: {e}")
        import traceback
        traceback.print_exc()
        logger.error(f"Erro ao criar regra de concilia��o: {e}")
        return jsonify({'success': False, 'error': 'Erro interno ao criar regra'}), 500


@app.route('/api/regras-conciliacao/<int:regra_id>', methods=['PUT'])
@require_permission('regras_conciliacao_edit')
def atualizar_regra_conciliacao(regra_id):
    """Atualiza uma regra de auto-concilia��o"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        dados = request.json
        
        sucesso = db.atualizar_regra_conciliacao(
            regra_id=regra_id,
            empresa_id=empresa_id,
            **dados
        )
        
        if sucesso:
            return jsonify({
                'success': True,
                'message': 'Regra atualizada com sucesso'
            }), 200
        else:
            return jsonify({'success': False, 'error': 'Regra n�o encontrada ou sem permiss�o'}), 404
        
    except Exception as e:
        logger.error(f"Erro ao atualizar regra de concilia��o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/regras-conciliacao/<int:regra_id>', methods=['DELETE'])
@require_permission('regras_conciliacao_delete')
def excluir_regra_conciliacao(regra_id):
    """Exclui uma regra de auto-concilia��o"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        sucesso = db.excluir_regra_conciliacao(
            regra_id=regra_id,
            empresa_id=empresa_id
        )
        
        if sucesso:
            return jsonify({
                'success': True,
                'message': 'Regra exclu�da com sucesso'
            }), 200
        else:
            return jsonify({'success': False, 'error': 'Regra n�o encontrada ou sem permiss�o'}), 404
        
    except Exception as e:
        logger.error(f"Erro ao excluir regra de concilia��o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# CONFIGURA��ES DE EXTRATO BANC�RIO
# ============================================================================

@app.route('/api/config-extrato', methods=['GET'])
@require_permission('config_extrato_bancario_view')
def obter_config_extrato():
    """
    Obt�m configura��es de extrato banc�rio da empresa
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        config = db.obter_config_extrato(empresa_id)
        
        return jsonify({
            'success': True,
            'data': config
        }), 200
    except Exception as e:
        logger.error(f"Erro ao obter configura��o de extrato: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config-extrato', methods=['PUT'])
@require_permission('config_extrato_bancario_edit')
def atualizar_config_extrato():
    """
    Atualiza configura��es de extrato banc�rio
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        dados = request.json
        integrar_folha = dados.get('integrar_folha_pagamento', False)
        
        sucesso = db.atualizar_config_extrato(
            empresa_id=empresa_id,
            integrar_folha=integrar_folha
        )
        
        if sucesso:
            return jsonify({
                'success': True,
                'message': 'Configura��o atualizada com sucesso'
            }), 200
        else:
            return jsonify({'success': False, 'error': 'Erro ao atualizar configura��o'}), 500
    except Exception as e:
        logger.error(f"Erro ao atualizar configura��o de extrato: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/regras-conciliacao/detectar', methods=['POST'])
@limiter.exempt  # Excluir do rate limiting (pode receber 694+ requisi��es paralelas)
@require_permission('lancamentos_view')
def detectar_regra_conciliacao():
    """
    Detecta regra aplic�vel e funcion�rio (se integra��o folha ativa)
    para uma descri��o de extrato
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        dados = request.json
        descricao = dados.get('descricao', '')
        
        if not descricao:
            return jsonify({'success': False, 'error': 'Descri��o � obrigat�ria'}), 400
        
        # Buscar regra aplic�vel
        regra = db.buscar_regra_aplicavel(empresa_id=empresa_id, descricao=descricao)
        
        resultado = {
            'success': True,
            'regra_encontrada': regra is not None,
            'regra': regra,
            'funcionario': None
        }
        
        # Se regra tem integra��o com folha, buscar CPF na descri��o
        if regra and regra.get('usa_integracao_folha'):
            import re
            # Buscar CPF na descri��o (11 d�gitos consecutivos)
            cpf_match = re.search(r'\b(\d{11})\b', descricao)
            
            if cpf_match:
                cpf = cpf_match.group(1)
                funcionario = db.buscar_funcionario_por_cpf(empresa_id=empresa_id, cpf=cpf)
                
                if funcionario:
                    resultado['funcionario'] = {
                        'nome': funcionario.get('nome'),
                        'cpf': funcionario.get('cpf'),
                        'cargo': funcionario.get('cargo')
                    }
        
        return jsonify(resultado), 200
        
    except Exception as e:
        logger.error(f"Erro ao detectar regra de concilia��o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/regras-conciliacao/detectar-batch', methods=['POST'])
@limiter.exempt
@require_permission('lancamentos_view')
def detectar_regras_batch():
    """
    Detecta regras aplic�veis em lote para m�ltiplas descri��es.
    Reduz 694 requisi��es para 1 �nica requisi��o = MUITO mais r�pido!
    
    Recebe: { "transacoes": [ { "id": 8745, "descricao": "PAGAMENTO PIX..." }, ... ] }
    Retorna: { "success": true, "resultados": [ { "id": 8745, "regra": {...}, "funcionario": {...} }, ... ] }
    """
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        dados = request.json
        transacoes = dados.get('transacoes', [])
        
        if not transacoes or not isinstance(transacoes, list):
            return jsonify({'success': False, 'error': 'Lista de transa��es � obrigat�ria'}), 400
        
        resultados = []
        import re
        
        # Processar cada transa��o
        for transacao in transacoes:
            transacao_id = transacao.get('id')
            descricao = transacao.get('descricao', '')
            
            if not descricao:
                resultados.append({
                    'id': transacao_id,
                    'regra_encontrada': False,
                    'regra': None,
                    'funcionario': None
                })
                continue
            
            # Buscar regra aplic�vel
            regra = db.buscar_regra_aplicavel(empresa_id=empresa_id, descricao=descricao)
            
            resultado = {
                'id': transacao_id,
                'regra_encontrada': regra is not None,
                'regra': regra,
                'funcionario': None
            }
            
            # Se regra tem integra��o com folha, buscar CPF na descri��o
            if regra and regra.get('usa_integracao_folha'):
                cpf_match = re.search(r'\b(\d{11})\b', descricao)
                
                if cpf_match:
                    cpf = cpf_match.group(1)
                    funcionario = db.buscar_funcionario_por_cpf(empresa_id=empresa_id, cpf=cpf)
                    
                    if funcionario:
                        resultado['funcionario'] = {
                            'nome': funcionario.get('nome'),
                            'cpf': funcionario.get('cpf'),
                            'cargo': funcionario.get('cargo')
                        }
            
            resultados.append(resultado)
        
        return jsonify({
            'success': True,
            'resultados': resultados,
            'total_processadas': len(resultados)
        }), 200
        
    except Exception as e:
        logger.error(f"Erro ao detectar regras em lote: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# === ROTAS DE FOLHA DE PAGAMENTO (FUNCION�RIOS) ===

@app.route('/api/funcionarios', methods=['GET'])
@require_permission('folha_pagamento_view')
def listar_funcionarios():
    """Listar todos os funcion�rios da empresa"""
    try:
        usuario = get_usuario_logado()
        
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        logger.info(f"?? [FUNCIONARIOS] Usuario: {usuario.get('username')}")
        logger.info(f"   cliente_id: {usuario.get('cliente_id')}")
        logger.info(f"   empresa_id: {usuario.get('empresa_id')}")
        logger.info(f"   empresas: {usuario.get('empresas', [])}")
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        logger.info(f"   ?? empresa_id final: {empresa_id}")
        
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT id, empresa_id, nome, cpf, email, celular, ativo, data_admissao, 
                   data_demissao, observacoes, nacionalidade, estado_civil, data_nascimento, 
                   idade, profissao, rua_av, numero_residencia, complemento, bairro, cidade, 
                   estado, cep, chave_pix, pis_pasep, data_criacao, data_atualizacao
            FROM funcionarios
            WHERE empresa_id = %s
            ORDER BY nome ASC
        """
        
        logger.info(f"?? [FUNCIONARIOS] Executando query com empresa_id = {empresa_id}")
        cursor.execute(query, (empresa_id,))
        rows = cursor.fetchall()
        logger.info(f"? [FUNCIONARIOS] Encontrados {len(rows)} funcion�rios")
        
        # Debug: Log primeiro funcion�rio
        if rows:
            logger.info(f"?? [DEBUG] Primeiro funcion�rio (tipo: {type(rows[0])})")
            if isinstance(rows[0], dict):
                logger.info(f"   Dict keys: {list(rows[0].keys())}")
            else:
                logger.info(f"   Tupla length: {len(rows[0])}")
        
        cursor.close()
        
        funcionarios = []
        for row in rows:
            # Verifica se row � dict ou tupla
            if isinstance(row, dict):
                funcionarios.append({
                    'id': row['id'],
                    'empresa_id': row['empresa_id'],
                    'nome': row['nome'],
                    'cpf': row.get('cpf'),
                    'email': row.get('email'),
                    'celular': row.get('celular'),
                    'ativo': row.get('ativo', True),
                    'data_admissao': row['data_admissao'].isoformat() if row.get('data_admissao') else None,
                    'data_demissao': row['data_demissao'].isoformat() if row.get('data_demissao') else None,
                    'observacoes': row.get('observacoes'),
                    'nacionalidade': row.get('nacionalidade'),
                    'estado_civil': row.get('estado_civil'),
                    'data_nascimento': row['data_nascimento'].isoformat() if row.get('data_nascimento') else None,
                    'idade': row.get('idade'),
                    'profissao': row.get('profissao'),
                    'rua_av': row.get('rua_av'),
                    'numero_residencia': row.get('numero_residencia'),
                    'complemento': row.get('complemento'),
                    'bairro': row.get('bairro'),
                    'cidade': row.get('cidade'),
                    'estado': row.get('estado'),
                    'cep': row.get('cep'),
                    'chave_pix': row.get('chave_pix'),
                    'pis_pasep': row.get('pis_pasep'),
                    'data_criacao': row['data_criacao'].isoformat() if row.get('data_criacao') else None,
                    'data_atualizacao': row['data_atualizacao'].isoformat() if row.get('data_atualizacao') else None
                })
            else:
                funcionarios.append({
                    'id': row[0],
                    'empresa_id': row[1],
                    'nome': row[2],
                    'cpf': row[3],
                    'email': row[4],
                    'celular': row[5],
                    'ativo': row[6] if row[6] is not None else True,
                    'data_admissao': row[7].isoformat() if row[7] else None,
                    'data_demissao': row[8].isoformat() if row[8] else None,
                    'observacoes': row[9],
                    'nacionalidade': row[10],
                    'estado_civil': row[11],
                    'data_nascimento': row[12].isoformat() if row[12] else None,
                    'idade': row[13],
                    'profissao': row[14],
                    'rua_av': row[15],
                    'numero_residencia': row[16],
                    'complemento': row[17],
                    'bairro': row[18],
                    'cidade': row[19],
                    'estado': row[20],
                    'cep': row[21],
                    'chave_pix': row[22],
                    'pis_pasep': row[23],
                    'data_criacao': row[24].isoformat() if row[24] else None,
                    'data_atualizacao': row[25].isoformat() if row[25] else None
                })
        
        # Log primeiro funcion�rio completo para debug
        if funcionarios:
            logger.info(f"?? [DEBUG] Primeiro funcion�rio sendo enviado:")
            logger.info(f"   Nome: {funcionarios[0].get('nome')}")
            logger.info(f"   CPF: {funcionarios[0].get('cpf')}")
            logger.info(f"   Nacionalidade: {funcionarios[0].get('nacionalidade')}")
            logger.info(f"   Estado Civil: {funcionarios[0].get('estado_civil')}")
            logger.info(f"   Email: {funcionarios[0].get('email')}")
        
        return jsonify({'funcionarios': funcionarios}), 200
    
    except Exception as e:
        logger.error(f"Erro ao listar funcion�rios: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios/cpf/relatorio', methods=['GET'])
@require_permission('folha_pagamento_view')
def relatorio_cpfs_invalidos():
    """
    ?? Relat�rio de CPFs Inv�lidos
    ================================
    
    Retorna lista de funcion�rios com CPFs inv�lidos ou ausentes.
    
    Resposta:
        - total_funcionarios: total de funcion�rios da empresa
        - total_cpfs_invalidos: quantidade de CPFs inv�lidos
        - total_cpfs_ausentes: quantidade de CPFs n�o informados
        - taxa_erro: percentual de erros (%)
        - funcionarios_invalidos: lista detalhada com erros
    """
    # Import local para evitar falha de carregamento do m�dulo
    from cpf_validator import CPFValidator
    
    try:
        print("\n?? [CPF RELATORIO] Iniciando an�lise...")
        
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Buscar todos os funcion�rios da empresa
        query = """
            SELECT id, nome, cpf, email, celular, ativo, data_admissao, data_demissao
            FROM funcionarios
            WHERE empresa_id = %s
            ORDER BY nome ASC
        """
        
        cursor.execute(query, (empresa_id,))
        rows = cursor.fetchall()
        cursor.close()
        
        # An�lise de CPFs
        total_funcionarios = len(rows)
        funcionarios_invalidos = []
        funcionarios_ausentes = []
        
        for row in rows:
            if isinstance(row, dict):
                func_id = row['id']
                nome = row['nome']
                cpf = row.get('cpf')
                email = row.get('email')
                celular = row.get('celular')
                ativo = row.get('ativo', True)
                data_admissao = row['data_admissao'].isoformat() if row.get('data_admissao') else None
                data_demissao = row['data_demissao'].isoformat() if row.get('data_demissao') else None
            else:
                func_id, nome, cpf, email, celular, ativo, data_admissao, data_demissao = row[:8]
                data_admissao = data_admissao.isoformat() if data_admissao else None
                data_demissao = data_demissao.isoformat() if data_demissao else None
            
            # Validar CPF
            if not cpf or cpf.strip() == '':
                funcionarios_ausentes.append({
                    'id': func_id,
                    'nome': nome,
                    'cpf': None,
                    'email': email,
                    'celular': celular,
                    'ativo': ativo,
                    'data_admissao': data_admissao,
                    'data_demissao': data_demissao,
                    'erro': 'CPF n�o informado',
                    'tipo_erro': 'ausente'
                })
            else:
                validacao = CPFValidator.validar_com_detalhes(cpf)
                if not validacao['valido']:
                    funcionarios_invalidos.append({
                        'id': func_id,
                        'nome': nome,
                        'cpf': cpf,
                        'cpf_formatado': CPFValidator.formatar(cpf) if len(CPFValidator.limpar(cpf)) == 11 else cpf,
                        'email': email,
                        'celular': celular,
                        'ativo': ativo,
                        'data_admissao': data_admissao,
                        'data_demissao': data_demissao,
                        'erro': validacao['erro'],
                        'tipo_erro': 'invalido'
                    })
        
        # Calcular estat�sticas
        total_invalidos = len(funcionarios_invalidos)
        total_ausentes = len(funcionarios_ausentes)
        total_problemas = total_invalidos + total_ausentes
        taxa_erro = round((total_problemas / total_funcionarios * 100), 2) if total_funcionarios > 0 else 0
        taxa_validos = round(((total_funcionarios - total_problemas) / total_funcionarios * 100), 2) if total_funcionarios > 0 else 0
        
        # Combinar listas
        todos_problemas = funcionarios_invalidos + funcionarios_ausentes
        
        print(f"? [CPF RELATORIO] An�lise conclu�da:")
        print(f"   Total: {total_funcionarios} funcion�rios")
        print(f"   Inv�lidos: {total_invalidos}")
        print(f"   Ausentes: {total_ausentes}")
        print(f"   Taxa de erro: {taxa_erro}%")
        
        return jsonify({
            'success': True,
            'total_funcionarios': total_funcionarios,
            'total_cpfs_validos': total_funcionarios - total_problemas,
            'total_cpfs_invalidos': total_invalidos,
            'total_cpfs_ausentes': total_ausentes,
            'total_cpfs_problemas': total_problemas,
            'taxa_erro': taxa_erro,
            'taxa_validos': taxa_validos,
            'funcionarios_com_problemas': todos_problemas,
            'data_analise': time.strftime('%Y-%m-%d %H:%M:%S')
        }), 200
        
    except Exception as e:
        logger.error(f"? Erro ao gerar relat�rio de CPFs: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios/cpf/correcao', methods=['GET'])
@require_permission('folha_pagamento_edit')
def gerar_correcoes_cpf():
    """Gera sugest�es de corre��o autom�tica para CPFs inv�lidos"""
    import traceback
    import sys
    
    # IMPORTAR M�DULOS CPF DENTRO DA FUN��O (para debug)
    try:
        logger.info("?? [CPF] Importando CPFValidator...")
        from cpf_validator import CPFValidator as CPFVal
        logger.info("? [CPF] CPFValidator importado com sucesso")
    except Exception as import_error:
        logger.error(f"? [CPF] ERRO AO IMPORTAR CPFValidator: {import_error}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Erro ao importar CPFValidator: {str(import_error)}',
            'traceback': traceback.format_exc()
        }), 500
    
    try:
        logger.info("?? [CPF] Importando CPFCorrector...")
        from cpf_corrector import CPFCorrector as CPFCorr
        logger.info("? [CPF] CPFCorrector importado com sucesso")
    except Exception as import_error:
        logger.error(f"? [CPF] ERRO AO IMPORTAR CPFCorrector: {import_error}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Erro ao importar CPFCorrector: {str(import_error)}',
            'traceback': traceback.format_exc()
        }), 500
    
    try:
        logger.info("=" * 80)
        logger.info("?? [CPF CORRETOR] === IN�CIO DA EXECU��O ===")
        logger.info("=" * 80)
        
        # Obter empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            logger.error("? [CPF CORRETOR] Empresa n�o selecionada")
            return jsonify({'error': 'Empresa n�o selecionada'}), 403
        
        logger.info(f"? [CPF CORRETOR] Empresa ID: {empresa_id}")
        
        # Buscar funcion�rios da empresa
        conn = None
        cursor = None
        try:
            logger.info("?? [CPF CORRETOR] Conectando ao banco de dados...")
            conn = db.get_connection()
            cursor = conn.cursor()
            
            logger.info("?? [CPF CORRETOR] Executando query...")
            query = """
                SELECT id, nome, cpf
                FROM funcionarios
                WHERE empresa_id = %s
                ORDER BY nome ASC
            """
            
            cursor.execute(query, (empresa_id,))
            rows = cursor.fetchall()
            
            logger.info(f"? [CPF CORRETOR] Encontrados {len(rows)} funcion�rios no banco")
            
        except Exception as db_error:
            logger.error(f"? [CPF CORRETOR] Erro na consulta ao banco: {db_error}")
            logger.error(traceback.format_exc())
            return jsonify({
                'success': False,
                'error': f'Erro ao acessar banco de dados: {str(db_error)}'
            }), 500
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            logger.info("?? [CPF CORRETOR] Conex�o com banco fechada")
        
        # Converter para lista de dicion�rios
        logger.info("?? [CPF CORRETOR] Convertendo dados...")
        funcionarios = []
        for row in rows:
            funcionarios.append({
                'id': row['id'],
                'nome': row['nome'],
                'cpf': row['cpf'] or ''
            })
        
        logger.info(f"? [CPF CORRETOR] {len(funcionarios)} funcion�rios convertidos")
        
        # Filtrar apenas funcion�rios com CPF inv�lido
        logger.info("?? [CPF CORRETOR] Iniciando valida��o de CPFs...")
        funcionarios_invalidos = []
        
        for i, func in enumerate(funcionarios):
            cpf = func.get('cpf', '')
            
            if not cpf:
                continue
                
            try:
                is_valid = CPFVal.validar(cpf)
                if not is_valid:
                    funcionarios_invalidos.append(func)
                    if len(funcionarios_invalidos) <= 5:  # Log apenas os 5 primeiros
                        logger.info(f"   ? CPF inv�lido [{i+1}]: {func['nome'][:30]} - '{cpf}'")
            except Exception as val_error:
                logger.error(f"? [CPF CORRETOR] Erro ao validar CPF de {func['nome']}: {val_error}")
                logger.error(traceback.format_exc())
        
        logger.info(f"? [CPF CORRETOR] Valida��o conclu�da: {len(funcionarios_invalidos)} CPFs inv�lidos")
        
        # Se n�o h� funcion�rios com CPF inv�lido, retornar resultado vazio
        if len(funcionarios_invalidos) == 0:
            logger.info("? [CPF CORRETOR] Nenhum CPF inv�lido - retornando sucesso")
            return jsonify({
                'success': True,
                'total_funcionarios': len(funcionarios),
                'total_cpfs_invalidos': 0,
                'total_corrigidos': 0,
                'total_nao_corrigidos': 0,
                'taxa_correcao': 0,
                'correcoes_por_tipo': {},
                'correcoes_sugeridas': []
            })
        
        # Aplicar corre��o autom�tica
        logger.info("?? [CPF CORRETOR] Iniciando corre��o autom�tica...")
        try:
            resultado_correcao = CPFCorr.corrigir_lista_funcionarios(funcionarios_invalidos)
            logger.info(f"? [CPF CORRETOR] Corre��o conclu�da: {resultado_correcao['total_corrigidos']}/{len(funcionarios_invalidos)}")
        except Exception as corrector_error:
            logger.error(f"? [CPF CORRETOR] ERRO NO CORRETOR: {corrector_error}")
            logger.error(f"Tipo do erro: {type(corrector_error).__name__}")
            logger.error(traceback.format_exc())
            return jsonify({
                'success': False,
                'error': f'Erro no sistema de corre��o: {str(corrector_error)}',
                'error_type': type(corrector_error).__name__,
                'traceback': traceback.format_exc()
            }), 500
        
        # Preparar resposta
        logger.info("?? [CPF CORRETOR] Preparando resposta...")
        resposta = {
            'success': True,
            'total_funcionarios': len(funcionarios),
            'total_cpfs_invalidos': len(funcionarios_invalidos),
            'total_corrigidos': resultado_correcao['total_corrigidos'],
            'total_nao_corrigidos': resultado_correcao['total_nao_corrigidos'],
            'taxa_correcao': round(resultado_correcao['total_corrigidos'] / len(funcionarios_invalidos) * 100, 1) if funcionarios_invalidos else 0,
            'correcoes_por_tipo': resultado_correcao['correcoes_por_tipo'],
            'correcoes_sugeridas': resultado_correcao['correcoes_sugeridas']
        }
        
        logger.info(f"? [CPF CORRETOR] === CONCLUS�O: {resultado_correcao['total_corrigidos']} corre��es ===")
        logger.info("=" * 80)
        
        return jsonify(resposta)
        
    except Exception as e:
        logger.error("=" * 80)
        logger.error(f"??? [CPF CORRETOR] ERRO CR�TICO N�O TRATADO: {e}")
        logger.error(f"Tipo do erro: {type(e).__name__}")
        logger.error(f"Args: {e.args}")
        logger.error("TRACEBACK COMPLETO:")
        logger.error(traceback.format_exc())
        logger.error("=" * 80)
        
        # Print para stderr tamb�m
        print("=" * 80, file=sys.stderr)
        print(f"ERRO CR�TICO CPF CORRETOR: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        
        return jsonify({
            'success': False, 
            'error': str(e),
            'error_type': type(e).__name__,
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/funcionarios/cpf/ping', methods=['GET'])
def ping_cpf_correcao():
    """Endpoint de teste puro - sem decorator, sem depend�ncias"""
    return jsonify({
        'success': True,
        'message': 'Endpoint CPF funcionando',
        'timestamp': str(datetime.now())
    })


@app.route('/api/funcionarios/cpf/corrigir-lote', methods=['PUT'])
@require_permission('folha_pagamento_edit')
def corrigir_cpf_lote():
    """
    ?? Corre��o em Lote de CPFs
    ============================
    
    Aplica corre��es de CPF em m�ltiplos funcion�rios de uma vez.
    Evita problemas de rate limit e melhora performance.
    
    Body:
        {
            "correcoes": [
                {"funcionario_id": 123, "cpf": "12345678901"},
                {"funcionario_id": 456, "cpf": "98765432100"}
            ]
        }
    
    Resposta:
        {
            "success": true,
            "total_processados": 2,
            "total_sucesso": 2,
            "total_falhas": 0,
            "resultados": [...]
        }
    """
    from cpf_validator import CPFValidator
    
    try:
        dados = request.get_json()
        correcoes = dados.get('correcoes', [])
        
        if not correcoes or not isinstance(correcoes, list):
            return jsonify({'error': 'Lista de corre��es n�o informada'}), 400
        
        if len(correcoes) > 500:
            return jsonify({'error': 'M�ximo de 500 corre��es por lote'}), 400
        
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o selecionada'}), 403
        
        logger.info(f"?? [LOTE CPF] Processando {len(correcoes)} corre��es para empresa {empresa_id}")
        
        resultados = []
        total_sucesso = 0
        total_falhas = 0
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            for i, correcao in enumerate(correcoes):
                funcionario_id = correcao.get('funcionario_id')
                novo_cpf = correcao.get('cpf', '').strip()
                
                if not funcionario_id or not novo_cpf:
                    resultados.append({
                        'funcionario_id': funcionario_id,
                        'success': False,
                        'error': 'Dados incompletos'
                    })
                    total_falhas += 1
                    continue
                
                # Validar novo CPF
                validacao = CPFValidator.validar_com_detalhes(novo_cpf)
                if not validacao['valido']:
                    resultados.append({
                        'funcionario_id': funcionario_id,
                        'success': False,
                        'error': f'CPF inv�lido: {validacao["erro"]}'
                    })
                    total_falhas += 1
                    continue
                
                # Atualizar CPF no banco
                cpf_limpo = CPFValidator.limpar(novo_cpf)
                
                cursor.execute("""
                    UPDATE funcionarios 
                    SET cpf = %s, data_atualizacao = CURRENT_TIMESTAMP
                    WHERE id = %s AND empresa_id = %s
                """, (cpf_limpo, funcionario_id, empresa_id))
                
                if cursor.rowcount > 0:
                    resultados.append({
                        'funcionario_id': funcionario_id,
                        'success': True,
                        'cpf_atualizado': CPFValidator.formatar(cpf_limpo)
                    })
                    total_sucesso += 1
                else:
                    resultados.append({
                        'funcionario_id': funcionario_id,
                        'success': False,
                        'error': 'Funcion�rio n�o encontrado ou sem permiss�o'
                    })
                    total_falhas += 1
                
                # Log a cada 50 corre��es
                if (i + 1) % 50 == 0:
                    logger.info(f"? [LOTE CPF] Processados {i + 1}/{len(correcoes)}")
            
            conn.commit()
            logger.info(f"? [LOTE CPF] Conclu�do: {total_sucesso} sucesso, {total_falhas} falhas")
            
            return jsonify({
                'success': True,
                'total_processados': len(correcoes),
                'total_sucesso': total_sucesso,
                'total_falhas': total_falhas,
                'resultados': resultados
            })
            
        except Exception as db_error:
            conn.rollback()
            logger.error(f"? [LOTE CPF] Erro no banco: {db_error}")
            raise db_error
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"? [LOTE CPF] Erro geral: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios/<int:funcionario_id>/cpf', methods=['PUT'])
@require_permission('folha_pagamento_edit')
def corrigir_cpf_funcionario(funcionario_id):
    """Aplica corre��o de CPF em um funcion�rio espec�fico"""
    # Import local para evitar falha de carregamento do m�dulo
    from cpf_validator import CPFValidator
    
    try:
        dados = request.get_json()
        novo_cpf = dados.get('cpf', '').strip()
        
        if not novo_cpf:
            return jsonify({'error': 'CPF n�o informado'}), 400
        
        # Validar novo CPF
        validacao = CPFValidator.validar_com_detalhes(novo_cpf)
        if not validacao['valido']:
            return jsonify({'error': f'CPF inv�lido: {validacao["erro"]}'}), 400
        
        # Obter empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o selecionada'}), 403
        
        # Atualizar CPF no banco
        cpf_formatado = validacao['cpf_formatado']
        
        # Fazer update direto no banco
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE funcionarios 
            SET cpf = %s, 
                data_atualizacao = CURRENT_TIMESTAMP
            WHERE id = %s 
              AND empresa_id = %s
              AND ativo = TRUE
        """, (cpf_formatado, funcionario_id, empresa_id))
        
        rows_affected = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
        
        success = rows_affected > 0
        
        if success:
            logger.info(f"? [CPF CORRETOR] CPF do funcion�rio {funcionario_id} atualizado para: {cpf_formatado}")
            return jsonify({
                'success': True,
                'cpf_novo': cpf_formatado,
                'message': 'CPF atualizado com sucesso'
            })
        else:
            return jsonify({'error': 'Funcion�rio n�o encontrado ou sem permiss�o'}), 404
        
    except Exception as e:
        logger.error(f"? Erro ao corrigir CPF: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios', methods=['POST'])
@require_permission('folha_pagamento_create')
def criar_funcionario():
    """Criar novo funcion�rio"""
    # Import local para evitar falha de carregamento do m�dulo
    from cpf_validator import CPFValidator
    
    try:
        import re
        
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        
        dados = request.get_json()
        
        # Valida��es obrigat�rias
        if not dados.get('nome'):
            return jsonify({'error': 'Nome � obrigat�rio'}), 400
        if not dados.get('cpf'):
            return jsonify({'error': 'CPF � obrigat�rio'}), 400
        
        # ?? NOVO: Validar CPF com CPFValidator
        validacao_cpf = CPFValidator.validar_com_detalhes(dados['cpf'])
        if not validacao_cpf['valido']:
            return jsonify({'error': f'CPF inv�lido: {validacao_cpf["erro"]}'}), 400
        
        # ?? Validar email se fornecido
        if dados.get('email'):
            try:
                from app.utils.validators import validate_email
                is_valid, error_msg = validate_email(dados['email'])
                if not is_valid:
                    return jsonify({'error': f'Email inv�lido: {error_msg}'}), 400
            except ImportError:
                # Valida��o simples caso validators n�o esteja dispon�vel
                import re
                email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_regex, dados['email']):
                    return jsonify({'error': 'Email inv�lido'}), 400
        
        # Limpar e formatar CPF
        cpf = CPFValidator.limpar(dados['cpf'])
        
        print(f"\n?? [POST /api/funcionarios]")
        print(f"   - empresa_id: {empresa_id}")
        print(f"   - nome: {dados.get('nome')}")
        print(f"   - cpf: {cpf}")
        print(f"   - cpf_formatado: {validacao_cpf['cpf_formatado']}")
        
        # ?? Usar get_db_connection com empresa_id para aplicar RLS
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            
            # Verificar se CPF j� existe
            cursor.execute("SELECT id FROM funcionarios WHERE cpf = %s AND empresa_id = %s", (cpf, empresa_id))
            if cursor.fetchone():
                cursor.close()
                return jsonify({'error': 'CPF j� cadastrado'}), 400
            
            query = """
                INSERT INTO funcionarios 
                (empresa_id, nome, cpf, email, celular, ativo, data_admissao, data_demissao, observacoes,
                 nacionalidade, estado_civil, data_nascimento, profissao, 
                 rua_av, numero_residencia, complemento, bairro, cidade, estado, cep, 
                 chave_pix, pis_pasep)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """
            
            cursor.execute(query, (
                empresa_id,
                dados['nome'],
                cpf,
                dados.get('email'),
                dados.get('celular'),
                dados.get('ativo', True),
                dados.get('data_admissao') if dados.get('data_admissao') else None,
                dados.get('data_demissao') if dados.get('data_demissao') else None,
                dados.get('observacoes'),
                dados.get('nacionalidade'),
                dados.get('estado_civil'),
                dados.get('data_nascimento') if dados.get('data_nascimento') else None,
                dados.get('profissao'),
                dados.get('rua_av'),
                dados.get('numero_residencia'),
                dados.get('complemento'),
                dados.get('bairro'),
                dados.get('cidade'),
                dados.get('estado'),
                dados.get('cep'),
                dados.get('chave_pix'),
                dados.get('pis_pasep')
            ))
            
            resultado = cursor.fetchone()
            funcionario_id = resultado['id'] if isinstance(resultado, dict) else resultado[0]
            conn.commit()
            cursor.close()
            
            print(f"   ? Funcion�rio criado com ID: {funcionario_id}")
            
            return jsonify({
                'success': True,
                'id': funcionario_id,
                'message': 'Funcion�rio cadastrado com sucesso'
            }), 201
    
    except Exception as e:
        logger.error(f"Erro ao criar funcion�rio: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios/<int:funcionario_id>', methods=['PUT'])
@require_permission('folha_pagamento_edit')
def atualizar_funcionario(funcionario_id):
    """Atualizar funcion�rio existente"""
    try:
        from app.utils.validators import validate_cpf, validate_email
        import re
        
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        
        dados = request.get_json()
        
        # ?? Validar CPF se fornecido
        if dados.get('cpf'):
            is_valid, error_msg = validate_cpf(dados['cpf'])
            if not is_valid:
                return jsonify({'error': f'CPF inv�lido: {error_msg}'}), 400
        
        # ?? Validar email se fornecido
        if dados.get('email'):
            is_valid, error_msg = validate_email(dados['email'])
            if not is_valid:
                return jsonify({'error': f'Email inv�lido: {error_msg}'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se funcion�rio existe e pertence � empresa
        cursor.execute("SELECT id FROM funcionarios WHERE id = %s AND empresa_id = %s", (funcionario_id, empresa_id))
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Funcion�rio n�o encontrado'}), 404
        
        # Construir query din�mica baseada nos campos fornecidos
        campos_update = []
        valores = []
        
        campos_permitidos = [
            'nome', 'email', 'celular', 'ativo', 'data_admissao', 'data_demissao', 'observacoes',
            'nacionalidade', 'estado_civil', 'data_nascimento', 'profissao',
            'rua_av', 'numero_residencia', 'complemento', 'bairro', 'cidade', 'estado', 'cep',
            'chave_pix', 'pis_pasep'
        ]
        
        for campo in campos_permitidos:
            if campo in dados:
                campos_update.append(f"{campo} = %s")
                valores.append(dados[campo])
        
        if 'cpf' in dados:
            cpf = dados['cpf'].replace('.', '').replace('-', '').replace('/', '')
            # Verificar se CPF j� existe em outro funcion�rio
            cursor.execute("SELECT id FROM funcionarios WHERE cpf = %s AND empresa_id = %s AND id != %s", 
                         (cpf, empresa_id, funcionario_id))
            if cursor.fetchone():
                cursor.close()
                return jsonify({'error': 'CPF j� cadastrado para outro funcion�rio'}), 400
            campos_update.append("cpf = %s")
            valores.append(cpf)
        
        if not campos_update:
            cursor.close()
            return jsonify({'error': 'Nenhum campo para atualizar'}), 400
        
        campos_update.append("data_atualizacao = CURRENT_TIMESTAMP")
        valores.append(funcionario_id)
        
        query = f"UPDATE funcionarios SET {', '.join(campos_update)} WHERE id = %s"
        cursor.execute(query, valores)
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Funcion�rio atualizado com sucesso'
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao atualizar funcion�rio: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios/<int:funcionario_id>', methods=['GET'])
@require_permission('folha_pagamento_view')
def obter_funcionario(funcionario_id):
    """Obter detalhes de um funcion�rio espec�fico"""
    try:
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Buscar funcion�rio da empresa
        cursor.execute("""
            SELECT id, empresa_id, nome, cpf, endereco, tipo_chave_pix, 
                   chave_pix, data_admissao, observacoes, ativo,
                   data_criacao, data_atualizacao
            FROM funcionarios 
            WHERE id = %s AND empresa_id = %s
        """, (funcionario_id, empresa_id))
        
        row = cursor.fetchone()
        cursor.close()
        
        if not row:
            return jsonify({'error': 'Funcion�rio n�o encontrado'}), 404
        
        # Verifica se row � dict ou tupla
        if isinstance(row, dict):
            funcionario = {
                'id': row['id'],
                'empresa_id': row['empresa_id'],
                'nome': row['nome'],
                'cpf': row['cpf'],
                'endereco': row['endereco'],
                'tipo_chave_pix': row['tipo_chave_pix'],
                'chave_pix': row['chave_pix'],
                'data_admissao': row['data_admissao'].isoformat() if row['data_admissao'] else None,
                'observacoes': row['observacoes'],
                'ativo': row['ativo'],
                'data_criacao': row['data_criacao'].isoformat() if row['data_criacao'] else None,
                'data_atualizacao': row['data_atualizacao'].isoformat() if row['data_atualizacao'] else None
            }
        else:
            funcionario = {
                'id': row[0],
                'empresa_id': row[1],
                'nome': row[2],
                'cpf': row[3],
                'endereco': row[4],
                'tipo_chave_pix': row[5],
                'chave_pix': row[6],
                'data_admissao': row[7].isoformat() if row[7] else None,
                'observacoes': row[8],
                'ativo': row[9],
                'data_criacao': row[10].isoformat() if row[10] else None,
                'data_atualizacao': row[11].isoformat() if row[11] else None
            }
        
        return jsonify(funcionario), 200
    
    except Exception as e:
        logger.error(f"Erro ao obter funcion�rio: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcionarios/<int:funcionario_id>', methods=['DELETE'])
@require_permission('folha_pagamento_edit')
def deletar_funcionario(funcionario_id):
    """Deletar um funcion�rio"""
    try:
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se funcion�rio existe e pertence � empresa
        cursor.execute("SELECT id FROM funcionarios WHERE id = %s AND empresa_id = %s", 
                      (funcionario_id, empresa_id))
        
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Funcion�rio n�o encontrado'}), 404
        
        # Deletar funcion�rio
        cursor.execute("DELETE FROM funcionarios WHERE id = %s AND empresa_id = %s", 
                      (funcionario_id, empresa_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Funcion�rio deletado com sucesso'
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao deletar funcion�rio: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


# === ROTAS DE EVENTOS ===

@app.route('/api/eventos', methods=['GET'])
@require_permission('eventos_view')
def listar_eventos():
    """Listar eventos com filtros opcionais"""
    try:
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # SEGURANCA MULTI-TENANT: Usar empresa_id da sessao
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa nao identificada'}), 403
        
        # Filtros opcionais
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        status = request.args.get('status')
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT id, empresa_id, nome_evento, data_evento, nf_associada,
                   valor_liquido_nf, custo_evento, margem, tipo_evento, status,
                   observacoes, data_criacao, data_atualizacao
            FROM eventos
            WHERE empresa_id = %s
        """
        params = [empresa_id]
        
        if data_inicio:
            query += " AND data_evento >= %s"
            params.append(data_inicio)
        
        if data_fim:
            query += " AND data_evento <= %s"
            params.append(data_fim)
        
        if status:
            query += " AND status = %s"
            params.append(status)
        
        query += " ORDER BY data_evento DESC"
        
        logger.info(f"?? [DEBUG LOAD] Query SQL: {query}")
        logger.info(f"?? [DEBUG LOAD] Params: {params}")
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        logger.info(f"?? [DEBUG LOAD] Linhas retornadas do DB: {len(rows)}")
        if rows:
            logger.info(f"?? [DEBUG LOAD] Primeira linha - data_evento: {rows[0].get('data_evento') if isinstance(rows[0], dict) else rows[0][3]}")
        
        cursor.close()
        
        eventos = []
        for row in rows:
            # Verifica se row � dict ou tupla
            if isinstance(row, dict):
                eventos.append({
                    'id': row['id'],
                    'empresa_id': row['empresa_id'],
                    'nome_evento': row['nome_evento'],
                    'data_evento': row['data_evento'].isoformat() if row['data_evento'] else None,
                    'nf_associada': row['nf_associada'],
                    'valor_liquido_nf': float(row['valor_liquido_nf']) if row['valor_liquido_nf'] else None,
                    'custo_evento': float(row['custo_evento']) if row['custo_evento'] else None,
                    'margem': float(row['margem']) if row['margem'] else None,
                    'tipo_evento': row['tipo_evento'],
                    'status': row['status'],
                    'observacoes': row['observacoes'],
                    'data_criacao': row['data_criacao'].isoformat() if row['data_criacao'] else None,
                    'data_atualizacao': row['data_atualizacao'].isoformat() if row['data_atualizacao'] else None
                })
            else:
                eventos.append({
                    'id': row[0],
                    'empresa_id': row[1],
                    'nome_evento': row[2],
                    'data_evento': row[3].isoformat() if row[3] else None,
                    'nf_associada': row[4],
                    'valor_liquido_nf': float(row[5]) if row[5] else None,
                    'custo_evento': float(row[6]) if row[6] else None,
                    'margem': float(row[7]) if row[7] else None,
                    'tipo_evento': row[8],
                    'status': row[9],
                    'observacoes': row[10],
                    'data_criacao': row[11].isoformat() if row[11] else None,
                    'data_atualizacao': row[12].isoformat() if row[12] else None
                })
        
        logger.info(f"[EVENTOS LIST] Total eventos: {len(eventos)}")
        for ev in eventos:
            logger.info(f"[EVENTOS LIST] id={ev['id']}, nome={ev['nome_evento']}, data={ev['data_evento']}")
        
        return jsonify({'eventos': eventos}), 200
    
    except Exception as e:
        logger.error(f"Erro ao listar eventos: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if conn:
                database.return_to_pool(conn)
        except:
            pass


@app.route('/api/eventos', methods=['POST'])
@require_permission('eventos_create')
def criar_evento():
    """Criar novo evento"""
    try:
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        
        dados = request.get_json()
        
        # Valida��es obrigat�rias
        if not dados.get('nome_evento'):
            return jsonify({'error': 'Nome do evento � obrigat�rio'}), 400
        if not dados.get('data_evento'):
            return jsonify({'error': 'Data do evento � obrigat�ria'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        query = """
            INSERT INTO eventos 
            (empresa_id, nome_evento, data_evento, nf_associada, valor_liquido_nf,
             custo_evento, margem, tipo_evento, status, observacoes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """
        
        cursor.execute(query, (
            empresa_id,
            dados['nome_evento'],
            dados['data_evento'],
            dados.get('nf_associada'),
            dados.get('valor_liquido_nf'),
            dados.get('custo_evento'),
            dados.get('margem'),
            dados.get('tipo_evento'),
            dados.get('status', 'PENDENTE'),
            dados.get('observacoes')
        ))
        
        resultado = cursor.fetchone()
        evento_id = resultado['id'] if isinstance(resultado, dict) else resultado[0]
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'id': evento_id,
            'message': 'Evento cadastrado com sucesso'
        }), 201
    
    except Exception as e:
        logger.error(f"Erro ao criar evento: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/eventos/<int:evento_id>', methods=['PUT'])
@require_permission('eventos_edit')
def atualizar_evento(evento_id):
    """Atualizar evento existente - usa autocommit=True para garantir persistencia"""
    conn = None
    cursor = None
    try:
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usuario nao autenticado'}), 401
        
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa nao identificada'}), 403
        
        dados = request.get_json()
        
        logger.info(f"[EVENTO UPDATE] === INICIO === evento_id={evento_id}, empresa_id={empresa_id}")
        logger.info(f"[EVENTO UPDATE] Dados recebidos: {dados}")
        
        # Usar autocommit=True (padrao do pool) - cada statement comita automaticamente
        conn = db.get_connection()
        # NAO alterar autocommit - manter True (padrao do pool)
        cursor = conn.cursor()
        
        # 1. Verificar se evento existe e pertence a empresa
        cursor.execute(
            "SELECT id, data_evento, nome_evento FROM eventos WHERE id = %s AND empresa_id = %s",
            (evento_id, empresa_id)
        )
        evento_atual = cursor.fetchone()
        if not evento_atual:
            logger.warning(f"[EVENTO UPDATE] Evento {evento_id} NAO encontrado para empresa {empresa_id}")
            return jsonify({'error': 'Evento nao encontrado'}), 404
        
        data_antes = evento_atual['data_evento'] if isinstance(evento_atual, dict) else evento_atual[1]
        nome_antes = evento_atual['nome_evento'] if isinstance(evento_atual, dict) else evento_atual[2]
        logger.info(f"[EVENTO UPDATE] Evento encontrado: '{nome_antes}', data_antes={data_antes}")
        
        # 2. Construir query dinamica
        campos_update = []
        valores = []
        
        campos_possiveis = [
            'nome_evento', 'data_evento', 'nf_associada', 'valor_liquido_nf',
            'custo_evento', 'margem', 'tipo_evento', 'status', 'observacoes'
        ]
        
        for campo in campos_possiveis:
            if campo in dados:
                campos_update.append(f"{campo} = %s")
                valores.append(dados[campo])
        
        if not campos_update:
            return jsonify({'error': 'Nenhum campo para atualizar'}), 400
        
        campos_update.append("data_atualizacao = CURRENT_TIMESTAMP")
        valores.append(evento_id)
        valores.append(empresa_id)
        
        query = f"UPDATE eventos SET {', '.join(campos_update)} WHERE id = %s AND empresa_id = %s"
        
        logger.info(f"[EVENTO UPDATE] SQL: {query}")
        logger.info(f"[EVENTO UPDATE] Valores: {valores}")
        
        # 3. Executar UPDATE (autocommit=True => commit automatico)
        cursor.execute(query, valores)
        linhas_afetadas = cursor.rowcount
        logger.info(f"[EVENTO UPDATE] Linhas afetadas pelo UPDATE: {linhas_afetadas}")
        
        if linhas_afetadas == 0:
            logger.error(f"[EVENTO UPDATE] NENHUMA LINHA AFETADA! UPDATE falhou silenciosamente.")
            return jsonify({'error': 'Falha ao atualizar evento - nenhuma linha afetada'}), 500
        
        # 4. Recalcular margem se necessario
        if 'valor_liquido_nf' in dados or 'custo_evento' in dados:
            cursor.execute(
                "SELECT valor_liquido_nf, custo_evento FROM eventos WHERE id = %s",
                (evento_id,)
            )
            evt = cursor.fetchone()
            if evt:
                vl_raw = evt['valor_liquido_nf'] if isinstance(evt, dict) else evt[0]
                ct_raw = evt['custo_evento'] if isinstance(evt, dict) else evt[1]
                vl = float(vl_raw) if vl_raw else 0
                ct = float(ct_raw) if ct_raw else 0
                margem_calc = vl - ct
                cursor.execute("UPDATE eventos SET margem = %s WHERE id = %s", (margem_calc, evento_id))
                logger.info(f"[EVENTO UPDATE] Margem recalculada: {margem_calc}")
        
        # 5. Commit explicito por seguranca (com autocommit=True eh redundante mas seguro)
        conn.commit()
        
        # 6. Verificacao pos-commit - ler de volta do banco
        cursor.execute(
            """SELECT id, nome_evento, data_evento, nf_associada, valor_liquido_nf,
                      custo_evento, margem, tipo_evento, status, observacoes
               FROM eventos WHERE id = %s""",
            (evento_id,)
        )
        evento_atualizado = cursor.fetchone()
        
        if evento_atualizado:
            if isinstance(evento_atualizado, dict):
                data_depois = evento_atualizado['data_evento']
                evento_resp = {
                    'id': evento_atualizado['id'],
                    'nome_evento': evento_atualizado['nome_evento'],
                    'data_evento': evento_atualizado['data_evento'].isoformat() if evento_atualizado['data_evento'] else None,
                    'nf_associada': evento_atualizado['nf_associada'],
                    'valor_liquido_nf': float(evento_atualizado['valor_liquido_nf']) if evento_atualizado['valor_liquido_nf'] else None,
                    'custo_evento': float(evento_atualizado['custo_evento']) if evento_atualizado['custo_evento'] else None,
                    'margem': float(evento_atualizado['margem']) if evento_atualizado['margem'] else None,
                    'tipo_evento': evento_atualizado['tipo_evento'],
                    'status': evento_atualizado['status'],
                    'observacoes': evento_atualizado['observacoes']
                }
            else:
                data_depois = evento_atualizado[2]
                evento_resp = {'id': evento_atualizado[0], 'data_evento': str(evento_atualizado[2])}
            
            logger.info(f"[EVENTO UPDATE] Data ANTES={data_antes} => DEPOIS={data_depois}")
            logger.info(f"[EVENTO UPDATE] === SUCESSO ===")
        else:
            evento_resp = {}
            logger.warning(f"[EVENTO UPDATE] Evento nao encontrado na verificacao pos-commit!")
        
        return jsonify({
            'success': True,
            'message': 'Evento atualizado com sucesso',
            'evento': evento_resp,
            'data_antes': str(data_antes),
            'data_depois': str(data_depois) if evento_atualizado else None
        }), 200
    
    except Exception as e:
        logger.error(f"[EVENTO UPDATE] ERRO: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if conn:
            try:
                database.return_to_pool(conn)
            except:
                pass


@app.route('/api/eventos/<int:evento_id>', methods=['DELETE'])
@require_permission('eventos_delete')
def deletar_evento(evento_id):
    """Deletar evento"""
    conn = None
    try:
        usuario = get_usuario_logado()
        if not usuario:
            return jsonify({'error': 'Usu�rio n�o autenticado'}), 401
        
        # ?? SEGURAN�A MULTI-TENANT: Usar empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'error': 'Empresa n�o identificada'}), 403
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se evento existe e pertence � empresa
        cursor.execute("SELECT id FROM eventos WHERE id = %s AND empresa_id = %s", (evento_id, empresa_id))
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Evento n�o encontrado'}), 404
        
        # Deletar evento
        cursor.execute("DELETE FROM eventos WHERE id = %s", (evento_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evento deletado com sucesso'
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao deletar evento: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500
    
    finally:
        if conn:
            database.return_to_pool(conn)


# === ROTAS DE ALOCA��O DE EQUIPE EM EVENTOS ===

@app.route('/api/funcoes-evento', methods=['GET'])
@require_permission('eventos_view')
def listar_funcoes_evento():
    """Listar fun��es dispon�veis para eventos"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nome, descricao, ativo
            FROM funcoes_evento
            WHERE ativo = TRUE
            ORDER BY nome
        """)
        
        funcoes = []
        for row in cursor.fetchall():
            funcoes.append({
                'id': row['id'],
                'nome': row['nome'],
                'descricao': row['descricao'],
                'ativo': row['ativo']
            })
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'funcoes': funcoes
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao listar fun��es: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcoes-evento', methods=['POST'])
@require_permission('eventos_create')
def criar_funcao_evento():
    """Cadastrar nova fun��o para eventos"""
    try:
        dados = request.get_json()
        
        print(f"\n?? [POST /api/funcoes-evento] Dados recebidos:")
        print(f"   - Raw JSON: {dados}")
        print(f"   - Tipo: {type(dados)}")
        print(f"   - Keys: {dados.keys() if dados else 'None'}")
        
        nome = dados.get('nome', '').strip() if dados else ''
        descricao = dados.get('descricao', '').strip() if dados else ''
        
        print(f"   - nome extra�do: '{nome}'")
        print(f"   - descricao extra�da: '{descricao}'")
        print(f"   - nome vazio? {not nome}")
        
        if not nome:
            print(f"   ? Rejeitando: nome vazio")
            return jsonify({'error': 'Nome da fun��o � obrigat�rio'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se j� existe
        cursor.execute("SELECT id FROM funcoes_evento WHERE UPPER(nome) = UPPER(%s)", (nome,))
        if cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'J� existe uma fun��o com este nome'}), 400
        
        # Inserir nova fun��o
        cursor.execute("""
            INSERT INTO funcoes_evento (nome, descricao, ativo)
            VALUES (%s, %s, TRUE)
            RETURNING id
        """, (nome, descricao))
        
        funcao_id = cursor.fetchone()['id']
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Fun��o cadastrada com sucesso',
            'funcao_id': funcao_id
        }), 201
    
    except Exception as e:
        logger.error(f"Erro ao criar fun��o: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcoes-evento/<int:funcao_id>', methods=['PUT'])
@require_permission('eventos_edit')
def atualizar_funcao_evento(funcao_id):
    """Atualizar fun��o de evento existente"""
    try:
        dados = request.get_json()
        nome = dados.get('nome', '').strip()
        descricao = dados.get('descricao', '').strip()
        ativo = dados.get('ativo', True)
        
        if not nome:
            return jsonify({'error': 'Nome da fun��o � obrigat�rio'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se fun��o existe
        cursor.execute("SELECT id FROM funcoes_evento WHERE id = %s", (funcao_id,))
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Fun��o n�o encontrada'}), 404
        
        # Verificar se nome j� existe em outra fun��o
        cursor.execute(
            "SELECT id FROM funcoes_evento WHERE UPPER(nome) = UPPER(%s) AND id != %s",
            (nome, funcao_id)
        )
        if cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'J� existe outra fun��o com este nome'}), 400
        
        # Atualizar fun��o
        cursor.execute("""
            UPDATE funcoes_evento 
            SET nome = %s, descricao = %s, ativo = %s
            WHERE id = %s
        """, (nome, descricao, ativo, funcao_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Fun��o atualizada com sucesso'
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao atualizar fun��o: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/funcoes-evento/<int:funcao_id>', methods=['DELETE'])
@require_permission('eventos_edit')
def deletar_funcao_evento(funcao_id):
    """Deletar fun��o de evento"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se fun��o existe
        cursor.execute("SELECT id FROM funcoes_evento WHERE id = %s", (funcao_id,))
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Fun��o n�o encontrada'}), 404
        
        # Verificar se h� funcion�rios usando esta fun��o
        cursor.execute(
            "SELECT COUNT(*) as total FROM evento_funcionarios WHERE funcao_id = %s",
            (funcao_id,)
        )
        result = cursor.fetchone()
        total = result['total'] if isinstance(result, dict) else result[0]
        
        if total > 0:
            cursor.close()
            return jsonify({
                'error': f'N�o � poss�vel excluir. Esta fun��o est� sendo usada por {total} aloca��o(�es) de funcion�rios.'
            }), 400
        
        # Deletar fun��o
        cursor.execute("DELETE FROM funcoes_evento WHERE id = %s", (funcao_id,))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Fun��o deletada com sucesso'
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao deletar fun��o: {e}")
        return jsonify({'error': str(e)}), 500


# ========================================
# ENDPOINTS: SETORES
# ========================================

@app.route('/api/setores', methods=['GET'])
@require_permission('eventos_view')
def listar_setores():
    """Listar setores cadastrados"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nome, ativo, created_at
            FROM setores
            ORDER BY nome
        """)
        
        setores = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'setores': setores
        })
    
    except Exception as e:
        logger.error(f"Erro ao listar setores: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/setores', methods=['POST'])
@require_permission('eventos_create')
def criar_setor():
    """Cadastrar novo setor"""
    try:
        dados = request.get_json()
        nome = dados.get('nome', '').strip()
        
        if not nome:
            return jsonify({'error': 'Nome do setor � obrigat�rio'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se j� existe
        cursor.execute("SELECT id FROM setores WHERE UPPER(nome) = UPPER(%s)", (nome,))
        if cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'J� existe um setor com este nome'}), 400
        
        # Inserir novo setor
        cursor.execute("""
            INSERT INTO setores (nome, ativo, created_at)
            VALUES (%s, TRUE, NOW())
            RETURNING id
        """, (nome,))
        
        setor_id = cursor.fetchone()['id']
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Setor cadastrado com sucesso',
            'setor_id': setor_id
        }), 201
    
    except Exception as e:
        logger.error(f"Erro ao criar setor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/setores/<int:setor_id>', methods=['PATCH'])
@require_permission('eventos_edit')
def atualizar_setor(setor_id):
    """Atualizar status do setor (ativar/desativar)"""
    try:
        dados = request.get_json()
        ativo = dados.get('ativo')
        
        if ativo is None:
            return jsonify({'error': 'Status ativo � obrigat�rio'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE setores
            SET ativo = %s
            WHERE id = %s
            RETURNING id
        """, (ativo, setor_id))
        
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Setor n�o encontrado'}), 404
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': f'Setor {"ativado" if ativo else "desativado"} com sucesso'
        })
    
    except Exception as e:
        logger.error(f"Erro ao criar fun��o: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500

@app.route('/api/setores/<int:setor_id>', methods=['DELETE'])
@require_permission('eventos_edit')
def excluir_setor(setor_id):
    """Excluir um setor"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se o setor existe
        cursor.execute("SELECT nome FROM setores WHERE id = %s", (setor_id,))
        setor = cursor.fetchone()
        
        if not setor:
            cursor.close()
            return jsonify({'error': 'Setor n�o encontrado'}), 404
        
        # Excluir o setor
        cursor.execute("DELETE FROM setores WHERE id = %s", (setor_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Setor exclu�do com sucesso'
        })
    
    except Exception as e:
        logger.error(f"Erro ao excluir setor: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/eventos/<int:evento_id>/equipe', methods=['GET'])
@require_permission('eventos_view')
def listar_equipe_evento(evento_id):
    """Listar equipe alocada em um evento"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                ef.id,
                ef.funcionario_id,
                f.nome as funcionario_nome,
                f.cpf as funcionario_cpf,
                f.email as funcionario_email,
                f.chave_pix as funcionario_chave_pix,
                ef.funcao_id,
                ef.funcao_nome,
                ef.setor_id,
                s.nome as setor_nome,
                ef.valor,
                ef.hora_inicio,
                ef.hora_fim,
                ef.observacoes
            FROM evento_funcionarios ef
            INNER JOIN funcionarios f ON f.id = ef.funcionario_id
            LEFT JOIN setores s ON s.id = ef.setor_id
            WHERE ef.evento_id = %s
            ORDER BY f.nome
        """, (evento_id,))
        
        equipe = []
        for row in cursor.fetchall():
            equipe.append({
                'id': row['id'],
                'funcionario_id': row['funcionario_id'],
                'funcionario_nome': row['funcionario_nome'],
                'funcionario_cpf': row['funcionario_cpf'],
                'funcionario_email': row['funcionario_email'],
                'funcionario_chave_pix': row['funcionario_chave_pix'],
                'funcao_id': row['funcao_id'],
                'funcao_nome': row['funcao_nome'],
                'setor_id': row['setor_id'],
                'setor_nome': row['setor_nome'],
                'valor': float(row['valor']) if row['valor'] else 0.0,
                'hora_inicio': str(row['hora_inicio']) if row['hora_inicio'] else None,
                'hora_fim': str(row['hora_fim']) if row['hora_fim'] else None,
                'observacoes': row['observacoes']
            })
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'equipe': equipe
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao listar equipe: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/eventos/<int:evento_id>/equipe', methods=['POST'])
@require_permission('eventos_create')
def adicionar_funcionario_evento(evento_id):
    """Adicionar funcion�rio � equipe do evento"""
    try:
        dados = request.get_json()
        print(f"[EQUIPE MASSA] Dados recebidos: {dados}", flush=True)  # DEBUG
        
        funcionario_id = dados.get('funcionario_id')
        funcao_id = dados.get('funcao_id')
        setor_id = dados.get('setor_id')  # Opcional
        valor = dados.get('valor', 0)
        hora_inicio = dados.get('hora_inicio')  # Opcional
        hora_fim = dados.get('hora_fim')  # Opcional
        
        print(f"[EQUIPE MASSA] funcionario_id={funcionario_id}, funcao_id={funcao_id}, valor={valor}", flush=True)  # DEBUG
        
        if not funcionario_id or not funcao_id:
            print(f"[EQUIPE MASSA] ? ERRO: Campos obrigat�rios ausentes", flush=True)  # DEBUG
            return jsonify({'error': 'Funcion�rio e fun��o s�o obrigat�rios'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Buscar nome da fun��o para hist�rico
        cursor.execute("SELECT nome FROM funcoes_evento WHERE id = %s", (funcao_id,))
        funcao_row = cursor.fetchone()
        if not funcao_row:
            print(f"[EQUIPE MASSA] ? ERRO: Fun��o {funcao_id} n�o encontrada", flush=True)  # DEBUG
            cursor.close()
            return jsonify({'error': 'Fun��o n�o encontrada'}), 404
        
        funcao_nome = funcao_row['nome']
        print(f"[EQUIPE MASSA] Fun��o encontrada: {funcao_nome}", flush=True)  # DEBUG
        
        # Verificar se j� existe aloca��o
        cursor.execute("""
            SELECT id FROM evento_funcionarios 
            WHERE evento_id = %s AND funcionario_id = %s AND funcao_id = %s
        """, (evento_id, funcionario_id, funcao_id))
        
        if cursor.fetchone():
            print(f"[EQUIPE MASSA] ?? DUPLICADO: Funcion�rio {funcionario_id} j� alocado no evento {evento_id} com fun��o {funcao_id}", flush=True)  # DEBUG
            cursor.close()
            return jsonify({'error': 'Este funcion�rio j� est� alocado com esta fun��o neste evento'}), 400
        
        # Inserir aloca��o (com setor_id, hora_inicio e hora_fim se fornecidos)
        cursor.execute("""
            INSERT INTO evento_funcionarios 
            (evento_id, funcionario_id, funcao_id, funcao_nome, setor_id, valor, hora_inicio, hora_fim)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (evento_id, funcionario_id, funcao_id, funcao_nome, setor_id, valor, hora_inicio, hora_fim))
        
        alocacao_id = cursor.fetchone()['id']
        
        # Calcular novo custo total do evento
        cursor.execute("""
            SELECT COALESCE(SUM(valor), 0) as total
            FROM evento_funcionarios
            WHERE evento_id = %s
        """, (evento_id,))
        
        custo_total = cursor.fetchone()['total']
        
        # Buscar valor_liquido_nf para recalcular margem
        cursor.execute("""
            SELECT valor_liquido_nf
            FROM eventos
            WHERE id = %s
        """, (evento_id,))
        
        evento_row = cursor.fetchone()
        valor_liquido = evento_row['valor_liquido_nf'] if evento_row and evento_row['valor_liquido_nf'] else 0
        
        # Calcular margem: Valor L�quido - Custo
        margem = float(valor_liquido) - float(custo_total)
        
        # Atualizar custo do evento E margem
        cursor.execute("""
            UPDATE eventos
            SET custo_evento = %s, margem = %s
            WHERE id = %s
        """, (custo_total, margem, evento_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Funcion�rio adicionado � equipe',
            'alocacao_id': alocacao_id,
            'custo_total': float(custo_total)
        }), 201
    
    except Exception as e:
        logger.error(f"Erro ao adicionar funcion�rio: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/eventos/equipe/<int:alocacao_id>', methods=['DELETE'])
@require_permission('eventos_delete')
def remover_funcionario_evento(alocacao_id):
    """Remover funcion�rio da equipe do evento"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Buscar evento_id antes de deletar
        cursor.execute("SELECT evento_id FROM evento_funcionarios WHERE id = %s", (alocacao_id,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            return jsonify({'error': 'Aloca��o n�o encontrada'}), 404
        
        evento_id = row['evento_id']
        
        # Deletar aloca��o
        cursor.execute("DELETE FROM evento_funcionarios WHERE id = %s", (alocacao_id,))
        
        # Recalcular custo total do evento
        cursor.execute("""
            SELECT COALESCE(SUM(valor), 0) as total
            FROM evento_funcionarios
            WHERE evento_id = %s
        """, (evento_id,))
        
        custo_total = cursor.fetchone()['total']
        
        # Buscar valor_liquido_nf para recalcular margem
        cursor.execute("""
            SELECT valor_liquido_nf
            FROM eventos
            WHERE id = %s
        """, (evento_id,))
        
        evento_row = cursor.fetchone()
        valor_liquido = evento_row['valor_liquido_nf'] if evento_row and evento_row['valor_liquido_nf'] else 0
        
        # Calcular margem: Valor L�quido - Custo
        margem = float(valor_liquido) - float(custo_total)
        
        # Atualizar custo do evento E margem
        cursor.execute("""
            UPDATE eventos
            SET custo_evento = %s, margem = %s
            WHERE id = %s
        """, (custo_total, margem, evento_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Funcion�rio removido da equipe',
            'custo_total': float(custo_total)
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao remover funcion�rio: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


# ============================================================================
# ROTAS DE FORNECEDORES DO EVENTO
# ============================================================================

@app.route('/api/eventos/<int:evento_id>/fornecedores', methods=['GET'])
@require_permission('eventos_view')
def listar_fornecedores_evento(evento_id):
    """Listar fornecedores vinculados ao evento"""
    try:
        from psycopg2.extras import RealDictCursor
        
        logger.info(f"?? GET /api/eventos/{evento_id}/fornecedores")
        
        conn = db.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verificar se a tabela existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'evento_fornecedores'
            );
        """)
        
        tabela_existe = cursor.fetchone()['exists']
        
        if not tabela_existe:
            logger.warning("   ??  Tabela evento_fornecedores n�o existe - Retornando lista vazia")
            cursor.close()
            return jsonify({
                'success': True,
                'fornecedores': [],
                'warning': 'Tabela evento_fornecedores n�o existe. Execute a migra��o.'
            }), 200
        
        cursor.execute("""
            SELECT 
                ef.id,
                ef.fornecedor_id,
                f.nome as fornecedor_nome,
                ef.categoria_id,
                c.nome as categoria_nome,
                ef.subcategoria_id,
                sc.nome as subcategoria_nome,
                ef.valor,
                ef.observacao,
                ef.created_at
            FROM evento_fornecedores ef
            JOIN fornecedores f ON ef.fornecedor_id = f.id
            LEFT JOIN categorias c ON ef.categoria_id = c.id
            LEFT JOIN subcategorias sc ON ef.subcategoria_id = sc.id
            WHERE ef.evento_id = %s
            ORDER BY ef.created_at DESC
        """, (evento_id,))
        
        fornecedores = cursor.fetchall()
        cursor.close()
        
        logger.info(f"   ?? Retornando {len(fornecedores)} fornecedores")
        
        return jsonify({
            'success': True,
            'fornecedores': fornecedores
        }), 200
    
    except Exception as e:
        logger.error(f"? Erro ao listar fornecedores do evento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/eventos/<int:evento_id>/fornecedores', methods=['POST'])
@require_permission('eventos_edit')
def adicionar_fornecedor_evento(evento_id):
    """Adicionar fornecedor ao evento"""
    try:
        dados = request.json
        fornecedor_id = dados.get('fornecedor_id')
        categoria_id = dados.get('categoria_id')
        subcategoria_id = dados.get('subcategoria_id')
        valor = dados.get('valor', 0)
        observacao = dados.get('observacao')
        
        if not fornecedor_id:
            return jsonify({'error': 'fornecedor_id � obrigat�rio'}), 400
        
        usuario = get_usuario_logado()
        usuario_id = usuario.get('id') if usuario else None
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Verificar se fornecedor j� est� vinculado ao evento
        cursor.execute("""
            SELECT id FROM evento_fornecedores
            WHERE evento_id = %s AND fornecedor_id = %s
        """, (evento_id, fornecedor_id))
        
        if cursor.fetchone():
            cursor.close()
            return jsonify({'error': 'Fornecedor j� est� vinculado a este evento'}), 400
        
        # Inserir fornecedor no evento
        cursor.execute("""
            INSERT INTO evento_fornecedores 
            (evento_id, fornecedor_id, categoria_id, subcategoria_id, valor, observacao, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (evento_id, fornecedor_id, categoria_id, subcategoria_id, valor, observacao, usuario_id))
        
        fornecedor_evento_id = cursor.fetchone()['id']
        
        # Recalcular custo total do evento (equipe + fornecedores)
        cursor.execute("""
            SELECT 
                COALESCE(SUM(ef.valor), 0) as custo_equipe,
                COALESCE((SELECT SUM(valor) FROM evento_fornecedores WHERE evento_id = %s), 0) as custo_fornecedores
            FROM evento_funcionarios ef
            WHERE ef.evento_id = %s
        """, (evento_id, evento_id))
        
        custos = cursor.fetchone()
        custo_equipe = float(custos['custo_equipe'])
        custo_fornecedores = float(custos['custo_fornecedores'])
        custo_total = custo_equipe + custo_fornecedores
        
        # Buscar valor_liquido_nf para recalcular margem
        cursor.execute("""
            SELECT valor_liquido_nf
            FROM eventos
            WHERE id = %s
        """, (evento_id,))
        
        evento_row = cursor.fetchone()
        valor_liquido = evento_row['valor_liquido_nf'] if evento_row and evento_row['valor_liquido_nf'] else 0
        
        # Calcular margem: Valor L�quido - Custo Total (Equipe + Fornecedores)
        margem = float(valor_liquido) - custo_total
        
        # Atualizar custo do evento E margem
        cursor.execute("""
            UPDATE eventos
            SET custo_evento = %s, margem = %s
            WHERE id = %s
        """, (custo_total, margem, evento_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Fornecedor adicionado ao evento',
            'id': fornecedor_evento_id,
            'custo_total': custo_total
        }), 201
    
    except Exception as e:
        logger.error(f"Erro ao adicionar fornecedor: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/eventos/fornecedores/<int:fornecedor_evento_id>', methods=['DELETE'])
@require_permission('eventos_delete')
def remover_fornecedor_evento(fornecedor_evento_id):
    """Remover fornecedor do evento"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Buscar evento_id antes de deletar
        cursor.execute("SELECT evento_id FROM evento_fornecedores WHERE id = %s", (fornecedor_evento_id,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            return jsonify({'error': 'V�nculo n�o encontrado'}), 404
        
        evento_id = row['evento_id']
        
        # Deletar v�nculo
        cursor.execute("DELETE FROM evento_fornecedores WHERE id = %s", (fornecedor_evento_id,))
        
        # Recalcular custo total do evento (equipe + fornecedores)
        cursor.execute("""
            SELECT 
                COALESCE(SUM(ef.valor), 0) as custo_equipe,
                COALESCE((SELECT SUM(valor) FROM evento_fornecedores WHERE evento_id = %s), 0) as custo_fornecedores
            FROM evento_funcionarios ef
            WHERE ef.evento_id = %s
        """, (evento_id, evento_id))
        
        custos = cursor.fetchone()
        custo_equipe = float(custos['custo_equipe'])
        custo_fornecedores = float(custos['custo_fornecedores'])
        custo_total = custo_equipe + custo_fornecedores
        
        # Buscar valor_liquido_nf para recalcular margem
        cursor.execute("""
            SELECT valor_liquido_nf
            FROM eventos
            WHERE id = %s
        """, (evento_id,))
        
        evento_row = cursor.fetchone()
        valor_liquido = evento_row['valor_liquido_nf'] if evento_row and evento_row['valor_liquido_nf'] else 0
        
        # Calcular margem: Valor L�quido - Custo Total (Equipe + Fornecedores)
        margem = float(valor_liquido) - custo_total
        
        # Atualizar custo do evento E margem
        cursor.execute("""
            UPDATE eventos
            SET custo_evento = %s, margem = %s
            WHERE id = %s
        """, (custo_total, margem, evento_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Fornecedor removido do evento',
            'custo_total': custo_total
        }), 200
    
    except Exception as e:
        logger.error(f"Erro ao remover fornecedor: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/api/subcategorias', methods=['GET'])
@require_permission('categorias_view')
def listar_subcategorias():
    """Lista subcategorias de uma categoria espec�fica"""
    try:
        from psycopg2.extras import RealDictCursor
        
        categoria_id = request.args.get('categoria_id')
        
        logger.info(f"?? GET /api/subcategorias - categoria_id={categoria_id}")
        
        if not categoria_id:
            return jsonify({'success': False, 'error': 'categoria_id � obrigat�rio'}), 400
        
        conn = db.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verificar se a coluna 'ativa' existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.columns 
                WHERE table_name = 'subcategorias' 
                AND column_name = 'ativa'
            );
        """)
        
        coluna_ativa_existe = cursor.fetchone()['exists']
        
        if coluna_ativa_existe:
            logger.info("   ? Coluna 'ativa' existe, filtrando por ativa=TRUE")
            cursor.execute("""
                SELECT id, nome, categoria_id, ativa
                FROM subcategorias
                WHERE categoria_id = %s AND ativa = TRUE
                ORDER BY nome
            """, (int(categoria_id),))
        else:
            logger.warning("   ??  Coluna 'ativa' n�o existe, listando todas")
            cursor.execute("""
                SELECT id, nome, categoria_id
                FROM subcategorias
                WHERE categoria_id = %s
                ORDER BY nome
            """, (int(categoria_id),))
        
        subcategorias = cursor.fetchall()
        cursor.close()
        
        logger.info(f"   ?? Retornando {len(subcategorias)} subcategorias")
        
        return jsonify({
            'success': True,
            'subcategorias': subcategorias
        }), 200
    
    except Exception as e:
        logger.error(f"? Erro ao listar subcategorias: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# === ROTAS DE RELAT�RIOS ===
# Todos os relat�rios movidos para app/routes/relatorios.py
# - dashboard, dashboard-completo, fluxo-projetado
# - analise-contas, resumo-parceiros, analise-categorias  
# - comparativo-periodos, indicadores, inadimplencia

@app.route('/api/relatorios/dashboard', methods=['GET'])
@require_permission('lancamentos_view')
def dashboard():
    """Dados para o dashboard - vers�o simplificada"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        # Pegar filtros opcionais
        ano = request.args.get('ano', type=int)
        mes = request.args.get('mes', type=int)
        
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
        
        # Filtrar lan�amentos por cliente se necess�rio
        usuario = request.usuario
        if usuario['tipo'] != 'admin' and usuario.get('cliente_id'):
            lancamentos = [l for l in lancamentos if getattr(l, 'pessoa', None) == usuario['cliente_id']]
        
        # Calcular saldos - USAR SALDO REAL DAS CONTAS (inclui extrato banc�rio)
        saldo_total = Decimal('0')
        
        # ?? Para cada conta, buscar saldo real (prioriza extrato banc�rio)
        for c in contas:
            try:
                with get_db_connection(empresa_id=empresa_id) as conn:
                    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                    
                    # Verificar se existem transa��es de extrato para esta conta
                    cursor.execute("""
                        SELECT saldo
                        FROM transacoes_extrato
                        WHERE empresa_id = %s
                        AND conta_bancaria = %s
                        ORDER BY data DESC, id DESC
                        LIMIT 1
                    """, (empresa_id, c.nome))
                    
                    ultima_transacao_extrato = cursor.fetchone()
                    
                    if ultima_transacao_extrato and ultima_transacao_extrato['saldo'] is not None:
                        # ? Usar saldo do extrato (fonte de verdade)
                        saldo_conta = Decimal(str(ultima_transacao_extrato['saldo']))
                        print(f"?? Dashboard - Conta {c.nome}: Saldo do extrato = R$ {saldo_conta}")
                    else:
                        # ?? Fallback: Calcular com lan�amentos manuais
                        cursor.execute("""
                            SELECT COALESCE(SUM(valor), 0) as total_receitas
                            FROM lancamentos
                            WHERE empresa_id = %s
                            AND conta_bancaria = %s
                            AND tipo = 'receita'
                            AND status = 'pago'
                        """, (empresa_id, c.nome))
                        total_receitas = Decimal(str(cursor.fetchone()['total_receitas'] or 0))
                        
                        cursor.execute("""
                            SELECT COALESCE(SUM(valor), 0) as total_despesas
                            FROM lancamentos
                            WHERE empresa_id = %s
                            AND conta_bancaria = %s
                            AND tipo = 'despesa'
                            AND status = 'pago'
                        """, (empresa_id, c.nome))
                        total_despesas = Decimal(str(cursor.fetchone()['total_despesas'] or 0))
                        
                        saldo_conta = Decimal(str(c.saldo_inicial)) + total_receitas - total_despesas
                        print(f"?? Dashboard - Conta {c.nome}: Saldo calculado = R$ {saldo_conta}")
                    
                    cursor.close()
                    saldo_total += saldo_conta
                    
            except Exception as e:
                print(f"?? Dashboard - Erro ao calcular saldo da conta {c.nome}: {e}")
                # Em caso de erro, usar saldo_inicial
                saldo_total += Decimal(str(c.saldo_inicial))
        
        # Contas pendentes
        hoje = date.today()
        contas_receber = Decimal('0')
        contas_pagar = Decimal('0')
        contas_vencidas = Decimal('0')
        
        for l in lancamentos:
            if l.tipo != TipoLancamento.TRANSFERENCIA:
                valor_decimal = Decimal(str(l.valor))
                if l.tipo == TipoLancamento.RECEITA and l.status == StatusLancamento.PENDENTE:
                    contas_receber += valor_decimal
                if l.tipo == TipoLancamento.DESPESA and l.status == StatusLancamento.PENDENTE:
                    contas_pagar += valor_decimal
                # Converter datetime para date se necess�rio
                data_venc = l.data_vencimento.date() if hasattr(l.data_vencimento, 'date') else l.data_vencimento
                if l.status == StatusLancamento.PENDENTE and data_venc < hoje:
                    contas_vencidas += valor_decimal
        
        # Dados para gr�fico - �ltimos 12 meses ou filtrado por ano/m�s
        from calendar import monthrange
        import locale
        
        # Tentar configurar locale para portugu�s
        try:
            locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
            except:
                pass
        
        meses_labels = []
        receitas_dados = []
        despesas_dados = []
        
        if ano and mes:
            # Apenas um m�s espec�fico
            _, ultimo_dia = monthrange(ano, mes)
            data_inicio = date(ano, mes, 1)
            data_fim = date(ano, mes, ultimo_dia)
            
            lancamentos_periodo = [
                l for l in lancamentos 
                if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA
                and data_inicio <= (l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento) <= data_fim
            ]
            
            receitas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_periodo if l.tipo == TipoLancamento.RECEITA)
            despesas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_periodo if l.tipo == TipoLancamento.DESPESA)
            
            meses_labels = [data_inicio.strftime('%b/%Y')]
            receitas_dados = [float(receitas_mes)]
            despesas_dados = [float(despesas_mes)]
        
        elif ano:
            # Todos os meses do ano
            for m in range(1, 13):
                _, ultimo_dia = monthrange(ano, m)
                data_inicio = date(ano, m, 1)
                data_fim = date(ano, m, ultimo_dia)
                
                lancamentos_periodo = [
                    l for l in lancamentos 
                    if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA
                    and data_inicio <= (l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento) <= data_fim
                ]
                
                receitas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_periodo if l.tipo == TipoLancamento.RECEITA)
                despesas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_periodo if l.tipo == TipoLancamento.DESPESA)
                
                meses_labels.append(data_inicio.strftime('%b/%Y'))
                receitas_dados.append(float(receitas_mes))
                despesas_dados.append(float(despesas_mes))
        
        else:
            # �ltimos 12 meses
            data_ref = hoje
            for i in range(11, -1, -1):
                mes_ref = data_ref.month - i
                ano_ref = data_ref.year
                
                while mes_ref <= 0:
                    mes_ref += 12
                    ano_ref -= 1
                
                _, ultimo_dia = monthrange(ano_ref, mes_ref)
                data_inicio = date(ano_ref, mes_ref, 1)
                data_fim = date(ano_ref, mes_ref, ultimo_dia)
                
                lancamentos_periodo = [
                    l for l in lancamentos 
                    if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA
                    and data_inicio <= (l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento) <= data_fim
                ]
                
                receitas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_periodo if l.tipo == TipoLancamento.RECEITA)
                despesas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_periodo if l.tipo == TipoLancamento.DESPESA)
                
                meses_labels.append(data_inicio.strftime('%b/%Y'))
                receitas_dados.append(float(receitas_mes))
                despesas_dados.append(float(despesas_mes))
        
        print(f"?? DADOS DO GR�FICO:")
        print(f"   Meses: {meses_labels}")
        print(f"   Receitas: {receitas_dados}")
        print(f"   Despesas: {despesas_dados}")
        print(f"?? CARDS:")
        print(f"   Contas a Receber: R$ {float(contas_receber):,.2f}")
        print(f"   Contas a Pagar: R$ {float(contas_pagar):,.2f}")
        print(f"   Contas Vencidas: R$ {float(contas_vencidas):,.2f}")
        print(f"   Saldo Total: R$ {float(saldo_total):,.2f}")
        print("=" * 80)
        
        return jsonify({
            'saldo_total': float(saldo_total),
            'contas_receber': float(contas_receber),
            'contas_pagar': float(contas_pagar),
            'contas_vencidas': float(contas_vencidas),
            'total_contas': len(contas),
            'total_lancamentos': len(lancamentos),
            'meses': meses_labels,
            'receitas': receitas_dados,
            'despesas': despesas_dados
        })
    except Exception as e:
        print(f"Erro no dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/relatorios/dashboard-completo', methods=['GET'])
@require_permission('relatorios_view')
def dashboard_completo():
    """Dashboard completo com an�lises detalhadas - apenas lan�amentos liquidados"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        if not data_inicio or not data_fim:
            return jsonify({'error': 'Datas obrigat�rias'}), 400
        
        data_inicio_obj = parse_date(data_inicio)
        data_fim_obj = parse_date(data_fim)
        
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        
        # Filtrar lan�amentos por cliente se necess�rio
        usuario = request.usuario
        if usuario['tipo'] != 'admin' and usuario.get('cliente_id'):
            lancamentos = [l for l in lancamentos if getattr(l, 'pessoa', None) == usuario['cliente_id']]
        
        # Filtrar apenas lan�amentos PAGOS/LIQUIDADOS no per�odo (baseado na data de pagamento)
        # Excluir transfer�ncias dos relat�rios
        lancamentos_periodo = []
        for l in lancamentos:
            if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA:
                data_pag = l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento
                if data_inicio_obj <= data_pag <= data_fim_obj:
                    lancamentos_periodo.append(l)
        
        # Evolu��o mensal (baseado na data de pagamento)
        evolucao = []
        current_date = data_inicio_obj
        
        while current_date <= data_fim_obj:
            mes_inicio = current_date.replace(day=1)
            if current_date.month == 12:
                mes_fim = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                mes_fim = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)
            
            # Filtrar por data de pagamento
            lancamentos_mes = [
                l for l in lancamentos_periodo
                if mes_inicio <= (l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento) <= mes_fim
            ]
            
            receitas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_mes if l.tipo == TipoLancamento.RECEITA)
            despesas_mes = sum(Decimal(str(l.valor)) for l in lancamentos_mes if l.tipo == TipoLancamento.DESPESA)
            saldo_mes = receitas_mes - despesas_mes
            
            evolucao.append({
                'periodo': current_date.strftime('%b/%y'),
                'receitas': float(receitas_mes),
                'despesas': float(despesas_mes),
                'saldo': float(saldo_mes)
            })
            
            # Avan�ar para o pr�ximo m�s
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1, day=1)
        
        # An�lise de Clientes
        clientes_resumo = {}
        for l in lancamentos_periodo:
            if l.tipo == TipoLancamento.RECEITA and l.pessoa:
                if l.pessoa not in clientes_resumo:
                    clientes_resumo[l.pessoa] = {'total': Decimal('0'), 'quantidade': 0}
                clientes_resumo[l.pessoa]['total'] += Decimal(str(l.valor))
                clientes_resumo[l.pessoa]['quantidade'] += 1
        
        melhor_cliente = None
        pior_cliente = None
        if clientes_resumo:
            melhor_cliente_nome = max(clientes_resumo, key=lambda x: clientes_resumo[x]['total'])
            pior_cliente_nome = min(clientes_resumo, key=lambda x: clientes_resumo[x]['total'])
            
            melhor_cliente = {
                'nome': melhor_cliente_nome,
                'total': float(clientes_resumo[melhor_cliente_nome]['total']),
                'quantidade': clientes_resumo[melhor_cliente_nome]['quantidade']
            }
            pior_cliente = {
                'nome': pior_cliente_nome,
                'total': float(clientes_resumo[pior_cliente_nome]['total']),
                'quantidade': clientes_resumo[pior_cliente_nome]['quantidade']
            }
        
        # An�lise de Fornecedores
        fornecedores_resumo = {}
        for l in lancamentos_periodo:
            if l.tipo == TipoLancamento.DESPESA and l.pessoa:
                if l.pessoa not in fornecedores_resumo:
                    fornecedores_resumo[l.pessoa] = {'total': Decimal('0'), 'quantidade': 0}
                fornecedores_resumo[l.pessoa]['total'] += Decimal(str(l.valor))
                fornecedores_resumo[l.pessoa]['quantidade'] += 1
        
        maior_fornecedor = None
        menor_fornecedor = None
        if fornecedores_resumo:
            maior_fornecedor_nome = max(fornecedores_resumo, key=lambda x: fornecedores_resumo[x]['total'])
            menor_fornecedor_nome = min(fornecedores_resumo, key=lambda x: fornecedores_resumo[x]['total'])
            
            maior_fornecedor = {
                'nome': maior_fornecedor_nome,
                'total': float(fornecedores_resumo[maior_fornecedor_nome]['total']),
                'quantidade': fornecedores_resumo[maior_fornecedor_nome]['quantidade']
            }
            menor_fornecedor = {
                'nome': menor_fornecedor_nome,
                'total': float(fornecedores_resumo[menor_fornecedor_nome]['total']),
                'quantidade': fornecedores_resumo[menor_fornecedor_nome]['quantidade']
            }
        
        # An�lise de Categorias - Receitas
        categorias_receita = {}
        for l in lancamentos_periodo:
            if l.tipo == TipoLancamento.RECEITA:
                cat = l.categoria or 'Sem categoria'
                subcat = l.subcategoria or 'Sem subcategoria'
                
                if cat not in categorias_receita:
                    categorias_receita[cat] = {}
                if subcat not in categorias_receita[cat]:
                    categorias_receita[cat][subcat] = Decimal('0')
                
                categorias_receita[cat][subcat] += Decimal(str(l.valor))
        
        melhor_categoria_receita = None
        if categorias_receita:
            melhor_cat = max(categorias_receita, key=lambda x: sum(categorias_receita[x].values()))
            melhor_subcat = max(categorias_receita[melhor_cat], key=lambda x: categorias_receita[melhor_cat][x])
            
            melhor_categoria_receita = {
                'categoria': melhor_cat,
                'subcategoria': melhor_subcat if melhor_subcat != 'Sem subcategoria' else None,
                'total': float(categorias_receita[melhor_cat][melhor_subcat])
            }
        
        # An�lise de Categorias - Despesas
        categorias_despesa = {}
        for l in lancamentos_periodo:
            if l.tipo == TipoLancamento.DESPESA:
                cat = l.categoria or 'Sem categoria'
                subcat = l.subcategoria or 'Sem subcategoria'
                
                if cat not in categorias_despesa:
                    categorias_despesa[cat] = {}
                if subcat not in categorias_despesa[cat]:
                    categorias_despesa[cat][subcat] = Decimal('0')
                
                categorias_despesa[cat][subcat] += Decimal(str(l.valor))
        
        maior_categoria_despesa = None
        if categorias_despesa:
            maior_cat = max(categorias_despesa, key=lambda x: sum(categorias_despesa[x].values()))
            maior_subcat = max(categorias_despesa[maior_cat], key=lambda x: categorias_despesa[maior_cat][x])
            
            maior_categoria_despesa = {
                'categoria': maior_cat,
                'subcategoria': maior_subcat if maior_subcat != 'Sem subcategoria' else None,
                'total': float(categorias_despesa[maior_cat][maior_subcat])
            }
        
        return jsonify({
            'evolucao': evolucao,
            'melhor_cliente': melhor_cliente,
            'pior_cliente': pior_cliente,
            'maior_fornecedor': maior_fornecedor,
            'menor_fornecedor': menor_fornecedor,
            'melhor_categoria_receita': melhor_categoria_receita,
            'maior_categoria_despesa': maior_categoria_despesa
        })
        
    except Exception as e:
        print(f"Erro no dashboard completo: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/relatorios/fluxo-projetado', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_fluxo_projetado():
    """Relat�rio de fluxo de caixa PROJETADO (incluindo lan�amentos pendentes futuros)"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        # Receber filtros - padr�o � projetar pr�ximos X dias
        dias = request.args.get('dias')
        dias = int(dias) if dias else 30
        
        hoje = date.today()
        data_inicial = hoje
        data_final = hoje + timedelta(days=dias)
        periodo_texto = f"PROJE��O - PR�XIMOS {dias} DIAS"
        
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
        
        # Filtrar lan�amentos por cliente se necess�rio
        usuario = request.usuario
        if usuario['tipo'] != 'admin' and usuario.get('cliente_id'):
            lancamentos = [l for l in lancamentos if getattr(l, 'pessoa', None) == usuario['cliente_id']]
        
        # ?? PRIORIDADE 1: Buscar saldo atual do EXTRATO BANC�RIO (fonte de verdade)
        saldo_atual = Decimal('0')
        
        try:
            with get_db_connection(empresa_id=empresa_id) as conn:
                cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                
                for c in contas:
                    # Verificar se existem transa��es de extrato para esta conta
                    cursor.execute("""
                        SELECT saldo, data, id
                        FROM transacoes_extrato
                        WHERE empresa_id = %s
                        AND conta_bancaria = %s
                        ORDER BY data DESC, id DESC
                        LIMIT 1
                    """, (empresa_id, c.nome))
                    
                    ultima_transacao_extrato = cursor.fetchone()
                    
                    if ultima_transacao_extrato and ultima_transacao_extrato['saldo'] is not None:
                        # ? USAR SALDO DO EXTRATO (mais recente e confi�vel)
                        saldo_conta = Decimal(str(ultima_transacao_extrato['saldo']))
                        print(f"?? Fluxo Projetado - Conta {c.nome}: Saldo do extrato = R$ {saldo_conta:.2f}")
                    else:
                        # ?? FALLBACK: Calcular com base nos lan�amentos manuais
                        print(f"?? Fluxo Projetado - Conta {c.nome}: Sem extrato, calculando com lan�amentos...")
                        
                        # Somar receitas pagas
                        cursor.execute("""
                            SELECT COALESCE(SUM(valor), 0) as total_receitas
                            FROM lancamentos
                            WHERE empresa_id = %s
                            AND conta_bancaria = %s
                            AND tipo = 'receita'
                            AND status = 'pago'
                        """, (empresa_id, c.nome))
                        resultado_receitas = cursor.fetchone()
                        total_receitas = Decimal(str(resultado_receitas['total_receitas'] or 0))
                        
                        # Somar despesas pagas
                        cursor.execute("""
                            SELECT COALESCE(SUM(valor), 0) as total_despesas
                            FROM lancamentos
                            WHERE empresa_id = %s
                            AND conta_bancaria = %s
                            AND tipo = 'despesa'
                            AND status = 'pago'
                        """, (empresa_id, c.nome))
                        resultado_despesas = cursor.fetchone()
                        total_despesas = Decimal(str(resultado_despesas['total_despesas'] or 0))
                        
                        # Calcular saldo
                        saldo_conta = Decimal(str(c.saldo_inicial)) + total_receitas - total_despesas
                        print(f"?? Fluxo Projetado - Conta {c.nome}: Saldo calculado = R$ {saldo_conta:.2f}")
                    
                    saldo_atual += saldo_conta
                
                cursor.close()
                
        except Exception as e:
            print(f"?? Erro ao calcular saldo atual no fluxo projetado: {e}")
            import traceback
            traceback.print_exc()
            # FALLBACK em caso de erro: usar saldo inicial
            for c in contas:
                saldo_atual += Decimal(str(c.saldo_inicial))
        
        # Buscar lan�amentos PENDENTES para proje��o (vencidos + futuros)
        lancamentos_futuros = []
        lancamentos_vencidos = []
        receitas_previstas = Decimal('0')
        despesas_previstas = Decimal('0')
        receitas_vencidas = Decimal('0')
        despesas_vencidas = Decimal('0')
        
        for l in lancamentos:
            if l.status == StatusLancamento.PENDENTE and l.tipo != TipoLancamento.TRANSFERENCIA:
                # Converter data_vencimento para date se for datetime
                if isinstance(l.data_vencimento, datetime):
                    data_venc = l.data_vencimento.date()
                else:
                    data_venc = l.data_vencimento
                
                valor_decimal = Decimal(str(l.valor))
                
                # Lan�amentos vencidos (j� passaram do vencimento)
                if data_venc < hoje:
                    lancamentos_vencidos.append(l)
                    if l.tipo == TipoLancamento.RECEITA:
                        receitas_vencidas += valor_decimal
                    else:
                        despesas_vencidas += valor_decimal
                
                # Lan�amentos futuros (dentro do per�odo de proje��o)
                elif data_inicial <= data_venc <= data_final:
                    lancamentos_futuros.append(l)
                    if l.tipo == TipoLancamento.RECEITA:
                        receitas_previstas += valor_decimal
                    else:
                        despesas_previstas += valor_decimal
        
        # Calcular saldo projetado (incluindo vencidos)
        saldo_projetado = saldo_atual + (receitas_previstas + receitas_vencidas) - (despesas_previstas + despesas_vencidas)
        
        # Ordenar por data de vencimento
        lancamentos_vencidos.sort(key=lambda x: x.data_vencimento)
        lancamentos_futuros.sort(key=lambda x: x.data_vencimento)
        
        # Montar fluxo projetado dia a dia (VENCIDOS PRIMEIRO, depois futuros)
        fluxo = []
        saldo_acumulado = saldo_atual
        
        # Adicionar vencidos primeiro
        for lanc in lancamentos_vencidos:
            valor_decimal = Decimal(str(lanc.valor))
            if lanc.tipo == TipoLancamento.RECEITA:
                saldo_acumulado += valor_decimal
            else:
                saldo_acumulado -= valor_decimal
            
            # Calcular dias de atraso
            data_venc_date = lanc.data_vencimento.date() if isinstance(lanc.data_vencimento, datetime) else lanc.data_vencimento
            dias_atraso = (hoje - data_venc_date).days
            
            fluxo.append({
                'data_vencimento': lanc.data_vencimento.isoformat(),
                'descricao': lanc.descricao,
                'tipo': lanc.tipo.value,
                'valor': float(lanc.valor),
                'categoria': lanc.categoria,
                'subcategoria': lanc.subcategoria,
                'pessoa': lanc.pessoa,
                'conta_bancaria': lanc.conta_bancaria,
                'saldo_acumulado': float(saldo_acumulado),
                'status': 'VENCIDO',
                'dias_atraso': dias_atraso
            })
        
        # Adicionar lan�amentos futuros
        for lanc in lancamentos_futuros:
            valor_decimal = Decimal(str(lanc.valor))
            if lanc.tipo == TipoLancamento.RECEITA:
                saldo_acumulado += valor_decimal
            else:
                saldo_acumulado -= valor_decimal
            
            fluxo.append({
                'data_vencimento': lanc.data_vencimento.isoformat(),
                'descricao': lanc.descricao,
                'tipo': lanc.tipo.value,
                'valor': float(lanc.valor),
                'categoria': lanc.categoria,
                'subcategoria': lanc.subcategoria,
                'pessoa': lanc.pessoa,
                'conta_bancaria': lanc.conta_bancaria,
                'saldo_acumulado': float(saldo_acumulado),
                'status': 'PENDENTE',
                'dias_atraso': 0
            })
        
        return jsonify({
            'periodo_texto': periodo_texto,
            'data_inicial': data_inicial.isoformat(),
            'data_final': data_final.isoformat(),
            'saldo_atual': float(saldo_atual),
            'receitas_previstas': float(receitas_previstas),
            'despesas_previstas': float(despesas_previstas),
            'receitas_vencidas': float(receitas_vencidas),
            'despesas_vencidas': float(despesas_vencidas),
            'saldo_projetado': float(saldo_projetado),
            'total_vencidos': len(lancamentos_vencidos),
            'total_futuros': len(lancamentos_futuros),
            'fluxo': fluxo
        })
    except Exception as e:
        print(f"Erro no fluxo projetado: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/relatorios/analise-contas', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_analise_contas():
    """Relat�rio de an�lise de contas a pagar e receber"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
    hoje = date.today()
    
    # Filtrar lan�amentos por cliente se necess�rio
    usuario = request.usuario
    if usuario['tipo'] != 'admin' and usuario.get('cliente_id'):
        lancamentos = [l for l in lancamentos if getattr(l, 'pessoa', None) == usuario['cliente_id']]
    
    # Fun��o auxiliar para converter datetime para date
    def get_date(data):
        return data.date() if hasattr(data, 'date') else data
    
    # Totais (excluindo transfer�ncias)
    total_receber = sum(l.valor for l in lancamentos 
                       if l.tipo == TipoLancamento.RECEITA and l.status == StatusLancamento.PENDENTE)
    total_pagar = sum(l.valor for l in lancamentos 
                     if l.tipo == TipoLancamento.DESPESA and l.status == StatusLancamento.PENDENTE)
    
    receber_vencidos = sum(l.valor for l in lancamentos 
                          if l.tipo == TipoLancamento.RECEITA and 
                          l.status == StatusLancamento.PENDENTE and 
                          get_date(l.data_vencimento) < hoje)
    pagar_vencidos = sum(l.valor for l in lancamentos 
                        if l.tipo == TipoLancamento.DESPESA and 
                        l.status == StatusLancamento.PENDENTE and 
                        get_date(l.data_vencimento) < hoje)
    
    # Aging (an�lise de vencimento) - excluindo transfer�ncias
    pendentes = [l for l in lancamentos if l.status == StatusLancamento.PENDENTE and l.tipo != TipoLancamento.TRANSFERENCIA]
    
    vencidos = sum(l.valor for l in pendentes if (get_date(l.data_vencimento) - hoje).days < 0)  # type: ignore
    ate_7 = sum(l.valor for l in pendentes if 0 <= (get_date(l.data_vencimento) - hoje).days <= 7)  # type: ignore
    ate_15 = sum(l.valor for l in pendentes if 7 < (get_date(l.data_vencimento) - hoje).days <= 15)  # type: ignore
    ate_30 = sum(l.valor for l in pendentes if 15 < (get_date(l.data_vencimento) - hoje).days <= 30)  # type: ignore
    ate_60 = sum(l.valor for l in pendentes if 30 < (get_date(l.data_vencimento) - hoje).days <= 60)  # type: ignore
    ate_90 = sum(l.valor for l in pendentes if 60 < (get_date(l.data_vencimento) - hoje).days <= 90)  # type: ignore
    acima_90 = sum(l.valor for l in pendentes if (get_date(l.data_vencimento) - hoje).days > 90)  # type: ignore
    
    return jsonify({
        'total_receber': float(total_receber),
        'total_pagar': float(total_pagar),
        'receber_vencidos': float(receber_vencidos),
        'pagar_vencidos': float(pagar_vencidos),
        'aging': {
            'vencidos': float(vencidos),
            'ate_7': float(ate_7),
            'ate_15': float(ate_15),
            'ate_30': float(ate_30),
            'ate_60': float(ate_60),
            'ate_90': float(ate_90),
            'acima_90': float(acima_90)
        }
    })


@app.route('/api/lancamentos/<int:lancamento_id>/pagar', methods=['PUT'])
@require_permission('lancamentos_edit')
def pagar_lancamento(lancamento_id):
    """Marca um lan�amento como pago"""
    try:
        # ?? Obter empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'empresa_id n�o encontrado na sess�o'}), 403
        
        data = request.json
        conta = data.get('conta_bancaria', '') if data else ''
        data_pagamento = datetime.fromisoformat(data.get('data_pagamento', datetime.now().isoformat())).date() if data else date.today()
        juros = float(data.get('juros', 0)) if data else 0
        desconto = float(data.get('desconto', 0)) if data else 0
        observacoes = data.get('observacoes', '') if data else ''
        
        success = db_pagar_lancamento(empresa_id, lancamento_id, conta, data_pagamento, juros, desconto, observacoes)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/lancamentos/<int:lancamento_id>/liquidar', methods=['POST'])
@require_permission('lancamentos_edit')
def liquidar_lancamento(lancamento_id):
    """Liquida um lan�amento (marca como pago com dados completos)"""
    try:
        print("\n" + "="*80)
        print(f"?? DEBUG LIQUIDA��O - ID: {lancamento_id}")
        print("="*80)
        
        # ?? Obter empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            print("? ERRO: empresa_id n�o encontrado na sess�o")
            return jsonify({'success': False, 'error': 'empresa_id n�o encontrado na sess�o'}), 403
        print(f"?? Empresa ID: {empresa_id}")
        
        data = request.json or {}
        print(f"?? Dados recebidos: {data}")
        
        conta = data.get('conta_bancaria', '')
        data_pagamento_str = data.get('data_pagamento', '')
        juros = float(data.get('juros', 0))
        desconto = float(data.get('desconto', 0))
        observacoes = data.get('observacoes', '')
        
        print(f"?? Par�metros extra�dos:")
        print(f"   - Conta: {conta}")
        print(f"   - Data: {data_pagamento_str}")
        print(f"   - Juros: {juros}")
        print(f"   - Desconto: {desconto}")
        print(f"   - Observa��es: {observacoes}")
        
        if not conta:
            print("? ERRO: Conta banc�ria vazia")
            return jsonify({'success': False, 'error': 'Conta banc�ria � obrigat�ria'}), 400
        
        if not data_pagamento_str or data_pagamento_str.strip() == '':
            print("? ERRO: Data de pagamento vazia")
            return jsonify({'success': False, 'error': 'Data de pagamento � obrigat�ria'}), 400
        
        data_pagamento = datetime.fromisoformat(data_pagamento_str).date()
        print(f"?? Data convertida: {data_pagamento} (tipo: {type(data_pagamento)})")
        
        print(f"?? Chamando db_pagar_lancamento...")
        print(f"   Argumentos: ({empresa_id}, {lancamento_id}, {conta}, {data_pagamento}, {juros}, {desconto}, {observacoes})")
        
        success = db_pagar_lancamento(empresa_id, lancamento_id, conta, data_pagamento, juros, desconto, observacoes)
        
        print(f"? Resultado: {success}")
        print("="*80 + "\n")
        
        return jsonify({'success': success})
    except Exception as e:
        print(f"? EXCE��O CAPTURADA:")
        print(f"   Tipo: {type(e).__name__}")
        print(f"   Mensagem: {str(e)}")
        import traceback
        print(f"   Traceback:\n{traceback.format_exc()}")
        print("="*80 + "\n")
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/lancamentos/<int:lancamento_id>/cancelar', methods=['PUT'])
@require_permission('lancamentos_edit')
def cancelar_lancamento_route(lancamento_id):
    """Cancela um lan�amento"""
    try:
        # ?? Obter empresa_id da sess�o
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'empresa_id n�o encontrado na sess�o'}), 403
        
        success = db_cancelar_lancamento(empresa_id, lancamento_id)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/lancamentos/<int:lancamento_id>/associacao', methods=['PATCH'])
@require_permission('lancamentos_edit')
def atualizar_associacao_lancamento(lancamento_id):
    """
    Atualiza apenas o campo de associa��o de um lan�amento (salvamento autom�tico)
    
    Security:
        ?? Validado empresa_id da sess�o
        ?? Verifica permiss�o lancamentos_edit
    """
    try:
        # ?? VALIDA��O DE SEGURAN�A
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403
        
        data = request.get_json()
        nova_associacao = data.get('associacao', '')
        
        # ?? Atualizar associacao E numero_documento simultaneamente (sincroniza��o bidirecional)
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE lancamentos 
                SET associacao = %s, numero_documento = %s
                WHERE id = %s AND empresa_id = %s
                RETURNING id
            """, (nova_associacao, nova_associacao, lancamento_id, empresa_id))
            
            resultado = cursor.fetchone()
            conn.commit()
            cursor.close()
            
            if not resultado:
                return jsonify({'success': False, 'error': 'Lan�amento n�o encontrado'}), 404
            
            return jsonify({'success': True, 'id': lancamento_id, 'associacao': nova_associacao})
            
    except Exception as e:
        print(f"? Erro ao atualizar associa��o: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400


# === ROTA PRINCIPAL ===

@app.route('/login')
def login_page():
    """P�gina de login"""
    return render_template('login.html')

@app.route('/api/csrf-token', methods=['GET'])
def get_csrf_token_endpoint():
    """
    Endpoint para obter CSRF token via API
    Gera e retorna um token CSRF v�lido
    """
    from flask_wtf.csrf import generate_csrf
    token = generate_csrf()
    print(f"?? CSRF Token gerado via API: {token[:20]}...")
    return jsonify({
        'csrf_token': token,
        'success': True
    })

@app.route('/admin')
@require_admin
def admin_page():
    """Painel administrativo - apenas para admins"""
    print(f"\n?????? ROTA /admin ALCAN�ADA - Decorador passou! ??????\n")
    return render_template('admin.html')

@app.route('/admin/fix-empresa-id', methods=['GET', 'POST'])
@require_admin
def admin_fix_empresa_id():
    """
    Rota administrativa para corrigir empresa_id em registros antigos
    
    ATENCAO: Esta rota atualiza TODOS os registros sem empresa_id!
    Use com cuidado!
    """
    from database_postgresql import get_db_connection
    
    if request.method == 'GET':
        # Mostrar p�gina de confirma��o
        return """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Corrigir empresa_id</title>
            <style>
                body { font-family: Arial; padding: 40px; max-width: 800px; margin: 0 auto; }
                .warning { background: #fff3cd; border: 2px solid #ffc107; padding: 20px; margin: 20px 0; border-radius: 8px; }
                .btn { background: #007bff; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; }
                .btn:hover { background: #0056b3; }
                .danger { background: #dc3545; }
                .danger:hover { background: #c82333; }
                pre { background: #f8f9fa; padding: 15px; border-radius: 5px; overflow-x: auto; }
            </style>
        </head>
        <body>
            <h1>?? Corrigir empresa_id em Registros Antigos</h1>
            
            <div class="warning">
                <h3>?? ATEN��O</h3>
                <p>Esta a��o ir� atualizar TODOS os registros sem <code>empresa_id</code> nas seguintes tabelas:</p>
                <ul>
                    <li>contratos</li>
                    <li>sessoes</li>
                    <li>lancamentos</li>
                    <li>clientes</li>
                    <li>fornecedores</li>
                    <li>categorias</li>
                </ul>
                <p><strong>Os registros ser�o associados � empresa ID 19.</strong></p>
            </div>
            
            <h3>O que ser� feito:</h3>
            <pre>
UPDATE contratos SET empresa_id = 19 WHERE empresa_id IS NULL;
UPDATE sessoes SET empresa_id = 19 WHERE empresa_id IS NULL;
UPDATE lancamentos SET empresa_id = 19 WHERE empresa_id IS NULL;
...
            </pre>
            
            <form method="POST" onsubmit="return confirm('Tem certeza? Esta a��o n�o pode ser desfeita!');">
                <button type="submit" class="btn danger">? Executar Corre��o</button>
                <a href="/admin" style="margin-left: 20px;">? Cancelar</a>
            </form>
        </body>
        </html>
        """
    
    # POST - Executar corre��o
    try:
        with get_db_connection(allow_global=True) as conn:
            cursor = conn.cursor()
            
            resultados = []
            
            # An�lise inicial
            tabelas = ['contratos', 'sessoes', 'lancamentos', 'clientes', 'fornecedores', 'categorias']
            analise_inicial = {}
            
            for tabela in tabelas:
                cursor.execute(f"""
                    SELECT 
                        COUNT(*) as total,
                        COUNT(CASE WHEN empresa_id IS NULL THEN 1 END) as sem_empresa_id
                    FROM {tabela}
                """)
                result = cursor.fetchone()
                analise_inicial[tabela] = {
                    'total': result['total'],
                    'sem_empresa_id': result['sem_empresa_id']
                }
            
            # Executar corre��es
            updates = {
                'contratos': "UPDATE contratos SET empresa_id = 19 WHERE empresa_id IS NULL",
                'sessoes': "UPDATE sessoes SET empresa_id = 19 WHERE empresa_id IS NULL",
                'lancamentos': "UPDATE lancamentos SET empresa_id = 19 WHERE empresa_id IS NULL",
                'clientes': "UPDATE clientes SET empresa_id = 19 WHERE empresa_id IS NULL",
                'fornecedores': "UPDATE fornecedores SET empresa_id = 19 WHERE empresa_id IS NULL",
                'categorias': "UPDATE categorias SET empresa_id = 19 WHERE empresa_id IS NULL"
            }
            
            for tabela, sql in updates.items():
                cursor.execute(sql)
                count = cursor.rowcount
                resultados.append(f"? {tabela}: {count} registro(s) atualizado(s)")
            
            conn.commit()
            cursor.close()
            
            # Retornar resultado
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Corre��o Conclu�da</title>
                <style>
                    body {{ font-family: Arial; padding: 40px; max-width: 800px; margin: 0 auto; }}
                    .success {{ background: #d4edda; border: 2px solid #28a745; padding: 20px; margin: 20px 0; border-radius: 8px; }}
                    .resultado {{ background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 10px 0; }}
                </style>
            </head>
            <body>
                <h1>? Corre��o Conclu�da com Sucesso!</h1>
                
                <div class="success">
                    <h3>Resultados:</h3>
                    {''.join(f'<div class="resultado">{r}</div>' for r in resultados)}
                </div>
                
                <a href="/admin">? Voltar ao Admin</a>
            </body>
            </html>
            """
            
            return html
            
    except Exception as e:
        return f"""
        <h1>? Erro ao executar corre��o</h1>
        <pre>{str(e)}</pre>
        <a href="/admin">? Voltar</a>
        """, 500

# ============================================================================
# ROTAS DE ADMINISTRA��O MOBILE
# ============================================================================

@app.route('/api/admin/mobile/config', methods=['GET'])
@require_admin
def admin_get_mobile_config():
    """
    Obt�m informa��es b�sicas sobre mobile (apenas detec��o de dispositivo)
    
    GET /api/admin/mobile/config
    
    Response: {
        "success": true,
        "device_info": {...}
    }
    """
    try:
        device_info = get_device_info()
        
        return jsonify({
            'success': True,
            'device_info': device_info,
            'message': 'Sistema usa detec��o b�sica de dispositivos mobile'
        }), 200
    except Exception as e:
        logger.error(f"Erro ao obter info mobile: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/mobile/config/<key>', methods=['PUT'])
@require_admin
def admin_update_mobile_config(key):
    """
    Atualiza uma configura��o mobile (admin apenas)
    
    PUT /api/admin/mobile/config/mobile_enabled
    Body: {
        "value": "true",
        "description": "Habilitar vers�o mobile"
    }
    
    Response: {
        "success": false,
        "message": "Configura��es mobile simplificadas - n�o h� configura��es para atualizar"
    }
    """
    return jsonify({
        'success': False,
        'message': 'Sistema usa detec��o b�sica de mobile - n�o h� configura��es din�micas',
        'info': 'Mobile detection baseado em User-Agent apenas'
    }), 400


@app.route('/api/device-info', methods=['GET'])
def get_device_info_route():
    """
    Retorna informa��es sobre o dispositivo atual
    �til para debug e UI
    
    GET /api/device-info
    
    Response: {
        "is_mobile": true,
        "is_mobile_app": false,
        "platform": "android",
        "os": "Android",
        "user_agent": "..."
    }
    """
    device_info = get_device_info()
    return jsonify(device_info), 200


@app.route('/api/device-preference', methods=['POST'])
@require_auth
def set_device_preference_route():
    """
    Define prefer�ncia de dispositivo do usu�rio
    
    POST /api/device-preference
    Body: {
        "preference": "web" | "mobile"
    }
    
    Response: {
        "success": true,
        "preference": "web"
    }
    """
    try:
        data = request.get_json()
        preference = data.get('preference', 'web')
        
        if preference not in ['web', 'mobile']:
            return jsonify({
                'success': False,
                'error': 'Prefer�ncia inv�lida. Use "web" ou "mobile".'
            }), 400
        
        from mobile_config import store_device_preference
        store_device_preference(preference)
        
        return jsonify({
            'success': True,
            'preference': preference
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/debug-usuario')
def debug_usuario():
    """Rota de debug para verificar dados do usu�rio atual"""
    usuario = get_usuario_logado()
    
    debug_info = {
        'encontrado': usuario is not None,
        'dados': {}
    }
    
    if usuario:
        debug_info['dados'] = {
            'id': usuario.get('id'),
            'username': usuario.get('username'),
            'tipo': usuario.get('tipo'),
            'tipo_python_type': str(type(usuario.get('tipo'))),
            'tipo_repr': repr(usuario.get('tipo')),
            'tipo_len': len(usuario.get('tipo', '')),
            'tipo_bytes': list(usuario.get('tipo', '').encode()) if usuario.get('tipo') else [],
            'tipo_comparacao_admin': usuario.get('tipo') == 'admin',
            'tipo_comparacao_cliente': usuario.get('tipo') == 'cliente',
            'nome_completo': usuario.get('nome_completo'),
            'email': usuario.get('email'),
            'cliente_id': usuario.get('cliente_id')
        }
    
    return jsonify(debug_info)

@app.route('/')
def index():
    """P�gina principal - Nova interface moderna"""
    # Verificar se est� autenticado
    usuario = get_usuario_logado()
    if not usuario:
        return render_template('login.html')
    
    # Passa o timestamp de build para o template
    return render_template('interface_nova.html', build_timestamp=BUILD_TIMESTAMP)

@app.route('/admin/import')
@require_permission('admin')
def admin_import_page():
    """P�gina de importa��o de banco de dados"""
    return render_template('admin_import.html')

# ============================================================================
# ROTAS DE IMPORTA��O DE BANCO DE DADOS
# ============================================================================
from werkzeug.utils import secure_filename
from database_import_manager import DatabaseImportManager
import tempfile
import csv

ALLOWED_EXTENSIONS = {'sql', 'dump', 'backup', 'csv', 'json', 'db', 'db-shm', 'db-wal', 'sqlite', 'sqlite3'}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/admin/import/upload', methods=['POST'])
@csrf.exempt
@require_permission('admin')
def upload_import_file():
    """Upload e processamento de arquivo para importa��o"""
    logger.info("?? [IMPORT] Fun��o upload_import_file() chamada")
    logger.info(f"?? [IMPORT] Request method: {request.method}")
    logger.info(f"?? [IMPORT] Content-Type: {request.content_type}")
    try:
        logger.info("?? Upload de arquivo iniciado")
        
        # Verificar se � upload m�ltiplo
        multiple_files = request.files.getlist('files[]')
        
        if multiple_files:
            logger.info(f"?? Upload m�ltiplo: {len(multiple_files)} arquivos")
            temp_dir = tempfile.gettempdir()
            db_file_path = None
            
            for file in multiple_files:
                if file.filename == '':
                    continue
                
                if not allowed_file(file.filename):
                    return jsonify({'error': f'Formato n�o suportado: {file.filename}'}), 400
                
                filename = secure_filename(file.filename)
                temp_path = os.path.join(temp_dir, f"import_{session.get('usuario_id')}_{filename}")
                file.save(temp_path)
                
                if filename.endswith('.db') or filename.endswith('.sqlite') or filename.endswith('.sqlite3'):
                    db_file_path = temp_path
                
                logger.info(f"? Arquivo salvo: {temp_path}")
            
            if not db_file_path:
                return jsonify({'error': 'Arquivo .db principal n�o encontrado'}), 400
            
            manager = DatabaseImportManager()
            schema = manager.parse_sqlite_database(db_file_path)
            
            return jsonify({
                'success': True,
                'schema': schema,
                'temp_file': db_file_path,
                'total_tabelas': len(schema),
                'total_registros': sum(t.get('total_registros', 0) for t in schema.values())
            })
        
        # Upload �nico
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': f'Formato n�o suportado. Use: {", ".join(ALLOWED_EXTENSIONS)}'}), 400
        
        # Validar tamanho
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': 'Arquivo muito grande (m�x: 100MB)'}), 400
        
        filename = secure_filename(file.filename)
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"import_{session.get('usuario_id')}_{filename}")
        file.save(temp_path)
        
        logger.info(f"? Arquivo salvo: {temp_path}")
        
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        manager = DatabaseImportManager()
        
        if file_ext in ['sql', 'dump', 'backup']:
            schema = manager.parse_sql_dump(temp_path)
        elif file_ext == 'csv':
            schema = manager.parse_csv_file(temp_path)
        elif file_ext == 'json':
            schema = manager.parse_json_file(temp_path)
        elif file_ext in ['db', 'db-shm', 'db-wal', 'sqlite', 'sqlite3']:
            schema = manager.parse_sqlite_database(temp_path)
        else:
            return jsonify({'error': 'Formato n�o reconhecido'}), 400
        
        return jsonify({
            'success': True,
            'schema': schema,
            'temp_file': temp_path,
            'total_tabelas': len(schema),
            'total_registros': sum(t.get('total_registros', 0) for t in schema.values())
        })
        
    except Exception as e:
        logger.error(f"? Erro no upload: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/import/schema/interno', methods=['GET'])
@csrf.exempt
@require_permission('admin')
def get_internal_schema():
    """Obt�m schema do banco interno usando a mesma conex�o do sistema"""
    try:
        db_instance = DatabaseManager()
        conn = db_instance.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # Buscar todas as tabelas do schema public
        cursor.execute("""
            SELECT 
                table_name,
                (SELECT COUNT(*) FROM information_schema.columns 
                 WHERE table_schema = 'public' AND table_name = t.table_name) as column_count
            FROM information_schema.tables t
            WHERE table_schema = 'public'
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)
        
        tables = cursor.fetchall()
        schema = {}
        
        for table in tables:
            table_name = table['table_name']
            
            # Buscar colunas da tabela
            cursor.execute("""
                SELECT 
                    column_name as name,
                    data_type as type,
                    is_nullable,
                    column_default as default_value
                FROM information_schema.columns
                WHERE table_schema = 'public'
                AND table_name = %s
                ORDER BY ordinal_position
            """, (table_name,))
            
            columns = cursor.fetchall()
            
            # Contar registros (com timeout de 5 segundos)
            try:
                cursor.execute(f"SELECT COUNT(*) as total FROM {table_name}")
                count_result = cursor.fetchone()
                total_registros = count_result['total'] if count_result else 0
            except:
                total_registros = 0
            
            schema[table_name] = {
                'columns': [dict(col) for col in columns],
                'total_registros': total_registros
            }
        
        conn.close()
        
        logger.info(f"? Schema interno carregado: {len(schema)} tabelas")
        
        return jsonify({
            'success': True,
            'schema': schema,
            'total_tabelas': len(schema)
        })
        
    except Exception as e:
        logger.error(f"? Erro ao obter schema interno: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/import/sugestao-mapeamento', methods=['POST'])
@csrf.exempt
@require_permission('admin')
def suggest_mapping():
    """Gera sugest�es de mapeamento entre tabelas"""
    try:
        data = request.json
        schema_externo = data.get('schema_externo')
        schema_interno = data.get('schema_interno')
        
        if not schema_externo or not schema_interno:
            return jsonify({'error': 'Schemas externo e interno s�o obrigat�rios'}), 400
        
        manager = DatabaseImportManager()
        sugestoes = manager.suggest_table_mapping(schema_externo, schema_interno)
        
        return jsonify({
            'success': True,
            'sugestoes': sugestoes,
            'total_mapeamentos': len(sugestoes)
        })
        
    except Exception as e:
        logger.error(f"? Erro ao gerar sugest�es: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/import/criar', methods=['POST'])
@csrf.exempt
@require_permission('admin')
def create_import():
    """Cria registro de importa��o com mapeamentos"""
    try:
        data = request.json
        empresa_id = data.get('empresa_id')
        mapeamentos = data.get('mapeamentos')
        schema_externo = data.get('schema_externo')
        
        if not empresa_id or not mapeamentos:
            return jsonify({'error': 'empresa_id e mapeamentos s�o obrigat�rios'}), 400
        
        manager = DatabaseImportManager()
        manager.connect()
        
        # Criar registro na tabela import_historico
        import_id = manager.create_import_record(
            empresa_id=empresa_id,
            usuario_id=session.get('usuario_id'),
            fonte_tipo='sqlite',
            mapeamentos=mapeamentos,
            schema_externo=schema_externo
        )
        
        manager.close()
        
        return jsonify({
            'success': True,
            'import_id': import_id,
            'message': 'Importa��o criada com sucesso'
        })
        
    except Exception as e:
        logger.error(f"? Erro ao criar importa��o: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/import/executar/<int:import_id>', methods=['POST'])
@csrf.exempt
@require_permission('admin')
def execute_import(import_id):
    """Executa a importa��o de dados"""
    try:
        data = request.json
        arquivo_path = data.get('arquivo_path')
        
        if not arquivo_path:
            return jsonify({'error': 'arquivo_path � obrigat�rio'}), 400
        
        manager = DatabaseImportManager()
        manager.connect()
        
        resultado = manager.execute_import(import_id, arquivo_path)
        
        manager.close()
        
        return jsonify({
            'success': True,
            'resultado': resultado,
            'message': 'Importa��o executada com sucesso'
        })
        
    except Exception as e:
        logger.error(f"? Erro ao executar importa��o: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/import/reverter/<int:import_id>', methods=['POST'])
@csrf.exempt
@require_permission('admin')
def rollback_import(import_id):
    """Reverte uma importa��o (rollback)"""
    try:
        manager = DatabaseImportManager()
        manager.connect()
        
        resultado = manager.rollback_import(import_id)
        
        manager.close()
        
        return jsonify({
            'success': True,
            'resultado': resultado,
            'message': 'Importa��o revertida com sucesso'
        })
        
    except Exception as e:
        logger.error(f"? Erro ao reverter importa��o: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================================

@app.route('/old')
@require_auth
def old_index():
    """P�gina antiga (backup)"""
    return render_template('interface.html')

@app.route('/teste')
def teste():
    """P�gina de teste JavaScript"""
    return render_template('teste.html')

@app.route('/teste-api')
def teste_api():
    """P�gina de teste API"""
    return render_template('teste_api.html')

# === ENDPOINTS DE RELAT�RIOS ===

@app.route('/api/relatorios/resumo-parceiros', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_resumo_parceiros():
    """Relat�rio de resumo por cliente/fornecedor"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        data_inicio = request.args.get('data_inicio', (date.today() - timedelta(days=30)).isoformat())
        data_fim = request.args.get('data_fim', date.today().isoformat())
        
        data_inicio = datetime.fromisoformat(data_inicio).date()
        data_fim = datetime.fromisoformat(data_fim).date()
        
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        
        # Agrupar por pessoa
        resumo_clientes = {}
        resumo_fornecedores = {}
        
        for l in lancamentos:
            if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA:
                data_pag = l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento
                if data_inicio <= data_pag <= data_fim and l.pessoa:
                    valor = float(l.valor)
                    categoria = l.categoria or 'Sem categoria'
                    
                    if l.tipo == TipoLancamento.RECEITA:
                        if l.pessoa not in resumo_clientes:
                            resumo_clientes[l.pessoa] = {'total': 0, 'quantidade': 0, 'categorias': {}}
                        resumo_clientes[l.pessoa]['total'] += valor
                        resumo_clientes[l.pessoa]['quantidade'] += 1
                        # Contar por categoria
                        if categoria not in resumo_clientes[l.pessoa]['categorias']:
                            resumo_clientes[l.pessoa]['categorias'][categoria] = 0
                        resumo_clientes[l.pessoa]['categorias'][categoria] += valor
                    else:
                        if l.pessoa not in resumo_fornecedores:
                            resumo_fornecedores[l.pessoa] = {'total': 0, 'quantidade': 0, 'categorias': {}}
                        resumo_fornecedores[l.pessoa]['total'] += valor
                        resumo_fornecedores[l.pessoa]['quantidade'] += 1
                        # Contar por categoria
                        if categoria not in resumo_fornecedores[l.pessoa]['categorias']:
                            resumo_fornecedores[l.pessoa]['categorias'][categoria] = 0
                        resumo_fornecedores[l.pessoa]['categorias'][categoria] += valor
        
        # Adicionar categoria principal para cada cliente/fornecedor
        for pessoa, dados in resumo_clientes.items():
            if dados['categorias']:
                categoria_principal = max(dados['categorias'].items(), key=lambda x: x[1])
                dados['categoria_principal'] = categoria_principal[0]
            else:
                dados['categoria_principal'] = 'Sem categoria'
        
        for pessoa, dados in resumo_fornecedores.items():
            if dados['categorias']:
                categoria_principal = max(dados['categorias'].items(), key=lambda x: x[1])
                dados['categoria_principal'] = categoria_principal[0]
            else:
                dados['categoria_principal'] = 'Sem categoria'
        
        return jsonify({
            'clientes': resumo_clientes,
            'fornecedores': resumo_fornecedores,
            'periodo': {
                'inicio': data_inicio.isoformat(),
                'fim': data_fim.isoformat()
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/relatorios/analise-categorias', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_analise_categorias():
    """Relat�rio de an�lise por categorias"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        data_inicio = request.args.get('data_inicio', (date.today() - timedelta(days=30)).isoformat())
        data_fim = request.args.get('data_fim', date.today().isoformat())
        
        data_inicio = datetime.fromisoformat(data_inicio).date()
        data_fim = datetime.fromisoformat(data_fim).date()
        
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        
        # Agrupar por categoria e subcategoria
        receitas = {}
        despesas = {}
        
        for l in lancamentos:
            if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA:
                data_pag = l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento
                if data_inicio <= data_pag <= data_fim:
                    categoria = l.categoria or 'Sem Categoria'
                    subcategoria = l.subcategoria or 'Sem Subcategoria'
                    valor = float(l.valor)
                    
                    if l.tipo == TipoLancamento.RECEITA:
                        if categoria not in receitas:
                            receitas[categoria] = {}
                        if subcategoria not in receitas[categoria]:
                            receitas[categoria][subcategoria] = {'total': 0, 'quantidade': 0}
                        receitas[categoria][subcategoria]['total'] += valor
                        receitas[categoria][subcategoria]['quantidade'] += 1
                    else:
                        if categoria not in despesas:
                            despesas[categoria] = {}
                        if subcategoria not in despesas[categoria]:
                            despesas[categoria][subcategoria] = {'total': 0, 'quantidade': 0}
                        despesas[categoria][subcategoria]['total'] += valor
                        despesas[categoria][subcategoria]['quantidade'] += 1
        
        return jsonify({
            'receitas': receitas,
            'despesas': despesas,
            'periodo': {
                'inicio': data_inicio.isoformat(),
                'fim': data_fim.isoformat()
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/relatorios/comparativo-periodos', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_comparativo_periodos():
    """Relat�rio comparativo entre per�odos"""
    try:
        print(f"[COMPARATIVO] Iniciando comparativo de per�odos")
        
        # Per�odo 1
        data_inicio1 = request.args.get('data_inicio1')
        data_fim1 = request.args.get('data_fim1')
        
        # Per�odo 2
        data_inicio2 = request.args.get('data_inicio2')
        data_fim2 = request.args.get('data_fim2')
        
        print(f"[COMPARATIVO] Par�metros recebidos:")
        print(f"  Per�odo 1: {data_inicio1} at� {data_fim1}")
        print(f"  Per�odo 2: {data_inicio2} at� {data_fim2}")
        
        if not all([data_inicio1, data_fim1, data_inicio2, data_fim2]):
            return jsonify({'error': 'Par�metros de datas obrigat�rios'}), 400
        
        # ?? VALIDA��O DE SEGURAN�A
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'erro': 'Empresa n�o selecionada'}), 403
        
        print(f"[COMPARATIVO] Empresa ID: {empresa_id}")
        
        try:
            data_inicio1 = datetime.fromisoformat(data_inicio1).date()
            data_fim1 = datetime.fromisoformat(data_fim1).date()
            data_inicio2 = datetime.fromisoformat(data_inicio2).date()
            data_fim2 = datetime.fromisoformat(data_fim2).date()
        except ValueError as e:
            print(f"[COMPARATIVO] Erro ao converter datas: {e}")
            return jsonify({'error': 'Formato de data inv�lido'}), 400
        
        print(f"[COMPARATIVO] Buscando lan�amentos...")
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        print(f"[COMPARATIVO] Total de lan�amentos: {len(lancamentos)}")
        
        def calcular_periodo(data_ini, data_fim):
            print(f"[COMPARATIVO] Calculando per�odo: {data_ini} at� {data_fim}")
            receitas = Decimal('0')
            despesas = Decimal('0')
            receitas_por_categoria = {}
            despesas_por_categoria = {}
            receitas_por_subcategoria = {}
            despesas_por_subcategoria = {}
            
            for l in lancamentos:
                try:
                    if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA:
                        data_pag = l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento
                        if data_ini <= data_pag <= data_fim:
                            valor = Decimal(str(l.valor))
                            categoria = l.categoria or 'Sem categoria'
                            subcategoria = l.subcategoria or 'Sem subcategoria'
                            chave_completa = f"{categoria} > {subcategoria}"
                            
                            if l.tipo == TipoLancamento.RECEITA:
                                receitas += valor
                                receitas_por_categoria[categoria] = receitas_por_categoria.get(categoria, Decimal('0')) + valor
                                receitas_por_subcategoria[chave_completa] = receitas_por_subcategoria.get(chave_completa, Decimal('0')) + valor
                            else:
                                despesas += valor
                                despesas_por_categoria[categoria] = despesas_por_categoria.get(categoria, Decimal('0')) + valor
                                despesas_por_subcategoria[chave_completa] = despesas_por_subcategoria.get(chave_completa, Decimal('0')) + valor
                except Exception as e:
                    print(f"[COMPARATIVO] Erro ao processar lan�amento {l.id}: {e}")
                    continue
            
            # Encontrar maiores por categoria
            maior_receita_cat = max(receitas_por_categoria.items(), key=lambda x: x[1]) if receitas_por_categoria else ('Nenhuma', Decimal('0'))
            maior_despesa_cat = max(despesas_por_categoria.items(), key=lambda x: x[1]) if despesas_por_categoria else ('Nenhuma', Decimal('0'))
            
            # Encontrar maiores por subcategoria
            maior_receita_sub = max(receitas_por_subcategoria.items(), key=lambda x: x[1]) if receitas_por_subcategoria else ('Nenhuma', Decimal('0'))
            maior_despesa_sub = max(despesas_por_subcategoria.items(), key=lambda x: x[1]) if despesas_por_subcategoria else ('Nenhuma', Decimal('0'))
            
            # Top 3 categorias
            top_receitas = sorted(receitas_por_categoria.items(), key=lambda x: x[1], reverse=True)[:3]
            top_despesas = sorted(despesas_por_categoria.items(), key=lambda x: x[1], reverse=True)[:3]
            
            return {
                'receitas': float(receitas),
                'despesas': float(despesas),
                'saldo': float(receitas - despesas),
                'maior_receita': {'categoria': maior_receita_cat[0], 'valor': float(maior_receita_cat[1])},
                'maior_despesa': {'categoria': maior_despesa_cat[0], 'valor': float(maior_despesa_cat[1])},
                'maior_receita_sub': {'subcategoria': maior_receita_sub[0], 'valor': float(maior_receita_sub[1])},
                'maior_despesa_sub': {'subcategoria': maior_despesa_sub[0], 'valor': float(maior_despesa_sub[1])},
                'top_receitas': [{'categoria': k, 'valor': float(v), 'percentual': float(v/receitas*100) if receitas > 0 else 0} for k, v in top_receitas],
                'top_despesas': [{'categoria': k, 'valor': float(v), 'percentual': float(v/despesas*100) if despesas > 0 else 0} for k, v in top_despesas],
                'qtd_categorias_receitas': len(receitas_por_categoria),
                'qtd_categorias_despesas': len(despesas_por_categoria)
            }
        
        print(f"[COMPARATIVO] Calculando per�odo 1...")
        periodo1 = calcular_periodo(data_inicio1, data_fim1)
        print(f"[COMPARATIVO] Per�odo 1 calculado - Receitas: {periodo1['receitas']}, Despesas: {periodo1['despesas']}")
        
        print(f"[COMPARATIVO] Calculando per�odo 2...")
        periodo2 = calcular_periodo(data_inicio2, data_fim2)
        print(f"[COMPARATIVO] Per�odo 2 calculado - Receitas: {periodo2['receitas']}, Despesas: {periodo2['despesas']}")
        
        # Calcular varia��es
        variacao_receitas = ((periodo2['receitas'] - periodo1['receitas']) / periodo1['receitas'] * 100) if periodo1['receitas'] > 0 else 0
        variacao_despesas = ((periodo2['despesas'] - periodo1['despesas']) / periodo1['despesas'] * 100) if periodo1['despesas'] > 0 else 0
        variacao_saldo = ((periodo2['saldo'] - periodo1['saldo']) / abs(periodo1['saldo']) * 100) if periodo1['saldo'] != 0 else 0
        
        print(f"[COMPARATIVO] Varia��es calculadas - Receitas: {variacao_receitas}%, Despesas: {variacao_despesas}%, Saldo: {variacao_saldo}%")
        
        resultado = {
            'periodo1': {
                'datas': {'inicio': data_inicio1.isoformat(), 'fim': data_fim1.isoformat()},
                'dados': periodo1
            },
            'periodo2': {
                'datas': {'inicio': data_inicio2.isoformat(), 'fim': data_fim2.isoformat()},
                'dados': periodo2
            },
            'variacoes': {
                'receitas': round(variacao_receitas, 2),
                'despesas': round(variacao_despesas, 2),
                'saldo': round(variacao_saldo, 2)
            }
        }
        
        print(f"[COMPARATIVO] Retornando resultado com sucesso")
        return jsonify(resultado)
        
    except Exception as e:
        print(f"[COMPARATIVO] ERRO CR�TICO: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/relatorios/indicadores', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_indicadores():
    """Relat�rio de indicadores financeiros"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        contas = db.listar_contas_por_empresa(empresa_id=empresa_id)
        
        # Obter filtros de data
        data_inicio_str = request.args.get('data_inicio')
        data_fim_str = request.args.get('data_fim')
        
        hoje = date.today()
        
        if data_inicio_str and data_fim_str:
            inicio_mes = parse_date(data_inicio_str)
            fim_periodo = parse_date(data_fim_str)
        else:
            # Padr�o: m�s atual
            inicio_mes = date(hoje.year, hoje.month, 1)
            fim_periodo = hoje
        
        # Totais do m�s atual
        receitas_mes = Decimal('0')
        despesas_mes = Decimal('0')
        
        # Totais a receber/pagar
        total_receber = Decimal('0')
        total_pagar = Decimal('0')
        
        # Vencidos
        vencidos_receber = Decimal('0')
        vencidos_pagar = Decimal('0')
        
        for l in lancamentos:
            if l.status == StatusLancamento.PAGO and l.data_pagamento and l.tipo != TipoLancamento.TRANSFERENCIA:
                data_pag = l.data_pagamento.date() if hasattr(l.data_pagamento, 'date') else l.data_pagamento
                if data_pag >= inicio_mes and data_pag <= fim_periodo:
                    valor = Decimal(str(l.valor))
                    if l.tipo == TipoLancamento.RECEITA:
                        receitas_mes += valor
                    else:
                        despesas_mes += valor
            
            if l.status == StatusLancamento.PENDENTE:
                valor = Decimal(str(l.valor))
                data_venc = l.data_vencimento.date() if hasattr(l.data_vencimento, 'date') else l.data_vencimento
                
                if l.tipo == TipoLancamento.RECEITA:
                    total_receber += valor
                    if data_venc < hoje:
                        vencidos_receber += valor
                else:
                    total_pagar += valor
                    if data_venc < hoje:
                        vencidos_pagar += valor
        
        # Saldo em caixa
        saldo_caixa = sum(Decimal(str(c.saldo_inicial)) for c in contas)
        
        # Liquidez = (Saldo + A Receber) / A Pagar
        liquidez = float((saldo_caixa + total_receber) / total_pagar) if total_pagar > 0 else 0
        
        # Margem l�quida = (Receitas - Despesas) / Receitas * 100
        margem = float((receitas_mes - despesas_mes) / receitas_mes * 100) if receitas_mes > 0 else 0
        
        return jsonify({
            'saldo_caixa': float(saldo_caixa),
            'mes_atual': {
                'receitas': float(receitas_mes),
                'despesas': float(despesas_mes),
                'saldo': float(receitas_mes - despesas_mes)
            },
            'pendentes': {
                'receber': float(total_receber),
                'pagar': float(total_pagar)
            },
            'vencidos': {
                'receber': float(vencidos_receber),
                'pagar': float(vencidos_pagar)
            },
            'indicadores': {
                'liquidez': round(liquidez, 2),
                'margem_liquida': round(margem, 2)
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/relatorios/inadimplencia', methods=['GET'])
@require_permission('relatorios_view')
def relatorio_inadimplencia():
    """Relat�rio de inadimpl�ncia"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        lancamentos = db.listar_lancamentos(empresa_id=empresa_id)
        hoje = date.today()
        
        inadimplentes = []
        
        for l in lancamentos:
            # Excluir transfer�ncias e considerar apenas PENDENTES
            if l.tipo == TipoLancamento.TRANSFERENCIA:
                continue
                
            if l.status == StatusLancamento.PENDENTE:
                # Converter data_vencimento para date se for datetime
                data_venc = l.data_vencimento.date() if hasattr(l.data_vencimento, 'date') else l.data_vencimento
                
                # Verificar se est� vencido (data anterior a hoje)
                if data_venc < hoje:
                    dias_atraso = (hoje - data_venc).days
                    inadimplentes.append({
                        'id': l.id,
                        'tipo': l.tipo.value.upper(),
                        'descricao': l.descricao,
                        'valor': float(l.valor),
                        'data_vencimento': data_venc.isoformat(),
                        'dias_atraso': dias_atraso,
                        'pessoa': l.pessoa or 'N�o informado',
                        'categoria': l.categoria or 'Sem categoria'
                    })
        
        # Ordenar por dias de atraso (maior para menor)
        inadimplentes.sort(key=lambda x: x['dias_atraso'], reverse=True)
        
        total_inadimplente = sum(i['valor'] for i in inadimplentes)
        
        return jsonify({
            'inadimplentes': inadimplentes,
            'total': total_inadimplente,
            'quantidade': len(inadimplentes)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/favicon.ico')
def favicon():
    """Retorna o favicon"""
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'favicon.svg',
        mimetype='image/svg+xml'
    )


# === EXPORTA��O DE CLIENTES E FORNECEDORES ===

@app.route('/api/clientes/exportar/pdf', methods=['GET'])
@require_permission('clientes_view')
def exportar_clientes_pdf():
    """Exporta clientes para PDF"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.pagesizes import A4, landscape  # type: ignore
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
        from reportlab.lib.units import cm  # type: ignore
        from reportlab.pdfgen import canvas  # type: ignore
        from io import BytesIO
        
        clientes = db.listar_clientes(empresa_id=empresa_id)
        
        # Criar PDF em mem�ria
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*cm, leftMargin=0.5*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # T�tulo
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2c3e50'), spaceAfter=15, alignment=1)
        elements.append(Paragraph(f'LISTA DE CLIENTES - {get_current_date_br()}', title_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Estilo de par�grafo para c�lulas
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=7, leading=9, wordWrap='CJK')
        
        # Dados da tabela com Paragraph para quebra de linha
        data = [['Raz�o Social', 'Nome Fantasia', 'CNPJ', 'Cidade', 'Est', 'Telefone', 'Email']]
        
        for cli in clientes:
            # Truncar textos longos e usar Paragraph para quebra autom�tica
            razao = cli.get('razao_social', '-') or '-'
            fantasia = cli.get('nome_fantasia', '-') or '-'
            cnpj = cli.get('cnpj', '-') or '-'
            cidade = cli.get('cidade', '-') or '-'
            estado = cli.get('estado', '-') or '-'
            telefone = cli.get('telefone', '-') or '-'
            email = cli.get('email', '-') or '-'
            
            # Limitar tamanho dos textos muito longos
            if len(razao) > 35:
                razao = razao[:32] + '...'
            if len(fantasia) > 30:
                fantasia = fantasia[:27] + '...'
            if len(cidade) > 20:
                cidade = cidade[:17] + '...'
            if len(email) > 25:
                email = email[:22] + '...'
            
            data.append([
                Paragraph(razao, cell_style),
                Paragraph(fantasia, cell_style),
                Paragraph(cnpj, cell_style),
                Paragraph(cidade, cell_style),
                estado,
                Paragraph(telefone, cell_style),
                Paragraph(email, cell_style)
            ])
        
        # Largura dispon�vel: A4 landscape = 29.7cm, menos margens = ~28.7cm
        # Criar tabela com larguras proporcionais
        table = Table(data, colWidths=[6*cm, 5*cm, 3.5*cm, 4*cm, 1.5*cm, 3.5*cm, 4.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'clientes_{get_current_date_filename()}.pdf')
    
    except Exception as e:
        print(f"Erro ao exportar PDF: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes/exportar/excel', methods=['GET'])
@require_permission('clientes_view')
def exportar_clientes_excel():
    """Exporta clientes para Excel"""
    # ?? VALIDA��O DE SEGURAN�A
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'erro': 'Empresa n�o selecionada'}), 403
    
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore
        from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
        from io import BytesIO
        
        clientes = db.listar_clientes(empresa_id=empresa_id)
        
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active  # type: ignore
        ws.title = "Clientes"
        
        headers = ['Raz�o Social', 'Nome Fantasia', 'CNPJ', 'IE', 'IM', 'Rua', 'N�mero', 'Complemento', 'Bairro', 'Cidade', 'Estado', 'CEP', 'Telefone', 'Email']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for cli in clientes:
            ws.append([
                cli.get('razao_social', ''), cli.get('nome_fantasia', ''), cli.get('cnpj', ''),
                cli.get('ie', ''), cli.get('im', ''), cli.get('rua', ''), cli.get('numero', ''),
                cli.get('complemento', ''), cli.get('bairro', ''), cli.get('cidade', ''),
                cli.get('estado', ''), cli.get('cep', ''), cli.get('telefone', ''), cli.get('email', '')
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'clientes_{get_current_date_filename()}.xlsx')
    
    except Exception as e:
        print(f"Erro ao exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/fornecedores/exportar/pdf', methods=['GET'])
@require_permission('fornecedores_view')
def exportar_fornecedores_pdf():
    """Exporta fornecedores para PDF"""
    try:
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.pagesizes import A4, landscape  # type: ignore
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
        from reportlab.lib.units import cm  # type: ignore
        from reportlab.pdfgen import canvas  # type: ignore
        from io import BytesIO
        
        fornecedores = db.listar_fornecedores()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*cm, leftMargin=0.5*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2c3e50'), spaceAfter=15, alignment=1)
        elements.append(Paragraph(f'LISTA DE FORNECEDORES - {get_current_date_br()}', title_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Estilo de par�grafo para c�lulas
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=7, leading=9, wordWrap='CJK')
        
        data = [['Raz�o Social', 'Nome Fantasia', 'CNPJ', 'Cidade', 'Est', 'Telefone', 'Email']]
        
        for forn in fornecedores:
            razao = forn.get('razao_social', '-') or '-'
            fantasia = forn.get('nome_fantasia', '-') or '-'
            cnpj = forn.get('cnpj', '-') or '-'
            cidade = forn.get('cidade', '-') or '-'
            estado = forn.get('estado', '-') or '-'
            telefone = forn.get('telefone', '-') or '-'
            email = forn.get('email', '-') or '-'
            
            # Limitar tamanho dos textos muito longos
            if len(razao) > 35:
                razao = razao[:32] + '...'
            if len(fantasia) > 30:
                fantasia = fantasia[:27] + '...'
            if len(cidade) > 20:
                cidade = cidade[:17] + '...'
            if len(email) > 25:
                email = email[:22] + '...'
            
            data.append([
                Paragraph(razao, cell_style),
                Paragraph(fantasia, cell_style),
                Paragraph(cnpj, cell_style),
                Paragraph(cidade, cell_style),
                estado,
                Paragraph(telefone, cell_style),
                Paragraph(email, cell_style)
            ])
        
        table = Table(data, colWidths=[6*cm, 5*cm, 3.5*cm, 4*cm, 1.5*cm, 3.5*cm, 4.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'fornecedores_{get_current_date_filename()}.pdf')
    
    except Exception as e:
        print(f"Erro ao exportar PDF: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/fornecedores/exportar/excel', methods=['GET'])
@require_permission('fornecedores_view')
def exportar_fornecedores_excel():
    """Exporta fornecedores para Excel"""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore
        from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
        from io import BytesIO
        
        fornecedores = db.listar_fornecedores()
        
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active  # type: ignore
        ws.title = "Fornecedores"
        
        headers = ['Raz�o Social', 'Nome Fantasia', 'CNPJ', 'IE', 'IM', 'Rua', 'N�mero', 'Complemento', 'Bairro', 'Cidade', 'Estado', 'CEP', 'Telefone', 'Email']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for forn in fornecedores:
            ws.append([
                forn.get('razao_social', ''), forn.get('nome_fantasia', ''), forn.get('cnpj', ''),
                forn.get('ie', ''), forn.get('im', ''), forn.get('rua', ''), forn.get('numero', ''),
                forn.get('complemento', ''), forn.get('bairro', ''), forn.get('cidade', ''),
                forn.get('estado', ''), forn.get('cep', ''), forn.get('telefone', ''), forn.get('email', '')
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'fornecedores_{get_current_date_filename()}.xlsx')
    
    except Exception as e:
        print(f"Erro ao exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500


# === ROTAS DO MENU OPERACIONAL ===
# Rotas de Contratos movidas para app/routes/contratos.py
# Rotas de Sess�es movidas para app/routes/sessoes.py

@app.route('/api/comissoes', methods=['GET', 'POST'])
@require_permission('operacional_view')
def comissoes():
    """Gerenciar comiss�es"""
    if request.method == 'GET':
        try:
            comissoes = db.listar_comissoes()
            return jsonify(comissoes)
        except Exception as e:
            logger.error(f"Erro ao listar comiss�es: {e}")
            return jsonify({'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            comissao_id = db.adicionar_comissao(data)
            return jsonify({'success': True, 'message': 'Comiss�o criada com sucesso', 'id': comissao_id}), 201
        except Exception as e:
            logger.error(f"Erro ao criar comiss�o: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/comissoes/<int:comissao_id>', methods=['PUT', 'DELETE'])
@require_permission('operacional_edit')
def comissao_detalhes(comissao_id):
    """Atualizar ou excluir comiss�o"""
    if request.method == 'PUT':
        try:
            data = request.json
            print(f"?? [COMISS�O PUT] ID: {comissao_id}, Dados: {data}")
            success = db.atualizar_comissao(comissao_id, data)
            if success:
                print(f"? [COMISS�O PUT] Atualizada com sucesso")
                return jsonify({'success': True, 'message': 'Comiss�o atualizada com sucesso'})
            print(f"?? [COMISS�O PUT] N�o encontrada")
            return jsonify({'success': False, 'error': 'Comiss�o n�o encontrada'}), 404
        except Exception as e:
            print(f"? [COMISS�O PUT] Erro: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # DELETE
        try:
            print(f"?? [COMISS�O DELETE] ID: {comissao_id}")
            success = db.deletar_comissao(comissao_id)
            if success:
                print(f"? [COMISS�O DELETE] Exclu�da com sucesso")
                return jsonify({'success': True, 'message': 'Comiss�o exclu�da com sucesso'})
            print(f"?? [COMISS�O DELETE] N�o encontrada")
            return jsonify({'success': False, 'error': 'Comiss�o n�o encontrada'}), 404
        except Exception as e:
            print(f"? [COMISS�O DELETE] Erro: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sessao-equipe', methods=['GET', 'POST', 'DELETE'])
@require_permission('operacional_view')
def sessao_equipe():
    """Gerenciar equipe de sess�o"""
    if request.method == 'DELETE':
        # Endpoint tempor�rio para FOR�AR limpeza da tabela
        import sys
        print(f"[CLEANUP] INICIANDO LIMPEZA DA TABELA sessao_equipe", flush=True)
        sys.stdout.flush()
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM sessao_equipe")
            deleted = cursor.rowcount
            print(f"[CLEANUP] Deletando {deleted} registros...", flush=True)
            sys.stdout.flush()
            conn.commit()
            print(f"[CLEANUP] COMMIT executado!", flush=True)
            sys.stdout.flush()
            cursor.close()
            conn.close()
            print(f"[CLEANUP] Tabela limpa! {deleted} registros removidos", flush=True)
            return jsonify({'success': True, 'deleted': deleted})
        except Exception as e:
            print(f"[CLEANUP ERROR] {e}", flush=True)
            sys.stdout.flush()
            return jsonify({'success': False, 'error': str(e)}), 500
    elif request.method == 'GET':
        try:
            lista = db.listar_sessao_equipe()
            return jsonify(lista)
        except Exception as e:
            logger.error(f"Erro ao listar equipe: {e}")
            return jsonify({'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            se_id = db.adicionar_sessao_equipe(data)
            return jsonify({'success': True, 'message': 'Membro adicionado com sucesso', 'id': se_id}), 201
        except Exception as e:
            logger.error(f"Erro ao adicionar membro equipe: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sessao-equipe/<int:membro_id>', methods=['PUT', 'DELETE'])
@require_permission('operacional_edit')
def sessao_equipe_detalhes(membro_id):
    """Atualizar ou excluir membro da equipe"""
    if request.method == 'PUT':
        try:
            data = request.json
            print(f"?? [EQUIPE PUT] ID: {membro_id}, Dados: {data}")
            success = db.atualizar_sessao_equipe(membro_id, data)
            if success:
                print(f"? [EQUIPE PUT] Membro atualizado com sucesso")
                return jsonify({'success': True, 'message': 'Membro atualizado com sucesso'})
            print(f"?? [EQUIPE PUT] Membro n�o encontrado")
            return jsonify({'success': False, 'error': 'Membro n�o encontrado'}), 404
        except Exception as e:
            print(f"? [EQUIPE PUT] Erro: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # DELETE
        try:
            print(f"?? [EQUIPE DELETE] ID: {membro_id}")
            success = db.deletar_sessao_equipe(membro_id)
            if success:
                print(f"? [EQUIPE DELETE] Membro removido com sucesso")
                return jsonify({'success': True, 'message': 'Membro removido com sucesso'})
            print(f"?? [EQUIPE DELETE] Membro n�o encontrado")
            return jsonify({'success': False, 'error': 'Membro n�o encontrado'}), 404
        except Exception as e:
            print(f"? [EQUIPE DELETE] Erro: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agenda', methods=['GET', 'POST'])
@require_permission('agenda_view')
def agenda():
    """Gerenciar agenda"""
    if request.method == 'GET':
        try:
            eventos = db.listar_agenda()
            return jsonify(eventos)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            agenda_id = db.adicionar_agenda(data)
            return jsonify({'message': 'Agendamento criado com sucesso', 'id': agenda_id}), 201
        except Exception as e:
            return jsonify({'error': str(e)}), 500


@app.route('/api/agenda/<int:agendamento_id>', methods=['PUT', 'DELETE'])
@require_permission('agenda_edit')
def agenda_detalhes(agendamento_id):
    """Atualizar ou excluir agendamento"""
    if request.method == 'PUT':
        try:
            data = request.json
            success = db.atualizar_agenda(agendamento_id, data)
            if success:
                return jsonify({'message': 'Agendamento atualizado com sucesso'})
            return jsonify({'error': 'Agendamento n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:  # DELETE
        try:
            success = db.deletar_agenda(agendamento_id)
            if success:
                return jsonify({'message': 'Agendamento exclu�do com sucesso'})
            return jsonify({'error': 'Agendamento n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500


@app.route('/api/estoque/produtos', methods=['GET', 'POST'])
@require_permission('estoque_view')
def produtos():
    """Gerenciar produtos do estoque"""
    if request.method == 'GET':
        try:
            produtos = db.listar_produtos()
            return jsonify(produtos)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            produto_id = db.adicionar_produto(data)
            return jsonify({'message': 'Produto criado com sucesso', 'id': produto_id}), 201
        except Exception as e:
            return jsonify({'error': str(e)}), 500


@app.route('/api/estoque/produtos/<int:produto_id>', methods=['PUT', 'DELETE'])
@require_permission('estoque_edit')
def produto_detalhes(produto_id):
    """Atualizar ou excluir produto"""
    if request.method == 'PUT':
        try:
            data = request.json
            success = db.atualizar_produto(produto_id, data)
            if success:
                return jsonify({'message': 'Produto atualizado com sucesso'})
            return jsonify({'error': 'Produto n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:  # DELETE
        try:
            success = db.deletar_produto(produto_id)
            if success:
                return jsonify({'message': 'Produto exclu�do com sucesso'})
            return jsonify({'error': 'Produto n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500


# ============================================================================
# RECURSOS HUMANOS - FUNCION�RIOS
# ============================================================================

@app.route('/api/rh/funcionarios', methods=['GET'])
def listar_funcionarios_rh():
    """Listar funcion�rios para uso em dropdowns (sem require_permission para permitir uso em modais)"""
    print("=" * 80)
    print("?? REQUISI��O RECEBIDA: /api/rh/funcionarios")
    print("=" * 80)
    try:
        print("?? Obtendo conex�o com banco...")
        conn = db.get_connection()
        cursor = conn.cursor()
        
        print("?? Verificando total de funcion�rios na tabela...")
        # Primeiro, verificar se a tabela existe e tem dados
        cursor.execute("SELECT COUNT(*) as total FROM funcionarios")
        result = cursor.fetchone()
        total = result['total'] if isinstance(result, dict) else (result[0] if result else 0)
        print(f"?? Total de funcion�rios na tabela: {total}")
        
        # Buscar apenas colunas que existem (id, nome, ativo)
        cursor.execute("""
            SELECT id, nome, ativo
            FROM funcionarios
            WHERE ativo = true
            ORDER BY nome
        """)
        
        rows = cursor.fetchall()
        
        print(f"?? Total de funcion�rios ativos encontrados: {len(rows)}")
        
        # Converter para dicion�rios (apenas id e nome para dropdown)
        funcionarios = []
        for row in rows:
            if isinstance(row, dict):
                funcionario = {
                    'id': row['id'],
                    'nome': row['nome']
                }
                print(f"  ? Funcion�rio: {row['nome']} (ID: {row['id']}, Ativo: {row.get('ativo', True)})")
            else:
                funcionario = {
                    'id': row[0],
                    'nome': row[1]
                }
                print(f"  ? Funcion�rio: {row[1]} (ID: {row[0]}, Ativo: {row[2] if len(row) > 2 else True})")
            funcionarios.append(funcionario)
        
        cursor.close()
        conn.close()
        
        print(f"? Retornando {len(funcionarios)} funcion�rios")
        return jsonify({'success': True, 'data': funcionarios})
    except Exception as e:
        print(f"? Erro ao listar funcion�rios RH: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ROTAS DE KITS - MOVIDAS PARA app/routes/kits.py
# ============================================================================
# As rotas de Kits foram extra�das para o Blueprint em app/routes/kits.py
# como parte da Fase 2 de otimiza��o (refatora��o modular)


# ============================================================================
# ENDPOINTS TEMPOR�RIOS PARA DEBUG E MIGRATIONS
# ?? ESTES ENDPOINTS S� FUNCIONAM EM DESENVOLVIMENTO
# ============================================================================

def _check_debug_endpoint_allowed():
    """Verifica se endpoints de debug podem ser executados"""
    if IS_PRODUCTION:
        return jsonify({
            'success': False,
            'error': 'Endpoints de debug n�o dispon�veis em produ��o',
            'message': 'Use migrations adequadas ou console admin'
        }), 403
    return None

@app.route('/api/debug/fix-kits-table', methods=['POST'])
@csrf_instance.exempt
def fix_kits_table():
    """
    Migration: Adiciona colunas 'descricao' e 'empresa_id' na tabela kits
    Bug descoberto na Fase 3 - c�digo usa campos que n�o existem
    
    ATENCAO: DISPON�VEL APENAS EM DESENVOLVIMENTO
    """
    # Bloquear em produ��o
    check = _check_debug_endpoint_allowed()
    if check:
        return check
    
    try:
        conn = database.get_connection()
        cursor = conn.cursor()
        
        results = {'steps': []}
        
        # 1. Adicionar coluna 'descricao'
        cursor.execute("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.columns 
                WHERE table_name='kits' AND column_name='descricao'
            ) as existe
        """)
        result = cursor.fetchone()
        descricao_existe = result[0] if isinstance(result, tuple) else result['existe']
        
        if not descricao_existe:
            cursor.execute("ALTER TABLE kits ADD COLUMN descricao TEXT")
            results['steps'].append('? Coluna descricao adicionada')
        else:
            results['steps'].append('?? Coluna descricao j� existe')
        
        # 2. Adicionar coluna 'empresa_id'
        cursor.execute("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.columns 
                WHERE table_name='kits' AND column_name='empresa_id'
            ) as existe
        """)
        result = cursor.fetchone()
        empresa_id_existe = result[0] if isinstance(result, tuple) else result['existe']
        
        if not empresa_id_existe:
            cursor.execute("ALTER TABLE kits ADD COLUMN empresa_id INTEGER DEFAULT 1")
            results['steps'].append('? Coluna empresa_id adicionada')
        else:
            results['steps'].append('?? Coluna empresa_id j� existe')
        
        # 3. Migrar dados de observacoes para descricao
        cursor.execute("""
            SELECT COUNT(*) FROM kits 
            WHERE observacoes IS NOT NULL 
            AND (descricao IS NULL OR descricao = '')
        """)
        result = cursor.fetchone()
        rows_to_migrate = result[0] if isinstance(result, tuple) else result['count']
        
        if rows_to_migrate > 0:
            cursor.execute("""
                UPDATE kits 
                SET descricao = observacoes 
                WHERE observacoes IS NOT NULL 
                AND (descricao IS NULL OR descricao = '')
            """)
            results['steps'].append(f'? {rows_to_migrate} registros migrados de observacoes ? descricao')
        else:
            results['steps'].append('?? Nenhum dado para migrar')
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Migration executada com sucesso',
            'results': results
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/debug/fix-p1-issues', methods=['POST'])
@csrf_instance.exempt
def fix_p1_issues():
    """
    Migration P1: Corrige bugs priorit�rios
    
    Funcionalidades:
    1. Adiciona empresa_id em todas as tabelas (multi-tenancy)
    2. Cria indexes para empresa_id
    3. Reporta campos que precisam de convers�o manual (VARCHAR ? FK)
    
    ATENCAO: DISPON�VEL APENAS EM DESENVOLVIMENTO
    
    Returns:
        JSON com resultados detalhados da migration
    """
    # Bloquear em produ��o
    check = _check_debug_endpoint_allowed()
    if check:
        return check
    
    try:
        conn = database.get_connection()
        cursor = conn.cursor()
        
        results = {
            'multi_tenancy': [],
            'indexes': [],
            'warnings': []
        }
        
        # Lista de tabelas que precisam de empresa_id
        tables_to_fix = [
            'lancamentos',
            'categorias', 
            'subcategorias',
            'clientes',
            'fornecedores',
            'contratos',
            'sessoes',
            'produtos',
            'contas_bancarias',
            'usuarios',
            'equipamentos',
            'projetos'
        ]
        
        # 1. Adicionar empresa_id em todas as tabelas
        for table_name in tables_to_fix:
            try:
                # Verifica se coluna j� existe
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT 1 FROM information_schema.columns 
                        WHERE table_name = %s AND column_name = 'empresa_id'
                    ) as existe
                """, (table_name,))
                
                result = cursor.fetchone()
                # Tenta diferentes formas de acessar o resultado
                if isinstance(result, dict):
                    empresa_id_existe = result['existe']
                elif isinstance(result, tuple):
                    empresa_id_existe = result[0]
                else:
                    empresa_id_existe = bool(result)
                
                if not empresa_id_existe:
                    # Adiciona coluna empresa_id
                    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN empresa_id INTEGER NOT NULL DEFAULT 1")
                    results['multi_tenancy'].append(f'? {table_name}: empresa_id adicionado')
                else:
                    results['multi_tenancy'].append(f'?? {table_name}: empresa_id j� existe')
                
                # Cria index para performance
                index_name = f'idx_{table_name}_empresa'
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT 1 FROM pg_indexes 
                        WHERE indexname = %s
                    ) as existe
                """, (index_name,))
                
                result = cursor.fetchone()
                if isinstance(result, dict):
                    index_existe = result['existe']
                elif isinstance(result, tuple):
                    index_existe = result[0]
                else:
                    index_existe = bool(result)
                
                if not index_existe:
                    cursor.execute(f"CREATE INDEX {index_name} ON {table_name}(empresa_id)")
                    results['indexes'].append(f'? Index {index_name} criado')
                else:
                    results['indexes'].append(f'?? Index {index_name} j� existe')
                    
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                results['warnings'].append(f'?? {table_name}: {type(e).__name__} - {str(e)}')
        
        # 2. Avisos sobre convers�es VARCHAR ? FK que precisam ser manuais
        fk_conversions_needed = [
            {
                'table': 'lancamentos',
                'column': 'categoria',
                'target': 'categorias',
                'reason': 'Campo VARCHAR precisa ser convertido para INTEGER FK'
            },
            {
                'table': 'lancamentos', 
                'column': 'subcategoria',
                'target': 'subcategorias',
                'reason': 'Campo VARCHAR precisa ser convertido para INTEGER FK'
            },
            {
                'table': 'lancamentos',
                'column': 'conta_bancaria', 
                'target': 'contas_bancarias',
                'reason': 'Campo VARCHAR precisa ser convertido para INTEGER FK'
            }
        ]
        
        results['warnings'].append('?? CONVERS�ES MANUAIS NECESS�RIAS:')
        for fk in fk_conversions_needed:
            results['warnings'].append(
                f"   � {fk['table']}.{fk['column']} ? {fk['target']}.id: {fk['reason']}"
            )
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Migration P1 executada com sucesso',
            'results': results,
            'summary': {
                'tables_updated': len([x for x in results['multi_tenancy'] if '?' in x]),
                'tables_skipped': len([x for x in results['multi_tenancy'] if '??' in x]),
                'indexes_created': len([x for x in results['indexes'] if '?' in x]),
                'warnings': len(results['warnings'])
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/debug/extrair-schema', methods=['GET'])
def extrair_schema_debug():
    """
    Endpoint tempor�rio para extrair schema do banco de dados
    Usado na Fase 3 da otimiza��o para documentar o banco
    """
    try:
        import json
        from datetime import datetime
        
        conn = database.get_connection()
        cursor = conn.cursor()
        
        schema_info = {
            'data_extracao': datetime.now().isoformat(),
            'tabelas': []
        }
        
        # 1. Obter lista de todas as tabelas
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)
        
        tabelas = cursor.fetchall()
        
        for tabela_row in tabelas:
            tabela_nome = tabela_row[0] if isinstance(tabela_row, tuple) else tabela_row['table_name']
            
            tabela_info = {
                'nome': tabela_nome,
                'colunas': [],
                'constraints': [],
                'foreign_keys': [],
                'indexes': []
            }
            
            # 2. Obter colunas da tabela
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    character_maximum_length,
                    is_nullable,
                    column_default
                FROM information_schema.columns
                WHERE table_name = %s
                ORDER BY ordinal_position
            """, (tabela_nome,))
            
            colunas = cursor.fetchall()
            
            for coluna in colunas:
                if isinstance(coluna, dict):
                    col_info = {
                        'nome': coluna['column_name'],
                        'tipo': coluna['data_type'],
                        'tamanho': coluna.get('character_maximum_length'),
                        'nullable': coluna['is_nullable'] == 'YES',
                        'default': coluna.get('column_default')
                    }
                else:
                    col_info = {
                        'nome': coluna[0],
                        'tipo': coluna[1],
                        'tamanho': coluna[2],
                        'nullable': coluna[3] == 'YES',
                        'default': coluna[4]
                    }
                
                tabela_info['colunas'].append(col_info)
            
            # 3. Obter constraints
            cursor.execute("""
                SELECT 
                    tc.constraint_name,
                    tc.constraint_type,
                    kcu.column_name
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage kcu 
                    ON tc.constraint_name = kcu.constraint_name
                WHERE tc.table_name = %s
                AND tc.constraint_type IN ('PRIMARY KEY', 'UNIQUE', 'CHECK')
            """, (tabela_nome,))
            
            constraints = cursor.fetchall()
            
            for constraint in constraints:
                if isinstance(constraint, dict):
                    const_info = {
                        'nome': constraint['constraint_name'],
                        'tipo': constraint['constraint_type'],
                        'coluna': constraint['column_name']
                    }
                else:
                    const_info = {
                        'nome': constraint[0],
                        'tipo': constraint[1],
                        'coluna': constraint[2]
                    }
                
                tabela_info['constraints'].append(const_info)
            
            # 4. Obter Foreign Keys
            cursor.execute("""
                SELECT
                    kcu.column_name,
                    ccu.table_name AS foreign_table_name,
                    ccu.column_name AS foreign_column_name
                FROM information_schema.table_constraints AS tc
                JOIN information_schema.key_column_usage AS kcu
                    ON tc.constraint_name = kcu.constraint_name
                JOIN information_schema.constraint_column_usage AS ccu
                    ON ccu.constraint_name = tc.constraint_name
                WHERE tc.constraint_type = 'FOREIGN KEY'
                AND tc.table_name = %s
            """, (tabela_nome,))
            
            fks = cursor.fetchall()
            
            for fk in fks:
                if isinstance(fk, dict):
                    fk_info = {
                        'coluna': fk['column_name'],
                        'referencia_tabela': fk['foreign_table_name'],
                        'referencia_coluna': fk['foreign_column_name']
                    }
                else:
                    fk_info = {
                        'coluna': fk[0],
                        'referencia_tabela': fk[1],
                        'referencia_coluna': fk[2]
                    }
                
                tabela_info['foreign_keys'].append(fk_info)
            
            schema_info['tabelas'].append(tabela_info)
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'schema': schema_info
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/tags', methods=['GET', 'POST'])
@require_permission('operacional_view')
def tags():
    """Gerenciar tags"""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
    
    if request.method == 'GET':
        try:
            tags = database.listar_tags(empresa_id)
            return jsonify({'success': True, 'data': tags})
        except Exception as e:
            logger.error(f"Erro ao listar tags: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            tag_id = database.adicionar_tag(empresa_id, data)
            return jsonify({'success': True, 'message': 'Tag criada com sucesso', 'id': tag_id}), 201
        except Exception as e:
            logger.error(f"Erro ao criar tag: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/tags/<int:tag_id>', methods=['GET', 'PUT', 'DELETE'])
def tag_detalhes(tag_id):
    """Buscar, atualizar ou excluir tag"""
    
    # Validar empresa_id
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
    
    # Validar permiss�es baseado no m�todo
    if request.method == 'GET':
        # GET requer apenas visualiza��o
        from auth_middleware import require_permission
        decorator = require_permission('operacional_view')
        # Aplicar valida��o manualmente
        usuario = session.get('usuario')
        if not usuario:
            return jsonify({'success': False, 'error': 'Usu�rio n�o autenticado'}), 401
        
        permissoes = usuario.get('permissoes', [])
        if 'operacional_view' not in permissoes and 'admin' not in permissoes:
            return jsonify({'success': False, 'error': 'Sem permiss�o para visualizar tags'}), 403
            
        try:
            tag = database.obter_tag(empresa_id, tag_id)
            if tag:
                return jsonify(tag)
            return jsonify({'success': False, 'error': 'Tag n�o encontrada'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
            
    elif request.method in ['PUT', 'DELETE']:
        # PUT e DELETE requerem permiss�o de edi��o
        usuario = session.get('usuario')
        if not usuario:
            return jsonify({'success': False, 'error': 'Usu�rio n�o autenticado'}), 401
        
        permissoes = usuario.get('permissoes', [])
        if 'operacional_edit' not in permissoes and 'admin' not in permissoes:
            return jsonify({'success': False, 'error': 'Sem permiss�o para editar/excluir tags'}), 403
        
        if request.method == 'PUT':
            try:
                data = request.json
                success = database.atualizar_tag(empresa_id, tag_id, data)
                if success:
                    return jsonify({'success': True, 'message': 'Tag atualizada com sucesso'})
                return jsonify({'success': False, 'error': 'Tag n�o encontrada'}), 404
            except Exception as e:
                return jsonify({'success': False, 'error': str(e)}), 500
        else:  # DELETE
            try:
                print(f"??? [DEBUG TAG] DELETE tag_id={tag_id}, empresa_id={empresa_id}")
                success = database.deletar_tag(empresa_id, tag_id)
                print(f"??? [DEBUG TAG] DELETE result: {success}")
                if success:
                    return jsonify({'success': True, 'message': 'Tag exclu�da com sucesso'})
                return jsonify({'success': False, 'error': 'Tag n�o encontrada'}), 404
            except Exception as e:
                print(f"? [DEBUG TAG] DELETE exception: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/funcoes-responsaveis', methods=['GET', 'POST'])
@require_permission('operacional_view')
def funcoes_responsaveis():
    """Gerenciar fun��es de respons�veis"""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
    
    if request.method == 'GET':
        try:
            apenas_ativas = request.args.get('apenas_ativas', 'true').lower() == 'true'
            funcoes = database.listar_funcoes_responsaveis(empresa_id, apenas_ativas)
            return jsonify({'success': True, 'data': funcoes})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            funcao_id = database.adicionar_funcao_responsavel(empresa_id, data)
            return jsonify({'success': True, 'message': 'Fun��o criada com sucesso', 'id': funcao_id}), 201
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/funcoes-responsaveis/<int:funcao_id>', methods=['PUT', 'DELETE'])
@require_permission('operacional_edit')
def funcao_responsavel_detalhes(funcao_id):
    """Atualizar ou excluir fun��o de respons�vel"""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
    
    if request.method == 'PUT':
        try:
            data = request.json
            success = database.atualizar_funcao_responsavel(empresa_id, funcao_id, data)
            if success:
                return jsonify({'success': True, 'message': 'Fun��o atualizada com sucesso'})
            return jsonify({'success': False, 'error': 'Fun��o n�o encontrada'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # DELETE
        try:
            success = database.deletar_funcao_responsavel(empresa_id, funcao_id)
            if success:
                return jsonify({'success': True, 'message': 'Fun��o exclu�da com sucesso'})
            return jsonify({'success': False, 'error': 'Fun��o n�o encontrada'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/custos-operacionais', methods=['GET', 'POST'])
@require_permission('operacional_view')
def custos_operacionais():
    """Gerenciar custos operacionais"""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
    
    if request.method == 'GET':
        try:
            apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
            categoria = request.args.get('categoria')
            custos = database.listar_custos_operacionais(empresa_id, apenas_ativos, categoria)
            return jsonify({'success': True, 'data': custos})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            custo_id = database.adicionar_custo_operacional(empresa_id, data)
            return jsonify({'success': True, 'message': 'Custo criado com sucesso', 'id': custo_id}), 201
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/custos-operacionais/<int:custo_id>', methods=['PUT', 'DELETE'])
@require_permission('operacional_edit')
def custo_operacional_detalhes(custo_id):
    """Atualizar ou excluir custo operacional"""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 400
    
    if request.method == 'PUT':
        try:
            data = request.json
            success = database.atualizar_custo_operacional(empresa_id, custo_id, data)
            if success:
                return jsonify({'success': True, 'message': 'Custo atualizado com sucesso'})
            return jsonify({'success': False, 'error': 'Custo n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # DELETE
        try:
            success = database.deletar_custo_operacional(empresa_id, custo_id)
            if success:
                return jsonify({'success': True, 'message': 'Custo exclu�do com sucesso'})
            return jsonify({'success': False, 'error': 'Custo n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/templates-equipe', methods=['GET', 'POST'])
@require_permission('operacional_view')
def templates_equipe():
    """Gerenciar templates de equipe"""
    if request.method == 'GET':
        try:
            templates = db.listar_templates_equipe()
            return jsonify(templates)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:  # POST
        try:
            data = request.json
            template_id = db.adicionar_template_equipe(data)
            return jsonify({'message': 'Template criado com sucesso', 'id': template_id}), 201
        except Exception as e:
            return jsonify({'error': str(e)}), 500


@app.route('/api/templates-equipe/<int:template_id>', methods=['PUT', 'DELETE'])
@require_permission('operacional_edit')
def template_equipe_detalhes(template_id):
    """Atualizar ou excluir template"""
    if request.method == 'PUT':
        try:
            data = request.json
            success = db.atualizar_template_equipe(template_id, data)
            if success:
                return jsonify({'message': 'Template atualizado com sucesso'})
            return jsonify({'error': 'Template n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:  # DELETE
        try:
            success = db.deletar_template_equipe(template_id)
            if success:
                return jsonify({'message': 'Template exclu�do com sucesso'})
            return jsonify({'error': 'Template n�o encontrado'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500


# ============================================================================
# EXPORTA��O DE DADOS POR CLIENTE (ADMIN)
# ============================================================================

@app.route('/api/admin/debug/schema', methods=['GET'])
@require_admin
def debug_database_schema():
    """
    Mostra todas as tabelas e colunas do banco de dados (DEBUG)
    """
    try:
        db = DatabaseManager()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        schema_info = {}
        
        # Buscar todas as tabelas
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)
        
        tabelas = cursor.fetchall()
        
        # Para cada tabela, buscar colunas
        for tabela_row in tabelas:
            tabela = tabela_row['table_name']
            
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default
                FROM information_schema.columns
                WHERE table_schema = 'public' 
                AND table_name = %s
                ORDER BY ordinal_position
            """, (tabela,))
            
            colunas = cursor.fetchall()
            
            schema_info[tabela] = [
                {
                    'nome': col['column_name'],
                    'tipo': col['data_type'],
                    'nullable': col['is_nullable'] == 'YES',
                    'default': col['column_default']
                }
                for col in colunas
            ]
        
        cursor.close()
        database.return_to_pool(conn)
        
        # Imprimir no console tamb�m
        print("\n" + "=" * 80)
        print("?? SCHEMA DO BANCO DE DADOS - TODAS AS TABELAS E COLUNAS")
        print("=" * 80)
        
        for tabela, colunas in sorted(schema_info.items()):
            print(f"\n?? Tabela: {tabela.upper()}")
            print("-" * 80)
            for col in colunas:
                nullable = "NULL" if col['nullable'] else "NOT NULL"
                print(f"  � {col['nome']:<30} {col['tipo']:<20} {nullable}")
        
        print("\n" + "=" * 80)
        
        return jsonify({
            'success': True,
            'schema': schema_info,
            'total_tabelas': len(schema_info)
        })
        
    except Exception as e:
        print(f"? Erro ao obter schema: {e}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/exportar-cliente/<int:cliente_id>', methods=['GET'])
@require_admin
def exportar_dados_cliente_admin(cliente_id):
    """
    Exporta todos os dados de um cliente espec�fico (apenas admin)
    
    Retorna um arquivo JSON com todos os dados do cliente:
    - Clientes
    - Fornecedores
    - Categorias
    - Contas Banc�rias
    - Lan�amentos
    """
    try:
        # Verificar se o usu�rio/cliente existe
        usuario = request.usuario
        usuario_cliente = auth_db.obter_usuario(cliente_id)
        
        if not usuario_cliente:
            return jsonify({
                'success': False,
                'error': f'Usu�rio com ID {cliente_id} n�o encontrado'
            }), 404
        
        # Exportar dados
        print(f"\n?? Iniciando exporta��o dos dados do cliente {cliente_id}")
        print(f"   ?? Usu�rio: {usuario_cliente.get('nome_completo', 'N/A')} ({usuario_cliente.get('email', 'N/A')})")
        export_data = database.exportar_dados_cliente(cliente_id)
        
        # Registrar log de auditoria
        auth_db.registrar_log_acesso(
            usuario_id=usuario['id'],
            acao='exportar_dados_cliente',
            descricao=f'Exportou dados do cliente_id {cliente_id}',
            ip_address=request.remote_addr,
            sucesso=True
        )
        
        print(f"? Exporta��o conclu�da para cliente {cliente_id}")
        
        # Retornar como arquivo TXT para download
        from flask import make_response
        response = make_response(export_data['texto'])
        response.headers['Content-Type'] = 'text/plain; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=export_cliente_{cliente_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        
        return response
        
    except Exception as e:
        print(f"? Erro ao exportar dados do cliente {cliente_id}: {e}")
        import traceback
        traceback.print_exc()
        
        # Registrar log de erro
        try:
            auth_db.registrar_log_acesso(
                usuario_id=request.usuario['id'],
                acao='exportar_dados_cliente',
                descricao=f'ERRO ao exportar cliente_id {cliente_id}: {str(e)}',
                ip_address=request.remote_addr,
                sucesso=False
            )
        except:
            pass
        
        return jsonify({
            'success': False,
            'error': f'Erro ao exportar dados: {str(e)}'
        }), 500


@app.route('/api/admin/listar-proprietarios', methods=['GET'])
@require_admin
def listar_proprietarios_disponiveis():
    """
    Lista todos os proprietario_id �nicos no sistema
    Para o admin selecionar qual cliente exportar
    """
    try:
        # Buscar todos os usu�rios do tipo 'cliente'
        usuarios = auth_db.listar_usuarios()
        
        proprietarios_info = []
        proprietarios_ids = set()
        
        for usuario in usuarios:
            # Adicionar todos os usu�rios com tipo 'cliente' ou que tenham cliente_id
            if usuario.get('tipo') == 'cliente' or usuario.get('cliente_id'):
                proprietario_id = usuario.get('cliente_id') or usuario.get('id')
                
                # Evitar duplicatas
                if proprietario_id in proprietarios_ids:
                    continue
                proprietarios_ids.add(proprietario_id)
                
                proprietarios_info.append({
                    'proprietario_id': proprietario_id,
                    'nome': usuario.get('nome_completo') or usuario.get('nome') or f'Usu�rio {proprietario_id}',
                    'email': usuario.get('email') or 'Sem email',
                    'tipo': usuario.get('tipo', 'cliente'),
                    'usuario_id': usuario.get('id')
                })
        
        # Tamb�m buscar proprietario_id �nicos das tabelas (para dados �rf�os)
        db_temp = DatabaseManager()
        conn = db_temp.get_connection()
        cursor = conn.cursor()
        
        # Buscar proprietario_id que n�o correspondem a usu�rios
        cursor.execute("""
            SELECT DISTINCT proprietario_id
            FROM (
                SELECT proprietario_id FROM clientes WHERE proprietario_id IS NOT NULL
                UNION
                SELECT proprietario_id FROM fornecedores WHERE proprietario_id IS NOT NULL
                UNION
                SELECT proprietario_id FROM lancamentos WHERE proprietario_id IS NOT NULL
                UNION
                SELECT proprietario_id FROM contas_bancarias WHERE proprietario_id IS NOT NULL
                UNION
                SELECT proprietario_id FROM categorias WHERE proprietario_id IS NOT NULL
            ) AS todos_proprietarios
            ORDER BY proprietario_id
        """)
        
        proprietarios_db = cursor.fetchall()
        
        for row in proprietarios_db:
            prop_id = row['proprietario_id']
            if prop_id not in proprietarios_ids:
                proprietarios_ids.add(prop_id)
                proprietarios_info.append({
                    'proprietario_id': prop_id,
                    'nome': f'Cliente ID {prop_id} (sem usu�rio)',
                    'email': 'N�o dispon�vel',
                    'tipo': 'orfao'
                })
        
        cursor.close()
        db_temp.return_to_pool(conn)
        
        # Ordenar por nome
        proprietarios_info.sort(key=lambda x: x['nome'])
        
        print(f"?? Encontrados {len(proprietarios_info)} propriet�rios �nicos")
        
        return jsonify({
            'success': True,
            'proprietarios': proprietarios_info,
            'total': len(proprietarios_info)
        })
        
    except Exception as e:
        print(f"? Erro ao listar propriet�rios: {e}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': f'Erro ao listar propriet�rios: {str(e)}'
        }), 500


@app.route('/api/admin/limpar-duplicatas-categorias', methods=['POST'])
@require_admin
def limpar_duplicatas_categorias():
    """
    Remove categorias duplicadas mantendo apenas a mais antiga (menor ID)
    Duplicata = mesmo nome + mesma empresa_id
    """
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        print('\n' + '='*80)
        print('?? ADMIN: Limpando categorias duplicadas')
        print('='*80)
        
        # Buscar todas as categorias
        cursor.execute("""
            SELECT id, nome, tipo, empresa_id 
            FROM categorias 
            ORDER BY empresa_id, nome, id
        """)
        categorias = cursor.fetchall()
        
        print(f'?? Total de categorias no banco: {len(categorias)}')
        
        # Agrupar por (nome normalizado, empresa_id)
        grupos = {}
        for cat in categorias:
            chave = (cat['nome'].strip().upper(), cat['empresa_id'])
            if chave not in grupos:
                grupos[chave] = []
            grupos[chave].append(cat)
        
        # Filtrar apenas grupos com duplicatas
        duplicatas = {k: v for k, v in grupos.items() if len(v) > 1}
        
        if not duplicatas:
            print('? Nenhuma duplicata encontrada!')
            cursor.close()
            db.return_to_pool(conn)
            return jsonify({
                'success': True,
                'message': 'Nenhuma duplicata encontrada',
                'removidas': 0
            })
        
        print(f'??  Encontradas {len(duplicatas)} categorias com duplicatas')
        
        ids_removidos = []
        detalhes = []
        
        for (nome, empresa), lista in duplicatas.items():
            # Ordenar por ID (manter o menor = mais antigo)
            lista_ordenada = sorted(lista, key=lambda x: x['id'])
            manter = lista_ordenada[0]
            excluir = lista_ordenada[1:]
            
            print(f'\n?? {nome} (Empresa: {empresa})')
            print(f'   ? MANTER: ID={manter["id"]}')
            
            for cat in excluir:
                print(f'   ? EXCLUIR: ID={cat["id"]}')
                cursor.execute('DELETE FROM categorias WHERE id = %s', (cat['id'],))
                ids_removidos.append(cat['id'])
            
            detalhes.append({
                'nome': nome,
                'empresa_id': empresa,
                'mantido': manter['id'],
                'removidos': [c['id'] for c in excluir]
            })
        
        conn.commit()
        cursor.close()
        db.return_to_pool(conn)
        
        print(f'\n? Removidas {len(ids_removidos)} duplicatas!')
        print('='*80 + '\n')
        
        return jsonify({
            'success': True,
            'message': f'{len(ids_removidos)} categoria(s) duplicada(s) removida(s)',
            'removidas': len(ids_removidos),
            'detalhes': detalhes
        })
        
    except Exception as e:
        print(f'? Erro ao limpar duplicatas: {str(e)}')
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/limpar-duplicatas-extrato', methods=['POST'])
@require_admin
def limpar_duplicatas_extrato():
    """
    Remove lan�amentos duplicados com [EXTRATO] mantendo apenas o mais recente
    Duplicata = mesma descri��o + mesmo valor + mesma data + mesmo tipo + mesma empresa
    """
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        print('\n' + '='*80)
        print('?? ADMIN: Limpando lan�amentos duplicados [EXTRATO]')
        print('='*80)
        
        # 1. AN�LISE INICIAL
        cursor.execute("SELECT COUNT(*) as total FROM lancamentos")
        total = cursor.fetchone()['total']
        print(f'?? Total de lan�amentos: {total:,}')
        
        cursor.execute("SELECT COUNT(*) as total FROM lancamentos WHERE descricao LIKE '[EXTRATO]%'")
        total_extrato = cursor.fetchone()['total']
        print(f'?? Lan�amentos com [EXTRATO]: {total_extrato:,}')
        
        # Contar registros duplicados (exceto o que ser� mantido)
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM lancamentos l
            WHERE l.descricao LIKE '[EXTRATO]%'
              AND l.id NOT IN (
                SELECT MAX(id) 
                FROM lancamentos
                WHERE descricao LIKE '[EXTRATO]%'
                GROUP BY descricao, valor, data_vencimento, tipo, empresa_id
              )
        """)
        registros_para_deletar = cursor.fetchone()['total']
        print(f'???  Registros duplicados a serem removidos: {registros_para_deletar:,}')
        
        if registros_para_deletar == 0:
            print('? Nenhuma duplicata encontrada!')
            cursor.close()
            db.return_to_pool(conn)
            return jsonify({
                'success': True,
                'message': 'Nenhuma duplicata encontrada',
                'removidas': 0,
                'total_antes': total,
                'total_depois': total
            })
        
        # 2. CRIAR BACKUP
        print('\n?? Criando backup...')
        cursor.execute("DROP TABLE IF EXISTS lancamentos_backup_duplicatas")
        cursor.execute("""
            CREATE TABLE lancamentos_backup_duplicatas AS
            SELECT l.*
            FROM lancamentos l
            WHERE l.descricao LIKE '[EXTRATO]%'
              AND l.id NOT IN (
                SELECT MAX(id) 
                FROM lancamentos
                WHERE descricao LIKE '[EXTRATO]%'
                GROUP BY descricao, valor, data_vencimento, tipo, empresa_id
              )
        """)
        conn.commit()
        
        cursor.execute("SELECT COUNT(*) as total FROM lancamentos_backup_duplicatas")
        backup_count = cursor.fetchone()['total']
        print(f'? Backup criado: {backup_count:,} registros salvos')
        
        # 3. DELETAR DUPLICATAS
        print('\n???  Removendo duplicatas...')
        cursor.execute("""
            DELETE FROM lancamentos
            WHERE descricao LIKE '[EXTRATO]%'
              AND id NOT IN (
                SELECT MAX(id) 
                FROM lancamentos
                WHERE descricao LIKE '[EXTRATO]%'
                GROUP BY descricao, valor, data_vencimento, tipo, empresa_id
              )
        """)
        deletados = cursor.rowcount
        conn.commit()
        
        print(f'? Removidos {deletados:,} lan�amentos duplicados')
        
        # 4. AN�LISE FINAL
        cursor.execute("SELECT COUNT(*) as total FROM lancamentos")
        total_apos = cursor.fetchone()['total']
        print(f'?? Total de lan�amentos ap�s limpeza: {total_apos:,}')
        
        # Verificar saldo da conta
        cursor.execute("""
            SELECT banco, agencia, conta, saldo_inicial, saldo_atual
            FROM contas_bancarias
            WHERE id = 6
        """)
        conta = cursor.fetchone()
        saldo_info = None
        if conta:
            saldo_info = {
                'banco': conta['banco'],
                'agencia': conta['agencia'],
                'conta': conta['conta'],
                'saldo_inicial': float(conta['saldo_inicial']),
                'saldo_atual': float(conta['saldo_atual'])
            }
            print(f'\n?? Saldo atualizado: R$ {saldo_info["saldo_atual"]:,.2f}')
        
        cursor.close()
        db.return_to_pool(conn)
        
        print('? Limpeza conclu�da com sucesso!')
        print('='*80 + '\n')
        
        return jsonify({
            'success': True,
            'message': f'{deletados:,} lan�amento(s) duplicado(s) removido(s)',
            'removidas': deletados,
            'total_antes': total,
            'total_depois': total_apos,
            'backup_table': 'lancamentos_backup_duplicatas',
            'saldo': saldo_info
        })
        
    except Exception as e:
        print(f'? Erro ao limpar duplicatas: {str(e)}')
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== ROTAS DE PREFER�NCIAS DO USU�RIO ====================
@app.route('/api/preferencias/menu-order', methods=['GET'])
@require_auth
def obter_ordem_menu():
    """Obt�m a ordem personalizada do menu do usu�rio"""
    try:
        # Usar session ao inv�s de request.usuario
        usuario_id = session.get('usuario_id')
        
        if not usuario_id:
            return jsonify({
                'success': False,
                'error': 'Usu�rio n�o autenticado'
            }), 401
        
        print(f"?? Obtendo ordem do menu para usuario_id={usuario_id}")
        
        # Ordem padr�o
        ordem_padrao = '["dashboard","financeiro","relatorios","cadastros","operacional"]'
        
        # Obter prefer�ncia do banco
        try:
            ordem = database.obter_preferencia_usuario(
                usuario_id, 
                'menu_order', 
                ordem_padrao
            )
        except Exception as db_error:
            print(f"?? Erro ao buscar prefer�ncia, usando padr�o: {db_error}")
            ordem = ordem_padrao
        
        # Parsear JSON
        import json
        menu_order = json.loads(ordem) if ordem else json.loads(ordem_padrao)
        
        return jsonify({
            'success': True,
            'menu_order': menu_order
        })
        
    except Exception as e:
        print(f"? Erro ao obter ordem do menu: {e}")
        import traceback
        traceback.print_exc()
        # Retornar ordem padr�o em caso de erro
        return jsonify({
            'success': True,
            'menu_order': ["dashboard","financeiro","relatorios","cadastros","operacional"]
        })


@app.route('/api/preferencias/menu-order', methods=['POST'])
@require_auth
def salvar_ordem_menu():
    """Salva a ordem personalizada do menu do usu�rio"""
    try:
        # Usar session ao inv�s de request.usuario
        usuario_id = session.get('usuario_id')
        
        if not usuario_id:
            return jsonify({
                'success': False,
                'error': 'Usu�rio n�o autenticado'
            }), 401
        
        print(f"?? Salvando ordem do menu para usuario_id={usuario_id}")
        
        data = request.json
        if not data:
            print("? Dados n�o fornecidos")
            return jsonify({
                'success': False,
                'error': 'Dados n�o fornecidos'
            }), 400
        
        menu_order = data.get('menu_order', [])
        print(f"?? Ordem recebida: {menu_order}")
        
        # Validar formato
        if not isinstance(menu_order, list):
            print("? menu_order n�o � lista")
            return jsonify({
                'success': False,
                'error': 'menu_order deve ser uma lista'
            }), 400
        
        # Validar itens permitidos
        itens_validos = ['dashboard', 'financeiro', 'relatorios', 'cadastros', 'operacional']
        for item in menu_order:
            if item not in itens_validos:
                print(f"? Item inv�lido: {item}")
                return jsonify({
                    'success': False,
                    'error': f'Item inv�lido: {item}'
                }), 400
        
        # Converter para JSON string
        import json
        menu_order_json = json.dumps(menu_order)
        
        # Salvar no banco
        print(f"?? Chamando salvar_preferencia_usuario...")
        sucesso = database.salvar_preferencia_usuario(
            usuario_id,
            'menu_order',
            menu_order_json
        )
        
        print(f"{'?' if sucesso else '?'} Resultado do save: {sucesso}")
        
        if sucesso:
            # Registrar log
            try:
                auth_db.registrar_log_acesso(
                    usuario_id=usuario_id,
                    acao='update_menu_order',
                    descricao=f'Ordem do menu atualizada: {menu_order}',
                    ip_address=request.remote_addr,
                    sucesso=True
                )
            except Exception as log_error:
                print(f"?? Erro ao registrar log (n�o cr�tico): {log_error}")
            
            return jsonify({
                'success': True,
                'message': 'Ordem do menu salva com sucesso'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Erro ao salvar no banco de dados'
            }), 500
        
    except Exception as e:
        print(f"? Erro ao salvar ordem do menu: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# ROTAS DE GEST�O DE EMPRESAS (MULTI-TENANT)
# ============================================================================
logger.info("="*80)
logger.info("INICIO DAS ROTAS DE EMPRESAS")
logger.info("="*80)

@app.route('/api/empresas', methods=['GET'])
@require_auth
def listar_empresas_api():
    """Lista empresas - admin v� todas, outros usu�rios v�em apenas as suas"""
    logger.info("\n" + "="*80)
    logger.info("[listar_empresas_api] FUNCAO INICIADA")
    logger.info(f"   Path: {request.path}")
    logger.info(f"   Metodo: {request.method}")
    logger.info("="*80)
    
    try:
        logger.info("GET /api/empresas - Iniciando processamento...")
        
        # Usa get_usuario_logado() para pegar usuario do token
        usuario = get_usuario_logado()
        logger.info(f"   Usuario autenticado: {usuario.get('username')} (tipo: {usuario.get('tipo')})")
        
        conn = database.get_connection()
        cursor = conn.cursor()
        
        # Admin pode listar todas as empresas
        if usuario['tipo'] == 'admin':
            logger.info("   ?? Admin: listando TODAS as empresas")
            
            query = "SELECT id, razao_social, cnpj, plano, ativo FROM empresas"
            params = []
            
            # Filtros opcionais
            filtros = []
            if request.args.get('ativo'):
                filtros.append("ativo = %s")
                params.append(request.args.get('ativo') == 'true')
            
            if request.args.get('plano'):
                filtros.append("plano = %s")
                params.append(request.args.get('plano'))
            
            if filtros:
                query += " WHERE " + " AND ".join(filtros)
            
            query += " ORDER BY razao_social"
            
            logger.info(f"   ?? Query: {query}")
            logger.info(f"   ?? Params: {params}")
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            empresas = []
            for row in rows:
                empresas.append({
                    'id': row['id'],
                    'razao_social': row['razao_social'],
                    'cnpj': row['cnpj'],
                    'plano': row['plano'],
                    'ativo': row['ativo']
                })
            
            cursor.close()
            
            logger.info(f"   ? Retornando {len(empresas)} empresas")
            logger.info("="*80 + "\n")
            
            return jsonify(empresas)
        
        # Usu�rios n�o-admin veem apenas empresas �s quais t�m acesso
        else:
            logger.info("   ?? Usu�rio: listando apenas empresas vinculadas")
            usuario_id = usuario.get('id')
            
            cursor.execute("""
                SELECT DISTINCT
                    e.id,
                    e.razao_social,
                    e.cnpj,
                    e.plano,
                    e.ativo,
                    e.criado_em,
                    ue.is_empresa_padrao
                FROM empresas e
                INNER JOIN usuario_empresas ue ON ue.empresa_id = e.id
                WHERE ue.usuario_id = %s AND ue.acesso_ativo = TRUE
                ORDER BY e.razao_social
            """, (usuario_id,))
            
            rows = cursor.fetchall()
            empresas = []
            for row in rows:
                empresas.append({
                    'id': row['id'],
                    'razao_social': row['razao_social'],
                    'cnpj': row['cnpj'],
                    'plano': row['plano'],
                    'ativo': row['ativo'],
                    'criado_em': row['criado_em'].isoformat() if row['criado_em'] else None,
                    'is_empresa_padrao': row['is_empresa_padrao']
                })
            
            cursor.close()
            
            logger.info(f"   ? Retornando {len(empresas)} empresas vinculadas")
            logger.info("="*80 + "\n")
            
            return jsonify(empresas)
        
    except Exception as e:
        logger.info(f"? Erro ao listar empresas: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        logger.info("="*80 + "\n")
        return jsonify({'error': str(e)}), 500

@app.route('/api/empresas/<int:empresa_id>', methods=['GET'])
@require_auth
def obter_empresa_api(empresa_id):
    """Obt�m dados de uma empresa espec�fica"""
    logger.info("\n" + "="*80)
    logger.info(f"[obter_empresa_api] FUNCAO CHAMADA - ID: {empresa_id}")
    try:
        logger.info(f"[obter_empresa_api] Obtendo usuario logado...")
        usuario = get_usuario_logado()
        logger.info(f"[obter_empresa_api] Usuario: {usuario.get('username')} (tipo: {usuario.get('tipo')})")
        
        # Admin pode ver qualquer empresa, usu�rio comum s� se tiver v�nculo ativo
        if usuario['tipo'] != 'admin':
            logger.info(f"[obter_empresa_api] Usuario nao e admin - verificando acesso...")
            from auth_functions import verificar_acesso_empresa
            tem_acesso = verificar_acesso_empresa(usuario['id'], empresa_id, auth_db)
            logger.info(f"[obter_empresa_api] Usuario tem acesso? {tem_acesso}")
            if not tem_acesso:
                logger.info(f"[obter_empresa_api] Acesso negado - sem vinculo ativo com empresa")
                return jsonify({'error': 'Acesso negado'}), 403
        
        logger.info(f"[obter_empresa_api] Chamando database.obter_empresa({empresa_id})...")
        empresa = database.obter_empresa(empresa_id)
        logger.info(f"[obter_empresa_api] Resultado: {empresa is not None}")
        
        if not empresa:
            logger.info(f"[obter_empresa_api] Empresa nao encontrada")
            logger.info("="*80 + "\n")
            return jsonify({'error': 'Empresa n�o encontrada'}), 404
        
        logger.info(f"[obter_empresa_api] Empresa encontrada: {empresa.get('razao_social')}")

        # Auto-popular CNPJ da empresa a partir do certificado ativo, se nao cadastrado
        if not empresa.get('cnpj'):
            try:
                with get_db_connection(empresa_id=empresa_id) as conn_cert:
                    cur_cert = conn_cert.cursor()
                    cur_cert.execute(
                        "SELECT cnpj FROM certificados_digitais WHERE empresa_id = %s AND ativo = TRUE AND cnpj IS NOT NULL AND cnpj != '' LIMIT 1",
                        (empresa_id,)
                    )
                    row = cur_cert.fetchone()
                    if row and row[0]:
                        cnpj_cert = row[0].strip()
                        cur_cert.execute(
                            "UPDATE empresas SET cnpj = %s WHERE id = %s",
                            (cnpj_cert, empresa_id)
                        )
                        conn_cert.commit()
                        empresa['cnpj'] = cnpj_cert
                        logger.info(f"[obter_empresa_api] CNPJ auto-populado do certificado ativo: {cnpj_cert}")
            except Exception as e_cert:
                logger.warning(f"[obter_empresa_api] Erro ao auto-popular CNPJ do cert: {e_cert}")

        logger.info(f"[obter_empresa_api] Obtendo estatisticas...")
        
        # Adicionar estat�sticas
        try:
            empresa['stats'] = database.obter_estatisticas_empresa(empresa_id)
            logger.info(f"[obter_empresa_api] Estatisticas obtidas")
        except Exception as e:
            logger.info(f"[obter_empresa_api] Erro ao obter stats: {e}")
            empresa['stats'] = {}
        
        logger.info(f"[obter_empresa_api] Retornando sucesso")
        logger.info("="*80 + "\n")
        return jsonify({
            'success': True,
            'empresa': empresa
        })
        
    except Exception as e:
        logger.info(f"[obter_empresa_api] EXCECAO: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        logger.info("="*80 + "\n")
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas', methods=['POST'])
@require_auth
def criar_empresa_api():
    """Cria uma nova empresa (apenas super admin)"""
    logger.info("\n" + "="*80)
    logger.info("[criar_empresa_api] FUNCAO CHAMADA")
    try:
        usuario = get_usuario_logado()
        logger.info(f"   Usuario: {usuario.get('username')} (tipo: {usuario.get('tipo')})")
        
        if usuario['tipo'] != 'admin':
            logger.info("   Acesso negado - usuario nao e admin")
            return jsonify({'error': 'Acesso negado'}), 403
        
        dados = request.json
        logger.info(f"   Dados recebidos: {dados}")
        
        if not dados:
            logger.info("   Erro: dados nao fornecidos")
            return jsonify({'error': 'Dados n�o fornecidos'}), 400
        
        logger.info("   Chamando database.criar_empresa()...")
        resultado = database.criar_empresa(dados)
        logger.info(f"   Resultado: {resultado}")
        
        if resultado['success']:
            # Registrar log
            try:
                auth_db.registrar_log_acesso(
                    usuario_id=usuario['id'],
                    acao='criar_empresa',
                    descricao=f"Empresa criada: {dados.get('razao_social')}",
                    sucesso=True
                )
            except:
                pass
            
            logger.info("   Empresa criada com sucesso!")
            logger.info("="*80 + "\n")
            return jsonify(resultado), 201
        else:
            logger.info(f"   Erro: {resultado.get('error')}")
            logger.info("="*80 + "\n")
            return jsonify(resultado), 400
        
    except Exception as e:
        logger.info(f"EXCECAO ao criar empresa: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        logger.info("="*80 + "\n")
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas/<int:empresa_id>', methods=['PUT'])
@require_auth
def atualizar_empresa_api(empresa_id):
    """Atualiza dados de uma empresa"""
    print(f"\n?? [atualizar_empresa_api] FUN��O CHAMADA - ID: {empresa_id}")
    try:
        usuario = get_usuario_logado()
        dados = request.json
        
        if not dados:
            return jsonify({'error': 'Dados n�o fornecidos'}), 400
        
        # ?? POL�TICA DE ACESSO:
        # - Admin pode editar QUALQUER campo de QUALQUER empresa
        # - Usu�rio comum pode editar APENAS o campo 'estado' da PR�PRIA empresa
        
        if usuario['tipo'] != 'admin':
            # Verifica se tem acesso � empresa
            from auth_functions import verificar_acesso_empresa
            tem_acesso = verificar_acesso_empresa(usuario['id'], empresa_id, auth_db)
            
            if not tem_acesso:
                return jsonify({'error': 'Acesso negado'}), 403
            
            # Permite editar 'estado' (UF) e 'cnpj' para auto-populacao via certificado
            campos_permitidos = ['estado', 'cnpj']
            campos_enviados = set(dados.keys())
            campos_proibidos = campos_enviados - set(campos_permitidos)
            
            if campos_proibidos:
                return jsonify({
                    'error': f'Usu�rio comum s� pode editar os campos: {", ".join(campos_permitidos)}. ' +
                             f'Campos n�o permitidos: {", ".join(campos_proibidos)}'
                }), 403
            
            print(f"? [atualizar_empresa_api] Usu�rio comum editando campo permitido: {list(dados.keys())}")
        
        resultado = database.atualizar_empresa(empresa_id, dados)
        
        if resultado['success']:
            # Registrar log
            try:
                campos_alterados = ', '.join(dados.keys())
                auth_db.registrar_log_acesso(
                    usuario_id=usuario['id'],
                    acao='atualizar_empresa',
                    descricao=f"Empresa {empresa_id} - campos atualizados: {campos_alterados}",
                    sucesso=True
                )
            except:
                pass
            
            return jsonify(resultado)
        else:
            return jsonify(resultado), 400
        
    except Exception as e:
        print(f"? Erro ao atualizar empresa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas/<int:empresa_id>/suspender', methods=['POST'])
@require_auth
def suspender_empresa_api(empresa_id):
    """Suspende uma empresa"""
    print(f"\n?? [suspender_empresa_api] FUN��O CHAMADA - ID: {empresa_id}")
    try:
        usuario = get_usuario_logado()
        
        if usuario['tipo'] != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403
        
        dados = request.json
        motivo = dados.get('motivo', 'N�o especificado')
        
        resultado = database.suspender_empresa(empresa_id, motivo)
        
        if resultado['success']:
            # Registrar log
            try:
                auth_db.registrar_log_acesso(
                    usuario_id=usuario['id'],
                    acao='suspender_empresa',
                    descricao=f"Empresa {empresa_id} suspensa: {motivo}",
                    sucesso=True
                )
            except:
                pass
            
            return jsonify(resultado)
        else:
            return jsonify(resultado), 400
        
    except Exception as e:
        print(f"? Erro ao suspender empresa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas/<int:empresa_id>/reativar', methods=['POST'])
@require_auth
def reativar_empresa_api(empresa_id):
    """Reativa uma empresa suspensa"""
    print(f"\n?? [reativar_empresa_api] FUN��O CHAMADA - ID: {empresa_id}")
    try:
        usuario = get_usuario_logado()
        
        if usuario['tipo'] != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403
        
        resultado = database.reativar_empresa(empresa_id)
        
        if resultado['success']:
            # Registrar log
            try:
                auth_db.registrar_log_acesso(
                    usuario_id=usuario['id'],
                    acao='reativar_empresa',
                    descricao=f"Empresa {empresa_id} reativada",
                    sucesso=True
                )
            except:
                pass
            
            return jsonify(resultado)
        else:
            return jsonify(resultado), 400
        
    except Exception as e:
        print(f"? Erro ao reativar empresa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas/<int:empresa_id>', methods=['DELETE'])
@require_auth
def deletar_empresa_api(empresa_id):
    """Deleta uma empresa (apenas admin e se n�o tiver usu�rios vinculados)"""
    print(f"\n? [deletar_empresa_api] FUN��O CHAMADA - ID: {empresa_id}")
    try:
        usuario = get_usuario_logado()
        
        if usuario['tipo'] != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403
        
        # Verificar se tem usu�rios vinculados
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) as count FROM usuarios WHERE empresa_id = %s", (empresa_id,))
            result = cursor.fetchone()
            cursor.close()
        
        if result and result['count'] > 0:
            return jsonify({
                'success': False,
                'error': f'N�o � poss�vel excluir. Existem {result["count"]} usu�rio(s) vinculado(s) a esta empresa.'
            }), 400
        
        # Excluir empresa
        with database.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM empresas WHERE id = %s", (empresa_id,))
            conn.commit()
            cursor.close()
        
        # Registrar log
        try:
            auth_db.registrar_log_acesso(
                usuario_id=usuario['id'],
                acao='deletar_empresa',
                descricao=f"Empresa {empresa_id} deletada",
                sucesso=True
            )
        except:
            pass
        
        return jsonify({'success': True, 'message': 'Empresa deletada com sucesso'})
        
    except Exception as e:
        print(f"? Erro ao deletar empresa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas/<int:empresa_id>/stats', methods=['GET'])
@require_auth
def estatisticas_empresa_api(empresa_id):
    """Obt�m estat�sticas de uma empresa"""
    print(f"\n?? [estatisticas_empresa_api] FUN��O CHAMADA - ID: {empresa_id}")
    try:
        usuario = auth_db.obter_usuario(session.get('usuario_id'))
        
        # Verificar acesso - admin ou usu�rio com v�nculo ativo
        if usuario['tipo'] != 'admin':
            from auth_functions import verificar_acesso_empresa
            tem_acesso = verificar_acesso_empresa(usuario['id'], empresa_id, auth_db)
            if not tem_acesso:
                return jsonify({'error': 'Acesso negado'}), 403
        
        stats = database.obter_estatisticas_empresa(empresa_id)
        
        return jsonify({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        print(f"? Erro ao obter estat�sticas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# LISTAR ROTAS (NIVEL DE MODULO - EXECUTA SEMPRE)
# ============================================================================
# DESABILITADO: Gera centenas de logs na inicializa��o, causando rate limit no Railway
# logger.info("="*80)
# logger.info("ROTAS REGISTRADAS:")
# logger.info("="*80)
# for rule in app.url_map.iter_rules():
#     if 'api' in rule.rule and 'static' not in rule.rule:
#         methods = ', '.join(sorted(rule.methods - {'HEAD', 'OPTIONS'}))
#         logger.info(f"  {rule.rule:<45} [{methods}]")
# logger.info("="*80)


# ============================================================================
# MONITORAMENTO DO POOL DE CONEX�ES
# ============================================================================

@app.route('/api/health/pool', methods=['GET'])
def pool_status():
    """Endpoint para monitorar status do pool de conex�es"""
    try:
        status = database.get_pool_status()
        
        # Adicionar informa��es extras
        status['status'] = 'healthy'
        status['pool_type'] = 'ThreadedConnectionPool'
        
        # Verificar se h� muitas conex�es em uso
        if 'in_use' in status and 'maxconn' in status:
            usage_percent = (status['in_use'] / status['maxconn']) * 100
            status['usage_percent'] = round(usage_percent, 2)
            
            if usage_percent > 90:
                status['warning'] = 'Pool quase esgotado! Considere aumentar maxconn.'
            elif usage_percent > 75:
                status['notice'] = 'Pool com alto uso.'
        
        return jsonify(status), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


# ============================================================================
# ENDPOINT TEMPOR�RIO PARA CRIAR USU�RIO ADMIN (RAILWAY)
# ============================================================================
@app.route('/api/debug/criar-admin', methods=['POST'])
@csrf_instance.exempt
def criar_admin_inicial():
    """
    Endpoint tempor�rio para criar usu�rio admin no Railway
    
    ATENCAO: DISPON�VEL APENAS EM DESENVOLVIMENTO
    Em produ��o, use: python criar_admin_railway.py
    """
    # Bloquear em produ��o
    check = _check_debug_endpoint_allowed()
    if check:
        return check
    
    try:
        from auth_functions import hash_password, verificar_senha
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        senha = "admin123"
        password_hash = hash_password(senha)
        
        # Testar se o hash funciona
        teste_verificacao = verificar_senha(senha, password_hash)
        
        # Atualizar ou criar admin
        cursor.execute("""
            INSERT INTO usuarios (username, password_hash, tipo, nome_completo, email, ativo)
            VALUES ('admin', %s, 'admin', 'Administrador do Sistema', 'admin@sistema.com', TRUE)
            ON CONFLICT (username) DO UPDATE 
            SET password_hash = EXCLUDED.password_hash, ativo = TRUE
            RETURNING id
        """, (password_hash,))
        
        result = cursor.fetchone()
        admin_id = result['id'] if result else None
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Admin atualizado com sucesso',
            'admin_id': admin_id,
            'username': 'admin',
            'senha': senha,
            'hash_preview': password_hash[:50],
            'teste_verificacao_interna': teste_verificacao
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT TEMPOR�RIO PARA ADICIONAR PERMISS�ES DE REGRAS (RAILWAY)
# ============================================================================
@app.route('/api/debug/adicionar-permissoes-regras', methods=['POST'])
@csrf_instance.exempt
def adicionar_permissoes_regras():
    """
    Endpoint tempor�rio para adicionar permiss�es de regras de concilia��o
    no campo JSONB permissoes_empresa da tabela usuario_empresas
    
    ATENCAO: DISPON�VEL APENAS EM DESENVOLVIMENTO
    """
    # Bloquear em produ��o
    check = _check_debug_endpoint_allowed()
    if check:
        return check
    
    try:
        import json
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Buscar todos os v�nculos usuario-empresa ativos
        cursor.execute("""
            SELECT usuario_id, empresa_id, permissoes_empresa
            FROM usuario_empresas
            WHERE ativo = TRUE
        """)
        vinculos = cursor.fetchall()
        
        # Permiss�es a adicionar
        novas_permissoes = [
            'regras_conciliacao_view',
            'regras_conciliacao_create', 
            'regras_conciliacao_edit',
            'regras_conciliacao_delete'
        ]
        
        atualizados = 0
        detalhes = []
        
        for vinculo in vinculos:
            usuario_id = vinculo['usuario_id']
            empresa_id = vinculo['empresa_id']
            permissoes_atual = vinculo['permissoes_empresa']
            
            # Converter JSONB para lista Python
            if permissoes_atual:
                if isinstance(permissoes_atual, str):
                    permissoes = json.loads(permissoes_atual)
                else:
                    permissoes = permissoes_atual
            else:
                permissoes = []
            
            # Adicionar novas permiss�es se n�o existirem
            permissoes_adicionadas = []
            for perm in novas_permissoes:
                if perm not in permissoes:
                    permissoes.append(perm)
                    permissoes_adicionadas.append(perm)
            
            if permissoes_adicionadas:
                # Atualizar no banco
                cursor.execute("""
                    UPDATE usuario_empresas
                    SET permissoes_empresa = %s::jsonb
                    WHERE usuario_id = %s AND empresa_id = %s
                """, (json.dumps(permissoes), usuario_id, empresa_id))
                
                atualizados += 1
                detalhes.append({
                    'usuario_id': usuario_id,
                    'empresa_id': empresa_id,
                    'permissoes_adicionadas': permissoes_adicionadas,
                    'total_permissoes': len(permissoes)
                })
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{atualizados} v�nculo(s) atualizado(s)',
            'vinculos_total': len(vinculos),
            'vinculos_atualizados': atualizados,
            'detalhes': detalhes
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT TEMPOR�RIO PARA FIX SUBCATEGORIAS (RAILWAY)
# ============================================================================
@app.route('/api/debug/fix-subcategorias-type', methods=['POST'])
@csrf.exempt
def fix_subcategorias_type():
    """
    Endpoint tempor�rio para corrigir tipo da coluna subcategorias
    Altera de TEXT para VARCHAR(255)
    """
    try:
        db_instance = DatabaseManager()
        conn = db_instance.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # Verificar tipo atual
        cursor.execute("""
            SELECT data_type, character_maximum_length
            FROM information_schema.columns
            WHERE table_name = 'categorias'
            AND column_name = 'subcategorias'
        """)
        
        result = cursor.fetchone()
        tipo_antes = result['data_type'] if result else 'n�o encontrado'
        tamanho_antes = result['character_maximum_length'] if result else None
        
        if tipo_antes == 'character varying':
            return jsonify({
                'success': True,
                'message': 'Coluna j� est� correta (character varying)',
                'tipo_atual': tipo_antes,
                'tamanho': tamanho_antes
            })
        
        # Alterar tipo
        cursor.execute("""
            ALTER TABLE categorias
            ALTER COLUMN subcategorias TYPE VARCHAR(255)
            USING subcategorias::VARCHAR(255)
        """)
        
        conn.commit()
        
        # Verificar resultado
        cursor.execute("""
            SELECT data_type, character_maximum_length
            FROM information_schema.columns
            WHERE table_name = 'categorias'
            AND column_name = 'subcategorias'
        """)
        
        result = cursor.fetchone()
        tipo_depois = result['data_type']
        tamanho_depois = result['character_maximum_length']
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Coluna subcategorias alterada com sucesso!',
            'tipo_antes': tipo_antes,
            'tamanho_antes': tamanho_antes,
            'tipo_depois': tipo_depois,
            'tamanho_depois': tamanho_depois
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT TEMPOR�RIO PARA VERIFICAR TABELA REGRAS_CONCILIACAO
# ============================================================================
@app.route('/api/debug/verificar-tabela-regras', methods=['GET'])
@csrf.exempt
def verificar_tabela_regras():
    """
    Endpoint tempor�rio para diagnosticar tabela regras_conciliacao
    """
    try:
        conn = db.get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # 1. Verificar se tabela existe
        cursor.execute("""
            SELECT COUNT(*) as existe
            FROM information_schema.tables
            WHERE table_schema = 'public'
            AND table_name = 'regras_conciliacao'
        """)
        tabela_existe = cursor.fetchone()['existe'] > 0
        
        resultado = {
            'tabela_existe': tabela_existe
        }
        
        if tabela_existe:
            # 2. Verificar estrutura da tabela
            cursor.execute("""
                SELECT column_name, data_type, is_nullable
                FROM information_schema.columns
                WHERE table_name = 'regras_conciliacao'
                ORDER BY ordinal_position
            """)
            colunas = cursor.fetchall()
            resultado['colunas'] = [dict(c) for c in colunas]
            
            # 3. Contar registros
            cursor.execute("SELECT COUNT(*) as total FROM regras_conciliacao")
            resultado['total_regras'] = cursor.fetchone()['total']
            
            # 4. Testar query simples
            try:
                cursor.execute("""
                    SELECT id, empresa_id, palavra_chave, ativo
                    FROM regras_conciliacao
                    LIMIT 5
                """)
                resultado['amostra'] = [dict(r) for r in cursor.fetchall()]
                resultado['query_ok'] = True
            except Exception as e:
                resultado['query_ok'] = False
                resultado['query_erro'] = str(e)
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': resultado
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT TEMPOR�RIO PARA VERIFICAR M�TODOS DO DatabaseManager
# ============================================================================
@app.route('/api/debug/verificar-metodos-db', methods=['GET'])
@csrf.exempt
def verificar_metodos_db():
    """
    Endpoint tempor�rio para verificar quais m�todos o objeto db possui
    """
    try:
        # Listar todos os m�todos do objeto db
        metodos_db = [m for m in dir(db) if not m.startswith('_')]
        
        # Verificar especificamente os m�todos de regras
        metodos_regras = {
            'listar_regras_conciliacao': hasattr(db, 'listar_regras_conciliacao'),
            'criar_regra_conciliacao': hasattr(db, 'criar_regra_conciliacao'),
            'atualizar_regra_conciliacao': hasattr(db, 'atualizar_regra_conciliacao'),
            'excluir_regra_conciliacao': hasattr(db, 'excluir_regra_conciliacao'),
        }
        
        # Informa��es sobre o objeto db
        info_db = {
            'tipo': str(type(db)),
            'modulo': db.__class__.__module__,
            'classe': db.__class__.__name__,
        }
        
        return jsonify({
            'success': True,
            'data': {
                'info_db': info_db,
                'total_metodos': len(metodos_db),
                'metodos_regras': metodos_regras,
                'sample_metodos': metodos_db[:50]  # Primeiros 50 m�todos
            }
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT TEMPOR�RIO PARA FOR�AR ATUALIZA��O DE PERMISS�ES
# ============================================================================
@app.route('/api/debug/adicionar-permissoes-config-extrato', methods=['POST'])
@csrf.exempt
def adicionar_permissoes_config_extrato():
    """
    Endpoint tempor�rio para for�ar adi��o de permiss�es de config_extrato
    """
    try:
        from database_postgresql import execute_query
        
        # 1. Garantir que as permiss�es existem
        execute_query("""
            INSERT INTO permissoes (codigo, nome, descricao, categoria) VALUES
            ('config_extrato_bancario_view', 'Visualizar Configura��es de Extrato', 'Permite visualizar configura��es de extrato banc�rio', 'configuracoes'),
            ('config_extrato_bancario_edit', 'Editar Configura��es de Extrato', 'Permite editar configura��es de extrato banc�rio', 'configuracoes')
            ON CONFLICT (codigo) DO NOTHING
        """, fetch_all=False, allow_global=True)
        
        # 2. Adicionar permiss�es aos usu�rios ativos e contar
        result = execute_query("""
            WITH atualizar AS (
                UPDATE usuario_empresas
                SET permissoes_empresa = permissoes_empresa || 
                    jsonb_build_array('config_extrato_bancario_view', 'config_extrato_bancario_edit')
                WHERE ativo = TRUE
                  AND NOT (permissoes_empresa @> '["config_extrato_bancario_view"]'::jsonb)
                RETURNING id
            )
            SELECT COUNT(*) as atualizados FROM atualizar
        """, fetch_one=True, allow_global=True)
        
        rows_updated = result['atualizados'] if result else 0
        
        # 3. Verificar total
        result_total = execute_query("""
            SELECT COUNT(*) as total
            FROM usuario_empresas
            WHERE ativo = TRUE
              AND permissoes_empresa @> '["config_extrato_bancario_view"]'::jsonb
        """, fetch_one=True, allow_global=True)
        
        total = result_total['total'] if result_total else 0
        
        return jsonify({
            'success': True,
            'message': 'Permiss�es adicionadas com sucesso',
            'data': {
                'usuarios_atualizados': rows_updated,
                'total_com_permissoes': total
            }
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT PARA EXECUTAR MIGRATION DE CONFIG EXTRATO
# ============================================================================
@app.route('/api/debug/executar-migration-config-extrato', methods=['POST'])
@csrf.exempt
def executar_migration_config_extrato():
    """
    Endpoint para for�ar execu��o da migration de config_extrato_bancario
    """
    try:
        from database_postgresql import execute_query
        import os
        
        # Ler arquivo SQL
        sql_file = os.path.join(os.path.dirname(__file__), 'migration_config_integracao_folha.sql')
        
        if not os.path.exists(sql_file):
            return jsonify({
                'success': False,
                'error': f'Arquivo n�o encontrado: {sql_file}'
            }), 404
        
        with open(sql_file, 'r', encoding='utf-8') as f:
            sql_content = f.read()
        
        # Executar SQL
        execute_query(sql_content, fetch_all=False, allow_global=True)
        
        # Verificar se tabela foi criada
        result = execute_query("""
            SELECT COUNT(*) as existe
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name = 'config_extrato_bancario'
        """, fetch_one=True, allow_global=True)
        
        tabela_criada = result['existe'] == 1 if result else False
        
        # Verificar se coluna foi removida
        result_coluna = execute_query("""
            SELECT COUNT(*) as existe
            FROM information_schema.columns 
            WHERE table_name = 'regras_conciliacao' 
            AND column_name = 'usa_integracao_folha'
        """, fetch_one=True, allow_global=True)
        
        coluna_removida = result_coluna['existe'] == 0 if result_coluna else False
        
        # Contar registros criados
        result_config = execute_query("""
            SELECT COUNT(*) as total
            FROM config_extrato_bancario
        """, fetch_one=True, allow_global=True)
        
        configs_criadas = result_config['total'] if result_config else 0
        
        return jsonify({
            'success': True,
            'message': 'Migration executada com sucesso',
            'data': {
                'tabela_criada': tabela_criada,
                'coluna_removida': coluna_removida,
                'configs_criadas': configs_criadas,
                'sql_size': len(sql_content)
            }
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/debug/listar-regras-raw', methods=['GET'])
@csrf.exempt
def listar_regras_raw():
    """
    Endpoint de debug para listar todas as regras diretamente do banco
    """
    try:
        from database_postgresql import execute_query
        
        empresa_id = request.args.get('empresa_id', type=int)
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'empresa_id � obrigat�rio'
            }), 400
        
        # Query direta no banco
        query = """
            SELECT 
                id,
                empresa_id,
                palavra_chave,
                categoria,
                subcategoria,
                cliente_padrao,
                descricao,
                ativo,
                created_at,
                updated_at
            FROM regras_conciliacao
            WHERE empresa_id = %s
            ORDER BY id
        """
        
        regras = execute_query(query, (empresa_id,), fetch_all=True, allow_global=True)
        
        return jsonify({
            'success': True,
            'empresa_id': empresa_id,
            'total': len(regras) if regras else 0,
            'regras': regras or []
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT DE STATUS DA MIGRA��O DE SENHAS
# ============================================================================
@app.route('/api/admin/passwords/migration-status', methods=['GET'])
@require_admin
def password_migration_status():
    """Retorna status da migra��o de senhas SHA-256 ? bcrypt"""
    try:
        from migration_upgrade_passwords import relatorio_hashes_pendentes
        
        stats = relatorio_hashes_pendentes(db)
        
        return jsonify({
            'success': True,
            'data': {
                'total_usuarios': stats['total_usuarios'],
                'usuarios_bcrypt': stats['usuarios_bcrypt'],
                'usuarios_sha256': stats['usuarios_sha256'],
                'usuarios_desconhecido': stats['usuarios_desconhecido'],
                'percentual_migrado': round(
                    (stats['usuarios_bcrypt'] / stats['total_usuarios'] * 100) 
                    if stats['total_usuarios'] > 0 else 100, 
                    2
                ),
                'pendentes': stats['pendentes']
            }
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/admin/passwords/force-upgrade', methods=['POST'])
@require_admin
def force_password_upgrade():
    """For�a upgrade de senha para um usu�rio espec�fico"""
    try:
        from migration_upgrade_passwords import forcar_upgrade_usuario
        
        data = request.json
        username = data.get('username')
        nova_senha = data.get('nova_senha')
        
        if not username or not nova_senha:
            return jsonify({
                'success': False,
                'error': 'username e nova_senha s�o obrigat�rios'
            }), 400
        
        sucesso = forcar_upgrade_usuario(username, nova_senha, db)
        
        if sucesso:
            return jsonify({
                'success': True,
                'message': f'Senha de {username} atualizada com sucesso'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Usu�rio n�o encontrado ou erro ao atualizar'
            }), 404
    
    except Exception as e:
        logger.error(f"Erro ao for�ar upgrade de senha: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# ENDPOINT PARA EXECUTAR MIGRATIONS
# ============================================================================
@app.route('/api/admin/migrations/evento-funcionarios', methods=['POST'])
@require_admin
def execute_migration_evento_funcionarios():
    """Executa migration para criar tabelas funcoes_evento e evento_funcionarios"""
    try:
        logger.info("?? Iniciando migration evento_funcionarios...")
        
        # Ler arquivo SQL
        sql_file = os.path.join(os.path.dirname(__file__), 'migration_evento_funcionarios.sql')
        
        if not os.path.exists(sql_file):
            return jsonify({
                'success': False,
                'error': f'Arquivo migration n�o encontrado: {sql_file}'
            }), 404
        
        with open(sql_file, 'r', encoding='utf-8') as f:
            sql_script = f.read()
        
        # Executar script
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute(sql_script)
            conn.commit()
            
            # Verificar tabelas criadas
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name IN ('funcoes_evento', 'evento_funcionarios')
                ORDER BY table_name
            """)
            tabelas = cursor.fetchall()
            
            # Verificar fun��es inseridas
            cursor.execute("SELECT COUNT(*) as total FROM funcoes_evento")
            total_funcoes = cursor.fetchone()['total']
            
            logger.info(f"? Migration executada: {len(tabelas)} tabelas, {total_funcoes} fun��es")
            
            return jsonify({
                'success': True,
                'message': 'Migration executada com sucesso',
                'data': {
                    'tabelas_criadas': [t['table_name'] for t in tabelas],
                    'funcoes_inseridas': total_funcoes
                }
            })
            
        except Exception as e:
            conn.rollback()
            raise e
        
        finally:
            cursor.close()
            conn.close()
    
    except Exception as e:
        logger.error(f"? Erro ao executar migration: {e}")
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENDPOINT DE ANALYTICS - PERFORMANCE MONITORING
# ============================================================================
@app.route('/api/analytics/lazy-loading', methods=['POST'])
@require_auth
def log_lazy_loading_performance():
    """Recebe e armazena m�tricas de performance do lazy loading"""
    try:
        data = request.json
        usuario_id = get_usuario_logado()['id']
        
        # Log estruturado das m�tricas
        logger.info("lazy_loading_metrics", extra={
            'usuario_id': usuario_id,
            'session_duration': data.get('summary', {}).get('sessionDuration'),
            'total_pages': data.get('summary', {}).get('totalPagesLoaded'),
            'cache_hit_rate': data.get('cache', {}).get('hitRate'),
            'avg_load_time': data.get('performance', {}).get('avgLoadTime'),
            'errors': len(data.get('errors', []))
        })
        
        # Opcionalmente, armazenar em tabela de m�tricas
        # (se quiser an�lise hist�rica mais complexa)
        
        return jsonify({
            'success': True,
            'message': 'M�tricas recebidas'
        })
        
    except Exception as e:
        logger.error(f"Erro ao processar m�tricas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/analytics/lazy-loading/summary', methods=['GET'])
@require_admin
def get_lazy_loading_summary():
    """Retorna resumo de m�tricas de performance do lazy loading (admin only)"""
    try:
        # Aqui voc� pode implementar agrega��o de m�tricas
        # Por enquanto, retorna instru��es de uso
        return jsonify({
            'success': True,
            'message': 'M�tricas dispon�veis nos logs estruturados',
            'instructions': {
                'log_query': 'Buscar por "lazy_loading_metrics" nos logs',
                'console_usage': [
                    'window.lazyLoadMonitors.default.printReport()',
                    'window.lazyLoadMonitors.default.sendToBackend()'
                ]
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ROTAS NFS-e (Nota Fiscal de Servi�o Eletr�nica)
# ============================================================================

@app.route('/api/nfse/config', methods=['GET'])
@require_auth
def get_config_nfse():
    """Lista configura��es de munic�pios da empresa"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        from nfse_functions import listar_municipios
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        configs = listar_municipios(db_params, empresa_id)
        
        return jsonify({
            'success': True,
            'configs': configs
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar configs NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/config', methods=['POST'])
@require_auth
@require_permission('nfse_config')
def add_config_nfse():
    """Adiciona configura��o de munic�pio"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        # Validar campos obrigat�rios
        required_fields = ['cnpj_cpf', 'codigo_municipio', 'nome_municipio', 'uf']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'Campo obrigat�rio: {field}'
                }), 400
        
        from nfse_functions import adicionar_municipio
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        sucesso, config_id, erro = adicionar_municipio(
            db_params=db_params,
            empresa_id=empresa_id,
            cnpj_cpf=data['cnpj_cpf'],
            codigo_municipio=data['codigo_municipio'],
            nome_municipio=data['nome_municipio'],
            uf=data['uf'],
            inscricao_municipal=data.get('inscricao_municipal', ''),
            provedor=data.get('provedor'),
            url_customizada=data.get('url_customizada')
        )
        
        if sucesso:
            # Log de auditoria
            from nfse_functions import registrar_operacao
            registrar_operacao(
                db_params=db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='CONFIG',
                detalhes={'municipio': data['nome_municipio']},
                ip_address=request.remote_addr
            )
            
            return jsonify({
                'success': True,
                'config_id': config_id,
                'message': 'Munic�pio configurado com sucesso'
            })
        else:
            return jsonify({
                'success': False,
                'error': erro
            }), 400
            
    except Exception as e:
        logger.error(f"Erro ao adicionar config NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/config/<int:config_id>', methods=['PUT'])
@require_auth
@require_permission('nfse_config')
def update_config_nfse(config_id):
    """Atualiza configura��o de munic�pio"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        # Validar campos obrigat�rios
        required_fields = ['cnpj_cpf', 'codigo_municipio', 'nome_municipio', 'uf']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'Campo obrigat�rio: {field}'
                }), 400
        
        from nfse_functions import atualizar_municipio
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        sucesso, erro = atualizar_municipio(
            db_params=db_params,
            empresa_id=empresa_id,
            config_id=config_id,
            cnpj_cpf=data['cnpj_cpf'],
            codigo_municipio=data['codigo_municipio'],
            nome_municipio=data['nome_municipio'],
            uf=data['uf'],
            inscricao_municipal=data.get('inscricao_municipal', ''),
            provedor=data.get('provedor'),
            url_customizada=data.get('url_customizada')
        )
        
        if sucesso:
            # Log de auditoria
            from nfse_functions import registrar_operacao
            registrar_operacao(
                db_params=db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='CONFIG_UPDATE',
                detalhes={'municipio': data['nome_municipio'], 'config_id': config_id},
                ip_address=request.remote_addr
            )
            
            return jsonify({
                'success': True,
                'config_id': config_id,
                'message': 'Munic�pio atualizado com sucesso'
            })
        else:
            return jsonify({
                'success': False,
                'error': erro
            }), 400
            
    except Exception as e:
        logger.error(f"Erro ao atualizar config NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/config/<int:config_id>', methods=['DELETE'])
@require_auth
@require_permission('nfse_config')
def delete_config_nfse(config_id):
    """Remove configura��o de munic�pio"""
    try:
        from nfse_functions import excluir_municipio
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        sucesso, erro = excluir_municipio(db_params, config_id)
        
        if sucesso:
            return jsonify({
                'success': True,
                'message': 'Configura��o exclu�da'
            })
        else:
            return jsonify({
                'success': False,
                'error': erro
            }), 400
            
    except Exception as e:
        logger.error(f"Erro ao excluir config NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/buscar', methods=['POST'])
@require_auth
@require_permission('nfse_buscar')
def buscar_nfse():
    """
    Proxy para busca pesada de NFS-e
    Redireciona requisi��o para microservi�o de busca
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        # Validar ordem das datas (apenas se ambas fornecidas)
        from datetime import datetime
        if data.get('data_inicial') and data.get('data_final'):
            try:
                dt_inicial = datetime.strptime(data['data_inicial'], '%Y-%m-%d')
                dt_final = datetime.strptime(data['data_final'], '%Y-%m-%d')
                if dt_final < dt_inicial:
                    return jsonify({
                        'success': False,
                        'error': 'Data final deve ser maior que data inicial'
                    }), 400
            except ValueError as e:
                return jsonify({
                    'success': False,
                    'error': f'Formato de data invalido: {e}'
                }), 400
        
# Obter URL do microservi�o de busca
        nfse_service_url = os.getenv('NFSE_SERVICE_URL')
        
        if not nfse_service_url:
            # Fallback: processar localmente (modo legacy)
            logger.warning("?? NFSE_SERVICE_URL n�o configurada - processando localmente")
            return _buscar_nfse_local(empresa_id, usuario, data, request.remote_addr)
        
        # Garantir que a URL tenha o protocolo https://
        if not nfse_service_url.startswith(('http://', 'https://')):
            nfse_service_url = f"https://{nfse_service_url}"
            logger.info(f"? Protocolo https:// adicionado automaticamente")
        
        # ========== CHAMADA AO MICROSERVI�O ==========
        logger.info(f"?? Redirecionando busca de NFS-e para microservi�o: {nfse_service_url}")
        
        import requests
        
        # Preparar headers para autentica��o no microservi�o
        # Como o usu�rio j� foi autenticado no ERP (@require_auth), 
        # criamos um token de servi�o interno para o microservi�o
        service_token = f"Bearer ERP-{empresa_id}-{usuario['id']}-{app.secret_key[:16]}"
        
        headers = {
            'Content-Type': 'application/json',
            'Authorization': service_token,
            'X-Empresa-ID': str(empresa_id),
            'X-Usuario-ID': str(usuario['id']),
            'X-Usuario-Nome': usuario.get('nome', 'Unknown')
        }
        
        # Injetar certificado ativo no payload para o microservico
        try:
            with get_db_connection(empresa_id=empresa_id) as _cert_conn:
                _cert_cur = _cert_conn.cursor()
                _cert_cur.execute("""
                    SELECT pfx_base64, senha_pfx
                    FROM certificados_digitais
                    WHERE empresa_id = %s AND ativo = TRUE
                    ORDER BY criado_em DESC LIMIT 1
                """, (empresa_id,))
                cert_row = _cert_cur.fetchone()
            if cert_row:
                pfx_b64 = cert_row['pfx_base64'] if isinstance(cert_row, dict) else cert_row[0]
                senha_cripto = cert_row['senha_pfx'] if isinstance(cert_row, dict) else cert_row[1]
                from relatorios.nfe import nfe_api as _nfe_api
                _fernet_key = os.getenv('FERNET_KEY', '').encode('utf-8')
                senha_plain = _nfe_api.descriptografar_senha(senha_cripto, _fernet_key)
                data = dict(data)  # copia para nao mutar o original
                data['pfx_base64'] = pfx_b64
                data['senha'] = senha_plain
                logger.info(f"?? Certificado ativo injetado no payload NFS-e (pfx_len={len(pfx_b64 or '')}b)")
            else:
                logger.warning("?? Nenhum certificado ativo encontrado para injetar no payload NFS-e")
        except Exception as _ce:
            logger.error(f"?? Erro ao injetar certificado no payload NFS-e: {_ce}")

        # Fazer requisição ao microserviço
        # Timeout 1700s (worker timeout = 1800s, deixa margem de 100s)
        try:
            response = requests.post(
                f"{nfse_service_url}/api/nfse/buscar",
                json=data,
                headers=headers,
                timeout=1700  # 1700 segundos (28min 20s) - worker timeout = 1800s (30min)
            )
            
            # Verificar se microservi�o retornou erro
            if response.status_code != 200:
                logger.error(f"? Microservi�o retornou status {response.status_code}")
                error_msg = f'Microservi�o retornou erro (status {response.status_code})'
                try:
                    erro_json = response.json()
                    logger.error(f"   Detalhes: {erro_json}")
                    # Repassar mensagem original do microservi�o ao front-end
                    if erro_json.get('error'):
                        error_msg = erro_json['error']
                except Exception:
                    logger.error(f"   Resposta: {response.text[:500]}")
                
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), response.status_code
            
            resultado = response.json()
            
            # Processar PDFs oficiais recebidos do microservi�o (se houver)
            from database_postgresql import get_nfse_db_params as _get_nfse_db_params
            _db_params_pdf = _get_nfse_db_params()
            try:
                if 'pdfs_oficiais' in resultado and resultado['pdfs_oficiais']:
                    logger.info(f"?? Processando {len(resultado['pdfs_oficiais'])} PDFs oficiais recebidos...")
                    
                    import base64
                    from nfse_functions import salvar_pdf_nfse
                    from nfse_database import NFSeDatabase
                    
                    pdfs_salvos = 0
                    for numero_nfse, pdf_info in resultado['pdfs_oficiais'].items():
                        try:
                            # Validar dados do PDF
                            if not pdf_info.get('pdf_base64'):
                                logger.warning(f"   ?? PDF {numero_nfse} sem base64, pulando")
                                continue
                            
                            # Decodificar base64
                            pdf_content = base64.b64decode(pdf_info['pdf_base64'])
                            logger.info(f"   ?? PDF {numero_nfse}: {len(pdf_content):,} bytes decodificados")
                            
                            # Salvar no storage LOCAL do ERP
                            pdf_path = salvar_pdf_nfse(
                                pdf_content=pdf_content,
                                numero_nfse=pdf_info['numero_nfse'],
                                cnpj_prestador=pdf_info['cnpj_prestador'],
                                codigo_municipio=pdf_info['codigo_municipio'],
                                data_emissao=pdf_info['data_emissao']
                            )
                            
                            logger.info(f"   ?? PDF salvo em: {pdf_path}")
                            
                            if pdf_path:
                                # Atualizar danfse_path no banco
                                logger.info(f"   ?? Atualizando banco: numero_nfse={numero_nfse}, codigo_municipio={pdf_info['codigo_municipio']}")
                                
                                with NFSeDatabase(_db_params_pdf) as db:
                                    cursor = db.conn.cursor()
                                    cursor.execute("""
                                        UPDATE nfse_baixadas 
                                        SET danfse_path = %s, atualizado_em = CURRENT_TIMESTAMP
                                        WHERE numero_nfse = %s AND codigo_municipio = %s
                                    """, (pdf_path, numero_nfse, pdf_info['codigo_municipio']))
                                    
                                    rows_affected = cursor.rowcount
                                    db.conn.commit()
                                    cursor.close()
                                    
                                    logger.info(f"   ? Banco atualizado: {rows_affected} linha(s) afetada(s)")
                                    
                                    if rows_affected == 0:
                                        logger.warning(f"   ?? NENHUMA linha foi atualizada! NFS-e pode n�o existir no banco")
                                    else:
                                        logger.info(f"   ?? danfse_path salvo: {pdf_path}")
                                
                                pdfs_salvos += 1
                                logger.info(f"   ? PDF {numero_nfse} processado com sucesso")
                            else:
                                logger.error(f"   ? salvar_pdf_nfse retornou None para {numero_nfse}")
                        
                        except Exception as e_pdf:
                            logger.error(f"   ? Erro ao processar PDF {numero_nfse}: {e_pdf}")
                            import traceback
                            logger.error(traceback.format_exc())
                    
                    logger.info(f"? {pdfs_salvos} PDFs oficiais salvos no storage do ERP")
                    
                    # Remover PDFs do resultado (n�o precisam ir pro frontend)
                    del resultado['pdfs_oficiais']
            
            except Exception as e_pdfs:
                logger.error(f"? Erro ao processar PDFs oficiais: {e_pdfs}")
                import traceback
                logger.error(traceback.format_exc())
                # Continua mesmo com erro nos PDFs (dados principais foram salvos)
            
            # Log de auditoria
            from nfse_functions import registrar_operacao
            from database_postgresql import get_nfse_db_params as _get_audit_db_params
            _audit_db_params = _get_audit_db_params()
            registrar_operacao(
                db_params=_audit_db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='BUSCA_VIA_MICROSERVICO',
                detalhes={
                    'data_inicial': data.get('data_inicial', 'auto'),
                    'data_final': data.get('data_final', 'auto'),
                    'metodo': data.get('metodo', 'ambiente_nacional'),
                    'total_nfse': resultado.get('total_nfse', 0)
                },
                ip_address=request.remote_addr
            )
            
            return jsonify(resultado), response.status_code
            
        except requests.exceptions.Timeout:
            logger.error("?? Timeout ao buscar NFS-e no microservi�o")
            return jsonify({
                'success': False,
                'error': 'A busca est� demorando muito. Tente reduzir o per�odo ou n�mero de munic�pios.'
            }), 504
            
        except requests.exceptions.ConnectionError:
            logger.error("? Erro de conex�o com microservi�o de busca")
            return jsonify({
                'success': False,
                'error': 'Servi�o de busca de notas temporariamente indispon�vel'
            }), 503
        
    except Exception as e:
        logger.error(f"Erro ao buscar NFS-e: {e}")
        capture_exception(e)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def _buscar_nfse_local(empresa_id, usuario, data, ip_address):
    """
    MODO LEGACY: Processamento local quando microservi�o n�o est� dispon�vel
    ?? Mantido para compatibilidade, mas n�o recomendado (busca pesada bloqueia o ERP)
    """
    logger.warning("?? Processando busca de NFS-e localmente (MODO LEGACY)")
    
    # Buscar CNPJ da empresa
    with get_db_connection(empresa_id=empresa_id) as conn:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute("SELECT cnpj FROM empresas WHERE id = %s", (empresa_id,))
        empresa = cursor.fetchone()
        cursor.close()
    
    if not empresa:
        return jsonify({
            'success': False,
            'error': 'Empresa n�o encontrada'
        }), 404
    
    if not empresa.get('cnpj'):
        return jsonify({
            'success': False,
            'error': 'CNPJ da empresa nao cadastrado. Acesse Dados da Empresa e preencha o CNPJ antes de buscar NFS-e.'
        }), 400
    
    if not empresa.get('cnpj'):
        return jsonify({
            'success': False,
            'error': 'CNPJ da empresa nao cadastrado. Acesse Dados da Empresa e preencha o CNPJ antes de buscar NFS-e.'
        }), 400
    
    cnpj_prestador = empresa['cnpj'].replace('.', '').replace('/', '').replace('-', '')
    
    # Buscar certificado
    from nfse_functions import get_certificado_para_soap
    from database_postgresql import get_nfse_db_params
    
    db_params = get_nfse_db_params()
    cert_data = get_certificado_para_soap(db_params, empresa_id)
    
    if cert_data:
        import tempfile
        pfx_bytes, cert_senha = cert_data
        temp_cert = tempfile.NamedTemporaryFile(delete=False, suffix='.pfx')
        temp_cert.write(pfx_bytes)
        temp_cert.close()
        certificado_path = temp_cert.name
        certificado_senha = cert_senha
    else:
        certificado_path = os.getenv('CERTIFICADO_A1_PATH', '/app/certificados/certificado.pfx')
        certificado_senha = os.getenv('CERTIFICADO_A1_SENHA', '')
        
        if not os.path.exists(certificado_path):
            return jsonify({
                'success': False,
                'error': 'Certificado A1 n�o configurado'
            }), 400
    
    # Auto-detectar datas quando não informadas (mesma lógica do microserviço)
    from datetime import date as _date, datetime
    if not data.get('data_inicial') or not data.get('data_final'):
        _today = _date.today()
        _ultima_data = None
        try:
            with get_db_connection(empresa_id=empresa_id) as _ac:
                _cur = _ac.cursor()
                _cur.execute(
                    "SELECT MAX(data_emissao) FROM nfse_baixadas WHERE empresa_id = %s",
                    (empresa_id,)
                )
                _r = _cur.fetchone()
                _cur.close()
                if _r and _r[0]:
                    _ultima_data = _r[0].date() if hasattr(_r[0], 'date') else _r[0]
        except Exception as _de:
            logger.warning(f"Auto-data local: erro ao consultar última NFS-e: {_de}")
        data = dict(data)
        if not data.get('data_final'):
            data['data_final'] = _today.strftime('%Y-%m-%d')
        if not data.get('data_inicial'):
            data['data_inicial'] = _ultima_data.strftime('%Y-%m-%d') if _ultima_data else '2016-01-01'

    # Converter datas
    data_inicial = datetime.strptime(data['data_inicial'], '%Y-%m-%d').date()
    data_final = datetime.strptime(data['data_final'], '%Y-%m-%d').date()
    
    codigos_municipios = data.get('codigos_municipios')
    metodo = data.get('metodo', 'ambiente_nacional')
    
    if metodo == 'ambiente_nacional':
        from nfse_functions import buscar_nfse_ambiente_nacional
        
        resultado = buscar_nfse_ambiente_nacional(
            db_params=db_params,
            empresa_id=empresa_id,
            cnpj_informante=cnpj_prestador,
            certificado_path=certificado_path,
            certificado_senha=certificado_senha,
            ambiente=data.get('ambiente', 'producao'),
            busca_completa=data.get('busca_completa', False),
            max_documentos=data.get('max_documentos', 50)
        )
    else:
        from nfse_functions import buscar_nfse_periodo
        
        resultado = buscar_nfse_periodo(
            db_params=db_params,
            empresa_id=empresa_id,
            cnpj_prestador=cnpj_prestador,
            data_inicial=data_inicial,
            data_final=data_final,
            certificado_path=certificado_path,
            certificado_senha=certificado_senha,
            codigos_municipios=codigos_municipios
        )
    
    # Limpar certificado tempor�rio
    if cert_data and os.path.exists(certificado_path):
        try:
            os.unlink(certificado_path)
        except:
            pass
    
    # Log de auditoria
    from nfse_functions import registrar_operacao
    registrar_operacao(
        db_params=db_params,
        empresa_id=empresa_id,
        usuario_id=usuario['id'],
        operacao='BUSCA_LOCAL',
        detalhes={
            'data_inicial': data.get('data_inicial', 'auto'),
            'data_final': data.get('data_final', 'auto'),
            'total_nfse': resultado.get('total_nfse', 0)
        },
        ip_address=ip_address
    )
    
    return jsonify(resultado)


@app.route('/api/nfse/consultar', methods=['POST'])
@require_auth
@require_permission('nfse_view')
def consultar_nfse():
    """Consulta NFS-e armazenadas localmente (sem buscar via API)"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        from nfse_functions import consultar_nfse_periodo
        from datetime import datetime
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        # Converter datas
        data_inicial = datetime.strptime(data['data_inicial'], '%Y-%m-%d').date()
        data_final = datetime.strptime(data['data_final'], '%Y-%m-%d').date()
        
        # Par�metros de pagina��o (None = sem limite)
        limit = data.get('limit')  # None por padr�o = busca todos
        offset = data.get('offset', 0)
        
        # Consultar banco local
        nfses = consultar_nfse_periodo(
            db_params=db_params,
            empresa_id=empresa_id,
            data_inicial=data_inicial,
            data_final=data_final,
            codigo_municipio=data.get('codigo_municipio'),
            limit=limit,
            offset=offset
        )
        
        # Converter objetos datetime para string (JSON serialization)
        for nfse in nfses:
            for key, value in nfse.items():
                if isinstance(value, (datetime, date)):
                    nfse[key] = value.isoformat()
                elif isinstance(value, Decimal):
                    nfse[key] = float(value)
        
        return jsonify({
            'success': True,
            'nfses': nfses,
            'total': len(nfses)
        })
        
    except Exception as e:
        logger.error(f"Erro ao consultar NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/diagnostico', methods=['POST'])
@require_auth
@require_permission('nfse_view')
def diagnostico_nfse():
    """Diagn�stico detalhado de NFS-e para identificar omiss�es"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        from datetime import datetime
        from database_postgresql import get_nfse_db_params
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        db_params = get_nfse_db_params()
        
        # Converter datas
        data_inicial = datetime.strptime(data['data_inicial'], '%Y-%m-%d').date()
        data_final = datetime.strptime(data['data_final'], '%Y-%m-%d').date()
        codigo_municipio = data.get('codigo_municipio')
        
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. Total por situa��o
        sql = """
            SELECT 
                situacao,
                COUNT(*) as total,
                SUM(valor_servico) as valor_total,
                SUM(valor_iss) as iss_total
            FROM nfse_baixadas
            WHERE empresa_id = %s
            AND data_competencia BETWEEN %s AND %s
        """
        params = [empresa_id, data_inicial, data_final]
        
        if codigo_municipio:
            sql += " AND codigo_municipio = %s"
            params.append(codigo_municipio)
        
        sql += " GROUP BY situacao ORDER BY total DESC"
        
        cursor.execute(sql, tuple(params))
        por_situacao = [dict(row) for row in cursor.fetchall()]
        
        # 2. Total geral SEM filtro de situa��o
        sql_total = """
            SELECT 
                COUNT(*) as total,
                SUM(valor_servico) as valor_total,
                SUM(valor_iss) as iss_total
            FROM nfse_baixadas
            WHERE empresa_id = %s
            AND data_competencia BETWEEN %s AND %s
        """
        params_total = [empresa_id, data_inicial, data_final]
        
        if codigo_municipio:
            sql_total += " AND codigo_municipio = %s"
            params_total.append(codigo_municipio)
        
        cursor.execute(sql_total, tuple(params_total))
        total_geral = dict(cursor.fetchone())
        
        # 3. Total APENAS NORMAL (o que a interface mostra)
        sql_normal = """
            SELECT 
                COUNT(*) as total,
                SUM(valor_servico) as valor_total,
                SUM(valor_iss) as iss_total
            FROM nfse_baixadas
            WHERE empresa_id = %s
            AND data_competencia BETWEEN %s AND %s
            AND situacao = 'NORMAL'
        """
        params_normal = [empresa_id, data_inicial, data_final]
        
        if codigo_municipio:
            sql_normal += " AND codigo_municipio = %s"
            params_normal.append(codigo_municipio)
        
        cursor.execute(sql_normal, tuple(params_normal))
        total_normal = dict(cursor.fetchone())
        
        # 4. Mostrar notas CANCELADAS/SUBSTITU�DAS se houver
        sql_outras = """
            SELECT 
                numero_nfse,
                data_emissao,
                razao_social_tomador,
                valor_servico,
                situacao
            FROM nfse_baixadas
            WHERE empresa_id = %s
            AND data_competencia BETWEEN %s AND %s
            AND situacao != 'NORMAL'
            ORDER BY data_emissao DESC
            LIMIT 20
        """
        params_outras = [empresa_id, data_inicial, data_final]
        
        if codigo_municipio:
            sql_outras += " AND codigo_municipio = %s"
            params_outras.append(codigo_municipio)
        
        cursor.execute(sql_outras, tuple(params_outras))
        notas_outras_situacoes = [dict(row) for row in cursor.fetchall()]
        
        cursor.close()
        conn.close()
        
        # Converter valores Decimal para float
        for item in por_situacao:
            for key, value in item.items():
                if isinstance(value, Decimal):
                    item[key] = float(value)
        
        for key, value in total_geral.items():
            if isinstance(value, Decimal):
                total_geral[key] = float(value)
        
        for key, value in total_normal.items():
            if isinstance(value, Decimal):
                total_normal[key] = float(value)
        
        for nota in notas_outras_situacoes:
            for key, value in nota.items():
                if isinstance(value, (datetime, date)):
                    nota[key] = value.isoformat()
                elif isinstance(value, Decimal):
                    nota[key] = float(value)
        
        omitidas = total_geral['total'] - total_normal['total']
        
        resultado = {
            'success': True,
            'total_banco': total_geral['total'],
            'total_interface': total_normal['total'],
            'total_omitidas': omitidas,
            'por_situacao': por_situacao,
            'total_geral': total_geral,
            'total_normal': total_normal,
            'exemplos_omitidas': notas_outras_situacoes if omitidas > 0 else []
        }
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro no diagn�stico: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/<int:nfse_id>', methods=['GET'])
@require_auth
@require_permission('nfse_view')
def get_nfse_detalhes(nfse_id):
    """Retorna detalhes completos de uma NFS-e"""
    try:
        from nfse_functions import get_detalhes_nfse
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        nfse = get_detalhes_nfse(db_params, nfse_id)
        
        if not nfse:
            return jsonify({
                'success': False,
                'error': 'NFS-e n�o encontrada'
            }), 404
        
        # Converter objetos datetime para string
        for key, value in nfse.items():
            if isinstance(value, (datetime, date)):
                nfse[key] = value.isoformat()
            elif isinstance(value, Decimal):
                nfse[key] = float(value)
        
        return jsonify({
            'success': True,
            'nfse': nfse
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/<int:nfse_id>', methods=['DELETE'])
@require_auth
@require_permission('nfse_delete')
def excluir_nfse(nfse_id):
    """Exclui uma NFS-e (banco de dados + arquivos XML e PDF)"""
    try:
        from nfse_database import NFSeDatabase
        from database_postgresql import get_nfse_db_params
        import os
        
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 403
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        with NFSeDatabase(db_params) as db:
            # Buscar informa��es da NFS-e antes de excluir (para deletar arquivos)
            sql_select = """
                SELECT xml_path, numero_nfse, cnpj_prestador, codigo_municipio, data_emissao
                FROM nfse_baixadas 
                WHERE id = %s AND empresa_id = %s
            """
            
            with db.conn.cursor() as cursor:
                cursor.execute(sql_select, (nfse_id, empresa_id))
                nfse_data = cursor.fetchone()
                
                if not nfse_data:
                    return jsonify({
                        'success': False,
                        'error': 'NFS-e n�o encontrada ou n�o pertence � empresa atual'
                    }), 404
                
                xml_path, numero_nfse, cnpj_prestador, codigo_municipio, data_emissao = nfse_data
                
                # Excluir do banco de dados
                success = db.excluir_nfse(nfse_id, empresa_id)
                
                if not success:
                    return jsonify({
                        'success': False,
                        'error': 'Erro ao excluir NFS-e do banco de dados'
                    }), 500
                
                # Tentar excluir arquivos XML e PDF
                arquivos_excluidos = []
                arquivos_nao_encontrados = []
                
                # Montar caminhos dos arquivos
                # Formato: storage/nfse/{CNPJ}/{CODIGO_MUNICIPIO}/{ANO}/{MES}/
                if data_emissao:
                    ano = data_emissao.year if hasattr(data_emissao, 'year') else data_emissao[:4]
                    mes = f"{data_emissao.month:02d}" if hasattr(data_emissao, 'month') else data_emissao[5:7]
                    
                    base_dir = f"storage/nfse/{cnpj_prestador}/{codigo_municipio}/{ano}/{mes}"
                    
                    # Nome dos arquivos
                    xml_filename = f"NFS-e_{numero_nfse}.xml"
                    pdf_filename = f"NFS-e_{numero_nfse}.pdf"
                    
                    xml_full_path = os.path.join(base_dir, xml_filename)
                    pdf_full_path = os.path.join(base_dir, pdf_filename)
                    
                    # Tentar excluir XML
                    if os.path.exists(xml_full_path):
                        try:
                            os.remove(xml_full_path)
                            arquivos_excluidos.append(xml_filename)
                            logger.info(f"??? XML exclu�do: {xml_full_path}")
                        except Exception as e:
                            logger.warning(f"?? Erro ao excluir XML: {e}")
                    else:
                        arquivos_nao_encontrados.append(xml_filename)
                        logger.warning(f"?? XML n�o encontrado: {xml_full_path}")
                    
                    # Tentar excluir PDF
                    if os.path.exists(pdf_full_path):
                        try:
                            os.remove(pdf_full_path)
                            arquivos_excluidos.append(pdf_filename)
                            logger.info(f"??? PDF exclu�do: {pdf_full_path}")
                        except Exception as e:
                            logger.warning(f"?? Erro ao excluir PDF: {e}")
                    else:
                        arquivos_nao_encontrados.append(pdf_filename)
                        logger.warning(f"?? PDF n�o encontrado: {pdf_full_path}")
        
        mensagem = f"NFS-e {numero_nfse} exclu�da com sucesso!"
        if arquivos_excluidos:
            mensagem += f" Arquivos removidos: {', '.join(arquivos_excluidos)}."
        if arquivos_nao_encontrados:
            mensagem += f" Arquivos n�o encontrados: {', '.join(arquivos_nao_encontrados)}."
        
        logger.info(f"? {mensagem}")
        
        return jsonify({
            'success': True,
            'message': mensagem,
            'arquivos_excluidos': arquivos_excluidos,
            'arquivos_nao_encontrados': arquivos_nao_encontrados
        })
        
    except Exception as e:
        logger.error(f"Erro ao excluir NFS-e: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/all', methods=['DELETE'])
@require_auth
@require_permission('nfse_delete')
def apagar_todas_nfse():
    """Apaga TODAS as NFS-e do per�odo selecionado (banco de dados + arquivos)"""
    try:
        from nfse_database import NFSeDatabase
        from database_postgresql import get_nfse_db_params
        import os
        from datetime import datetime
        
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 403
        
        # Pegar par�metros
        data_inicial = request.args.get('data_inicial')
        data_final = request.args.get('data_final')
        codigo_municipio = request.args.get('codigo_municipio', '')
        
        if not data_inicial or not data_final:
            return jsonify({
                'success': False,
                'error': 'Data inicial e final s�o obrigat�rias'
            }), 400
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        with NFSeDatabase(db_params) as db:
            # Buscar TODAS as NFS-e do per�odo
            sql_select = """
                SELECT id, xml_path, numero_nfse, cnpj_prestador, codigo_municipio, data_emissao
                FROM nfse_baixadas 
                WHERE empresa_id = %s 
                AND data_emissao >= %s 
                AND data_emissao <= %s
            """
            params = [empresa_id, data_inicial, data_final]
            
            if codigo_municipio:
                sql_select += " AND codigo_municipio = %s"
                params.append(codigo_municipio)
            
            with db.conn.cursor() as cursor:
                cursor.execute(sql_select, params)
                nfse_list = cursor.fetchall()
                
                if not nfse_list:
                    return jsonify({
                        'success': True,
                        'message': 'Nenhuma NFS-e encontrada no per�odo',
                        'total_excluidas': 0,
                        'total_arquivos_excluidos': 0
                    })
                
                total_excluidas = 0
                total_arquivos_excluidos = 0
                erros = []
                
                logger.info(f"??? Iniciando exclus�o de {len(nfse_list)} NFS-e(s) do per�odo {data_inicial} a {data_final}")
                
                # Excluir cada NFS-e
                for nfse_data in nfse_list:
                    nfse_id, xml_path, numero_nfse, cnpj_prestador, codigo_mun, data_emissao = nfse_data
                    
                    try:
                        # Excluir do banco
                        success = db.excluir_nfse(nfse_id, empresa_id)
                        
                        if success:
                            total_excluidas += 1
                            
                            # Tentar excluir arquivos
                            if data_emissao:
                                ano = data_emissao.year if hasattr(data_emissao, 'year') else data_emissao[:4]
                                mes = f"{data_emissao.month:02d}" if hasattr(data_emissao, 'month') else data_emissao[5:7]
                                
                                base_dir = f"storage/nfse/{cnpj_prestador}/{codigo_mun}/{ano}/{mes}"
                                xml_full_path = os.path.join(base_dir, f"NFS-e_{numero_nfse}.xml")
                                pdf_full_path = os.path.join(base_dir, f"NFS-e_{numero_nfse}.pdf")
                                
                                # Excluir XML
                                if os.path.exists(xml_full_path):
                                    try:
                                        os.remove(xml_full_path)
                                        total_arquivos_excluidos += 1
                                    except Exception as e:
                                        logger.warning(f"?? Erro ao excluir XML {numero_nfse}: {e}")
                                
                                # Excluir PDF
                                if os.path.exists(pdf_full_path):
                                    try:
                                        os.remove(pdf_full_path)
                                        total_arquivos_excluidos += 1
                                    except Exception as e:
                                        logger.warning(f"?? Erro ao excluir PDF {numero_nfse}: {e}")
                        
                    except Exception as e:
                        erro_msg = f"NFS-e {numero_nfse}: {str(e)}"
                        erros.append(erro_msg)
                        logger.error(f"? Erro ao excluir NFS-e {numero_nfse}: {e}")
                
                mensagem = f"{total_excluidas} NFS-e(s) exclu�da(s) com sucesso! {total_arquivos_excluidos} arquivo(s) removido(s)."
                
                if erros:
                    mensagem += f" Erros: {len(erros)}"
                
                logger.info(f"? {mensagem}")
                
                return jsonify({
                    'success': True,
                    'message': mensagem,
                    'total_excluidas': total_excluidas,
                    'total_arquivos_excluidos': total_arquivos_excluidos,
                    'erros': erros if erros else []
                })
        
    except Exception as e:
        logger.error(f"Erro ao apagar todas as NFS-e: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/importar-xml', methods=['POST'])
@require_auth
@require_permission('nfse_view')
def importar_xml_nfse():
    """Importa NFS-e a partir de arquivos XML enviados pelo usu�rio"""
    try:
        import xml.etree.ElementTree as ET

        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')

        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 403

        files = request.files.getlist('xmls')
        if not files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo XML enviado'}), 400

        # Obt�m conex�o direta s/ passar pelo NFSeDatabase para simplificar
        from database_postgresql import get_nfse_db_params
        import psycopg2

        db_params = get_nfse_db_params()
        conn = psycopg2.connect(**db_params)

        importadas = 0
        atualizadas = 0
        erros = []

        import re as _re
        import zipfile as _zf
        import io as _io
        import xml.etree.ElementTree as ET

        # cStat (Padrao Nacional SPED) -> coluna situacao
        CSTAT_MAP = {
            '100': 'NORMAL', '101': 'NORMAL', '103': 'NORMAL',
            '104': 'NORMAL', '105': 'NORMAL', '106': 'NORMAL',
            '102': 'CANCELADA',
            '107': 'SUBSTITUIDA',
        }

        def _strip_ns(xml_text):
            # Remove declaracoes de namespace e prefixos para simplificar XPath
            cleaned = _re.sub(r'\s+xmlns(?::\w+)?="[^"]*"', '', xml_text)
            cleaned = _re.sub(r'<(/?)(\w+):(\w+)', r'<\1\3', cleaned)
            return cleaned

        def _t(el, *paths):
            # Primeiro texto nao-vazio dentre os paths (busca recursiva .//)
            for p in paths:
                found = el.find('.//' + p)
                if found is not None and found.text and found.text.strip():
                    return found.text.strip()
            return None

        def _tf(el, *paths):
            v = _t(el, *paths)
            if not v:
                return 0.0
            try:
                return float(v.replace(',', '.'))
            except Exception:
                return 0.0

        def _parse_nfse(xml_text):
            # Parse NFS-e Padrao Nacional SPED 1.x ou ABRASF. Retorna dict.
            try:
                root = ET.fromstring(_strip_ns(xml_text))
            except ET.ParseError:
                root = ET.fromstring(_strip_ns(xml_text.encode('ascii', errors='ignore').decode()))

            root_tag = root.tag.lower()
            data = {}

            # Padrao Nacional SPED (tag raiz: NFSe)
            if root_tag == 'nfse' or root_tag.endswith('}nfse'):
                cstat = _t(root, 'cStat') or ''
                data['situacao']       = CSTAT_MAP.get(cstat, 'NORMAL')
                data['numero_nfse']    = _t(root, 'nNFSe')
                data['nome_municipio'] = _t(root, 'xLocEmi', 'xLocPrestacao', 'xLocIncid') or ''

                emit = root.find('.//emit')
                data['cnpj_prestador'] = _t(emit or root, 'CNPJ')
                data['uf']             = _t(emit or root, 'UF') or ''

                toma = root.find('.//toma')
                if toma is not None:
                    data['cnpj_tomador']         = _t(toma, 'CNPJ', 'CPF', 'NIF')
                    data['razao_social_tomador']  = _t(toma, 'xNome')
                else:
                    data['cnpj_tomador']         = None
                    data['razao_social_tomador']  = None

                data['data_emissao']     = _t(root, 'dhEmi')
                data['data_competencia'] = _t(root, 'dCompet') or data['data_emissao']
                data['codigo_municipio'] = _t(root, 'cLocEmi', 'cLocIncid', 'cMun') or '0000000'

                # Valores do infNFSe (acima do DPS): vBC, pAliqAplic, vISSQN, vLiq
                nfse_vals = root.find('.//infNFSe/valores') or root.find('.//valores')
                if nfse_vals is not None:
                    data['valor_servico'] = _tf(nfse_vals, 'vBC')
                    data['aliquota_iss']  = _tf(nfse_vals, 'pAliqAplic')
                    data['valor_iss']     = _tf(nfse_vals, 'vISSQN')
                    data['valor_liquido'] = _tf(nfse_vals, 'vLiq')
                else:
                    data['valor_servico'] = _tf(root, 'vBC', 'vServ')
                    data['aliquota_iss']  = _tf(root, 'pAliqAplic')
                    data['valor_iss']     = _tf(root, 'vISSQN')
                    data['valor_liquido'] = _tf(root, 'vLiq')

                data['valor_deducoes'] = _tf(root, 'vDeducao')
                data['tp_ret_issqn']   = _t(root, 'tpRetISSQN')
                data['discriminacao']  = (_t(root, 'xDescServ', 'xDesc') or '')[:500]

            # ABRASF / prefeitura
            else:
                data['numero_nfse']          = _t(root, 'Numero', 'NumeroNFSe', 'NfseNumero')
                data['situacao']             = _t(root, 'SituacaoNfse', 'Situacao') or 'NORMAL'
                data['data_emissao']         = _t(root, 'DataEmissaoNfse', 'DataEmissao')
                data['data_competencia']     = _t(root, 'Competencia', 'DataCompetencia') or data['data_emissao']
                data['cnpj_prestador']       = _t(root, 'CnpjPrestador')
                data['cnpj_tomador']         = _t(root, 'CnpjTomador')
                data['razao_social_tomador'] = _t(root, 'RazaoSocialTomador', 'RazaoSocial')
                data['valor_servico']        = _tf(root, 'ValorServicos', 'ValorServico')
                data['valor_deducoes']       = _tf(root, 'ValorDeducoes')
                data['valor_iss']            = _tf(root, 'ValorIss', 'ValorISSQN')
                data['aliquota_iss']         = _tf(root, 'Aliquota')
                data['valor_liquido']        = _tf(root, 'ValorLiquidoNfse', 'ValorLiquido')
                data['tp_ret_issqn']         = _t(root, 'TipoRecolhimento', 'RetencaoIssqn')
                data['codigo_municipio']     = _t(root, 'CodigoMunicipio', 'CodigoMun') or '0000000'
                data['nome_municipio']       = _t(root, 'Municipio', 'NomeMunicipio') or ''
                data['uf']                   = _t(root, 'Uf', 'UF') or ''
                data['discriminacao']        = (_t(root, 'Discriminacao', 'xDescServ') or '')[:500]

            # Normalizar datas -> YYYY-MM-DD
            for campo in ('data_emissao', 'data_competencia'):
                v = data.get(campo)
                if v:
                    data[campo] = v[:10]
            return data

        def _salvar_nfse(cur, nd, xml_text):
            nonlocal importadas, atualizadas
            cur.execute("""
                INSERT INTO nfse_baixadas
                    (empresa_id, numero_nfse, cnpj_prestador, cnpj_tomador,
                     razao_social_tomador, data_emissao, data_competencia,
                     valor_servico, valor_deducoes, valor_iss, aliquota_iss, valor_liquido,
                     situacao, tp_ret_issqn, codigo_municipio, nome_municipio, uf,
                     discriminacao, xml_content, provedor)
                VALUES
                    (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s,%s, %s,%s,%s,%s,%s, %s,%s,%s)
                ON CONFLICT (numero_nfse, codigo_municipio)
                DO UPDATE SET
                    cnpj_prestador       = EXCLUDED.cnpj_prestador,
                    cnpj_tomador         = EXCLUDED.cnpj_tomador,
                    razao_social_tomador = EXCLUDED.razao_social_tomador,
                    data_emissao         = EXCLUDED.data_emissao,
                    data_competencia     = EXCLUDED.data_competencia,
                    valor_servico        = EXCLUDED.valor_servico,
                    valor_deducoes       = EXCLUDED.valor_deducoes,
                    valor_iss            = EXCLUDED.valor_iss,
                    aliquota_iss         = EXCLUDED.aliquota_iss,
                    valor_liquido        = EXCLUDED.valor_liquido,
                    situacao             = EXCLUDED.situacao,
                    tp_ret_issqn         = EXCLUDED.tp_ret_issqn,
                    discriminacao        = EXCLUDED.discriminacao,
                    xml_content          = EXCLUDED.xml_content,
                    atualizado_em        = CURRENT_TIMESTAMP
            """, (
                empresa_id,
                nd['numero_nfse'],
                nd.get('cnpj_prestador'),
                nd.get('cnpj_tomador'),
                nd.get('razao_social_tomador'),
                nd.get('data_emissao'),
                nd.get('data_competencia'),
                nd.get('valor_servico', 0),
                nd.get('valor_deducoes', 0),
                nd.get('valor_iss', 0),
                nd.get('aliquota_iss', 0),
                nd.get('valor_liquido', 0),
                nd.get('situacao', 'NORMAL'),
                nd.get('tp_ret_issqn'),
                nd.get('codigo_municipio', '0000000'),
                nd.get('nome_municipio', ''),
                nd.get('uf', ''),
                nd.get('discriminacao', ''),
                xml_text,
                'xml_importado',
            ))
            if cur.statusmessage == 'INSERT 0 1':
                importadas += 1
            else:
                atualizadas += 1

        # Tomadores capturados para auto-cadastro: {cnpj: razao_social}
        tomadores_capturados = {}

        def _processar_xml_bytes(cur, xml_bytes, fname):
            xml_text = xml_bytes.decode('utf-8', errors='replace')
            nd = _parse_nfse(xml_text)
            if not nd.get('numero_nfse'):
                raise ValueError('Numero NFS-e nao encontrado no XML')
            _salvar_nfse(cur, nd, xml_text)
            # Registrar tomador para auto-cadastro posterior
            cnpj_t = (nd.get('cnpj_tomador') or '').strip()
            nome_t = (nd.get('razao_social_tomador') or '').strip()
            if cnpj_t and cnpj_t not in tomadores_capturados:
                tomadores_capturados[cnpj_t] = nome_t

        def _consultar_brasilapi(cnpj_limpo):
            """Consulta BrasilAPI para obter dados do CNPJ. Retorna dict ou None."""
            import urllib.request
            import urllib.error
            import json as _json
            try:
                url = f'https://brasilapi.com.br/api/cnpj/v1/{cnpj_limpo}'
                req = urllib.request.Request(url, headers={'User-Agent': 'SistemaFinanceiro/1.0'})
                with urllib.request.urlopen(req, timeout=6) as resp:
                    return _json.loads(resp.read().decode('utf-8'))
            except Exception:
                return None

        def _formatar_cnpj(cnpj):
            c = ''.join(filter(str.isdigit, cnpj))
            if len(c) == 14:
                return f'{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}'
            if len(c) == 11:
                return f'{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}'
            return cnpj

        def _auto_cadastrar_tomadores(cur_reg):
            """Para cada tomador capturado, verifica se já existe como cliente e insere se não."""
            registrados = 0
            for cnpj_t, nome_xml in tomadores_capturados.items():
                cnpj_digits = ''.join(filter(str.isdigit, cnpj_t))
                # Padroniza: formata apenas se 11 ou 14 dígitos
                cnpj_fmt = _formatar_cnpj(cnpj_t) if len(cnpj_digits) in (11, 14) else cnpj_t
                try:
                    # Verifica se já existe na tabela clientes (qualquer formato)
                    cur_reg.execute("""
                        SELECT id FROM clientes
                        WHERE empresa_id = %s
                          AND REGEXP_REPLACE(cpf_cnpj, '[^0-9]', '', 'g') = %s
                    """, (empresa_id, cnpj_digits))
                    if cur_reg.fetchone():
                        continue  # já cadastrado

                    # Tentar enriquecer via BrasilAPI (apenas CNPJ 14 dígitos)
                    nome_final = nome_xml or cnpj_fmt
                    email_final = None
                    telefone_final = None
                    endereco_final = None

                    if len(cnpj_digits) == 14:
                        dados = _consultar_brasilapi(cnpj_digits)
                        if dados and not dados.get('message'):  # message = erro da API
                            nome_final = (
                                dados.get('razao_social') or
                                dados.get('nome_fantasia') or
                                nome_xml or cnpj_fmt
                            ).strip()
                            email_final = dados.get('email') or None
                            ddd   = dados.get('ddd_telefone_1') or ''
                            tel   = dados.get('telefone') or ''
                            if ddd and tel:
                                telefone_final = f'({ddd}) {tel}'
                            # Montar endereço
                            partes = []
                            for campo in ('logradouro', 'numero', 'complemento', 'bairro', 'municipio'):
                                v = (dados.get(campo) or '').strip()
                                if v:
                                    partes.append(v)
                            uf  = (dados.get('uf') or '').strip()
                            cep = (dados.get('cep') or '').strip()
                            if uf:
                                partes.append(uf)
                            if cep:
                                partes.append(f'CEP: {cep}')
                            if partes:
                                endereco_final = ', '.join(partes)

                    cur_reg.execute("""
                        INSERT INTO clientes (nome, cpf_cnpj, email, telefone, endereco, empresa_id)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (nome_final, cnpj_fmt, email_final, telefone_final, endereco_final, empresa_id))
                    registrados += 1
                    logger.info(f"[NFS-e] Novo cliente auto-cadastrado: {nome_final} ({cnpj_fmt})")
                except Exception as e_cli:
                    logger.warning(f"[NFS-e] Nao foi possivel auto-cadastrar {cnpj_t}: {e_cli}")
            return registrados

        try:
            cur = conn.cursor()
            for f in files:
                fname = f.filename or 'desconhecido'
                file_bytes = f.read()
                try:
                    # ZIP: extrai e processa cada XML interno
                    if fname.lower().endswith('.zip') or file_bytes[:2] == b'PK':
                        with _zf.ZipFile(_io.BytesIO(file_bytes)) as zf:
                            xml_names = [
                                n for n in zf.namelist()
                                if n.lower().endswith('.xml')
                                and not n.startswith('__')
                                and not n.startswith('.')
                            ]
                            if not xml_names:
                                erros.append({'arquivo': fname,
                                              'erro': 'Nenhum XML encontrado dentro do ZIP'})
                                continue
                            for xml_name in xml_names:
                                try:
                                    _processar_xml_bytes(cur, zf.read(xml_name), xml_name)
                                except Exception as ex:
                                    erros.append({'arquivo': f'{fname}/{xml_name}',
                                                  'erro': str(ex)})
                    # XML direto
                    else:
                        _processar_xml_bytes(cur, file_bytes, fname)

                except _zf.BadZipFile:
                    erros.append({'arquivo': fname,
                                  'erro': 'Arquivo ZIP invalido ou corrompido'})
                except Exception as e_inner:
                    erros.append({'arquivo': fname, 'erro': str(e_inner)})

            conn.commit()

            # Auto-cadastrar novos clientes (tomadores) detectados nos XMLs
            novos_clientes = 0
            try:
                novos_clientes = _auto_cadastrar_tomadores(cur)
                conn.commit()
            except Exception as e_ac:
                logger.warning(f"[NFS-e] Erro no auto-cadastro de clientes: {e_ac}")
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Reconciliação automática com Contas a Receber
        reconcil = {'reconciliadas': 0}
        try:
            reconcil = _reconciliar_nfse_lancamentos(empresa_id)
        except Exception as e_rec:
            logger.warning(f"[NFS-e] Erro na reconciliação automática: {e_rec}")

        return jsonify({
            'success': True,
            'importadas': importadas,
            'atualizadas': atualizadas,
            'total': importadas + atualizadas,
            'novos_clientes': novos_clientes,
            'reconciliadas': reconcil.get('reconciliadas', 0),
            'erros': erros
        })

    except Exception as e:
        logger.error(f"Erro ao importar XMLs NFS-e: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# RECONCILIAÇÃO NFS-e ↔ CONTAS A RECEBER
# ============================================================================

def _reconciliar_nfse_lancamentos(empresa_id, conn_nfse=None, conn_main=None):  # noqa: conn_nfse/conn_main unused, kept for compat
    """
    Reconcilia automaticamente NFS-es com lançamentos de receita.

    Lógica de matching (por empresa_id):
      - Para cada NFS-e sem situacao_recebimento='PAGO':
        1. Busca lancamentos tipo='receita' com |valor - valor_liquido| <= 0.02
        2. Filtra pelo CNPJ/CPF do tomador OU pelo nome (razao_social_tomador)
           comparando com lancamento.pessoa (sem maiúsculas/pontuação)
        3. Ao encontrar: atualiza lancamento.numero_documento = 'NF {numero_nfse}'
                         e nfse_baixadas.data_pagamento = lancamento.data_vencimento
                         e nfse_baixadas.situacao_recebimento = 'PAGO'

    Retorna: dict com contagens
    """
    import psycopg2
    import re as _re

    def _digits(s):
        return _re.sub(r'\D', '', s or '')

    def _normalizar(s):
        return _re.sub(r'\s+', ' ', _re.sub(r'[^a-z0-9 ]', '', (s or '').lower())).strip()

    resultado = {'reconciliadas': 0, 'ja_pagas': 0}

    try:
        # Conexão NFS-e (psycopg2 direto)
        from database_postgresql import get_nfse_db_params
        nfse_params = get_nfse_db_params()
        conn_nfse_local = psycopg2.connect(**nfse_params)

        # Conexão principal via get_db_connection (context manager)
        with get_db_connection(empresa_id=empresa_id) as conn_main_local:
            cur_nfse = conn_nfse_local.cursor()
            cur_main = conn_main_local.cursor()

            try:
                # Buscar NFS-es ainda não pagas (ou sem situacao_recebimento)
                cur_nfse.execute("""
                    SELECT id, numero_nfse, valor_liquido, cnpj_tomador, razao_social_tomador
                    FROM nfse_baixadas
                    WHERE empresa_id = %s
                      AND situacao = 'NORMAL'
                      AND (situacao_recebimento IS NULL OR situacao_recebimento <> 'PAGO')
                      AND valor_liquido > 0
                """, (empresa_id,))
            nfses = cur_nfse.fetchall()

            for nfse_row in nfses:
                nfse_id, numero_nfse, valor_liq, cnpj_tom, nome_tom = nfse_row
                valor_liq = float(valor_liq or 0)

                cnpj_digits = _digits(cnpj_tom)
                nome_norm   = _normalizar(nome_tom)

                # Buscar receitas com valor próximo (±2 centavos)
                cur_main.execute("""
                    SELECT id, pessoa, descricao, data_vencimento, numero_documento
                    FROM lancamentos
                    WHERE empresa_id = %s
                      AND tipo = 'receita'
                      AND ABS(valor - %s) <= 0.02
                """, (empresa_id, valor_liq))
                candidatos = cur_main.fetchall()

                lancamento_match = None
                for lanc_id, pessoa, descricao, data_venc, num_doc in candidatos:
                    # Ignora se já tem numero_documento preenchido com outra NF
                    if num_doc and num_doc.strip() and num_doc.strip() != f'NF {numero_nfse}':
                        continue

                    desc_str      = (descricao or '')
                    desc_digits   = _digits(desc_str)
                    desc_norm     = _normalizar(desc_str)
                    pessoa_digits = _digits(pessoa)
                    pessoa_norm   = _normalizar(pessoa)

                    # 1. Match por CNPJ/CPF na descrição (principal - ex: "...PIX_CRED 52177416000183...")
                    if cnpj_digits and len(cnpj_digits) >= 11 and cnpj_digits in desc_digits:
                        lancamento_match = (lanc_id, data_venc)
                        break

                    # 2. Match por CNPJ/CPF no campo pessoa
                    if cnpj_digits and len(cnpj_digits) >= 11 and cnpj_digits == pessoa_digits:
                        lancamento_match = (lanc_id, data_venc)
                        break

                    # 3. Match por nome do tomador na descrição
                    if nome_norm and len(nome_norm) >= 6 and nome_norm in desc_norm:
                        lancamento_match = (lanc_id, data_venc)
                        break

                    # 4. Match por nome normalizado no campo pessoa
                    if nome_norm and len(nome_norm) >= 6 and pessoa_norm:
                        s1, s2 = sorted([nome_norm, pessoa_norm], key=len)
                        if len(s1) >= 6 and s1 in s2:
                            lancamento_match = (lanc_id, data_venc)
                            break

                if lancamento_match:
                    lanc_id, data_venc = lancamento_match
                    num_doc_novo = f'NF {numero_nfse}'

                    # Atualizar lancamento: preencher numero_documento
                    cur_main.execute("""
                        UPDATE lancamentos
                        SET numero_documento = %s, associacao = %s
                        WHERE id = %s AND empresa_id = %s
                    """, (num_doc_novo, num_doc_novo, lanc_id, empresa_id))

                    # Atualizar NFS-e: data_pagamento + situacao_recebimento
                    cur_nfse.execute("""
                        UPDATE nfse_baixadas
                        SET data_pagamento         = %s,
                            situacao_recebimento   = 'PAGO'
                        WHERE id = %s AND empresa_id = %s
                    """, (data_venc, nfse_id, empresa_id))

                    resultado['reconciliadas'] += 1
                    logger.info(f"[Reconciliar] NF {numero_nfse} ↔ lancamento {lanc_id} | data_pg={data_venc}")

                conn_main_local.commit()
                conn_nfse_local.commit()
            finally:
                cur_nfse.close()
                cur_main.close()

        try:
            conn_nfse_local.close()
        except Exception:
            pass

    except Exception as e:
        logger.error(f"[Reconciliar NFS-e] Erro: {e}")
        import traceback
        traceback.print_exc()

    return resultado


@app.route('/api/nfse/reconciliar', methods=['POST'])
@require_auth
@require_permission('nfse_view')
def reconciliar_nfse():
    """Dispara reconciliação manual entre NFS-e e Contas a Receber"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa não selecionada'}), 403

        result = _reconciliar_nfse_lancamentos(empresa_id)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/nfse/resumo-mensal', methods=['POST'])
@require_auth
@require_permission('nfse_view')
def get_resumo_mensal_nfse():
    """Retorna resumo mensal de NFS-e"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        ano = data.get('ano')
        mes = data.get('mes')
        
        if not ano or not mes:
            return jsonify({
                'success': False,
                'error': 'Ano e m�s s�o obrigat�rios'
            }), 400
        
        from nfse_functions import get_resumo_mensal
        
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        resumo = get_resumo_mensal(db_params, empresa_id, ano, mes)
        
        # Converter Decimal para float
        for key, value in resumo.items():
            if isinstance(value, Decimal):
                resumo[key] = float(value)
        
        return jsonify({
            'success': True,
            'resumo': resumo
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar resumo mensal: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/export/excel', methods=['POST'])
@require_auth
@require_permission('nfse_export')
def export_nfse_excel():
    """Exporta NFS-e para Excel/CSV"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        from nfse_functions import exportar_nfse_excel
        from datetime import datetime
        from database_postgresql import get_nfse_db_params
        import tempfile
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        # Converter datas
        data_inicial = datetime.strptime(data['data_inicial'], '%Y-%m-%d').date()
        data_final = datetime.strptime(data['data_final'], '%Y-%m-%d').date()
        
        # Criar arquivo tempor�rio
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        caminho_arquivo = temp_file.name
        temp_file.close()
        
        # Exportar
        sucesso, erro = exportar_nfse_excel(
            db_params=db_params,
            empresa_id=empresa_id,
            data_inicial=data_inicial,
            data_final=data_final,
            caminho_arquivo=caminho_arquivo
        )
        
        if sucesso:
            # Log de auditoria
            from nfse_functions import registrar_operacao
            registrar_operacao(
                db_params=db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='EXPORT',
                detalhes={'formato': 'CSV'},
                ip_address=request.remote_addr
            )
            
            return send_file(
                caminho_arquivo,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'nfse_{data_inicial}_{data_final}.csv'
            )
        else:
            return jsonify({
                'success': False,
                'error': erro
            }), 400
            
    except Exception as e:
        logger.error(f"Erro ao exportar NFS-e: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/nfse/export/xml', methods=['POST'])
@require_auth
@require_permission('nfse_export')
def export_nfse_xml():
    """Exporta XMLs de NFS-e para arquivo ZIP"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({
                'success': False,
                'error': 'Empresa n�o selecionada'
            }), 400
        
        data = request.json
        
        from nfse_functions import exportar_xmls_zip
        from datetime import datetime
        from database_postgresql import get_nfse_db_params
        import tempfile
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        
        # Converter datas
        data_inicial = datetime.strptime(data['data_inicial'], '%Y-%m-%d').date()
        data_final = datetime.strptime(data['data_final'], '%Y-%m-%d').date()
        
        # Criar arquivo tempor�rio
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        caminho_arquivo = temp_file.name
        temp_file.close()
        
        # Exportar
        sucesso, erro = exportar_xmls_zip(
            db_params=db_params,
            empresa_id=empresa_id,
            data_inicial=data_inicial,
            data_final=data_final,
            caminho_arquivo=caminho_arquivo
        )
        
        if sucesso:
            # Log de auditoria
            from nfse_functions import registrar_operacao
            registrar_operacao(
                db_params=db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='EXPORT',
                detalhes={'formato': 'XML_ZIP'},
                ip_address=request.remote_addr
            )
            
            return send_file(
                caminho_arquivo,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'nfse_xmls_{data_inicial}_{data_final}.zip'
            )
        else:
            return jsonify({
                'success': False,
                'error': erro
            }), 400
            
    except Exception as e:
        logger.error(f"Erro ao exportar XMLs: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# ROTAS NFS-e - CERTIFICADO DIGITAL A1
# ============================================================================

@app.route('/api/nfse/certificado/upload', methods=['POST'])
@require_auth
@require_permission('nfse_config')
def upload_certificado_nfse():
    """Upload de certificado digital A1 (.pfx/.p12)"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        # Verificar se o arquivo foi enviado
        if 'certificado' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo de certificado enviado'}), 400
        
        arquivo = request.files['certificado']
        senha = request.form.get('senha', '')
        
        if not arquivo.filename:
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400
        
        if not senha:
            return jsonify({'success': False, 'error': 'Senha do certificado � obrigat�ria'}), 400
        
        # Validar extens�o
        extensao = arquivo.filename.rsplit('.', 1)[-1].lower() if '.' in arquivo.filename else ''
        if extensao not in ('pfx', 'p12'):
            return jsonify({'success': False, 'error': 'Formato inv�lido. Use arquivo .pfx ou .p12'}), 400
        
        # Ler bytes do arquivo
        pfx_bytes = arquivo.read()
        
        if len(pfx_bytes) == 0:
            return jsonify({'success': False, 'error': 'Arquivo vazio'}), 400
        
        if len(pfx_bytes) > 10 * 1024 * 1024:  # 10MB max
            return jsonify({'success': False, 'error': 'Arquivo muito grande (m�ximo 10MB)'}), 400
        
        from nfse_functions import upload_certificado, registrar_operacao
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()
        # Remover 'dsn' se existir (NFSeDatabase espera par�metros individuais)
        
        sucesso, info, erro = upload_certificado(db_params, empresa_id, pfx_bytes, senha)
        
        if sucesso:
            # Log de auditoria
            registrar_operacao(
                db_params=db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='CONFIG',
                detalhes={
                    'acao': 'upload_certificado',
                    'cnpj': info.get('cnpj'),
                    'municipio': info.get('nome_municipio'),
                    'validade_fim': info.get('validade_fim'),
                    'config_criada': info.get('config_criada', False)
                },
                ip_address=request.remote_addr
            )
            
            # Mensagem personalizada se criou configura��o automaticamente
            message = 'Certificado carregado com sucesso!'
            if info.get('config_criada'):
                message += f' Munic�pio {info.get("nome_municipio")} configurado automaticamente. Complete a Inscri��o Municipal em Configura��es.'
            elif info.get('codigo_municipio'):
                message += ' Lembre-se de configurar o munic�pio em Configura��es.'
            
            return jsonify({
                'success': True,
                'message': message,
                'certificado': {
                    'id': info.get('cert_id'),
                    'cnpj': info.get('cnpj'),
                    'razao_social': info.get('razao_social'),
                    'emitente': info.get('emitente'),
                    'validade_inicio': info.get('validade_inicio'),
                    'validade_fim': info.get('validade_fim'),
                    'codigo_municipio': info.get('codigo_municipio'),
                    'nome_municipio': info.get('nome_municipio'),
                    'uf': info.get('uf'),
                    'config_criada': info.get('config_criada', False),
                    'config_id': info.get('config_id')
                }
            })
        else:
            return jsonify({'success': False, 'error': erro}), 400
        
    except Exception as e:
        logger.error(f"Erro ao fazer upload de certificado: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/nfse/certificado', methods=['GET'])
@require_auth
def get_certificado_nfse():
    """Retorna informa��es do certificado ativo da empresa"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        from nfse_functions import get_certificado_info
        from database_postgresql import get_nfse_db_params
        
        # Usar configura��o centralizada do banco
        db_params = get_nfse_db_params()  # Remover DSN se existir
        
        cert = get_certificado_info(db_params, empresa_id)
        
        return jsonify({
            'success': True,
            'certificado': cert
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar certificado: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/nfse/certificado/<int:cert_id>', methods=['DELETE'])
@require_auth
@require_permission('nfse_config')
def delete_certificado_nfse(cert_id):
    """Remove certificado digital"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        from nfse_functions import excluir_certificado_empresa, registrar_operacao
        from database_postgresql import get_nfse_db_params
        
        db_params = get_nfse_db_params()
        
        sucesso = excluir_certificado_empresa(db_params, cert_id)
        
        if sucesso:
            registrar_operacao(
                db_params=db_params,
                empresa_id=empresa_id,
                usuario_id=usuario['id'],
                operacao='CONFIG',
                detalhes={'acao': 'excluir_certificado', 'cert_id': cert_id},
                ip_address=request.remote_addr
            )
            return jsonify({'success': True, 'message': 'Certificado removido'})
        else:
            return jsonify({'success': False, 'error': 'Certificado n�o encontrado'}), 404
        
    except Exception as e:
        logger.error(f"Erro ao excluir certificado: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ROTAS NFS-e - GERA��O DE PDF (DANFSE)
# ============================================================================

@app.route('/api/nfse/<int:nfse_id>/pdf', methods=['GET'])
@require_auth
@require_permission('nfse_view')
def gerar_pdf_nfse_route(nfse_id):
    """Gera e retorna o PDF (DANFSE) de uma NFS-e"""
    try:
        logger.info(f"?? API /api/nfse/{nfse_id}/pdf chamada")
        
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        from nfse_functions import gerar_pdf_nfse
        from database_postgresql import get_nfse_db_params
        
        db_params = get_nfse_db_params()

        # Resolve o número real da NFS-e para o nome do arquivo
        numero_nfse_str = str(nfse_id)  # fallback = db id
        try:
            from nfse_functions import NFSeDatabase
            with NFSeDatabase(db_params) as _db_tmp:
                _nfse_row = _db_tmp.get_nfse_by_id(nfse_id)
            if _nfse_row and _nfse_row.get('numero_nfse'):
                numero_nfse_str = str(_nfse_row['numero_nfse'])
        except Exception:
            pass
        
        logger.info(f"?? Chamando gerar_pdf_nfse(nfse_id={nfse_id})...")
        pdf_bytes = gerar_pdf_nfse(db_params, nfse_id)
        
        if pdf_bytes:
            logger.info(f"? PDF gerado com sucesso: {len(pdf_bytes):,} bytes")
            
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            temp_file.write(pdf_bytes)
            temp_file.close()
            
            logger.info(f"?? Enviando PDF para cliente: {temp_file.name}")
            
            return send_file(
                temp_file.name,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'nfse_{numero_nfse_str}.pdf'
            )
        else:
            logger.error(f"? gerar_pdf_nfse retornou None")
            return jsonify({
                'success': False,
                'error': 'N�o foi poss�vel gerar o PDF desta NFS-e'
            }), 400
        
    except Exception as e:
        logger.error(f"? Erro ao gerar PDF NFS-e: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS CONTABILIDADE - PLANO DE CONTAS
# ============================================================================

@app.route('/api/contabilidade/versoes', methods=['GET'])
@require_auth
def listar_versoes_plano():
    """Lista vers�es do plano de contas"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        from contabilidade_functions import listar_versoes
        versoes = listar_versoes(empresa_id)
        
        logger.info(f"?? Vers�es para empresa {empresa_id}: {len(versoes)} encontrada(s)")
        if versoes:
            logger.info(f"   ?? Primeira vers�o: {versoes[0]}")
        else:
            logger.warning(f"   ?? Nenhuma vers�o encontrada para empresa {empresa_id}")
        
        return jsonify({'success': True, 'versoes': versoes})
    except Exception as e:
        logger.error(f"Erro ao listar vers�es: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/versoes', methods=['POST'])
@require_auth
def criar_versao_plano():
    """Cria nova vers�o do plano de contas"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json()
        if not dados or not dados.get('nome_versao') or not dados.get('exercicio_fiscal'):
            return jsonify({'success': False, 'error': 'nome_versao e exercicio_fiscal s�o obrigat�rios'}), 400
        
        from contabilidade_functions import criar_versao
        versao_id = criar_versao(empresa_id, dados)
        return jsonify({'success': True, 'id': versao_id, 'message': 'Vers�o criada com sucesso'})
    except Exception as e:
        logger.error(f"Erro ao criar vers�o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/versoes/<int:versao_id>', methods=['PUT'])
@require_auth
def atualizar_versao_plano(versao_id):
    """Atualiza vers�o do plano de contas"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json()
        from contabilidade_functions import atualizar_versao
        atualizar_versao(empresa_id, versao_id, dados)
        return jsonify({'success': True, 'message': 'Vers�o atualizada'})
    except Exception as e:
        logger.error(f"Erro ao atualizar vers�o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/versoes/<int:versao_id>', methods=['DELETE'])
@require_auth
def excluir_versao_plano(versao_id):
    """Exclui vers�o do plano de contas"""
    try:
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        from contabilidade_functions import excluir_versao
        excluir_versao(empresa_id, versao_id)
        return jsonify({'success': True, 'message': 'Vers�o exclu�da'})
    except Exception as e:
        logger.error(f"Erro ao excluir vers�o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas', methods=['GET'])
@require_auth
def listar_plano_contas():
    """Lista contas do plano com filtros"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        versao_id = request.args.get('versao_id', type=int)
        classificacao = request.args.get('classificacao')
        tipo_conta = request.args.get('tipo_conta')
        busca = request.args.get('busca')
        
        from contabilidade_functions import listar_contas
        contas = listar_contas(empresa_id, versao_id=versao_id, classificacao=classificacao,
                               tipo_conta=tipo_conta, busca=busca)
        return jsonify({'success': True, 'contas': contas, 'total': len(contas)})
    except Exception as e:
        logger.error(f"Erro ao listar plano de contas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/tree', methods=['GET'])
@require_auth
def arvore_plano_contas():
    """Retorna plano de contas em estrutura de �rvore"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        versao_id = request.args.get('versao_id', type=int)
        if not versao_id:
            # Tentar vers�o ativa
            from contabilidade_functions import obter_versao_ativa
            versao_ativa = obter_versao_ativa(empresa_id)
            if versao_ativa:
                versao_id = versao_ativa['id']
            else:
                return jsonify({'success': True, 'tree': [], 'message': 'Nenhuma vers�o encontrada'})
        
        from contabilidade_functions import obter_arvore_contas
        tree = obter_arvore_contas(empresa_id, versao_id)
        return jsonify({'success': True, 'tree': tree, 'versao_id': versao_id})
    except Exception as e:
        logger.error(f"Erro ao obter �rvore: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas', methods=['POST'])
@require_auth
def criar_conta_plano():
    """Cria nova conta no plano"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json()
        if not dados or not dados.get('codigo') or not dados.get('descricao'):
            return jsonify({'success': False, 'error': 'c�digo e descri��o s�o obrigat�rios'}), 400
        if not dados.get('versao_id'):
            return jsonify({'success': False, 'error': 'versao_id � obrigat�rio'}), 400
        if not dados.get('classificacao'):
            return jsonify({'success': False, 'error': 'classificacao � obrigat�ria'}), 400
        
        from contabilidade_functions import criar_conta
        conta_id = criar_conta(empresa_id, dados)
        return jsonify({'success': True, 'id': conta_id, 'message': 'Conta criada com sucesso'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        error_msg = str(e)
        # Tratar erros de constraint de banco de dados
        if 'unique constraint' in error_msg.lower() or 'duplicate key' in error_msg.lower():
            if 'codigo' in error_msg:
                return jsonify({'success': False, 'error': f'O c�digo "{dados.get("codigo")}" j� existe nesta vers�o do plano de contas'}), 400
            return jsonify({'success': False, 'error': 'J� existe um registro com estes dados'}), 400
        logger.error(f"Erro ao criar conta: {e}")
        return jsonify({'success': False, 'error': 'Erro ao criar conta. Verifique os dados e tente novamente.'}), 500


@app.route('/api/contabilidade/plano-contas/<int:conta_id>', methods=['PUT'])
@require_auth
def atualizar_conta_plano(conta_id):
    """Atualiza conta do plano"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json()
        from contabilidade_functions import atualizar_conta
        atualizar_conta(empresa_id, conta_id, dados)
        return jsonify({'success': True, 'message': 'Conta atualizada'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        error_msg = str(e)
        # Tratar erros de constraint de banco de dados
        if 'unique constraint' in error_msg.lower() or 'duplicate key' in error_msg.lower():
            if 'codigo' in error_msg:
                return jsonify({'success': False, 'error': f'O c�digo "{dados.get("codigo")}" j� existe nesta vers�o do plano de contas'}), 400
            return jsonify({'success': False, 'error': 'J� existe um registro com estes dados'}), 400
        logger.error(f"Erro ao atualizar conta: {e}")
        return jsonify({'success': False, 'error': 'Erro ao atualizar conta. Verifique os dados e tente novamente.'}), 500


@app.route('/api/contabilidade/plano-contas/<int:conta_id>', methods=['DELETE'])
@require_auth
def excluir_conta_plano(conta_id):
    """Soft delete da conta e subcontas"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        from contabilidade_functions import excluir_conta
        deleted = excluir_conta(empresa_id, conta_id)
        return jsonify({'success': True, 'message': f'{deleted} conta(s) exclu�da(s)'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Erro ao excluir conta: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/mover', methods=['POST'])
@require_auth
def mover_conta_plano():
    """Move conta para outro pai (drag-and-drop)"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json()
        conta_id = dados.get('conta_id')
        novo_parent_id = dados.get('novo_parent_id')  # None = raiz
        
        from contabilidade_functions import mover_conta
        mover_conta(empresa_id, conta_id, novo_parent_id)
        return jsonify({'success': True, 'message': 'Conta movida com sucesso'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Erro ao mover conta: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/import', methods=['POST'])
@require_auth
def importar_plano_contas():
    """Importa contas de CSV"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json()
        versao_id = dados.get('versao_id')
        linhas = dados.get('linhas', [])
        
        if not versao_id or not linhas:
            return jsonify({'success': False, 'error': 'versao_id e linhas s�o obrigat�rios'}), 400
        
        from contabilidade_functions import importar_contas_csv
        resultado = importar_contas_csv(empresa_id, versao_id, linhas)
        return jsonify({'success': True, **resultado})
    except Exception as e:
        logger.error(f"Erro ao importar: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/export', methods=['GET'])
@require_auth
def exportar_plano_contas():
    """Exporta contas em JSON (para CSV/Excel no frontend)"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        versao_id = request.args.get('versao_id', type=int)
        if not versao_id:
            return jsonify({'success': False, 'error': 'versao_id � obrigat�rio'}), 400
        
        from contabilidade_functions import exportar_contas
        contas = exportar_contas(empresa_id, versao_id)
        return jsonify({'success': True, 'contas': contas})
    except Exception as e:
        logger.error(f"Erro ao exportar: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/importar-padrao', methods=['POST'])
@require_auth
def importar_plano_padrao_route():
    """Importa o plano de contas padr�o para a empresa"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        dados = request.get_json() or {}
        ano_fiscal = dados.get('ano_fiscal')
        
        from contabilidade_functions import importar_plano_padrao
        resultado = importar_plano_padrao(empresa_id, ano_fiscal)
        
        # Se a fun��o j� retorna success, apenas retornar o resultado
        if resultado.get('success'):
            return jsonify(resultado)
        
        # Formato antigo (retrocompatibilidade)
        return jsonify({
            'success': True,
            'versao_id': resultado.get('versao_id'),
            'contas_importadas': resultado.get('contas_importadas', resultado.get('contas_criadas', 0)),
            'erros': resultado.get('erros', []),
            'message': resultado.get('message')
        })
    except Exception as e:
        logger.error(f"Erro ao importar plano padr�o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/exportar-speed', methods=['GET'])
@require_auth
def exportar_plano_speed():
    """Exporta plano de contas no formato Speed (TXT)"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        versao_id = request.args.get('versao_id', type=int)
        if not versao_id:
            return jsonify({'success': False, 'error': 'versao_id � obrigat�rio'}), 400
        
        from contabilidade_functions import listar_contas
        from speed_integration import exportar_plano_contas_speed, estatisticas_mapeamento
        
        contas = listar_contas(empresa_id, versao_id=versao_id)
        
        if not contas:
            return jsonify({'success': False, 'error': 'Nenhuma conta encontrada'}), 404
        
        # Gerar arquivo TXT
        conteudo_txt = exportar_plano_contas_speed(contas)
        
        # Estat�sticas
        stats = estatisticas_mapeamento(contas)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo_txt,
            'formato': 'txt',
            'total_contas': len(contas),
            'estatisticas': stats
        })
    except Exception as e:
        logger.error(f"Erro ao exportar para Speed: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/contabilidade/plano-contas/mapeamento-referencial', methods=['GET'])
@require_auth
def exportar_mapeamento_referencial():
    """Exporta mapeamento com Referencial Cont�bil (CSV)"""
    try:
        usuario = request.usuario
        empresa_id = usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o selecionada'}), 400
        
        versao_id = request.args.get('versao_id', type=int)
        if not versao_id:
            return jsonify({'success': False, 'error': 'versao_id � obrigat�rio'}), 400
        
        from contabilidade_functions import listar_contas
        from speed_integration import exportar_plano_contas_referencial
        
        contas = listar_contas(empresa_id, versao_id=versao_id)
        
        if not contas:
            return jsonify({'success': False, 'error': 'Nenhuma conta encontrada'}), 404
        
        # Gerar CSV com mapeamento
        conteudo_csv = exportar_plano_contas_referencial(contas)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo_csv,
            'formato': 'csv'
        })
    except Exception as e:
        logger.error(f"Erro ao exportar mapeamento: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# LAN�AMENTOS CONT�BEIS - FASE 2 SPEED
# =============================================================================

@app.route('/api/lancamentos-contabeis', methods=['GET'])
@require_auth
def listar_lancamentos_contabeis():
    """Lista lan�amentos cont�beis com filtros"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        # Par�metros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo_lancamento = request.args.get('tipo_lancamento')
        origem = request.args.get('origem')
        busca = request.args.get('busca')
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))
        
        # Converter datas
        from datetime import datetime
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        # Importar fun��o
        from lancamentos_functions import listar_lancamentos
        
        conn = get_db_connection()
        resultado = listar_lancamentos(
            conn=conn,
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            tipo_lancamento=tipo_lancamento,
            origem=origem,
            busca=busca,
            limit=limit,
            offset=offset
        )
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao listar lan�amentos: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos-contabeis/<int:lancamento_id>', methods=['GET'])
@require_auth
def obter_lancamento_detalhado_api(lancamento_id):
    """Obt�m detalhes completos de um lan�amento"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        from lancamentos_functions import obter_lancamento_detalhado
        
        conn = get_db_connection()
        resultado = obter_lancamento_detalhado(conn, lancamento_id, empresa_id)
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao obter lan�amento: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos-contabeis', methods=['POST'])
@require_auth
def criar_lancamento_api():
    """Cria novo lan�amento cont�bil"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        user_id = user.get('id')
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('data_lancamento'):
            return jsonify({'success': False, 'error': 'Data do lan�amento � obrigat�ria'}), 400
        if not data.get('historico'):
            return jsonify({'success': False, 'error': 'Hist�rico � obrigat�rio'}), 400
        if not data.get('itens') or len(data['itens']) < 2:
            return jsonify({'success': False, 'error': 'Lan�amento deve ter pelo menos 2 itens'}), 400
        
        # Converter data
        from datetime import datetime
        data_lancamento = datetime.strptime(data['data_lancamento'], '%Y-%m-%d').date()
        
        from lancamentos_functions import criar_lancamento
        
        conn = get_db_connection()
        resultado = criar_lancamento(
            conn=conn,
            empresa_id=empresa_id,
            data_lancamento=data_lancamento,
            historico=data['historico'],
            itens=data['itens'],
            tipo_lancamento=data.get('tipo_lancamento', 'manual'),
            origem=data.get('origem'),
            origem_id=data.get('origem_id'),
            versao_plano_id=data.get('versao_plano_id'),
            observacoes=data.get('observacoes'),
            created_by=user_id
        )
        conn.close()
        
        if resultado['success']:
            return jsonify(resultado), 201
        else:
            return jsonify(resultado), 400
    except Exception as e:
        logger.error(f"Erro ao criar lan�amento: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos-contabeis/<int:lancamento_id>/estornar', methods=['POST'])
@require_auth
def estornar_lancamento_api(lancamento_id):
    """Estorna um lan�amento criando lan�amento inverso"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        user_id = user.get('id')
        
        data = request.get_json()
        historico_estorno = data.get('historico_estorno', 'Estorno de lan�amento')
        
        from lancamentos_functions import estornar_lancamento
        
        conn = get_db_connection()
        resultado = estornar_lancamento(
            conn=conn,
            lancamento_id=lancamento_id,
            empresa_id=empresa_id,
            historico_estorno=historico_estorno,
            created_by=user_id
        )
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao estornar lan�amento: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos-contabeis/<int:lancamento_id>', methods=['DELETE'])
@require_auth
def deletar_lancamento_api(lancamento_id):
    """Deleta um lan�amento cont�bil"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        from lancamentos_functions import deletar_lancamento
        
        conn = get_db_connection()
        resultado = deletar_lancamento(conn, lancamento_id, empresa_id)
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao deletar lan�amento: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos-contabeis/estatisticas', methods=['GET'])
@require_auth
def estatisticas_lancamentos_api():
    """Obt�m estat�sticas dos lan�amentos"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        ano = request.args.get('ano', type=int)
        
        from lancamentos_functions import obter_estatisticas_lancamentos
        
        conn = get_db_connection()
        resultado = obter_estatisticas_lancamentos(conn, empresa_id, ano)
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao obter estat�sticas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/lancamentos-contabeis/exportar-speed', methods=['POST'])
@require_auth
def exportar_lancamentos_speed_api():
    """Exporta lan�amentos para formato Speed (TXT ou XML)"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        formato = data.get('formato', 'txt')  # 'txt' ou 'xml'
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        # Converter datas
        from datetime import datetime
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        
        # Buscar lan�amentos com itens detalhados
        from lancamentos_functions import listar_lancamentos, obter_lancamento_detalhado
        
        conn = get_db_connection()
        
        # Listar todos os lan�amentos do per�odo
        resultado_lista = listar_lancamentos(
            conn=conn,
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limit=10000  # Limite alto para exporta��o
        )
        
        if not resultado_lista['success']:
            conn.close()
            return jsonify(resultado_lista), 400
        
        # Buscar detalhes de cada lan�amento (incluindo itens)
        lancamentos_completos = []
        for lanc in resultado_lista['lancamentos']:
            detalhe = obter_lancamento_detalhado(conn, lanc['id'], empresa_id)
            if detalhe['success']:
                lancamentos_completos.append(detalhe['lancamento'])
        
        conn.close()
        
        # Validar antes de exportar
        from speed_integration import validar_lancamentos_exportacao, exportar_lancamentos_speed, exportar_lancamentos_speed_xml
        
        validacao = validar_lancamentos_exportacao(lancamentos_completos)
        
        if not validacao['valido']:
            return jsonify({
                'success': False,
                'error': 'Valida��o falhou',
                'validacao': validacao
            }), 400
        
        # Exportar no formato escolhido
        if formato == 'xml':
            conteudo = exportar_lancamentos_speed_xml(lancamentos_completos)
        else:
            conteudo = exportar_lancamentos_speed(lancamentos_completos)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo,
            'formato': formato,
            'total_lancamentos': len(lancamentos_completos),
            'validacao': validacao
        })
        
    except Exception as e:
        logger.error(f"Erro ao exportar lan�amentos: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# RELAT�RIOS CONT�BEIS - FASE 3 SPEED
# =============================================================================

@app.route('/api/relatorios/balancete', methods=['POST'])
@require_auth
def gerar_balancete_api():
    """Gera Balancete de Verifica��o"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('data_inicio') or not data.get('data_fim'):
            return jsonify({'success': False, 'error': 'Per�odo � obrigat�rio'}), 400
        
        # Converter datas
        from datetime import datetime
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_balancete_verificacao
        
        conn = get_db_connection()
        resultado = gerar_balancete_verificacao(
            conn=conn,
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            versao_plano_id=data.get('versao_plano_id'),
            nivel_minimo=data.get('nivel_minimo'),
            nivel_maximo=data.get('nivel_maximo'),
            classificacao=data.get('classificacao'),
            apenas_com_movimento=data.get('apenas_com_movimento', False)
        )
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao gerar balancete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/dre', methods=['POST'])
@require_auth
def gerar_dre_api():
    """Gera DRE (Demonstra��o do Resultado do Exerc�cio) COMPLETA"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('data_inicio') or not data.get('data_fim'):
            return jsonify({'success': False, 'error': 'Per�odo � obrigat�rio'}), 400
        
        # Converter datas
        from datetime import datetime
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_dre
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            resultado = gerar_dre(
                conn=conn,
                empresa_id=empresa_id,
                data_inicio=data_inicio,
                data_fim=data_fim,
                versao_plano_id=data.get('versao_plano_id'),
                comparar_periodo_anterior=data.get('comparar_periodo_anterior', False)
            )
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao gerar DRE: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/dre/pdf', methods=['POST'])
@require_auth
def gerar_dre_pdf_api():
    """Exporta DRE em formato PDF"""
    try:
        from datetime import datetime
        from flask import send_file
        from relatorios_contabeis_functions import gerar_dre
        from pdf_export import gerar_dre_pdf
        
        user = request.usuario
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('data_inicio') or not data.get('data_fim'):
            return jsonify({'success': False, 'error': 'Per�odo � obrigat�rio'}), 400
        
        # Converter datas
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        # Buscar nome da empresa e gerar DRE
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute("SELECT razao_social FROM empresas WHERE id = %s", (empresa_id,))
            empresa = cursor.fetchone()
            nome_empresa = empresa['razao_social'] if empresa else "Empresa"
            
            # Gerar dados da DRE
            dados_dre = gerar_dre(
                conn=conn,
                empresa_id=empresa_id,
                data_inicio=data_inicio,
                data_fim=data_fim,
                versao_plano_id=data.get('versao_plano_id'),
                comparar_periodo_anterior=data.get('comparar_periodo_anterior', False)
            )
        
        if not dados_dre.get('success'):
            return jsonify({'success': False, 'error': 'Erro ao gerar dados da DRE'}), 400
        
        # Formatar per�odo para o PDF
        periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        
        # Gerar PDF
        pdf_buffer = gerar_dre_pdf(
            dados_dre=dados_dre,
            nome_empresa=nome_empresa,
            periodo=periodo
        )
        
        # Nome do arquivo PDF
        filename = f"DRE_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.pdf"
        
        # Retornar PDF
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar PDF da DRE: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/dre/excel', methods=['POST'])
@require_auth
def gerar_dre_excel_api():
    """Exporta DRE em formato Excel"""
    try:
        from datetime import datetime
        from flask import send_file
        from relatorios_contabeis_functions import gerar_dre
        from pdf_export import gerar_dre_excel
        
        user = request.usuario
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('data_inicio') or not data.get('data_fim'):
            return jsonify({'success': False, 'error': 'Per�odo � obrigat�rio'}), 400
        
        # Converter datas
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        # Buscar nome da empresa e gerar DRE
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute("SELECT razao_social FROM empresas WHERE id = %s", (empresa_id,))
            empresa = cursor.fetchone()
            nome_empresa = empresa['razao_social'] if empresa else "Empresa"
            
            # Gerar dados da DRE
            dados_dre = gerar_dre(
                conn=conn,
                empresa_id=empresa_id,
                data_inicio=data_inicio,
                data_fim=data_fim,
                versao_plano_id=data.get('versao_plano_id'),
                comparar_periodo_anterior=data.get('comparar_periodo_anterior', False)
            )
        
        if not dados_dre.get('success'):
            return jsonify({'success': False, 'error': 'Erro ao gerar dados da DRE'}), 400
        
        # Formatar per�odo para o Excel
        periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        
        # Gerar Excel
        excel_buffer = gerar_dre_excel(
            dados_dre=dados_dre,
            nome_empresa=nome_empresa,
            periodo=periodo
        )
        
        # Nome do arquivo Excel
        filename = f"DRE_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
        
        # Retornar Excel
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar Excel da DRE: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# CONFIGURA��O DRE - MAPEAMENTO DE SUBCATEGORIAS
# ============================================================================

@app.route('/api/dre/configuracao/mapeamentos', methods=['GET'])
@require_auth
def listar_mapeamentos_dre():
    """Lista todos os mapeamentos de subcategorias para o DRE da empresa"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Buscar mapeamentos com informa��es das subcategorias e contas
            cursor.execute("""
                SELECT 
                    m.id,
                    m.subcategoria_id,
                    s.nome as subcategoria_nome,
                    c.nome as categoria_nome,
                    c.tipo as categoria_tipo,
                    m.plano_contas_id,
                    pc.codigo as plano_contas_codigo,
                    pc.descricao as plano_contas_descricao,
                    pc.classificacao as plano_contas_classificacao,
                    m.ativo,
                    m.created_at,
                    m.updated_at
                FROM dre_mapeamento_subcategoria m
                INNER JOIN subcategorias s ON s.id = m.subcategoria_id
                INNER JOIN categorias c ON c.id = s.categoria_id
                INNER JOIN plano_contas pc ON pc.id = m.plano_contas_id
                WHERE m.empresa_id = %s
                ORDER BY c.nome, s.nome
            """, (empresa_id,))
            
            mapeamentos = []
            for row in cursor.fetchall():
                mapeamentos.append({
                    'id': row['id'],
                    'subcategoria': {
                        'id': row['subcategoria_id'],
                        'nome': row['subcategoria_nome'],
                        'categoria': {
                            'nome': row['categoria_nome'],
                            'tipo': row['categoria_tipo']
                        }
                    },
                    'plano_contas': {
                        'id': row['plano_contas_id'],
                        'codigo': row['plano_contas_codigo'],
                        'descricao': row['plano_contas_descricao'],
                        'classificacao': row['plano_contas_classificacao']
                    },
                    'ativo': row['ativo'],
                    'created_at': row['created_at'].isoformat() if row['created_at'] else None,
                    'updated_at': row['updated_at'].isoformat() if row['updated_at'] else None
                })
            
            cursor.close()
        
        return jsonify({
            'success': True,
            'mapeamentos': mapeamentos,
            'total': len(mapeamentos)
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar mapeamentos DRE: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dre/configuracao/mapeamentos', methods=['POST'])
@require_auth
def criar_mapeamento_dre():
    """Cria um novo mapeamento de subcategoria para conta do DRE"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('subcategoria_id') or not data.get('plano_contas_id'):
            return jsonify({
                'success': False,
                'error': 'subcategoria_id e plano_contas_id s�o obrigat�rios'
            }), 400
        
        subcategoria_id = data['subcategoria_id']
        plano_contas_id = data['plano_contas_id']
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Verificar se a subcategoria pertence � empresa
            cursor.execute("""
                SELECT s.id, s.nome, c.nome as categoria_nome, c.empresa_id
                FROM subcategorias s
                INNER JOIN categorias c ON c.id = s.categoria_id
                WHERE s.id = %s
            """, (subcategoria_id,))
            
            subcategoria = cursor.fetchone()
            if not subcategoria:
                return jsonify({'success': False, 'error': 'Subcategoria n�o encontrada'}), 404
            
            if subcategoria['empresa_id'] != empresa_id:
                return jsonify({'success': False, 'error': 'Subcategoria n�o pertence a esta empresa'}), 403
            
            # Verificar se a conta do plano pertence � empresa e � DRE (c�digos 4, 5, 6, 7)
            cursor.execute("""
                SELECT id, codigo, descricao, classificacao, empresa_id
                FROM plano_contas
                WHERE id = %s
                  AND empresa_id = %s
                  AND tipo_conta = 'analitica'
                  AND (codigo LIKE '4%%' OR codigo LIKE '5%%' OR codigo LIKE '6%%' OR codigo LIKE '7%%')
                  AND deleted_at IS NULL
            """, (plano_contas_id, empresa_id))
            
            plano_contas = cursor.fetchone()
            if not plano_contas:
                return jsonify({
                    'success': False,
                    'error': 'Conta do plano n�o encontrada ou n�o � v�lida para DRE (deve ser c�digo 4.x, 5.x, 6.x ou 7.x)'
                }), 404
            
            # Verificar se j� existe mapeamento para esta subcategoria
            cursor.execute("""
                SELECT id FROM dre_mapeamento_subcategoria
                WHERE empresa_id = %s AND subcategoria_id = %s
            """, (empresa_id, subcategoria_id))
            
            if cursor.fetchone():
                return jsonify({
                    'success': False,
                    'error': 'J� existe um mapeamento para esta subcategoria. Atualize o existente ou exclua-o primeiro.'
                }), 409
            
            # Criar o mapeamento
            cursor.execute("""
                INSERT INTO dre_mapeamento_subcategoria 
                (empresa_id, subcategoria_id, plano_contas_id, ativo)
                VALUES (%s, %s, %s, TRUE)
                RETURNING id, created_at
            """, (empresa_id, subcategoria_id, plano_contas_id))
            
            result = cursor.fetchone()
            mapeamento_id = result['id']
            created_at = result['created_at']
            
            conn.commit()
            cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Mapeamento criado com sucesso',
            'mapeamento': {
                'id': mapeamento_id,
                'subcategoria': {
                    'id': subcategoria_id,
                    'nome': subcategoria['nome'],
                    'categoria': subcategoria['categoria_nome']
                },
                'plano_contas': {
                    'id': plano_contas_id,
                    'codigo': plano_contas['codigo'],
                    'descricao': plano_contas['descricao']
                },
                'created_at': created_at.isoformat()
            }
        }), 201
        
    except Exception as e:
        logger.error(f"Erro ao criar mapeamento DRE: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dre/configuracao/mapeamentos/<int:mapeamento_id>', methods=['PUT'])
@require_auth
def atualizar_mapeamento_dre(mapeamento_id):
    """Atualiza um mapeamento existente"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Verificar se o mapeamento existe e pertence � empresa
            cursor.execute("""
                SELECT id FROM dre_mapeamento_subcategoria
                WHERE id = %s AND empresa_id = %s
            """, (mapeamento_id, empresa_id))
            
            if not cursor.fetchone():
                return jsonify({'success': False, 'error': 'Mapeamento n�o encontrado'}), 404
            
            # Campos atualiz�veis
            updates = []
            params = []
            
            if 'plano_contas_id' in data:
                # Validar nova conta
                cursor.execute("""
                    SELECT id, codigo, descricao
                    FROM plano_contas
                    WHERE id = %s
                      AND empresa_id = %s
                      AND tipo_conta = 'analitica'
                      AND (codigo LIKE '4%%' OR codigo LIKE '5%%' OR codigo LIKE '6%%' OR codigo LIKE '7%%')
                      AND deleted_at IS NULL
                """, (data['plano_contas_id'], empresa_id))
                
                if not cursor.fetchone():
                    return jsonify({
                        'success': False,
                        'error': 'Conta do plano inv�lida para DRE'
                    }), 400
                
                updates.append('plano_contas_id = %s')
                params.append(data['plano_contas_id'])
            
            if 'ativo' in data:
                updates.append('ativo = %s')
                params.append(data['ativo'])
            
            if not updates:
                return jsonify({'success': False, 'error': 'Nenhum campo para atualizar'}), 400
            
            # Executar update
            params.extend([mapeamento_id, empresa_id])
            query = f"""
                UPDATE dre_mapeamento_subcategoria
                SET {', '.join(updates)}
                WHERE id = %s AND empresa_id = %s
                RETURNING updated_at
            """
            
            cursor.execute(query, params)
            result = cursor.fetchone()
            
            conn.commit()
            cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Mapeamento atualizado com sucesso',
            'updated_at': result['updated_at'].isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro ao atualizar mapeamento DRE: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dre/configuracao/mapeamentos/<int:mapeamento_id>', methods=['DELETE'])
@require_auth
def excluir_mapeamento_dre(mapeamento_id):
    """Exclui um mapeamento"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            
            # Verificar se o mapeamento existe e pertence � empresa
            cursor.execute("""
                DELETE FROM dre_mapeamento_subcategoria
                WHERE id = %s AND empresa_id = %s
                RETURNING id
            """, (mapeamento_id, empresa_id))
            
            result = cursor.fetchone()
            
            if not result:
                return jsonify({'success': False, 'error': 'Mapeamento n�o encontrado'}), 404
            
            conn.commit()
            cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Mapeamento exclu�do com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao excluir mapeamento DRE: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dre/configuracao/subcategorias-disponiveis', methods=['GET'])
@require_auth
def listar_subcategorias_disponiveis_dre():
    """Lista subcategorias que ainda n�o t�m mapeamento para o DRE"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Buscar subcategorias sem mapeamento
            cursor.execute("""
                SELECT 
                    s.id,
                    s.nome as subcategoria_nome,
                    c.id as categoria_id,
                    c.nome as categoria_nome,
                    c.tipo as categoria_tipo
                FROM subcategorias s
                INNER JOIN categorias c ON c.id = s.categoria_id
                LEFT JOIN dre_mapeamento_subcategoria m ON m.subcategoria_id = s.id AND m.empresa_id = %s
                WHERE c.empresa_id = %s
                  AND s.ativa = TRUE
                  AND m.id IS NULL
                ORDER BY c.nome, s.nome
            """, (empresa_id, empresa_id))
            
            subcategorias = []
            for row in cursor.fetchall():
                subcategorias.append({
                    'id': row['id'],
                    'nome': row['subcategoria_nome'],
                    'categoria': {
                        'id': row['categoria_id'],
                        'nome': row['categoria_nome'],
                        'tipo': row['categoria_tipo']
                    }
                })
            
            cursor.close()
        
        return jsonify({
            'success': True,
            'subcategorias': subcategorias,
            'total': len(subcategorias)
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar subcategorias dispon�veis: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dre/configuracao/plano-contas-dre', methods=['GET'])
@require_auth
def listar_plano_contas_dre():
    """Lista contas do plano de contas v�lidas para DRE (c�digos 4, 5, 6, 7)"""
    try:
        user = request.usuario
        empresa_id = user['empresa_id']
        
        # Par�metro opcional para filtrar por classifica��o
        classificacao = request.args.get('classificacao')  # 'receita' ou 'despesa'
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            query = """
                SELECT 
                    id,
                    codigo,
                    descricao,
                    classificacao,
                    nivel
                FROM plano_contas
                WHERE empresa_id = %s
                  AND tipo_conta = 'analitica'
                  AND (codigo LIKE '4%%' OR codigo LIKE '5%%' OR codigo LIKE '6%%' OR codigo LIKE '7%%')
                  AND deleted_at IS NULL
            """
            params = [empresa_id]
            
            if classificacao:
                query += " AND classificacao = %s"
                params.append(classificacao)
            
            query += " ORDER BY codigo"
            
            cursor.execute(query, params)
            
            contas = []
            for row in cursor.fetchall():
                # Determinar grupo DRE baseado no c�digo
                codigo = row['codigo']
                if codigo.startswith('4.9'):
                    grupo_dre = 'Dedu��es da Receita'
                elif codigo.startswith('4'):
                    grupo_dre = 'Receita Bruta'
                elif codigo.startswith('5'):
                    grupo_dre = 'Custos'
                elif codigo.startswith('6'):
                    grupo_dre = 'Despesas Operacionais'
                elif codigo.startswith('7.1'):
                    grupo_dre = 'Receitas Financeiras'
                elif codigo.startswith('7.2'):
                    grupo_dre = 'Despesas Financeiras'
                else:
                    grupo_dre = 'Outros'
                
                contas.append({
                    'id': row['id'],
                    'codigo': row['codigo'],
                    'descricao': row['descricao'],
                    'classificacao': row['classificacao'],
                    'nivel': row['nivel'],
                    'grupo_dre': grupo_dre
                })
            
            cursor.close()
        
        return jsonify({
            'success': True,
            'contas': contas,
            'total': len(contas)
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar plano de contas DRE: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# FIM - CONFIGURA��O DRE
# ============================================================================


@app.route('/api/dashboard/gerencial', methods=['GET'])
@require_auth
def dashboard_gerencial_api():
    """
    Dashboard Gerencial Completo
    
    Retorna:
    - KPIs do m�s (receita, despesas, lucro, margem)
    - Evolu��o mensal (12 meses)
    - Ponto de equil�brio
    - Compara��o com m�s anterior
    """
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        # Par�metros opcionais
        data_referencia_str = request.args.get('data_referencia')
        versao_plano_id = request.args.get('versao_plano_id', type=int)
        
        # Converter data se fornecida
        from datetime import datetime
        data_referencia = None
        if data_referencia_str:
            data_referencia = datetime.strptime(data_referencia_str, '%Y-%m-%d').date()
        
        from dashboard_functions import gerar_dashboard_gerencial
        
        conn = get_db_connection()
        resultado = gerar_dashboard_gerencial(
            conn=conn,
            empresa_id=empresa_id,
            data_referencia=data_referencia,
            versao_plano_id=versao_plano_id
        )
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao gerar dashboard gerencial: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dashboard/gerencial/pdf', methods=['GET'])
@require_auth
def gerar_dashboard_pdf_api():
    """Exporta Dashboard Gerencial em formato PDF"""
    try:
        from datetime import datetime
        from flask import send_file
        from dashboard_functions import gerar_dashboard_gerencial
        from pdf_export import gerar_dashboard_pdf
        
        user = request.usuario
        empresa_id = user['empresa_id']
        
        # Par�metros
        data_referencia_str = request.args.get('data_referencia')
        versao_plano_id = request.args.get('versao_plano_id', type=int)
        
        # Converter data
        data_referencia = None
        if data_referencia_str:
            data_referencia = datetime.strptime(data_referencia_str, '%Y-%m-%d').date()
        
        # Buscar nome da empresa e gerar dashboard
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute("SELECT razao_social FROM empresas WHERE id = %s", (empresa_id,))
            empresa = cursor.fetchone()
            nome_empresa = empresa['razao_social'] if empresa else "Empresa"
            
            # Gerar dados do dashboard
            dados_dashboard = gerar_dashboard_gerencial(
                conn=conn,
                empresa_id=empresa_id,
                data_referencia=data_referencia,
                versao_plano_id=versao_plano_id
            )
        
        if not dados_dashboard.get('success'):
            return jsonify({'success': False, 'error': 'Erro ao gerar dados do dashboard'}), 400
        
        # Formatar m�s de refer�ncia
        mes_ref = dados_dashboard['dashboard'].get('mes_referencia', '')
        
        # Gerar PDF
        pdf_buffer = gerar_dashboard_pdf(
            dados_dashboard=dados_dashboard,
            nome_empresa=nome_empresa,
            mes_referencia=mes_ref
        )
        
        # Nome do arquivo
        mes_ano = data_referencia.strftime('%Y%m') if data_referencia else datetime.now().strftime('%Y%m')
        filename = f"Dashboard_{mes_ano}.pdf"
        
        # Retornar PDF
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar PDF do dashboard: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dashboard/gerencial/excel', methods=['GET'])
@require_auth
def gerar_dashboard_excel_api():
    """Exporta Dashboard Gerencial em formato Excel"""
    try:
        from datetime import datetime
        from flask import send_file
        from dashboard_functions import gerar_dashboard_gerencial
        from pdf_export import gerar_dashboard_excel
        
        user = request.usuario
        empresa_id = user['empresa_id']
        
        # Par�metros
        data_referencia_str = request.args.get('data_referencia')
        versao_plano_id = request.args.get('versao_plano_id', type=int)
        
        # Converter data
        data_referencia = None
        if data_referencia_str:
            data_referencia = datetime.strptime(data_referencia_str, '%Y-%m-%d').date()
        
        # Buscar nome da empresa e gerar dashboard
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor.execute("SELECT razao_social FROM empresas WHERE id = %s", (empresa_id,))
            empresa = cursor.fetchone()
            nome_empresa = empresa['razao_social'] if empresa else "Empresa"
            
            # Gerar dados do dashboard
            dados_dashboard = gerar_dashboard_gerencial(
                conn=conn,
                empresa_id=empresa_id,
                data_referencia=data_referencia,
                versao_plano_id=versao_plano_id
            )
        
        if not dados_dashboard.get('success'):
            return jsonify({'success': False, 'error': 'Erro ao gerar dados do dashboard'}), 400
        
        # Formatar m�s de refer�ncia
        mes_ref = dados_dashboard['dashboard'].get('mes_referencia', '')
        
        # Gerar Excel
        excel_buffer = gerar_dashboard_excel(
            dados_dashboard=dados_dashboard,
            nome_empresa=nome_empresa,
            mes_referencia=mes_ref
        )
        
        # Nome do arquivo
        mes_ano = data_referencia.strftime('%Y%m') if data_referencia else datetime.now().strftime('%Y%m')
        filename = f"Dashboard_{mes_ano}.xlsx"
        
        # Retornar Excel
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar Excel do dashboard: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/balanco-patrimonial', methods=['POST'])
@require_auth
def gerar_balanco_patrimonial_api():
    """Gera Balan�o Patrimonial"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('data_referencia'):
            return jsonify({'success': False, 'error': 'Data de refer�ncia � obrigat�ria'}), 400
        
        # Converter data
        from datetime import datetime
        data_referencia = datetime.strptime(data['data_referencia'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_balanco_patrimonial
        
        conn = get_db_connection()
        resultado = gerar_balanco_patrimonial(
            conn=conn,
            empresa_id=empresa_id,
            data_referencia=data_referencia,
            versao_plano_id=data.get('versao_plano_id')
        )
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao gerar balan�o patrimonial: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/razao-contabil', methods=['POST'])
@require_auth
def gerar_razao_contabil_api():
    """Gera Raz�o Cont�bil (extrato de uma conta)"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Validar campos obrigat�rios
        if not data.get('conta_id'):
            return jsonify({'success': False, 'error': 'ID da conta � obrigat�rio'}), 400
        if not data.get('data_inicio') or not data.get('data_fim'):
            return jsonify({'success': False, 'error': 'Per�odo � obrigat�rio'}), 400
        
        # Converter datas
        from datetime import datetime
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_razao_contabil
        
        conn = get_db_connection()
        resultado = gerar_razao_contabil(
            conn=conn,
            empresa_id=empresa_id,
            conta_id=data['conta_id'],
            data_inicio=data_inicio,
            data_fim=data_fim
        )
        conn.close()
        
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao gerar raz�o cont�bil: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/balancete/exportar', methods=['POST'])
@require_auth
def exportar_balancete_api():
    """Exporta Balancete em TXT ou CSV"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        formato = data.get('formato', 'txt')  # 'txt' ou 'csv'
        
        # Gerar balancete
        from datetime import datetime
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_balancete_verificacao
        from speed_integration import exportar_balancete_speed_txt, exportar_balancete_speed_csv
        
        conn = get_db_connection()
        balancete = gerar_balancete_verificacao(
            conn=conn,
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            versao_plano_id=data.get('versao_plano_id'),
            nivel_minimo=data.get('nivel_minimo'),
            nivel_maximo=data.get('nivel_maximo'),
            classificacao=data.get('classificacao'),
            apenas_com_movimento=data.get('apenas_com_movimento', False)
        )
        conn.close()
        
        if not balancete['success']:
            return jsonify(balancete), 400
        
        # Exportar no formato escolhido
        if formato == 'csv':
            conteudo = exportar_balancete_speed_csv(balancete)
        else:
            conteudo = exportar_balancete_speed_txt(balancete)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo,
            'formato': formato,
            'total_contas': balancete.get('total_contas', 0)
        })
    except Exception as e:
        logger.error(f"Erro ao exportar balancete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/dre/exportar', methods=['POST'])
@require_auth
def exportar_dre_api():
    """Exporta DRE em TXT"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Gerar DRE
        from datetime import datetime
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_dre
        from speed_integration import exportar_dre_speed_txt
        
        conn = get_db_connection()
        dre = gerar_dre(
            conn=conn,
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            versao_plano_id=data.get('versao_plano_id')
        )
        conn.close()
        
        if not dre['success']:
            return jsonify(dre), 400
        
        # Exportar em TXT
        conteudo = exportar_dre_speed_txt(dre)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo,
            'formato': 'txt',
            'resultado_liquido': dre.get('dre', {}).get('resultado_liquido', 0)
        })
    except Exception as e:
        logger.error(f"Erro ao exportar DRE: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/balanco-patrimonial/exportar', methods=['POST'])
@require_auth
def exportar_balanco_patrimonial_api():
    """Exporta Balan�o Patrimonial em TXT"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Gerar Balan�o
        from datetime import datetime
        data_referencia = datetime.strptime(data['data_referencia'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_balanco_patrimonial
        from speed_integration import exportar_balanco_patrimonial_speed_txt
        
        conn = get_db_connection()
        balanco = gerar_balanco_patrimonial(
            conn=conn,
            empresa_id=empresa_id,
            data_referencia=data_referencia,
            versao_plano_id=data.get('versao_plano_id')
        )
        conn.close()
        
        if not balanco['success']:
            return jsonify(balanco), 400
        
        # Exportar em TXT
        conteudo = exportar_balanco_patrimonial_speed_txt(balanco)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo,
            'formato': 'txt',
            'balanco_fechado': balanco.get('validacao', {}).get('balanco_fechado', False)
        })
    except Exception as e:
        logger.error(f"Erro ao exportar balan�o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorios/razao-contabil/exportar', methods=['POST'])
@require_auth
def exportar_razao_contabil_api():
    """Exporta Raz�o Cont�bil em TXT"""
    try:
        user = request.user
        empresa_id = user['empresa_id']
        
        data = request.get_json()
        
        # Gerar Raz�o
        from datetime import datetime
        data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
        
        from relatorios_contabeis_functions import gerar_razao_contabil
        from speed_integration import exportar_razao_contabil_speed_txt
        
        conn = get_db_connection()
        razao = gerar_razao_contabil(
            conn=conn,
            empresa_id=empresa_id,
            conta_id=data['conta_id'],
            data_inicio=data_inicio,
            data_fim=data_fim
        )
        conn.close()
        
        if not razao['success']:
            return jsonify(razao), 400
        
        # Exportar em TXT
        conteudo = exportar_razao_contabil_speed_txt(razao)
        
        return jsonify({
            'success': True,
            'conteudo': conteudo,
            'formato': 'txt',
            'total_movimentacoes': razao.get('total_movimentacoes', 0)
        })
    except Exception as e:
        logger.error(f"Erro ao exportar raz�o: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# SPED ECD - FASE 4 SPEED
# =============================================================================

@app.route('/api/sped/ecd/gerar', methods=['POST'])
@require_auth
def sped_ecd_gerar():
    """
    Gera arquivo SPED ECD (Escritura��o Cont�bil Digital)
    
    Body:
    {
        "data_inicio": "YYYY-MM-DD",
        "data_fim": "YYYY-MM-DD",
        "versao_plano_id": <opcional>
    }
    
    Returns:
    {
        "success": true,
        "total_linhas": 1234,
        "hash": "ABC123...",
        "data_geracao": "17/02/2026 10:30:00",
        "periodo": "01012026 a 31122026",
        "preview": "primeiras 50 linhas do arquivo"
    }
    """
    try:
        from sped_ecd_functions import gerar_arquivo_ecd
        
        data = request.get_json()
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        # Valida��es
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        versao_plano_id = data.get('versao_plano_id')
        
        if not data_inicio or not data_fim:
            return jsonify({
                'success': False,
                'error': 'data_inicio e data_fim s�o obrigat�rios'
            }), 400
        
        # Gerar ECD
        resultado = gerar_arquivo_ecd(
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            versao_plano_id=versao_plano_id
        )
        
        if not resultado['success']:
            return jsonify(resultado), 400
        
        # Retornar preview (primeiras 50 linhas)
        linhas = resultado['conteudo'].split('\n')
        preview = '\n'.join(linhas[:50])
        if len(linhas) > 50:
            preview += f"\n\n... (mais {len(linhas) - 50} linhas)"
        
        return jsonify({
            'success': True,
            'total_linhas': resultado['total_linhas'],
            'hash': resultado['hash'],
            'data_geracao': resultado['data_geracao'],
            'periodo': resultado['periodo'],
            'preview': preview
        })
        
    except Exception as e:
        logger.error(f"Erro ao gerar ECD: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sped/ecd/exportar', methods=['POST'])
@require_auth
def sped_ecd_exportar():
    """
    Exporta arquivo SPED ECD completo
    
    Body:
    {
        "data_inicio": "YYYY-MM-DD",
        "data_fim": "YYYY-MM-DD",
        "versao_plano_id": <opcional>
    }
    
    Returns:
    {
        "success": true,
        "conteudo": "conte�do completo do arquivo TXT",
        "total_linhas": 1234,
        "hash": "ABC123...",
        "nome_arquivo": "ECD_CNPJ_AAAAMMDD.txt"
    }
    """
    try:
        from sped_ecd_functions import gerar_arquivo_ecd
        from database_postgresql import get_connection
        
        data = request.get_json()
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        # Valida��es
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        versao_plano_id = data.get('versao_plano_id')
        
        if not data_inicio or not data_fim:
            return jsonify({
                'success': False,
                'error': 'data_inicio e data_fim s�o obrigat�rios'
            }), 400
        
        # Gerar ECD
        resultado = gerar_arquivo_ecd(
            empresa_id=empresa_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            versao_plano_id=versao_plano_id
        )
        
        if not resultado['success']:
            return jsonify(resultado), 400
        
        # Buscar CNPJ para nome do arquivo
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT cnpj FROM empresas WHERE id = %s", (empresa_id,))
            empresa = cursor.fetchone()
            cnpj_limpo = ''.join(filter(str.isdigit, empresa[0] if empresa and empresa[0] else '00000000000000'))
        finally:
            cursor.close()
            conn.close()
        
        # Nome do arquivo: ECD_CNPJ_AAAAMMDD.txt
        data_ref = data_fim.replace('-', '')[:8]  # AAAAMMDD
        nome_arquivo = f"ECD_{cnpj_limpo}_{data_ref}.txt"
        
        return jsonify({
            'success': True,
            'conteudo': resultado['conteudo'],
            'total_linhas': resultado['total_linhas'],
            'hash': resultado['hash'],
            'nome_arquivo': nome_arquivo,
            'data_geracao': resultado['data_geracao']
        })
        
    except Exception as e:
        logger.error(f"Erro ao exportar ECD: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# SPED EFD-CONTRIBUI��ES - FASE 5 SPEED
# =============================================================================

@app.route('/api/sped/efd-contribuicoes/calcular', methods=['POST'])
@require_auth
def sped_efd_contribuicoes_calcular():
    """
    Calcula apura��o mensal de PIS/COFINS sem gerar arquivo
    
    Body:
    {
        "mes": 1-12,
        "ano": 2026
    }
    
    Returns:
    {
        "success": true,
        "periodo": "01/2026",
        "regime": "Lucro Presumido (Cumulativo)",
        "receitas": {
            "total": 100000.00,
            "tributavel": 100000.00
        },
        "pis": {
            "aliquota": 0.65,
            "base_calculo": 100000.00,
            "valor": 650.00
        },
        "cofins": {
            "aliquota": 3.0,
            "base_calculo": 100000.00,
            "valor": 3000.00
        },
        "total_tributos": 3650.00
    }
    """
    try:
        from sped_efd_contribuicoes_functions import calcular_apuracao_mensal
        
        data = request.get_json()
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        # Valida��es
        mes = data.get('mes')
        ano = data.get('ano')
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        if not isinstance(mes, int) or mes < 1 or mes > 12:
            return jsonify({
                'success': False,
                'error': 'mes deve estar entre 1 e 12'
            }), 400
        
        if not isinstance(ano, int) or ano < 2000 or ano > 2100:
            return jsonify({
                'success': False,
                'error': 'ano inv�lido'
            }), 400
        
        # Calcular apura��o
        resultado = calcular_apuracao_mensal(
            empresa_id=empresa_id,
            mes=mes,
            ano=ano
        )
        
        if not resultado['success']:
            return jsonify(resultado), 400
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao calcular PIS/COFINS: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sped/efd-contribuicoes/gerar', methods=['POST'])
@require_auth
def sped_efd_contribuicoes_gerar():
    """
    Gera arquivo EFD-Contribui��es com preview
    
    Body:
    {
        "mes": 1-12,
        "ano": 2026,
        "usar_creditos_reais": true  # Opcional, padr�o: true
    }
    
    Returns:
    {
        "success": true,
        "total_linhas": 450,
        "hash": "ABC123...",
        "data_geracao": "17/02/2026 14:30:00",
        "periodo": "01/2026",
        "modo": "creditos_reais" ou "simplificado",
        "totais": {
            "receitas": 100000.00,
            "pis": 650.00,
            "cofins": 3000.00,
            "total_tributos": 3650.00,
            "credito_pis": 150.00,
            "credito_cofins": 690.00
        },
        "preview": "primeiras 50 linhas do arquivo"
    }
    """
    try:
        from sped_efd_contribuicoes_functions import gerar_arquivo_efd_contribuicoes
        
        data = request.get_json()
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        # Valida��es
        mes = data.get('mes')
        ano = data.get('ano')
        usar_creditos_reais = data.get('usar_creditos_reais', True)
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        if not isinstance(mes, int) or mes < 1 or mes > 12:
            return jsonify({
                'success': False,
                'error': 'mes deve estar entre 1 e 12'
            }), 400
        
        if not isinstance(ano, int) or ano < 2000 or ano > 2100:
            return jsonify({
                'success': False,
                'error': 'ano inv�lido'
            }), 400
        
        # Gerar EFD-Contribui��es
        resultado = gerar_arquivo_efd_contribuicoes(
            empresa_id=empresa_id,
            mes=mes,
            ano=ano,
            usar_creditos_reais=usar_creditos_reais
        )
        
        if not resultado['success']:
            return jsonify(resultado), 400
        
        # Retornar preview (primeiras 50 linhas)
        linhas = resultado['conteudo'].split('\n')
        preview = '\n'.join(linhas[:50])
        if len(linhas) > 50:
            preview += f"\n\n... (mais {len(linhas) - 50} linhas)"
        
        return jsonify({
            'success': True,
            'total_linhas': resultado['total_linhas'],
            'hash': resultado['hash'],
            'data_geracao': resultado['data_geracao'],
            'periodo': resultado['periodo'],
            'modo': resultado.get('modo', 'simplificado'),
            'totais': resultado['totais'],
            'preview': preview
        })
        
    except Exception as e:
        logger.error(f"Erro ao gerar EFD-Contribui��es: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sped/efd-contribuicoes/exportar', methods=['POST'])
@require_auth
def sped_efd_contribuicoes_exportar():
    """
    Exporta arquivo EFD-Contribui��es completo
    
    Body:
    {
        "mes": 1-12,
        "ano": 2026,
        "usar_creditos_reais": true  # Opcional, padr�o: true
    }
    
    Returns:
    {
        "success": true,
        "conteudo": "conte�do completo do arquivo TXT",
        "total_linhas": 450,
        "hash": "ABC123...",
        "nome_arquivo": "EFD_Contribuicoes_CNPJ_AAAAMM.txt",
        "modo": "creditos_reais" ou "simplificado",
        "totais": {
            "receitas": 100000.00,
            "pis": 650.00,
            "cofins": 3000.00
        }
    }
    """
    try:
        from sped_efd_contribuicoes_functions import gerar_arquivo_efd_contribuicoes
        from database_postgresql import get_connection
        
        data = request.get_json()
        usuario = get_usuario_logado()
        empresa_id = usuario.get('empresa_id')
        
        # Valida��es
        mes = data.get('mes')
        ano = data.get('ano')
        usar_creditos_reais = data.get('usar_creditos_reais', True)
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        if not isinstance(mes, int) or mes < 1 or mes > 12:
            return jsonify({
                'success': False,
                'error': 'mes deve estar entre 1 e 12'
            }), 400
        
        if not isinstance(ano, int) or ano < 2000 or ano > 2100:
            return jsonify({
                'success': False,
                'error': 'ano inv�lido'
            }), 400
        
        # Gerar EFD-Contribui��es
        resultado = gerar_arquivo_efd_contribuicoes(
            empresa_id=empresa_id,
            mes=mes,
            ano=ano,
            usar_creditos_reais=usar_creditos_reais
        )
        
        if not resultado['success']:
            return jsonify(resultado), 400
        
        # Buscar CNPJ para nome do arquivo
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT cnpj FROM empresas WHERE id = %s", (empresa_id,))
            empresa = cursor.fetchone()
            cnpj_limpo = ''.join(filter(str.isdigit, empresa[0] if empresa and empresa[0] else '00000000000000'))
        finally:
            cursor.close()
            conn.close()
        
        # Nome do arquivo: EFD_Contribuicoes_CNPJ_AAAAMM.txt
        nome_arquivo = f"EFD_Contribuicoes_{cnpj_limpo}_{ano}{mes:02d}.txt"
        
        return jsonify({
            'success': True,
            'conteudo': resultado['conteudo'],
            'total_linhas': resultado['total_linhas'],
            'hash': resultado['hash'],
            'nome_arquivo': nome_arquivo,
            'data_geracao': resultado['data_geracao'],
            'modo': resultado.get('modo', 'simplificado'),
            'totais': resultado['totais']
        })
        
    except Exception as e:
        logger.error(f"Erro ao exportar EFD-Contribui��es: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# INTEGRA CONTADOR - API SERPRO
# =============================================================================

@app.route('/api/integra-contador/enviar', methods=['POST'])
@require_auth
def integra_contador_enviar():
    """Envia requisi��o para a API Integra Contador do SERPRO"""
    try:
        data = request.get_json()
        
        tipo_operacao = data.get('tipoOperacao')
        payload = data.get('payload')
        
        if not tipo_operacao or not payload:
            return jsonify({
                'success': False,
                'error': 'tipoOperacao e payload s�o obrigat�rios'
            }), 400
        
        # Importar fun��es
        from integra_contador_functions import enviar_requisicao, validar_payload
        
        # Validar payload
        valido, mensagem = validar_payload(payload)
        if not valido:
            return jsonify({
                'success': False,
                'error': f'Valida��o falhou: {mensagem}'
            }), 400
        
        # Enviar requisi��o
        resultado = enviar_requisicao(tipo_operacao, payload)
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao enviar requisi��o Integra Contador: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/integra-contador/testar', methods=['GET'])
@require_auth
def integra_contador_testar():
    """Testa conex�o com a API Integra Contador"""
    try:
        from integra_contador_functions import testar_conexao
        resultado = testar_conexao()
        return jsonify(resultado)
    except Exception as e:
        logger.error(f"Erro ao testar conex�o: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/integra-contador/token', methods=['GET'])
@require_auth
def integra_contador_token():
    """Obt�m token de acesso (apenas para debug)"""
    try:
        from integra_contador_functions import obter_token
        token = obter_token()
        return jsonify({
            'success': True,
            'token': token[:50] + '...',  # Mostrar apenas primeiros 50 caracteres
            'message': 'Token obtido com sucesso'
        })
    except Exception as e:
        logger.error(f"Erro ao obter token: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== M�DULO FISCAL FEDERAL (Integra Contador SERPRO) =====

def _fiscal_get_db_empresa():
    """Helper: retorna (db, empresa_id) ou raises ValueError."""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        raise ValueError('Empresa n�o identificada')
    return db, empresa_id


@app.route('/api/fiscal/dashboard', methods=['GET'])
@require_auth
def fiscal_dashboard():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import gerar_dashboard
        resultado = gerar_dashboard(database, empresa_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_dashboard] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/cnpj/consultar', methods=['POST'])
@require_auth
def fiscal_cnpj_consultar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_cnpj
        resultado = consultar_cnpj(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            cnpj_consultar=data.get('contribuinte_doc', data.get('cnpj_consulta', ''))
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_cnpj_consultar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/cnpj/historico', methods=['GET'])
@require_auth
def fiscal_cnpj_historico():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import listar_historico_cnpj
        resultado = listar_historico_cnpj(database, empresa_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_cnpj_historico] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/cnd/solicitar', methods=['POST'])
@require_auth
def fiscal_cnd_solicitar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import solicitar_cnd
        resultado = solicitar_cnd(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_cnd_solicitar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/cnd/consultar', methods=['POST'])
@require_auth
def fiscal_cnd_consultar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_cnd
        resultado = consultar_cnd(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_cnd_consultar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/cnd/lista', methods=['GET'])
@require_auth
def fiscal_cnd_lista():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import listar_certidoes
        resultado = listar_certidoes(database, empresa_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_cnd_lista] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/dctfweb/consultar', methods=['POST'])
@require_auth
def fiscal_dctfweb_consultar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_dctfweb
        resultado = consultar_dctfweb(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            competencia=data.get('competencia', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_dctfweb_consultar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/dctfweb/lista', methods=['GET'])
@require_auth
def fiscal_dctfweb_lista():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import listar_dctfweb
        resultado = listar_dctfweb(database, empresa_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_dctfweb_lista] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/mit/incluir', methods=['POST'])
@require_auth
def fiscal_mit_incluir():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import incluir_mit
        resultado = incluir_mit(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            dados_tributo=data.get('dados_mit', {})
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_mit_incluir] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/mit/consultar', methods=['POST'])
@require_auth
def fiscal_mit_consultar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_mit
        resultado = consultar_mit(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            competencia=data.get('competencia', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_mit_consultar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/reinf/consultar', methods=['POST'])
@require_auth
def fiscal_reinf_consultar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_reinf
        resultado = consultar_reinf(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            evento=data.get('evento', 'R-1000'),
            competencia=data.get('competencia', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_reinf_consultar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/reinf/lista', methods=['GET'])
@require_auth
def fiscal_reinf_lista():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import listar_reinf
        resultado = listar_reinf(database, empresa_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_reinf_lista] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/darf/emitir', methods=['POST'])
@require_auth
def fiscal_darf_emitir():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import emitir_darf
        resultado = emitir_darf(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            codigo_receita=data.get('codigo_receita', ''),
            competencia=data.get('competencia', ''),
            valor=data.get('valor', 0),
            data_vencimento=data.get('data_vencimento', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_darf_emitir] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/darf/consultar-pagamento', methods=['POST'])
@require_auth
def fiscal_darf_consultar_pagamento():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_pagamento_darf
        resultado = consultar_pagamento_darf(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            codigo_receita=data.get('codigo_receita', ''),
            competencia=data.get('competencia', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_darf_consultar_pagamento] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/darf/lista', methods=['GET'])
@require_auth
def fiscal_darf_lista():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import listar_darfs
        resultado = listar_darfs(database, empresa_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_darf_lista] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/pagamentos/consultar', methods=['POST'])
@require_auth
def fiscal_pagamentos_consultar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        from fiscal_federal_service import consultar_pagamentos
        resultado = consultar_pagamentos(
            database, empresa_id,
            contratante_cnpj=data.get('contratante_cnpj', ''),
            autor_doc=data.get('autor_doc', ''),
            contribuinte_doc=data.get('contribuinte_doc', ''),
            competencia=data.get('competencia', '')
        )
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_pagamentos_consultar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/logs', methods=['GET'])
@require_auth
def fiscal_logs_listar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        limite = int(request.args.get('limite', 50))
        tipo = request.args.get('tipo')
        from fiscal_federal_service import listar_logs
        resultado = listar_logs(database, empresa_id, limit=limite, tipo=tipo)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_logs_listar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/logs/<int:log_id>', methods=['GET'])
@require_auth
def fiscal_log_detalhe(log_id):
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        from fiscal_federal_service import obter_log_detalhe
        resultado = obter_log_detalhe(database, empresa_id, log_id)
        return jsonify({'success': True, 'data': resultado})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_log_detalhe] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/fila', methods=['GET'])
@require_auth
def fiscal_fila_listar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        status_filtro = request.args.get('status')
        with database.get_connection() as conn:
            cur = conn.cursor()
            if status_filtro:
                cur.execute(
                    """SELECT id, tipo, parametros, status, tentativas, resultado,
                              criado_em, processado_em
                       FROM fiscal_fila WHERE empresa_id=%s AND status=%s
                       ORDER BY criado_em DESC LIMIT 100""",
                    (empresa_id, status_filtro)
                )
            else:
                cur.execute(
                    """SELECT id, tipo, parametros, status, tentativas, resultado,
                              criado_em, processado_em
                       FROM fiscal_fila WHERE empresa_id=%s
                       ORDER BY criado_em DESC LIMIT 100""",
                    (empresa_id,)
                )
            rows = cur.fetchall()
            cur.close()
        data = []
        for r in rows:
            row = dict(r) if hasattr(r, 'keys') else {
                'id': r[0], 'tipo': r[1], 'parametros': r[2],
                'status': r[3], 'tentativas': r[4], 'resultado': r[5],
                'criado_em': r[6], 'processado_em': r[7]
            }
            for k in ('criado_em', 'processado_em'):
                if row.get(k):
                    row[k] = row[k].isoformat()
            data.append(row)
        return jsonify({'success': True, 'data': data})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_fila_listar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fiscal/fila', methods=['POST'])
@require_auth
def fiscal_fila_adicionar():
    try:
        database, empresa_id = _fiscal_get_db_empresa()
        data = request.get_json() or {}
        tipo = data.get('tipo', '')
        parametros = data.get('parametros', {})
        if not tipo:
            return jsonify({'success': False, 'error': 'Campo tipo � obrigat�rio'}), 400
        with database.get_connection() as conn:
            cur = conn.cursor()
            cur.execute(
                """INSERT INTO fiscal_fila (empresa_id, tipo, parametros)
                   VALUES (%s, %s, %s) RETURNING id""",
                (empresa_id, tipo, json.dumps(parametros))
            )
            row = cur.fetchone()
            novo_id = row['id'] if hasattr(row, 'keys') else row[0]
            conn.commit()
            cur.close()
        return jsonify({'success': True, 'id': novo_id, 'message': 'Adicionado � fila com sucesso'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[fiscal_fila_adicionar] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# EFD-REINF � M�dulo Completo
# ============================================================================

def _reinf_get_db_empresa():
    """Helper: retorna (db, empresa_id) ou lan�a ValueError."""
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        raise ValueError('empresa_id n�o encontrado na sess�o')
    return db, int(empresa_id)


@app.route('/api/reinf/competencias', methods=['GET'])
@require_auth
def reinf_listar_competencias():
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import listar_competencias
        return jsonify({'success': True, 'data': listar_competencias(db, eid)})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_listar_competencias] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/competencia/<comp>', methods=['GET'])
@require_auth
def reinf_carregar_competencia(comp):
    """Lista todos os eventos de uma compet�ncia."""
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import listar_eventos
        eventos = listar_eventos(db, eid, competencia=comp)
        return jsonify({'success': True, 'data': eventos})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_carregar_competencia] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/dashboard/<comp>', methods=['GET'])
@require_auth
def reinf_dashboard(comp):
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import dashboard_reinf
        return jsonify({'success': True, 'data': dashboard_reinf(db, eid, comp)})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_dashboard] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/motor-sugestoes/<comp>', methods=['GET'])
@require_auth
def reinf_motor_sugestoes(comp):
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import gerar_sugestoes
        return jsonify({'success': True, 'data': gerar_sugestoes(db, eid, comp)})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_motor_sugestoes] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/evento/criar', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_criar_evento():
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        competencia  = body.get('competencia', '')
        evento       = (body.get('evento') or '').upper()
        payload_dados = body.get('dados', {})
        identificador = body.get('identificador')

        if not evento or not competencia:
            return jsonify({'success': False, 'error': 'Campos obrigat�rios: evento, competencia'}), 400

        from reinf_service import criar_evento, validar_evento
        erros = validar_evento(db, eid, evento, payload_dados, competencia)
        if erros:
            return jsonify({'success': False, 'error': erros[0], 'erros': erros}), 422

        resultado = criar_evento(db, eid, competencia, evento, payload_dados, identificador)
        code = 201 if resultado.get('success') else 409
        return jsonify(resultado), code
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_criar_evento] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/evento/<evento_id>/enviar', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_enviar_evento(evento_id):
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        contratante_cnpj = body.get('contratante_cnpj', '')
        autor_doc        = body.get('autor_doc', '')

        if not contratante_cnpj or not autor_doc:
            return jsonify({'success': False, 'error': 'contratante_cnpj e autor_doc s�o obrigat�rios'}), 400

        from reinf_service import enviar_evento
        resultado = enviar_evento(db, eid, evento_id, contratante_cnpj, autor_doc)
        return jsonify(resultado), 200 if resultado.get('success') else 422
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_enviar_evento] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/evento/<evento_id>', methods=['GET'])
@require_auth
def reinf_detalhe_evento(evento_id):
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import obter_evento
        ev = obter_evento(db, eid, evento_id)
        if not ev:
            return jsonify({'success': False, 'error': 'Evento n�o encontrado'}), 404
        return jsonify({'success': True, 'data': ev})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_detalhe_evento] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/evento/<evento_id>/consultar-status', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_consultar_status(evento_id):
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        contratante_cnpj = body.get('contratante_cnpj', '')
        autor_doc        = body.get('autor_doc', '')
        from reinf_service import consultar_status_evento
        resultado = consultar_status_evento(db, eid, evento_id, contratante_cnpj, autor_doc)
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_consultar_status] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/evento/<evento_id>/excluir', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_excluir_evento(evento_id):
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        motivo = body.get('motivo', 'Exclus�o manual')
        from reinf_service import excluir_evento
        resultado = excluir_evento(db, eid, evento_id, motivo)
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_excluir_evento] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/fechar', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_fechar_competencia():
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        competencia      = body.get('competencia', '')
        tipo_fechamento  = body.get('tipo', 'R-2099')   # 'R-2099' ou 'R-4099'
        contratante_cnpj = body.get('contratante_cnpj', '')
        autor_doc        = body.get('autor_doc', '')

        if not competencia or not contratante_cnpj or not autor_doc:
            return jsonify({'success': False, 'error': 'competencia, contratante_cnpj e autor_doc obrigat�rios'}), 400

        from reinf_service import fechar_competencia
        resultado = fechar_competencia(db, eid, competencia, tipo_fechamento, contratante_cnpj, autor_doc)
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_fechar_competencia] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/reabrir', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_reabrir_competencia():
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        competencia      = body.get('competencia', '')
        contratante_cnpj = body.get('contratante_cnpj', '')
        autor_doc        = body.get('autor_doc', '')

        from reinf_service import reabrir_competencia
        resultado = reabrir_competencia(db, eid, competencia, contratante_cnpj, autor_doc)
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_reabrir_competencia] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/totalizadores/<comp>', methods=['GET'])
@require_auth
def reinf_totalizadores(comp):
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import calcular_totalizadores
        totais = calcular_totalizadores(db, eid, comp)
        return jsonify({'success': True, 'data': totais})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_totalizadores] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/sincronizar-dctfweb', methods=['POST'])
@require_auth
@csrf_instance.exempt
def reinf_sincronizar_dctfweb():
    try:
        db, eid = _reinf_get_db_empresa()
        body = request.get_json() or {}
        competencia      = body.get('competencia', '')
        contratante_cnpj = body.get('contratante_cnpj', '')
        autor_doc        = body.get('autor_doc', '')
        from reinf_service import sincronizar_dctfweb
        resultado = sincronizar_dctfweb(db, eid, competencia, contratante_cnpj, autor_doc)
        return jsonify(resultado)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_sincronizar_dctfweb] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reinf/exportar-xml/<evento_id>', methods=['GET'])
@require_auth
def reinf_exportar_xml(evento_id):
    try:
        db, eid = _reinf_get_db_empresa()
        from reinf_service import exportar_xml_evento
        xml = exportar_xml_evento(db, eid, evento_id)
        if xml is None:
            return jsonify({'success': False, 'error': 'Evento n�o encontrado'}), 404
        from flask import Response
        return Response(xml, mimetype='application/xml',
                        headers={'Content-Disposition': f'attachment; filename=reinf_{evento_id}.xml'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 403
    except Exception as e:
        logger.error(f"[reinf_exportar_xml] {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== NOTAS FISCAIS (NF-e / NFS-e) =====

@app.route('/api/notas-fiscais/importar', methods=['POST'])
@require_auth
def importar_nota_fiscal():
    """
    Importa XML de NF-e ou NFS-e
    
    Body: {
        "tipo": "NFE" ou "NFSE",
        "xml_content": "conte�do do XML"
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        tipo = data.get('tipo', '').upper()
        xml_content = data.get('xml_content', '')
        
        if not xml_content:
            return jsonify({
                'success': False,
                'error': 'XML n�o fornecido'
            }), 400
        
        # Importar fun��o apropriada
        from nfe_import_functions import importar_xml_nfe, importar_xml_nfse
        
        if tipo == 'NFE':
            resultado = importar_xml_nfe(empresa_id, xml_content, usuario['id'])
        elif tipo == 'NFSE':
            resultado = importar_xml_nfse(empresa_id, xml_content, usuario['id'])
        else:
            return jsonify({
                'success': False,
                'error': 'Tipo inv�lido. Use NFE ou NFSE'
            }), 400
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao importar nota fiscal: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/notas-fiscais/upload', methods=['POST'])
@require_auth
def upload_xml_nota_fiscal():
    """
    Upload de arquivo XML de NF-e ou NFS-e
    
    Form-data:
        file: arquivo XML
        tipo: NFE ou NFSE
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        # Verificar se foi enviado arquivo
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'Nenhum arquivo enviado'
            }), 400
        
        file = request.files['file']
        tipo = request.form.get('tipo', '').upper()
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'Nome de arquivo vazio'
            }), 400
        
        # Ler conte�do do XML
        xml_content = file.read().decode('utf-8')
        
        # Importar fun��o apropriada
        from nfe_import_functions import importar_xml_nfe, importar_xml_nfse
        
        if tipo == 'NFE':
            resultado = importar_xml_nfe(empresa_id, xml_content, usuario['id'])
        elif tipo == 'NFSE':
            resultado = importar_xml_nfse(empresa_id, xml_content, usuario['id'])
        else:
            # Tentar detectar automaticamente
            if 'NFe' in xml_content or 'nfe:' in xml_content:
                resultado = importar_xml_nfe(empresa_id, xml_content, usuario['id'])
            else:
                resultado = importar_xml_nfse(empresa_id, xml_content, usuario['id'])
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao fazer upload de nota fiscal: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/notas-fiscais/listar', methods=['POST'])
@require_auth
def listar_notas_fiscais():
    """
    Lista notas fiscais da empresa
    
    Body: {
        "tipo": "NFE" ou "NFSE" (opcional),
        "data_inicio": "2026-01-01" (opcional),
        "data_fim": "2026-01-31" (opcional),
        "limit": 100 (opcional)
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        tipo = data.get('tipo')
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        limit = data.get('limit', 100)
        
        from nfe_import_functions import listar_notas_fiscais
        notas = listar_notas_fiscais(empresa_id, tipo, data_inicio, data_fim, limit)
        
        return jsonify({
            'success': True,
            'notas': notas,
            'quantidade': len(notas)
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar notas fiscais: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/notas-fiscais/<int:nota_id>', methods=['GET'])
@require_auth
def obter_nota_fiscal(nota_id):
    """Obt�m detalhes completos de uma nota fiscal"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        from nfe_import_functions import obter_detalhes_nota_fiscal
        detalhes = obter_detalhes_nota_fiscal(nota_id)
        
        if not detalhes:
            return jsonify({
                'success': False,
                'error': 'Nota fiscal n�o encontrada'
            }), 404
        
        # Verificar se a nota pertence � empresa
        if detalhes['nota']['empresa_id'] != empresa_id:
            return jsonify({
                'success': False,
                'error': 'Acesso negado'
            }), 403
        
        return jsonify({
            'success': True,
            **detalhes
        })
        
    except Exception as e:
        logger.error(f"Erro ao obter detalhes da nota fiscal: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/notas-fiscais/totais', methods=['POST'])
@require_auth
def calcular_totais_notas():
    """
    Calcula totais de notas fiscais no per�odo
    
    Body: {
        "data_inicio": "2026-01-01",
        "data_fim": "2026-01-31"
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        if not data_inicio or not data_fim:
            return jsonify({
                'success': False,
                'error': 'data_inicio e data_fim s�o obrigat�rios'
            }), 400
        
        from nfe_import_functions import calcular_totais_periodo
        totais = calcular_totais_periodo(empresa_id, data_inicio, data_fim)
        
        return jsonify({
            'success': True,
            'periodo': {
                'data_inicio': data_inicio,
                'data_fim': data_fim
            },
            'totais': totais
        })
        
    except Exception as e:
        logger.error(f"Erro ao calcular totais de notas: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ===== CR�DITOS TRIBUT�RIOS =====

@app.route('/api/creditos-tributarios/calcular', methods=['POST'])
@require_auth
def calcular_creditos_tributarios():
    """
    Calcula cr�ditos tribut�rios de PIS/COFINS
    
    Body: {
        "mes": 1,
        "ano": 2026,
        "tipos": ["INSUMOS", "ENERGIA", "ALUGUEL"] (opcional - calcula todos se n�o informado)
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        mes = data.get('mes')
        ano = data.get('ano')
        tipos = data.get('tipos', [])
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        from creditos_tributarios_functions import (
            calcular_todos_creditos,
            calcular_creditos_insumos,
            calcular_creditos_energia,
            calcular_creditos_aluguel
        )
        
        # Se n�o especificou tipos, calcular todos
        if not tipos:
            resultado = calcular_todos_creditos(empresa_id, mes, ano)
        else:
            # Calcular apenas os tipos especificados
            resultados = {}
            total_pis = 0
            total_cofins = 0
            
            if 'INSUMOS' in tipos:
                r = calcular_creditos_insumos(empresa_id, mes, ano)
                if r.get('success'):
                    resultados['insumos'] = r
                    total_pis += r.get('credito_pis', 0)
                    total_cofins += r.get('credito_cofins', 0)
            
            if 'ENERGIA' in tipos:
                r = calcular_creditos_energia(empresa_id, mes, ano)
                if r.get('success'):
                    resultados['energia'] = r
                    total_pis += r.get('credito_pis', 0)
                    total_cofins += r.get('credito_cofins', 0)
            
            if 'ALUGUEL' in tipos:
                r = calcular_creditos_aluguel(empresa_id, mes, ano)
                if r.get('success'):
                    resultados['aluguel'] = r
                    total_pis += r.get('credito_pis', 0)
                    total_cofins += r.get('credito_cofins', 0)
            
            resultado = {
                'success': True,
                'mes': mes,
                'ano': ano,
                'detalhamento': resultados,
                'resumo': {
                    'total_credito_pis': total_pis,
                    'total_credito_cofins': total_cofins,
                    'total_geral': total_pis + total_cofins
                }
            }
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao calcular cr�ditos tribut�rios: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/creditos-tributarios/listar', methods=['POST'])
@require_auth
def listar_creditos_tributarios():
    """
    Lista cr�ditos tribut�rios calculados
    
    Body: {
        "mes": 1,
        "ano": 2026,
        "tributo": "PIS" ou "COFINS" (opcional)
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        mes = data.get('mes')
        ano = data.get('ano')
        tributo = data.get('tributo')
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        from creditos_tributarios_functions import listar_creditos_periodo
        creditos = listar_creditos_periodo(empresa_id, mes, ano, tributo)
        
        return jsonify({
            'success': True,
            'mes': mes,
            'ano': ano,
            'creditos': creditos,
            'quantidade': len(creditos)
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar cr�ditos tribut�rios: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/creditos-tributarios/resumo', methods=['POST'])
@require_auth
def resumo_creditos_tributarios():
    """
    Obt�m resumo dos cr�ditos tribut�rios
    
    Body: {
        "mes": 1,
        "ano": 2026
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        mes = data.get('mes')
        ano = data.get('ano')
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        from creditos_tributarios_functions import obter_resumo_creditos
        resumo = obter_resumo_creditos(empresa_id, mes, ano)
        
        return jsonify({
            'success': True,
            'mes': mes,
            'ano': ano,
            **resumo
        })
        
    except Exception as e:
        logger.error(f"Erro ao obter resumo de cr�ditos: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ===== DCTF (Declara��o de D�bitos Federais) =====

@app.route('/api/dctf/gerar', methods=['POST'])
@require_auth
def gerar_dctf():
    """
    Gera arquivo DCTF mensal
    
    Body: {
        "mes": 1,
        "ano": 2026
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        mes = data.get('mes')
        ano = data.get('ano')
        
        if not mes or not ano:
            return jsonify({
                'success': False,
                'error': 'mes e ano s�o obrigat�rios'
            }), 400
        
        from dctf_functions import gerar_arquivo_dctf
        resultado = gerar_arquivo_dctf(empresa_id, mes, ano)
        
        if resultado.get('success'):
            logger.info(f"DCTF gerado para empresa {empresa_id}: {mes}/{ano}")
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao gerar DCTF: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ===== DIRF (Declara��o de IR Retido na Fonte) =====

@app.route('/api/dirf/gerar', methods=['POST'])
@require_auth
def gerar_dirf():
    """
    Gera arquivo DIRF anual
    
    Body: {
        "ano": 2025
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        ano = data.get('ano')
        
        if not ano:
            return jsonify({
                'success': False,
                'error': 'ano � obrigat�rio'
            }), 400
        
        from dirf_functions import gerar_arquivo_dirf
        resultado = gerar_arquivo_dirf(empresa_id, ano)
        
        if resultado.get('success'):
            logger.info(f"DIRF gerado para empresa {empresa_id}: {ano}")
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao gerar DIRF: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/dirf/resumo', methods=['POST'])
@require_auth
def resumo_dirf():
    """
    Obt�m resumo da DIRF antes de gerar
    
    Body: {
        "ano": 2025
    }
    """
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        ano = data.get('ano')
        
        if not ano:
            return jsonify({
                'success': False,
                'error': 'ano � obrigat�rio'
            }), 400
        
        from dirf_functions import obter_resumo_dirf
        resultado = obter_resumo_dirf(empresa_id, ano)
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao obter resumo DIRF: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ============================================================================
# RELAT�RIOS FISCAIS - NF-e / CT-e
# ============================================================================

@app.route('/relatorios/fiscal')
@require_auth
@require_permission('relatorios_view')
def relatorios_fiscal():
    """Dashboard de relat�rios fiscais (NF-e, CT-e)"""
    return render_template('relatorios_fiscais.html')


# ===== GERENCIAMENTO DE CERTIFICADOS DIGITAIS =====

@app.route('/api/relatorios/certificados', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def listar_certificados():
    """Lista todos os certificados digitais da empresa"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            cursor.execute("""
                SELECT 
                    id, cnpj, nome_certificado, cuf, ambiente, ativo,
                    ultimo_nsu, max_nsu, data_ultima_busca,
                    valido_de, valido_ate,
                    total_documentos_baixados, total_nfes, total_ctes, total_eventos,
                    criado_em, atualizado_em, senha_pfx
                FROM certificados_digitais
                WHERE empresa_id = %s
                ORDER BY ativo DESC, criado_em DESC
            """, (empresa_id,))
            
            certificados = cursor.fetchall()
            
            # ? Valida se a senha pode ser descriptografada
            from relatorios.nfe import nfe_api
            import os
            chave_cripto = os.environ.get('FERNET_KEY', '').encode('utf-8')
            
            for cert in certificados:
                # Verifica se a senha est� em formato v�lido
                senha_cripto = cert.get('senha_pfx', '')
                if len(senha_cripto) < 50:
                    cert['senha_valida'] = False
                    cert['erro_senha'] = 'Certificado precisa ser recadastrado (senha em formato inv�lido)'
                else:
                    try:
                        # Tenta descriptografar para validar
                        if chave_cripto:
                            nfe_api.descriptografar_senha(senha_cripto, chave_cripto)
                            cert['senha_valida'] = True
                        else:
                            cert['senha_valida'] = False
                            cert['erro_senha'] = 'Chave de criptografia n�o configurada'
                    except Exception as e:
                        cert['senha_valida'] = False
                        cert['erro_senha'] = 'Certificado precisa ser recadastrado'
                
                # Remove senha_pfx do retorno (seguran�a)
                cert.pop('senha_pfx', None)
        
        return jsonify({
            'success': True,
            'certificados': certificados
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar certificados: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/certificado/validar', methods=['POST'])
@require_auth
def validar_certificado():
    """Valida e extrai informa��es de um certificado digital"""
    try:
        dados = request.get_json()
        pfx_base64 = dados.get('pfx_base64')
        senha = dados.get('senha')
        
        if not pfx_base64 or not senha:
            return jsonify({
                'sucesso': False,
                'erro': 'Arquivo e senha s�o obrigat�rios'
            })
        
        # Decodificar base64
        import base64
        pfx_bytes = base64.b64decode(pfx_base64)
        
        # Processar certificado usando fun��o do NFS-e
        from nfse_functions import processar_certificado
        info = processar_certificado(pfx_bytes, senha)
        
        if not info.get('success'):
            return jsonify({
                'sucesso': False,
                'erro': info.get('error', 'Erro ao processar certificado')
            })
        
        # Se n�o conseguiu extrair UF do certificado, tentar consultar ReceitaWS
        if not info.get('uf') and info.get('cnpj'):
            try:
                logger.info(f"?? UF n�o encontrada no certificado, consultando ReceitaWS para CNPJ {info['cnpj']}")
                import requests
                cnpj_limpo = info['cnpj'].replace('.', '').replace('/', '').replace('-', '')
                url = f"https://www.receitaws.com.br/v1/cnpj/{cnpj_limpo}"
                
                response = requests.get(url, timeout=5)
                if response.status_code == 200:
                    dados_empresa = response.json()
                    if dados_empresa.get('status') == 'OK':
                        uf = dados_empresa.get('uf', '').strip().upper()
                        if len(uf) == 2 and uf.isalpha():
                            info['uf'] = uf
                            logger.info(f"? UF obtida via ReceitaWS: {uf}")
                        else:
                            logger.warning(f"?? UF inv�lida retornada pela ReceitaWS: {uf}")
                    else:
                        logger.warning(f"?? ReceitaWS retornou status: {dados_empresa.get('status')}")
                else:
                    logger.warning(f"?? ReceitaWS retornou status code: {response.status_code}")
            except requests.Timeout:
                logger.warning("?? Timeout ao consultar ReceitaWS (5s)")
            except Exception as e:
                logger.warning(f"?? Erro ao consultar ReceitaWS: {str(e)}")
        
        # Retornar informa��es extra�das
        return jsonify({
            'sucesso': True,
            'certificado': {
                'cnpj': info.get('cnpj'),
                'razao_social': info.get('razao_social'),
                'emitente': info.get('emitente'),
                'validade_inicio': info.get('validade_inicio'),
                'validade_fim': info.get('validade_fim'),
                'serial_number': info.get('serial_number'),
                'uf': info.get('uf')
            }
        })
        
    except ValueError as e:
        return jsonify({
            'sucesso': False,
            'erro': 'Senha incorreta ou arquivo inv�lido'
        })
    except Exception as e:
        print(f"Erro ao validar certificado: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': f'Erro ao processar certificado: {str(e)}'
        })


@app.route('/api/relatorios/certificados/extrair-dados', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def extrair_dados_certificado():
    """Extrai dados automaticamente do arquivo .pfx (CNPJ, nome, validades)"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'sucesso': False, 'erro': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        
        if not data.get('pfx_base64') or not data.get('senha'):
            return jsonify({
                'sucesso': False,
                'erro': 'PFX e senha s�o obrigat�rios'
            }), 400
        
        # Importa m�dulo de busca (onde est� CertificadoA1)
        from relatorios.nfe import nfe_busca
        
        # Tenta carregar o certificado
        try:
            cert = nfe_busca.CertificadoA1(
                pfx_base64=data['pfx_base64'],
                senha=data['senha']
            )
            
            if not cert.esta_valido():
                return jsonify({
                    'sucesso': False,
                    'erro': 'Certificado fora do prazo de validade',
                    'valido_de': cert.cert_data.get('valido_de').isoformat() if cert.cert_data.get('valido_de') else None,
                    'valido_ate': cert.cert_data.get('valido_ate').isoformat() if cert.cert_data.get('valido_ate') else None
                })
            
            # Busca dados da empresa para completar campos (UF)
            with get_db_connection(empresa_id=empresa_id) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT estado
                    FROM empresas
                    WHERE id = %s
                """, (empresa_id,))
                empresa = cursor.fetchone()
            
            estado_db = (empresa[0] if empresa else '') or ''
            
            # Mapa de UF (sigla) para c�digo IBGE (cUF)
            uf_para_codigo = {
                'AC': '12', 'AL': '27', 'AP': '16', 'AM': '13', 'BA': '29',
                'CE': '23', 'DF': '53', 'ES': '32', 'GO': '52', 'MA': '21',
                'MT': '51', 'MS': '50', 'MG': '31', 'PA': '15', 'PB': '25',
                'PR': '41', 'PE': '26', 'PI': '22', 'RJ': '33', 'RN': '24',
                'RS': '43', 'RO': '11', 'RR': '14', 'SC': '42', 'SP': '35',
                'SE': '28', 'TO': '17'
            }
            
            # --- Extrai nome do certificado a partir do Subject CN ---
            # Formato t�pico BR: CN=NOME EMPRESA:CNPJ14DIGITS ou CN=NOME:CNPJ-TITULAR
            import re
            subject_str = cert.cert_data.get('subject', '')
            nome_cert_extraido = ''
            
            # Tenta extrair de CN=
            cn_match = re.search(r'(?:^|,)CN=([^,]+)', subject_str)
            if cn_match:
                cn_value = cn_match.group(1).strip()
                # Remove parte ap�s ":" (que geralmente � CNPJ ou c�digo)
                nome_cert_extraido = cn_value.split(':')[0].strip()
            
            # Se n�o encontrou nome v�lido no CN, usa O= (Organization)
            if not nome_cert_extraido or len(nome_cert_extraido) < 3:
                o_match = re.search(r'(?:^|,)O=([^,]+)', subject_str)
                if o_match:
                    nome_cert_extraido = o_match.group(1).strip()

            # Fallback: usa nome do arquivo
            if not nome_cert_extraido or nome_cert_extraido.upper() in ('ICP-BRASIL', 'ICP-BRASIL'):
                nome_cert_extraido = 'Certificado Digital'
            
            # --- UF: tenta extrair do Subject, fallback para empresa ---
            uf_sigla = ''
            # Alguns certs t�m ST= (State) no subject
            st_match = re.search(r'(?:^|,)ST=([A-Z]{2})', subject_str)
            if st_match:
                uf_sigla = st_match.group(1).strip()
            
            # Fallback: estado da empresa (somente se for uma sigla v�lida de 2 letras)
            if not uf_sigla and len(estado_db.strip()) == 2 and estado_db.strip().upper() in uf_para_codigo:
                uf_sigla = estado_db.strip().upper()
            
            cuf = uf_para_codigo.get(uf_sigla, '')  # Deixa vazio se n�o souber
            
            # Extrai CNPJ do certificado
            cnpj_cert = cert.cert_data.get('cnpj', '')
            
            logger.info(f"[CERTIFICADO DEBUG] Subject: {subject_str[:100]}")
            logger.info(f"[CERTIFICADO DEBUG] Nome extra�do do cert: '{nome_cert_extraido}'")
            logger.info(f"[CERTIFICADO DEBUG] CNPJ cert: '{cnpj_cert}'")
            logger.info(f"[CERTIFICADO DEBUG] UF: sigla='{uf_sigla}' cuf='{cuf}'")
            
            resultado = {
                'sucesso': True,
                'dados': {
                    'cnpj': cnpj_cert,
                    'nome_certificado': nome_cert_extraido,
                    'valido_de': cert.cert_data.get('valido_de').isoformat() if cert.cert_data.get('valido_de') else None,
                    'valido_ate': cert.cert_data.get('valido_ate').isoformat() if cert.cert_data.get('valido_ate') else None,
                    'cuf': cuf,
                    'uf_sigla': uf_sigla,
                    'subject': cert.cert_data.get('subject', ''),
                    'issuer': cert.cert_data.get('issuer', '')
                }
            }
            
            return jsonify(resultado)
            
        except Exception as e:
            return jsonify({
                'sucesso': False,
                'erro': f'Erro ao ler certificado: {str(e)}'
            })
        
    except Exception as e:
        logger.error(f"Erro ao extrair dados do certificado: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/certificados/novo', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def cadastrar_certificado():
    """Cadastra um novo certificado digital"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id') or usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'sucesso': False, 'erro': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        
        # Valida��o
        required = ['nome_certificado', 'pfx_base64', 'senha', 'cuf']
        missing = [f for f in required if not data.get(f)]
        if missing:
            return jsonify({
                'sucesso': False,
                'erro': f'Campos obrigat�rios faltando: {", ".join(missing)}'
            }), 400
        # CNPJ pode vir do form ou ser derivado da empresa
        if not data.get('cnpj'):
            try:
                with get_db_connection(allow_global=True) as conn:
                    cur = conn.cursor()
                    cur.execute("SELECT cnpj FROM empresas WHERE id = %s", (empresa_id,))
                    row = cur.fetchone()
                    data['cnpj'] = row[0] if row else ''
            except Exception:
                data['cnpj'] = ''
        
        # Importa m�dulo de API
        from relatorios.nfe import nfe_api
        
        # Salva certificado
        logger.info(f"[CERTIFICADO] Salvando certificado para empresa {empresa_id}, CNPJ {data['cnpj']}")
        resultado = nfe_api.salvar_certificado(
            empresa_id=empresa_id,
            cnpj=data['cnpj'],
            nome_certificado=data['nome_certificado'],
            pfx_base64=data['pfx_base64'],
            senha=data['senha'],
            cuf=int(data['cuf']),
            ambiente=data.get('ambiente', 'producao'),
            usuario_id=usuario['id']
        )
        
        logger.info(f"[CERTIFICADO] Resultado: {resultado}")
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao cadastrar certificado: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/certificados/<int:certificado_id>/desativar', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def desativar_certificado(certificado_id):
    """Desativa um certificado digital"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE certificados_digitais
                SET ativo = FALSE,
                    atualizado_em = NOW(),
                    atualizado_por = %s
                WHERE id = %s AND empresa_id = %s
            """, (usuario['id'], certificado_id, empresa_id))
            
            conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Certificado desativado com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao desativar certificado: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/certificados/<int:certificado_id>/recadastrar', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def recadastrar_certificado(certificado_id):
    """Atualiza pfx e senha de um certificado existente pelo ID (sem depender de CNPJ)"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403

        data = request.get_json()
        pfx_base64 = data.get('pfx_base64', '').strip()
        senha = data.get('senha', '').strip()

        if not pfx_base64 or not senha:
            return jsonify({'success': False, 'error': 'Arquivo PFX e senha s�o obrigat�rios'}), 400

        from relatorios.nfe import nfe_api, nfe_busca
        import os

        # Valida certificado
        try:
            cert = nfe_busca.CertificadoA1(pfx_base64=pfx_base64, senha=senha)
            if not cert.esta_valido():
                return jsonify({'success': False, 'error': 'Certificado fora do prazo de validade'}), 400
            dados_cert = cert.cert_data
        except Exception as e:
            return jsonify({'success': False, 'error': f'Certificado inv�lido ou senha incorreta: {str(e)}'}), 400

        # Criptografa senha
        chave_str = os.environ.get('FERNET_KEY', '').strip()
        if not chave_str:
            return jsonify({'success': False, 'error': 'FERNET_KEY n�o configurada no servidor'}), 500

        try:
            senha_cripto = nfe_api.criptografar_senha(senha, chave_str.encode('utf-8'))
        except Exception as e:
            return jsonify({'success': False, 'error': f'Erro ao criptografar senha: {str(e)}'}), 500

        logger.info(f"[RECADASTRAR] Cert ID {certificado_id} - senha criptografada: {len(senha_cripto)} chars")

        # Atualiza diretamente pelo ID (sem depender de CNPJ)
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            # Verifica que o cert pertence � empresa
            cursor.execute(
                "SELECT id FROM certificados_digitais WHERE id = %s AND empresa_id = %s",
                (certificado_id, empresa_id)
            )
            if not cursor.fetchone():
                return jsonify({'success': False, 'error': 'Certificado n�o encontrado'}), 404

            cursor.execute("""
                UPDATE certificados_digitais
                SET pfx_base64   = %s,
                    senha_pfx    = %s,
                    valido_de    = %s,
                    valido_ate   = %s,
                    ativo        = TRUE,
                    atualizado_em = NOW(),
                    atualizado_por = %s
                WHERE id = %s AND empresa_id = %s
            """, (
                pfx_base64,
                senha_cripto,
                dados_cert.get('valido_de'),
                dados_cert.get('valido_ate'),
                usuario['id'],
                certificado_id,
                empresa_id,
            ))
            conn.commit()

        logger.info(f"[RECADASTRAR] ? Certificado ID {certificado_id} atualizado com sucesso")
        return jsonify({'success': True, 'message': 'Certificado recadastrado com sucesso!'})

    except Exception as e:
        logger.error(f"Erro ao recadastrar certificado {certificado_id}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== BUSCA E PROCESSAMENTO DE DOCUMENTOS =====

def _auto_obter_certificado_id(empresa_id):
    """
    Retorna o certificado_id do certificado cadastrado em
    'Dados da Empresa e Certificado Digital' (tabela certificados_digitais).
    Reativa automaticamente se estiver inativo.
    """
    try:
        with get_db_connection(allow_global=True) as conn:
            cursor = conn.cursor()
            # Prioriza cert ativo; se n�o existir, pega qualquer um para reativar
            cursor.execute("""
                SELECT id, ativo, senha_pfx
                FROM certificados_digitais
                WHERE empresa_id = %s AND pfx_base64 IS NOT NULL AND pfx_base64 != ''
                ORDER BY ativo DESC, id DESC
                LIMIT 1
            """, (empresa_id,))
            row = cursor.fetchone()
            if not row:
                logger.warning(f"[_auto_cert] Nenhum certificado encontrado para empresa {empresa_id}")
                return None

            cert_id = row[0] if not isinstance(row, dict) else row['id']
            cert_ativo = row[1] if not isinstance(row, dict) else row['ativo']
            senha_pfx = row[2] if not isinstance(row, dict) else row['senha_pfx']

            if not cert_ativo:
                logger.info(f"[_auto_cert] Reativando cert ID {cert_id} (empresa {empresa_id})")
                cursor.execute("UPDATE certificados_digitais SET ativo = TRUE WHERE id = %s", (cert_id,))
                conn.commit()

            logger.info(f"[_auto_cert] Usando cert ID {cert_id} (empresa {empresa_id}, ativo={cert_ativo}, senha_len={len(senha_pfx or '')})")
            return cert_id

    except Exception as e:
        logger.error(f"[_auto_obter_certificado_id] empresa {empresa_id}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


@app.route('/api/relatorios/buscar-documentos', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def buscar_documentos():
    """Inicia busca autom�tica de documentos na SEFAZ"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id') or usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'sucesso': False, 'erro': 'Empresa n�o identificada'}), 403

        data = request.get_json() or {}
        certificado_id = data.get('certificado_id') or _auto_obter_certificado_id(empresa_id)

        if not certificado_id:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum certificado digital cadastrado para esta empresa. '
                        'Acesse ?? Dados da Empresa e Certificado Digital para cadastrar.'
            }), 400

        # -- Diagn�stico do certificado antes de tentar usar ------------------
        diag = {'cert_id': certificado_id}
        try:
            with get_db_connection(allow_global=True) as _conn:
                _cur = _conn.cursor()
                _cur.execute("""
                    SELECT id, ativo, empresa_id, cnpj, cuf, ambiente,
                           LENGTH(pfx_base64) AS pfx_len,
                           LENGTH(senha_pfx)  AS senha_len,
                           valido_ate
                    FROM certificados_digitais WHERE id = %s
                """, (certificado_id,))
                row = _cur.fetchone()
                if row:
                    keys = ['id','ativo','empresa_id','cnpj','cuf','ambiente',
                            'pfx_len','senha_len','valido_ate']
                    diag.update(dict(zip(keys, row)) if not isinstance(row, dict) else row)
                else:
                    diag['erro_diag'] = f'ID {certificado_id} n�o existe na tabela'
        except Exception as de:
            diag['erro_diag'] = str(de)

        logger.info(f"[buscar_documentos] Diagn�stico cert: {diag}")

        if diag.get('erro_diag'):
            return jsonify({
                'sucesso': False,
                'erro': f'Certificado ID {certificado_id} n�o encontrado no banco.',
                'diagnostico': diag
            })

        if not diag.get('ativo'):
            return jsonify({
                'sucesso': False,
                'erro': f'Certificado ID {certificado_id} est� inativo (ativo=False). Recadastre em ?? Dados da Empresa.',
                'diagnostico': diag
            })

        if not diag.get('pfx_len'):
            return jsonify({
                'sucesso': False,
                'erro': f'Certificado ID {certificado_id} n�o tem PFX gravado (pfx_base64 vazio). Recadastre em ?? Dados da Empresa.',
                'diagnostico': diag
            })

        # Detecta senha Fernet sem FERNET_KEY configurada
        import os as _os
        fernet_key_presente = bool(_os.environ.get('FERNET_KEY', ''))
        senha_parece_fernet = diag.get('senha_len', 0) >= 50
        if senha_parece_fernet and not fernet_key_presente:
            return jsonify({
                'sucesso': False,
                'erro': '?? A senha do certificado est� criptografada (Fernet, {} chars), mas FERNET_KEY n�o est� configurada no servidor. '
                        'SOLU��O: Recadastre o certificado agora em ?? Dados da Empresa � a nova vers�o salva sem criptografia quando FERNET_KEY est� ausente.'.format(diag['senha_len']),
                'diagnostico': diag
            })
        # ---------------------------------------------------------------------

        from relatorios.nfe import nfe_api

        # nsu_override permite re-buscar desde um NSU espec�fico (ex: '000000000000000')
        nsu_override = data.get('nsu_override')  # optional, string or null

        resultado = nfe_api.buscar_e_processar_novos_documentos(
            certificado_id=certificado_id,
            usuario_id=usuario['id'],
            nsu_override=nsu_override
        )

        # Se ainda falhou, inclui diagn�stico na resposta para debug
        if not resultado.get('sucesso'):
            resultado['diagnostico'] = diag

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Erro ao buscar documentos: {e}")
        import traceback as _tb
        return jsonify({
            'sucesso': False,
            'erro': f'Erro no servidor: {str(e)}',
            'traceback': _tb.format_exc()
        }), 500


@app.route('/api/relatorios/consultar-chave', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def consultar_por_chave():
    """Consulta uma NF-e espec�fica por chave de acesso"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id') or usuario.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json() or {}
        chave = data.get('chave')
        certificado_id = data.get('certificado_id') or _auto_obter_certificado_id(empresa_id)

        if not chave:
            return jsonify({'success': False, 'error': 'chave � obrigat�ria'}), 400

        if not certificado_id:
            return jsonify({
                'success': False,
                'error': 'Nenhum certificado digital encontrado para esta empresa.'
            }), 400
        
        # Importa m�dulos
        from relatorios.nfe import nfe_api, nfe_busca
        
        # Carrega certificado
        cert = nfe_api.obter_certificado(certificado_id)
        if not cert:
            return jsonify({
                'success': False,
                'error': 'Certificado n�o encontrado ou inv�lido'
            }), 404
        
        # Consulta por chave (auto-detecta NF-e/CT-e pelo modelo na chave)
        resultado = nfe_busca.consultar_documento_por_chave(
            certificado=cert,
            chave=chave,
            ambiente=data.get('ambiente', 'producao')
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"Erro ao consultar chave: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ===== LISTAGEM E CONSULTA DE DOCUMENTOS =====

@app.route('/api/relatorios/documentos', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def listar_documentos():
    """Lista documentos fiscais com filtros"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        # Par�metros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo = request.args.get('tipo')  # NFe, CTe, Evento
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Query base
            sql = """
                SELECT 
                    id, nsu, chave, tipo_documento, schema_name,
                    numero_documento, serie, valor_total,
                    cnpj_emitente, nome_emitente,
                    cnpj_destinatario, nome_destinatario,
                    data_emissao, data_busca, processado,
                    COALESCE(cancelado, FALSE) AS cancelado,
                    cancelamento_motivo, cancelamento_data
                FROM documentos_fiscais_log
                WHERE empresa_id = %s
            """
            
            params = [empresa_id]
            
            # Filtros opcionais
            if data_inicio:
                sql += " AND data_busca >= %s"
                params.append(data_inicio)
            
            if data_fim:
                sql += " AND data_busca <= %s"
                params.append(data_fim)
            
            if tipo:
                sql += " AND tipo_documento = %s"
                params.append(tipo)
            else:
                # Por padrão exibe apenas NF-e e CT-e (não exibe eventos internos)
                sql += " AND tipo_documento IN ('NFe', 'CTe')"
            
            # Pagina��o
            sql += " ORDER BY data_busca DESC"
            sql += " LIMIT %s OFFSET %s"
            params.extend([per_page, (page - 1) * per_page])
            
            cursor.execute(sql, params)
            documentos = cursor.fetchall()
            
            # Total de registros
            sql_count = """
                SELECT COUNT(*) as total
                FROM documentos_fiscais_log
                WHERE empresa_id = %s
            """
            count_params = [empresa_id]
            
            if data_inicio:
                sql_count += " AND data_busca >= %s"
                count_params.append(data_inicio)
            
            if data_fim:
                sql_count += " AND data_busca <= %s"
                count_params.append(data_fim)
            
            if tipo:
                sql_count += " AND tipo_documento = %s"
                count_params.append(tipo)
            else:
                sql_count += " AND tipo_documento IN ('NFe', 'CTe')"
            
            cursor.execute(sql_count, count_params)
            total = cursor.fetchone()['total']
        
        return jsonify({
            'success': True,
            'documentos': documentos,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })
        
    except Exception as e:
        logger.error(f"Erro ao listar documentos: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/documento/<int:doc_id>', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def obter_documento(doc_id):
    """Obt�m detalhes de um documento espec�fico"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            cursor.execute("""
                SELECT *
                FROM documentos_fiscais_log
                WHERE id = %s AND empresa_id = %s
            """, (doc_id, empresa_id))
            
            documento = cursor.fetchone()
        
        if not documento:
            return jsonify({
                'success': False,
                'error': 'Documento n�o encontrado'
            }), 404
        
        return jsonify({
            'success': True,
            'documento': documento
        })
        
    except Exception as e:
        logger.error(f"Erro ao obter documento: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/documento/<int:doc_id>/xml', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def download_xml(doc_id):
    """Download do XML de um documento"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            # Garante coluna xml_content (migração lazy)
            try:
                cursor.execute("ALTER TABLE documentos_fiscais_log ADD COLUMN IF NOT EXISTS xml_content TEXT")
                conn.commit()
            except Exception:
                conn.rollback()
            cursor.execute("""
                SELECT chave, caminho_xml, tipo_documento, xml_content
                FROM documentos_fiscais_log
                WHERE id = %s AND empresa_id = %s
            """, (doc_id, empresa_id))
            
            row = cursor.fetchone()
        
        if not row:
            return jsonify({
                'success': False,
                'error': 'Documento n�o encontrado'
            }), 404
        
        # Pool usa RealDictCursor — row é dict, nunca tupla
        chave         = row['chave']
        caminho_xml   = row['caminho_xml']
        tipo_doc      = row['tipo_documento']
        xml_content_db = row['xml_content']

        # Tenta filesystem; senao usa xml_content do banco (Railway ephemeral)
        from io import BytesIO as _BytesIO
        if caminho_xml and os.path.exists(str(caminho_xml)):
            return send_file(
                caminho_xml,
                mimetype='application/xml',
                as_attachment=True,
                download_name=f'{tipo_doc}_{chave}.xml'
            )
        elif xml_content_db:
            xml_bytes = xml_content_db.encode("utf-8") if isinstance(xml_content_db, str) else xml_content_db
            buf = _BytesIO(xml_bytes)
            buf.seek(0)
            return send_file(
                buf,
                mimetype='application/xml',
                as_attachment=True,
                download_name=f'{tipo_doc}_{chave}.xml'
            )
        else:
            return jsonify({
                'success': False,
                'error': 'XML nao encontrado (nem no storage, nem no banco)'
            }), 404
        
    except Exception as e:
        logger.error(f"Erro ao fazer download do XML: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/documento/<int:doc_id>/pdf', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def download_pdf_documento(doc_id):
    """Gera DANFE (NF-e) ou DACTE (CT-e) em PDF usando BrazilFiscalReport"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa nao identificada'}), 403

        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            # Garante coluna xml_content (migração lazy)
            try:
                cursor.execute("ALTER TABLE documentos_fiscais_log ADD COLUMN IF NOT EXISTS xml_content TEXT")
                conn.commit()
            except Exception:
                conn.rollback()
            cursor.execute("""
                SELECT chave, caminho_xml, tipo_documento, xml_content, schema_name,
                       cnpj_destinatario
                FROM documentos_fiscais_log
                WHERE id = %s AND empresa_id = %s
            """, (doc_id, empresa_id))
            row = cursor.fetchone()

        if not row:
            return jsonify({'success': False, 'error': 'Documento nao encontrado'}), 404

        # Pool usa RealDictCursor — row é dict, nunca tupla
        chave          = row['chave']
        caminho_xml    = row['caminho_xml']
        tipo_doc       = row['tipo_documento']
        xml_content_db = row['xml_content']
        schema_name    = (row.get('schema_name') or '').lower()
        cnpj_dest_db   = (row.get('cnpj_destinatario') or '').replace('.', '').replace('/', '').replace('-', '')

        # Tenta filesystem primeiro; senao usa xml_content do banco (Railway ephemeral)
        if caminho_xml and os.path.exists(str(caminho_xml)):
            with open(caminho_xml, 'rb') as f:
                xml_bytes = f.read()
        elif xml_content_db:
            xml_bytes = xml_content_db.encode('utf-8') if isinstance(xml_content_db, str) else xml_content_db
        else:
            return jsonify({
                'success': False,
                'error': 'XML nao encontrado. Este documento foi importado antes do armazenamento de XML ser ativado. '
                         'Acesse "Documentos Fiscais" > "Buscar Documentos" para re-sincronizar e o PDF passara a funcionar.'
            }), 404

        # Detecta XML de resumo (resNFe / resCTe) — não tem dados suficientes para DANFE
        xml_head = xml_bytes[:500].decode('utf-8', errors='ignore') if xml_bytes else ''
        is_resumo = (
            schema_name.startswith('res')
            or '<resNFe' in xml_head
            or '<resCTe' in xml_head
            or '<retDistDFeInt' in xml_head
        )
        if is_resumo:
            # ----------------------------------------------------------------
            # Auto-manifest "Ciência da Operação" + download procNFe completo
            # Só aplicável a NF-e (resNFe). CT-e resumo não tem manifestação.
            # ----------------------------------------------------------------
            tipo_norm_check = (tipo_doc or '').upper().replace('-','').replace('_','').replace(' ','')
            if tipo_norm_check not in ('NFE', '55', 'NFCE', '65'):
                return jsonify({
                    'success': False,
                    'error': 'Este documento e um XML de resumo (resCTe) sem dados suficientes para PDF. '
                             'Re-sincronize em "Buscar Documentos" para obter o XML completo.'
                }), 422

            cert_id_pdf = _auto_obter_certificado_id(empresa_id)
            if not cert_id_pdf:
                return jsonify({
                    'success': False,
                    'error': 'Certificado digital nao encontrado. Cadastre em Dados da Empresa para gerar o PDF.'
                }), 422

            try:
                from relatorios.nfe import nfe_api, nfe_busca

                cert_pdf = nfe_api.obter_certificado(cert_id_pdf)
                if not cert_pdf:
                    return jsonify({
                        'success': False,
                        'error': 'Certificado invalido ou expirado. Recadastre em Dados da Empresa.'
                    }), 422

                # Ambiente do certificado cadastrado
                with get_db_connection(allow_global=True) as _cc:
                    _ccur = _cc.cursor()
                    _ccur.execute(
                        "SELECT ambiente FROM certificados_digitais WHERE id = %s",
                        (cert_id_pdf,)
                    )
                    _crow = _ccur.fetchone()
                    cert_ambiente = (
                        (_crow[0] if not isinstance(_crow, dict) else _crow.get('ambiente', 'producao'))
                        or 'producao'
                    )

                # CNPJ do destinatário: usa o armazenado na linha ou cai para o CNPJ da empresa
                cnpj_dest = cnpj_dest_db
                if not cnpj_dest:
                    with get_db_connection(empresa_id=empresa_id) as _ce:
                        _ceu = _ce.cursor()
                        _ceu.execute("SELECT cnpj FROM empresas WHERE id = %s", (empresa_id,))
                        _cer = _ceu.fetchone()
                        raw_cnpj = (_cer[0] if not isinstance(_cer, dict) else (_cer.get('cnpj') or '')) or ''
                        import re as _re2
                        cnpj_dest = _re2.sub(r'\D', '', raw_cnpj)

                # 1. Manifesta Ciência da Operação
                logger.info(f"[PDF] Manifestando Ciencia da Operacao: chave={chave} cnpj={cnpj_dest} amb={cert_ambiente}")
                manif = nfe_busca.manifestar_ciencia_operacao(cert_pdf, chave, cnpj_dest, cert_ambiente)
                logger.info(f"[PDF] Manifestacao result: {manif}")

                if not manif.get('sucesso'):
                    return jsonify({
                        'success': False,
                        'error': (
                            f'Falha ao manifestar NF-e (cStat={manif.get("codigo_sefaz","?")}: '
                            f'{manif.get("mensagem") or manif.get("erro","?")}). '
                            'Tente re-sincronizar em "Buscar Documentos".'
                        )
                    }), 422

                # 2. Aguarda processamento no SEFAZ
                import time as _time
                _time.sleep(5)

                # 3. Baixa procNFe completo
                logger.info(f"[PDF] Baixando procNFe: chave={chave}")
                down = nfe_busca.baixar_procnfe_completo(cert_pdf, chave, cert_ambiente)
                logger.info(f"[PDF] Download procNFe: sucesso={down.get('sucesso')} erro={down.get('erro','')}")

                if not down.get('sucesso'):
                    return jsonify({
                        'success': False,
                        'error': (
                            f'Falha ao baixar NF-e completa: {down.get("erro","?")}. '
                            'Tente re-sincronizar em "Buscar Documentos".'
                        )
                    }), 422

                xml_bytes = down['xml_bytes']

                # 4. Persiste procNFe no banco para próximas consultas
                try:
                    xml_str_save = xml_bytes.decode('utf-8') if isinstance(xml_bytes, bytes) else xml_bytes
                    with get_db_connection(empresa_id=empresa_id) as _cs:
                        _cs_cur = _cs.cursor()
                        _cs_cur.execute("""
                            UPDATE documentos_fiscais_log
                            SET xml_content = %s, schema_name = 'procNFe_v4.00'
                            WHERE id = %s AND empresa_id = %s
                        """, (xml_str_save, doc_id, empresa_id))
                        _cs.commit()
                    logger.info(f"[PDF] procNFe salvo no banco para doc_id={doc_id}")
                except Exception as _save_err:
                    logger.warning(f"[PDF] Falha ao persistir procNFe: {_save_err}")

                # Continua com geração do DANFE abaixo (xml_bytes agora é procNFe)

            except Exception as _me:
                logger.error(f"[PDF] Erro no auto-manifest: {_me}")
                import traceback as _tb2
                logger.error(_tb2.format_exc())
                return jsonify({
                    'success': False,
                    'error': f'Erro ao processar manifestacao automatica: {str(_me)}'
                }), 500

        from io import BytesIO as _BytesIO
        buffer = _BytesIO()

        # Normaliza tipo para aceitar variantes: NF-e, nfe, 55, NFe etc.
        tipo_norm = (tipo_doc or '').upper().replace('-', '').replace('_', '').replace(' ', '')
        eh_nfe = tipo_norm in ('NFE', '55', 'NFCE', '65')
        eh_cte = tipo_norm in ('CTE', '57', 'CTEOS', '67')

        if eh_nfe:
            try:
                from brazilfiscalreport.danfe import Danfe
            except ImportError:
                return jsonify({'success': False, 'error': 'Biblioteca brazilfiscalreport nao instalada'}), 500
            try:
                danfe = Danfe(xml=xml_bytes)
                danfe.output(buffer)
            except Exception as danfe_err:
                logger.error(f"Erro brazilfiscalreport Danfe: {danfe_err}")
                return jsonify({
                    'success': False,
                    'error': f'Erro ao processar XML para DANFE: {danfe_err}. '
                             'Verifique se o XML e um procNFe completo com protocolo de autorizacao.'
                }), 422
            download_name = f'DANFE_{chave}.pdf'
        elif eh_cte:
            try:
                from brazilfiscalreport.dacte import Dacte
            except ImportError:
                return jsonify({'success': False, 'error': 'Biblioteca brazilfiscalreport nao instalada'}), 500
            try:
                dacte = Dacte(xml=xml_bytes)
                dacte.output(buffer)
            except Exception as dacte_err:
                logger.error(f"Erro brazilfiscalreport Dacte: {dacte_err}")
                return jsonify({
                    'success': False,
                    'error': f'Erro ao processar XML para DACTE: {dacte_err}. '
                             'Verifique se o XML e um procCTe completo com protocolo de autorizacao.'
                }), 422
            download_name = f'DACTE_{chave}.pdf'
        else:
            return jsonify({'success': False, 'error': f'Tipo {tipo_doc!r} nao suporta geracao de PDF (esperado NFe ou CTe)'}), 400

        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=download_name
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF do documento: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': f'Erro ao gerar PDF: {str(e)}'}), 500


@app.route('/api/relatorios/documentos/apagar-tudo', methods=['DELETE'])
@require_auth
@require_permission('relatorios_delete')
def apagar_documentos_fiscais():
    """Apaga todos os documentos fiscais da empresa e reseta o NSU dos certificados"""
    try:
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa nao identificada'}), 403

        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()

            # Conta documentos antes de deletar
            cursor.execute(
                "SELECT COUNT(*) FROM documentos_fiscais_log WHERE empresa_id = %s",
                (empresa_id,)
            )
            total = cursor.fetchone()[0]

            # Apaga todos os documentos
            cursor.execute(
                "DELETE FROM documentos_fiscais_log WHERE empresa_id = %s",
                (empresa_id,)
            )

            # Reseta o NSU de todos os certificados da empresa
            cursor.execute("""
                UPDATE certificados_digitais
                SET ultimo_nsu = '000000000000000',
                    proximo_nsu = '000000000000000',
                    total_documentos_baixados = 0,
                    total_nfes = 0,
                    total_ctes = 0,
                    total_eventos = 0
                WHERE empresa_id = %s
            """, (empresa_id,))

            conn.commit()

        logger.info(f"[APAGAR-DOCS] Empresa {empresa_id}: {total} documentos apagados e NSU resetado")
        return jsonify({
            'success': True,
            'message': f'{total} documentos apagados e NSU resetado com sucesso.'
        })

    except Exception as e:
        logger.error(f"Erro ao apagar documentos fiscais: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ===== ESTAT�STICAS E DASHBOARDS =====

@app.route('/api/relatorios/estatisticas', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def obter_estatisticas():
    """Obt�m estat�sticas de documentos fiscais da empresa"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        from relatorios.nfe import nfe_api
        
        stats = nfe_api.obter_estatisticas_empresa(empresa_id)
        
        return jsonify({
            'success': True,
            'estatisticas': stats
        })
        
    except Exception as e:
        logger.error(f"Erro ao obter estat�sticas: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


@app.route('/api/relatorios/nsu-status', methods=['GET'])
@require_auth
@require_permission('relatorios_view')
def obter_nsu_status():
    """Obt�m status dos NSUs de todos os certificados"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        with get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            cursor.execute("""
                SELECT 
                    id, nome_certificado, cnpj, ambiente,
                    ultimo_nsu, max_nsu, data_ultima_busca,
                    total_documentos_baixados,
                    CASE 
                        WHEN max_nsu IS NOT NULL AND ultimo_nsu::BIGINT < max_nsu::BIGINT 
                        THEN (max_nsu::BIGINT - ultimo_nsu::BIGINT)
                        ELSE 0
                    END as nsus_pendentes
                FROM certificados_digitais
                WHERE empresa_id = %s AND ativo = TRUE
                ORDER BY data_ultima_busca DESC NULLS LAST
            """, (empresa_id,))
            
            certificados = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'certificados': certificados
        })
        
    except Exception as e:
        logger.error(f"Erro ao obter status NSU: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ===== EXPORTA��O =====

@app.route('/api/relatorios/exportar-excel', methods=['POST'])
@require_auth
@require_permission('relatorios_view')
def exportar_excel():
    """Exporta documentos fiscais para Excel"""
    try:
        usuario = get_usuario_logado()
        empresa_id = session.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa n�o identificada'}), 403
        
        data = request.get_json()
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        tipo = data.get('tipo')
        
        from relatorios.nfe import nfe_api
        
        # Lista documentos
        documentos = nfe_api.listar_documentos_periodo(
            empresa_id=empresa_id,
            data_inicio=datetime.fromisoformat(data_inicio) if data_inicio else datetime.now() - timedelta(days=30),
            data_fim=datetime.fromisoformat(data_fim) if data_fim else datetime.now(),
            tipo=tipo
        )
        
        # Gera Excel
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Documentos Fiscais"
        
        # Cabe�alho
        headers = ['NSU', 'Chave', 'Tipo', 'N�mero', 'S�rie', 'Valor', 
                  'Emitente CNPJ', 'Emitente Nome', 'Destinat�rio CNPJ', 
                  'Destinat�rio Nome', 'Data Emiss�o', 'Data Busca']
        
        ws.append(headers)
        
        # Estilos do cabe�alho
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for doc in documentos:
            ws.append([
                doc.get('nsu'),
                doc.get('chave'),
                doc.get('tipo'),
                doc.get('numero'),
                doc.get('serie'),
                doc.get('valor'),
                doc.get('emitente_cnpj'),
                doc.get('emitente_nome'),
                doc.get('destinatario_cnpj'),
                doc.get('destinatario_nome'),
                doc.get('data_emissao'),
                doc.get('data_busca')
            ])
        
        # Ajusta largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Salva em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Envia arquivo
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'documentos_fiscais_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {e}")
        return jsonify({
            'success': False,
            'error': f'Erro no servidor: {str(e)}'
        }), 500


# ===== INTERFACE WEB SPED =====

@app.route('/sped')
@require_auth
def sped_interface():
    """Interface web para gera��o de arquivos SPED"""
    return render_template('sped_interface.html')


# ==============================================================================
# ?? ENDPOINT ADMINISTRATIVO TEMPOR�RIO: Corrigir cliente_id nos contratos
# ==============================================================================
@app.route('/api/admin/fix-contratos-cliente-id', methods=['POST'])
@require_admin
def fix_contratos_cliente_id():
    """
    Endpoint administrativo para corrigir cliente_id NULL nos contratos
    Atualiza baseado nas sess�es vinculadas ou busca pelo cliente_nome
    """
    try:
        print("\n" + "="*80)
        print("?? INICIANDO CORRE��O DE CLIENTE_ID NOS CONTRATOS")
        print("="*80)
        
        # Buscar todos os contratos
        query_contratos = """
            SELECT id, numero, nome, cliente_id, cliente_nome, empresa_id
            FROM contratos
            ORDER BY id
        """
        
        contratos = database.execute_query(query_contratos)
        print(f"\n?? Total de contratos encontrados: {len(contratos)}")
        
        contratos_corrigidos = 0
        contratos_com_problemas = []
        detalhes = []
        
        for contrato in contratos:
            contrato_id = contrato['id']
            numero = contrato['numero']
            cliente_id_atual = contrato['cliente_id']
            cliente_nome = contrato['cliente_nome']
            
            print(f"\n{'='*80}")
            print(f"?? Contrato ID {contrato_id} - {numero}")
            print(f"   Nome contrato: {contrato['nome']}")
            print(f"   Cliente ID atual: {cliente_id_atual}")
            print(f"   Cliente Nome: {cliente_nome}")
            
            # Se j� tem cliente_id, pular
            if cliente_id_atual:
                print(f"   ? J� tem cliente_id, pulando...")
                detalhes.append({
                    'contrato_id': contrato_id,
                    'numero': numero,
                    'status': 'OK',
                    'cliente_id': cliente_id_atual
                })
                continue
            
            # Buscar sess�es deste contrato
            query_sessoes = """
                SELECT DISTINCT cliente_id, cliente_nome
                FROM sessoes
                WHERE contrato_id = %s AND cliente_id IS NOT NULL
            """
            
            sessoes = database.execute_query(query_sessoes, (contrato_id,))
            
            if not sessoes:
                print(f"   ?? Sem sess�es com cliente_id para este contrato")
                
                # Se tem cliente_nome, tentar buscar pelo nome
                if cliente_nome:
                    query_cliente = """
                        SELECT id, razao_social
                        FROM clientes
                        WHERE razao_social ILIKE %s
                        LIMIT 1
                    """
                    clientes = database.execute_query(query_cliente, (cliente_nome,))
                    
                    if clientes:
                        novo_cliente_id = clientes[0]['id']
                        print(f"   ?? Cliente encontrado pelo nome: ID {novo_cliente_id}")
                        
                        # Atualizar contrato
                        update_query = """
                            UPDATE contratos
                            SET cliente_id = %s
                            WHERE id = %s
                        """
                        database.execute_update(update_query, (novo_cliente_id, contrato_id))
                        print(f"   ? Contrato atualizado com cliente_id {novo_cliente_id}")
                        contratos_corrigidos += 1
                        detalhes.append({
                            'contrato_id': contrato_id,
                            'numero': numero,
                            'status': 'CORRIGIDO_POR_NOME',
                            'cliente_id': novo_cliente_id,
                            'cliente_nome': cliente_nome
                        })
                    else:
                        print(f"   ? Cliente n�o encontrado com nome '{cliente_nome}'")
                        contratos_com_problemas.append({
                            'contrato_id': contrato_id,
                            'numero': numero,
                            'cliente_nome': cliente_nome,
                            'motivo': 'Cliente n�o encontrado'
                        })
                        detalhes.append({
                            'contrato_id': contrato_id,
                            'numero': numero,
                            'status': 'ERRO',
                            'motivo': 'Cliente n�o encontrado'
                        })
                else:
                    print(f"   ? Contrato sem cliente_nome para buscar")
                    contratos_com_problemas.append({
                        'contrato_id': contrato_id,
                        'numero': numero,
                        'cliente_nome': None,
                        'motivo': 'Sem cliente_nome'
                    })
                    detalhes.append({
                        'contrato_id': contrato_id,
                        'numero': numero,
                        'status': 'ERRO',
                        'motivo': 'Sem cliente_nome'
                    })
                continue
            
            # Se tem m�ltiplos clientes nas sess�es, usar o primeiro e avisar
            if len(sessoes) > 1:
                print(f"   ?? ATEN��O: Contrato tem sess�es de {len(sessoes)} clientes diferentes!")
                for sessao in sessoes:
                    print(f"      - Cliente ID {sessao['cliente_id']}: {sessao['cliente_nome']}")
                print(f"   ?? Usando o primeiro cliente encontrado")
            
            # Atualizar com o cliente_id da sess�o
            novo_cliente_id = sessoes[0]['cliente_id']
            novo_cliente_nome = sessoes[0]['cliente_nome']
            
            print(f"   ?? Atualizando com cliente_id {novo_cliente_id} ({novo_cliente_nome})")
            
            update_query = """
                UPDATE contratos
                SET cliente_id = %s
                WHERE id = %s
            """
            
            database.execute_update(update_query, (novo_cliente_id, contrato_id))
            print(f"   ? Contrato atualizado com sucesso!")
            contratos_corrigidos += 1
            detalhes.append({
                'contrato_id': contrato_id,
                'numero': numero,
                'status': 'CORRIGIDO_POR_SESSAO',
                'cliente_id': novo_cliente_id,
                'cliente_nome': novo_cliente_nome,
                'multiplos_clientes': len(sessoes) > 1
            })
        
        # Resumo final
        print(f"\n{'='*80}")
        print(f"?? RESUMO DA CORRE��O")
        print(f"{'='*80}")
        print(f"? Contratos corrigidos: {contratos_corrigidos}")
        print(f"?? Contratos com problemas: {len(contratos_com_problemas)}")
        print(f"{'='*80}\n")
        
        return jsonify({
            'success': True,
            'message': f'Corre��o conclu�da',
            'contratos_corrigidos': contratos_corrigidos,
            'contratos_com_problemas': len(contratos_com_problemas),
            'problemas': contratos_com_problemas,
            'detalhes': detalhes
        })
        
    except Exception as e:
        logger.error(f"Erro ao corrigir contratos: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/test/db-connection', methods=['GET'])
def test_db_connection():
    """Endpoint de teste para verificar conex�o ao banco"""
    try:
        empresa_id = session.get('empresa_id', 20)
        
        logger.info(f"?? TEST: Testando conex�o com empresa_id={empresa_id}")
        
        # Testar se o m�todo existe
        if not hasattr(database, 'get_db_connection'):
            return jsonify({
                'success': False,
                'error': 'M�todo get_db_connection N�O ENCONTRADO na classe DatabaseManager',
                'available_methods': [m for m in dir(database) if not m.startswith('_')]
            }), 500
        
        logger.info(f"? TEST: M�todo get_db_connection existe!")
        
        # Testar a conex�o
        with database.get_db_connection(empresa_id=empresa_id) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM transacoes_extrato WHERE empresa_id = %s", (empresa_id,))
            _row_t = cursor.fetchone()
            total = _row_t['count'] if isinstance(_row_t, dict) else _row_t[0]
            cursor.close()
            
            logger.info(f"? TEST: Conex�o funcionou! {total} transa��es encontradas")
            
            return jsonify({
                'success': True,
                'message': 'Conex�o funcionando corretamente',
                'empresa_id': empresa_id,
                'total_transacoes': total
            })
            
    except Exception as e:
        logger.error(f"? TEST: Erro ao testar conex�o: {e}")
        import traceback
        logger.error(traceback.format_exc())
        
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


if __name__ == '__main__':
    # Inicializar tabelas de importa��o
    try:
        from database_import_manager import DatabaseImportManager
        import_manager = DatabaseImportManager()
        import_manager.create_import_tables()
        print("? Tabelas de importa��o inicializadas")
    except Exception as e:
        print(f"?? Erro ao inicializar tabelas de importa��o: {e}")
    
    # Configurar logging para produ��o (WARNING/ERROR apenas)
    import logging
    log_level = logging.WARNING if os.getenv('RAILWAY_ENVIRONMENT') else logging.INFO
    logging.basicConfig(level=log_level)
    app.logger.setLevel(log_level)
    
    # Porta configur�vel (Railway usa vari�vel de ambiente PORT)
    port = int(os.getenv('PORT', 5000))
    
    print("="*60)
    print("Sistema Financeiro - Vers�o Web")
    print("="*60)
    print(f"Servidor iniciado em: http://0.0.0.0:{port}")
    print(f"Banco de dados: {os.getenv('DATABASE_TYPE', 'sqlite')}")
    print(f"Log level: {logging.getLevelName(log_level)}")
    print("="*60)
    
    # Habilitar debug apenas em desenvolvimento local
    is_production = bool(os.getenv('RAILWAY_ENVIRONMENT') or os.getenv('RAILWAY_PROJECT_ID'))
    debug_mode = not is_production
    
    logger.info(f"Iniciando servidor - Modo: {'DESENVOLVIMENTO' if debug_mode else 'PRODU��O'}")
    app.run(debug=debug_mode, host='0.0.0.0', port=port, use_reloader=False)



