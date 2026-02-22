"""
Módulo para exportação de relatórios em PDF e Excel
Utiliza ReportLab para PDFs e openpyxl para Excel
"""

from io import BytesIO
from datetime import datetime
from typing import Dict, Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def formatar_moeda_pdf(valor: float) -> str:
    """Formata valor para display em moeda brasileira"""
    if valor < 0:
        return f"(R$ {abs(valor):,.2f})".replace(',', '_').replace('.', ',').replace('_', '.')
    return f"R$ {valor:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')


def formatar_percentual_pdf(valor: float) -> str:
    """Formata percentual para display"""
    return f"{valor:.2f}%"


def gerar_dre_pdf(
    dados_dre: Dict,
    nome_empresa: str = "Empresa",
    periodo: str = ""
) -> BytesIO:
    """
    Gera PDF da DRE (Demonstração do Resultado do Exercício)
    
    Args:
        dados_dre: Dicionário com dados da DRE retornados por gerar_dre()
        nome_empresa: Nome da empresa para o cabeçalho
        periodo: String com período (ex: "01/01/2026 a 31/01/2026")
    
    Returns:
        BytesIO com o PDF gerado
    """
    
    # Criar buffer para o PDF
    buffer = BytesIO()
    
    # Criar documento PDF (A4, margens)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo para título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulo
    subtitulo_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Estilo para data de geração
    data_style = ParagraphStyle(
        'DataStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_RIGHT
    )
    
    # Elementos do documento
    elements = []
    
    # Cabeçalho
    elements.append(Paragraph(nome_empresa, titulo_style))
    elements.append(Paragraph("DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO", titulo_style))
    
    if periodo:
        elements.append(Paragraph(f"Período: {periodo}", subtitulo_style))
    
    elements.append(Spacer(1, 10*mm))
    
    # Extrair dados da DRE
    dre = dados_dre.get('dre', {})
    comparacao = dados_dre.get('dre_anterior')
    
    # Preparar dados da tabela
    dados_tabela = []
    
    # Cabeçalho da tabela
    if comparacao:
        dados_tabela.append([
            'DESCRIÇÃO',
            'VALOR ATUAL (R$)',
            '% RECEITA',
            'VALOR ANTERIOR (R$)',
            'VARIAÇÃO'
        ])
    else:
        dados_tabela.append([
            'DESCRIÇÃO',
            'VALOR (R$)',
            '% RECEITA'
        ])
    
    def adicionar_linha(descricao: str, valor: float, percentual: float, 
                       nivel: int = 0, negrito: bool = False, final: bool = False,
                       valor_anterior: Optional[float] = None, variacao: Optional[float] = None):
        """Adiciona linha na tabela com formatação"""
        # Indentação baseada no nível
        desc_formatada = "  " * nivel + descricao
        
        if comparacao and valor_anterior is not None:
            linha = [
                desc_formatada,
                formatar_moeda_pdf(valor),
                formatar_percentual_pdf(percentual),
                formatar_moeda_pdf(valor_anterior),
                f"{variacao:+.2f}%" if variacao is not None else "-"
            ]
        else:
            linha = [
                desc_formatada,
                formatar_moeda_pdf(valor),
                formatar_percentual_pdf(percentual)
            ]
        
        dados_tabela.append(linha)
        return len(dados_tabela) - 1  # Retorna índice da linha para formatação
    
    indices_formato = []
    
    # 1. RECEITA BRUTA
    receita_bruta = dre.get('receita_bruta', {})
    idx = adicionar_linha(
        "1. RECEITA BRUTA",
        receita_bruta.get('total', 0),
        receita_bruta.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('receita_bruta', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('receita_bruta') if comparacao else None
    )
    indices_formato.append(('subtotal', idx))
    
    # Sub-contas de receita bruta
    for item in receita_bruta.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,  # Percentual não exibido para sub-contas
            nivel=1
        )
    
    # 2. DEDUÇÕES
    deducoes = dre.get('deducoes', {})
    idx = adicionar_linha(
        "2. (-) DEDUÇÕES DA RECEITA",
        -abs(deducoes.get('total', 0)),
        deducoes.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=-abs(comparacao.get('deducoes', {}).get('total', 0)) if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('deducoes') if comparacao else None
    )
    indices_formato.append(('deducao', idx))
    
    # Sub-contas de deduções
    for item in deducoes.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            -abs(item.get('valor', 0)),
            0,
            nivel=1
        )
    
    # 3. RECEITA LÍQUIDA
    receita_liquida = dre.get('receita_liquida', {})
    idx = adicionar_linha(
        "3. = RECEITA LÍQUIDA",
        receita_liquida.get('total', 0),
        receita_liquida.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('receita_liquida', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('receita_liquida') if comparacao else None
    )
    indices_formato.append(('resultado', idx))
    
    # 4. CUSTOS
    custos = dre.get('custos', {})
    idx = adicionar_linha(
        "4. (-) CUSTOS",
        custos.get('total', 0),
        custos.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('custos', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('custos') if comparacao else None
    )
    indices_formato.append(('deducao', idx))
    
    # Sub-contas de custos
    for item in custos.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=1
        )
    
    # 5. LUCRO BRUTO
    lucro_bruto = dre.get('lucro_bruto', {})
    idx = adicionar_linha(
        "5. = LUCRO BRUTO",
        lucro_bruto.get('total', 0),
        lucro_bruto.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('lucro_bruto', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('lucro_bruto') if comparacao else None
    )
    indices_formato.append(('resultado', idx))
    
    # 6. DESPESAS OPERACIONAIS
    desp_op = dre.get('despesas_operacionais', {})
    idx = adicionar_linha(
        "6. (-) DESPESAS OPERACIONAIS",
        desp_op.get('total', 0),
        desp_op.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('despesas_operacionais', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('despesas_operacionais') if comparacao else None
    )
    indices_formato.append(('deducao', idx))
    
    # Sub-contas de despesas operacionais
    for item in desp_op.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=1
        )
    
    # 7. RESULTADO OPERACIONAL
    res_op = dre.get('resultado_operacional', {})
    idx = adicionar_linha(
        "7. = RESULTADO OPERACIONAL",
        res_op.get('total', 0),
        res_op.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('resultado_operacional', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('resultado_operacional') if comparacao else None
    )
    indices_formato.append(('resultado', idx))
    
    # 8. RESULTADO FINANCEIRO
    res_fin = dre.get('resultado_financeiro', {})
    
    # 8.1 Receitas Financeiras
    rec_fin = res_fin.get('receitas', {})
    adicionar_linha(
        "8.1 (+) Receitas Financeiras",
        rec_fin.get('total', 0),
        rec_fin.get('percentual', 0),
        nivel=1
    )
    for item in rec_fin.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=2
        )
    
    # 8.2 Despesas Financeiras
    desp_fin = res_fin.get('despesas', {})
    adicionar_linha(
        "8.2 (-) Despesas Financeiras",
        desp_fin.get('total', 0),
        desp_fin.get('percentual', 0),
        nivel=1
    )
    for item in desp_fin.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=2
        )
    
    # 8. Total Resultado Financeiro
    idx = adicionar_linha(
        "8. = RESULTADO FINANCEIRO",
        res_fin.get('total', 0),
        res_fin.get('percentual', 0),
        nivel=0,
        negrito=True,
        valor_anterior=comparacao.get('resultado_financeiro', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('resultado_financeiro') if comparacao else None
    )
    indices_formato.append(('resultado', idx))
    
    # 9. LUCRO LÍQUIDO DO EXERCÍCIO
    lucro_liq = dre.get('lucro_liquido', {})
    idx = adicionar_linha(
        "9. = LUCRO LÍQUIDO DO EXERCÍCIO",
        lucro_liq.get('total', 0),
        lucro_liq.get('percentual', 0),
        nivel=0,
        negrito=True,
        final=True,
        valor_anterior=comparacao.get('lucro_liquido', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('lucro_liquido') if comparacao else None
    )
    indices_formato.append(('final', idx))
    
    # Criar tabela
    if comparacao:
        col_widths = [100*mm, 30*mm, 20*mm, 30*mm, 20*mm]
    else:
        col_widths = [120*mm, 40*mm, 30*mm]
    
    tabela = Table(dados_tabela, colWidths=col_widths)
    
    # Estilo da tabela
    estilo_tabela = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Corpo da tabela
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ])
    
    # Aplicar formatação especial para linhas específicas
    for tipo, idx in indices_formato:
        if tipo == 'subtotal':
            estilo_tabela.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#e8f5e9'))
            estilo_tabela.add('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold')
            estilo_tabela.add('TEXTCOLOR', (0, idx), (-1, idx), colors.HexColor('#27ae60'))
        elif tipo == 'deducao':
            estilo_tabela.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#ffebee'))
            estilo_tabela.add('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold')
            estilo_tabela.add('TEXTCOLOR', (0, idx), (-1, idx), colors.HexColor('#e74c3c'))
        elif tipo == 'resultado':
            estilo_tabela.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#e3f2fd'))
            estilo_tabela.add('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold')
            estilo_tabela.add('TEXTCOLOR', (0, idx), (-1, idx), colors.HexColor('#3498db'))
        elif tipo == 'final':
            estilo_tabela.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#2c3e50'))
            estilo_tabela.add('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold')
            estilo_tabela.add('TEXTCOLOR', (0, idx), (-1, idx), colors.white)
            estilo_tabela.add('FONTSIZE', (0, idx), (-1, idx), 11)
            estilo_tabela.add('TOPPADDING', (0, idx), (-1, idx), 8)
            estilo_tabela.add('BOTTOMPADDING', (0, idx), (-1, idx), 8)
    
    tabela.setStyle(estilo_tabela)
    elements.append(tabela)
    
    # Rodapé
    elements.append(Spacer(1, 10*mm))
    data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
    elements.append(Paragraph(f"Relatório gerado em: {data_geracao}", data_style))
    
    # Gerar PDF
    doc.build(elements)
    
    # Resetar buffer para leitura
    buffer.seek(0)
    
    return buffer


def gerar_dashboard_pdf(
    dados_dashboard: Dict,
    nome_empresa: str = "Empresa",
    mes_referencia: str = ""
) -> BytesIO:
    """
    Gera PDF do Dashboard Gerencial
    
    Args:
        dados_dashboard: Dicionário retornado por gerar_dashboard_gerencial()
        nome_empresa: Nome da empresa
        mes_referencia: Mês de referência formatado (ex: "Fevereiro/2026")
    
    Returns:
        BytesIO com o PDF gerado
    """
    
    buffer = BytesIO()
    
    # Criar documento
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitulo_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=16,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'Heading',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    elements = []
    
    # Cabeçalho
    elements.append(Paragraph(nome_empresa, titulo_style))
    elements.append(Paragraph("DASHBOARD GERENCIAL", titulo_style))
    if mes_referencia:
        elements.append(Paragraph(f"Mês de Referência: {mes_referencia}", subtitulo_style))
    elements.append(Spacer(1, 8*mm))
    
    # Extrair dados
    dashboard = dados_dashboard.get('dashboard', {})
    kpis = dashboard.get('kpis', {})
    ponto_eq = dashboard.get('ponto_equilibrio', {})
    comparacao = dashboard.get('comparacao', {})
    
    # ===== KPIs PRINCIPAIS =====
    elements.append(Paragraph("📊 INDICADORES PRINCIPAIS", heading_style))
    
    kpi_data = [
        ['INDICADOR', 'VALOR', 'VARIAÇÃO'],
        [
            'Receita do Mês',
            formatar_moeda_pdf(kpis.get('receita_mes', {}).get('valor', 0)),
            f"{kpis.get('receita_mes', {}).get('variacao_percentual', 0):+.2f}%"
        ],
        [
            'Despesas do Mês',
            formatar_moeda_pdf(kpis.get('despesas_mes', {}).get('valor', 0)),
            f"{kpis.get('despesas_mes', {}).get('percentual_receita', 0):.2f}% da receita"
        ],
        [
            'Lucro Líquido',
            formatar_moeda_pdf(kpis.get('lucro_liquido_mes', {}).get('valor', 0)),
            f"{kpis.get('lucro_liquido_mes', {}).get('variacao_percentual', 0):+.2f}%"
        ],
        [
            'Margem Líquida',
            f"{kpis.get('margem_liquida', {}).get('percentual', 0):.2f}%",
            kpis.get('margem_liquida', {}).get('status', '').upper()
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[60*mm, 50*mm, 50*mm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    elements.append(kpi_table)
    elements.append(Spacer(1, 8*mm))
    
    # ===== PONTO DE EQUILÍBRIO =====
    elements.append(Paragraph("⚖️ PONTO DE EQUILÍBRIO", heading_style))
    
    pe_data = [
        ['MÉTRICA', 'VALOR'],
        ['Valor de Equilíbrio', formatar_moeda_pdf(ponto_eq.get('valor', 0))],
        ['Percentual Atingido', f"{ponto_eq.get('percentual_atingido', 0):.2f}%"],
        ['Falta para Equilíbrio', formatar_moeda_pdf(ponto_eq.get('falta_para_equilibrio', 0))],
        ['Status', '✅ ATINGIDO' if ponto_eq.get('atingiu') else '⚠️ NÃO ATINGIDO'],
        ['Custos Fixos', formatar_moeda_pdf(ponto_eq.get('custos_fixos', 0))],
        ['Margem de Contribuição', f"{ponto_eq.get('margem_contribuicao_percentual', 0):.2f}%"]
    ]
    
    pe_table = Table(pe_data, colWidths=[80*mm, 80*mm])
    pe_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        
        # Destacar status
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#d5f4e6') if ponto_eq.get('atingiu') else colors.HexColor('#fadbd8')),
        ('TEXTCOLOR', (0, 4), (-1, 4), colors.HexColor('#27ae60') if ponto_eq.get('atingiu') else colors.HexColor('#e74c3c')),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
    ]))
    
    elements.append(pe_table)
    elements.append(Spacer(1, 8*mm))
    
    # ===== COMPARAÇÃO MÊS ANTERIOR =====
    elements.append(Paragraph("📅 COMPARAÇÃO COM MÊS ANTERIOR", heading_style))
    
    comp_data = [
        ['MÉTRICA', 'MÊS ANTERIOR', 'MÊS ATUAL', 'VARIAÇÃO'],
        [
            'Receita',
            formatar_moeda_pdf(comparacao.get('mes_anterior', {}).get('receita', 0)),
            formatar_moeda_pdf(kpis.get('receita_mes', {}).get('valor', 0)),
            f"{comparacao.get('variacoes', {}).get('receita', 0):+.2f}%"
        ],
        [
            'Lucro Líquido',
            formatar_moeda_pdf(comparacao.get('mes_anterior', {}).get('lucro', 0)),
            formatar_moeda_pdf(kpis.get('lucro_liquido_mes', {}).get('valor', 0)),
            f"{comparacao.get('variacoes', {}).get('lucro', 0):+.2f}%"
        ]
    ]
    
    comp_table = Table(comp_data, colWidths=[50*mm, 35*mm, 35*mm, 40*mm])
    comp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    elements.append(comp_table)
    elements.append(Spacer(1, 8*mm))
    
    # ===== DESPESAS POR CATEGORIA =====
    graficos_adicionais = dashboard.get('graficos_adicionais', {})
    despesas_cat = graficos_adicionais.get('despesas_por_categoria', [])
    
    if despesas_cat:
        elements.append(Paragraph("🔴 DESPESAS POR CATEGORIA", heading_style))
        
        desp_data = [['CATEGORIA', 'VALOR', '% DO TOTAL']]
        total_despesas = sum(item['valor'] for item in despesas_cat)
        
        for item in despesas_cat:
            percentual = (item['valor'] / total_despesas * 100) if total_despesas > 0 else 0
            desp_data.append([
                item['categoria'],
                formatar_moeda_pdf(item['valor']),
                f"{percentual:.2f}%"
            ])
        
        desp_table = Table(desp_data, colWidths=[80*mm, 50*mm, 30*mm])
        desp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffebee')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPadding', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(desp_table)
        elements.append(Spacer(1, 8*mm))
    
    # ===== RECEITAS POR CATEGORIA =====
    receitas_cat = graficos_adicionais.get('receitas_por_categoria', [])
    
    if receitas_cat:
        elements.append(Paragraph("🟢 RECEITAS POR CATEGORIA (Top 10)", heading_style))
        
        rec_data = [['CATEGORIA', 'VALOR', '% DO TOTAL']]
        total_receitas = sum(item['valor'] for item in receitas_cat)
        
        for item in receitas_cat[:10]:  # Top 10
            percentual = (item['valor'] / total_receitas * 100) if total_receitas > 0 else 0
            rec_data.append([
                item['categoria'],
                formatar_moeda_pdf(item['valor']),
                f"{percentual:.2f}%"
            ])
        
        rec_table = Table(rec_data, colWidths=[80*mm, 50*mm, 30*mm])
        rec_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f5e9')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(rec_table)
    
    # Rodapé
    elements.append(Spacer(1, 10*mm))
    data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
    rodape_style = ParagraphStyle(
        'Rodape',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#7F8C8D'),
        alignment=TA_RIGHT,
        fontName='Helvetica-Oblique'
    )
    elements.append(Paragraph(f"Relatório gerado em: {data_geracao}", rodape_style))
    
    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


def gerar_dre_excel(
    dados_dre: Dict,
    nome_empresa: str = "Empresa",
    periodo: str = ""
) -> BytesIO:
    """
    Gera Excel da DRE (Demonstração do Resultado do Exercício)
    
    Args:
        dados_dre: Dicionário com dados da DRE retornados por gerar_dre()
        nome_empresa: Nome da empresa para o cabeçalho
        periodo: String com período (ex: "01/01/2026 a 31/01/2026")
    
    Returns:
        BytesIO com o arquivo Excel gerado
    """
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DRE"
    
    # Extrair dados
    dre = dados_dre.get('dre', {})
    comparacao = dados_dre.get('dre_anterior')
    
    # ===== ESTILOS =====
    
    # Fonte título
    font_titulo = Font(name='Arial', size=14, bold=True, color='2C3E50')
    
    # Fonte subtítulo
    font_subtitulo = Font(name='Arial', size=11, color='34495E')
    
    # Fonte cabeçalho tabela
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    
    # Bordas
    borda_fina = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    borda_grossa = Border(
        left=Side(style='medium', color='2C3E50'),
        right=Side(style='medium', color='2C3E50'),
        top=Side(style='medium', color='2C3E50'),
        bottom=Side(style='medium', color='2C3E50')
    )
    
    # Cores de fundo por tipo
    fill_receita = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_deducao = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    fill_resultado = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    fill_final = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    fill_alternado = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # Fontes por tipo
    font_receita = Font(name='Arial', size=10, bold=True, color='27AE60')
    font_deducao = Font(name='Arial', size=10, bold=True, color='E74C3C')
    font_resultado = Font(name='Arial', size=10, bold=True, color='3498DB')
    font_final = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    font_normal = Font(name='Arial', size=10, color='2C3E50')
    font_subconta = Font(name='Arial', size=9, color='34495E', italic=True)
    
    # Alinhamentos
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # ===== CABEÇALHO =====
    
    # Linha 1: Nome da empresa
    ws.merge_cells('A1:E1' if comparacao else 'A1:C1')
    cell_empresa = ws['A1']
    cell_empresa.value = nome_empresa
    cell_empresa.font = font_titulo
    cell_empresa.alignment = align_center
    
    # Linha 2: Título
    ws.merge_cells('A2:E2' if comparacao else 'A2:C2')
    cell_titulo = ws['A2']
    cell_titulo.value = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO"
    cell_titulo.font = font_titulo
    cell_titulo.alignment = align_center
    
    # Linha 3: Período
    if periodo:
        ws.merge_cells('A3:E3' if comparacao else 'A3:C3')
        cell_periodo = ws['A3']
        cell_periodo.value = f"Período: {periodo}"
        cell_periodo.font = font_subtitulo
        cell_periodo.alignment = align_center
    
    # Linha 4: Vazia
    linha_atual = 5
    
    # ===== CABEÇALHO DA TABELA =====
    
    if comparacao:
        headers = ['DESCRIÇÃO', 'VALOR ATUAL (R$)', '% RECEITA', 'VALOR ANTERIOR (R$)', 'VARIAÇÃO (%)']
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
    else:
        headers = ['DESCRIÇÃO', 'VALOR (R$)', '% RECEITA']
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=linha_atual, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda_fina
    
    linha_atual += 1
    linha_dados_inicio = linha_atual
    
    # ===== FUNÇÃO AUXILIAR PARA ADICIONAR LINHAS =====
    
    def adicionar_linha(descricao: str, valor: float, percentual: float,
                       nivel: int = 0, tipo: str = 'normal',
                       valor_anterior: Optional[float] = None,
                       variacao: Optional[float] = None):
        """Adiciona linha na planilha com formatação"""
        nonlocal linha_atual
        
        # Indentação
        desc_formatada = "  " * nivel + descricao
        
        # Coluna A: Descrição
        cell_desc = ws.cell(row=linha_atual, column=1)
        cell_desc.value = desc_formatada
        cell_desc.alignment = align_left
        cell_desc.border = borda_fina
        
        # Coluna B: Valor
        cell_valor = ws.cell(row=linha_atual, column=2)
        cell_valor.value = valor
        cell_valor.number_format = '#,##0.00'
        cell_valor.alignment = align_right
        cell_valor.border = borda_fina
        
        # Coluna C: Percentual
        cell_perc = ws.cell(row=linha_atual, column=3)
        cell_perc.value = percentual / 100  # Excel espera 0.1 para 10%
        cell_perc.number_format = '0.00%'
        cell_perc.alignment = align_right
        cell_perc.border = borda_fina
        
        # Colunas D e E: Comparação (se houver)
        if comparacao and valor_anterior is not None:
            cell_ant = ws.cell(row=linha_atual, column=4)
            cell_ant.value = valor_anterior
            cell_ant.number_format = '#,##0.00'
            cell_ant.alignment = align_right
            cell_ant.border = borda_fina
            
            cell_var = ws.cell(row=linha_atual, column=5)
            if variacao is not None:
                cell_var.value = variacao / 100
                cell_var.number_format = '+0.00%;-0.00%'
            else:
                cell_var.value = '-'
            cell_var.alignment = align_right
            cell_var.border = borda_fina
        
        # Aplicar formatação por tipo
        if tipo == 'receita':
            cell_desc.fill = fill_receita
            cell_valor.fill = fill_receita
            cell_perc.fill = fill_receita
            cell_desc.font = font_receita
            cell_valor.font = font_receita
            cell_perc.font = font_receita
            if comparacao:
                cell_ant.fill = fill_receita
                cell_var.fill = fill_receita
                cell_ant.font = font_receita
                cell_var.font = font_receita
        elif tipo == 'deducao':
            cell_desc.fill = fill_deducao
            cell_valor.fill = fill_deducao
            cell_perc.fill = fill_deducao
            cell_desc.font = font_deducao
            cell_valor.font = font_deducao
            cell_perc.font = font_deducao
            if comparacao:
                cell_ant.fill = fill_deducao
                cell_var.fill = fill_deducao
                cell_ant.font = font_deducao
                cell_var.font = font_deducao
        elif tipo == 'resultado':
            cell_desc.fill = fill_resultado
            cell_valor.fill = fill_resultado
            cell_perc.fill = fill_resultado
            cell_desc.font = font_resultado
            cell_valor.font = font_resultado
            cell_perc.font = font_resultado
            if comparacao:
                cell_ant.fill = fill_resultado
                cell_var.fill = fill_resultado
                cell_ant.font = font_resultado
                cell_var.font = font_resultado
        elif tipo == 'final':
            cell_desc.fill = fill_final
            cell_valor.fill = fill_final
            cell_perc.fill = fill_final
            cell_desc.font = font_final
            cell_valor.font = font_final
            cell_perc.font = font_final
            cell_desc.border = borda_grossa
            cell_valor.border = borda_grossa
            cell_perc.border = borda_grossa
            if comparacao:
                cell_ant.fill = fill_final
                cell_var.fill = fill_final
                cell_ant.font = font_final
                cell_var.font = font_final
                cell_ant.border = borda_grossa
                cell_var.border = borda_grossa
        elif tipo == 'subconta':
            cell_desc.font = font_subconta
            cell_valor.font = font_subconta
            cell_perc.font = font_subconta
            if linha_atual % 2 == 0:
                cell_desc.fill = fill_alternado
                cell_valor.fill = fill_alternado
                cell_perc.fill = fill_alternado
                if comparacao:
                    cell_ant.fill = fill_alternado
                    cell_var.fill = fill_alternado
        else:  # normal
            cell_desc.font = font_normal
            cell_valor.font = font_normal
            cell_perc.font = font_normal
            if linha_atual % 2 == 0:
                cell_desc.fill = fill_alternado
                cell_valor.fill = fill_alternado
                cell_perc.fill = fill_alternado
                if comparacao:
                    cell_ant.fill = fill_alternado
                    cell_var.fill = fill_alternado
        
        linha_atual += 1
    
    # ===== POPULAR DADOS DA DRE =====
    
    # 1. RECEITA BRUTA
    receita_bruta = dre.get('receita_bruta', {})
    adicionar_linha(
        "1. RECEITA BRUTA",
        receita_bruta.get('total', 0),
        receita_bruta.get('percentual', 0),
        nivel=0,
        tipo='receita',
        valor_anterior=comparacao.get('receita_bruta', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('receita_bruta') if comparacao else None
    )
    for item in receita_bruta.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=1,
            tipo='subconta'
        )
    
    # 2. DEDUÇÕES
    deducoes = dre.get('deducoes', {})
    adicionar_linha(
        "2. (-) DEDUÇÕES DA RECEITA",
        -abs(deducoes.get('total', 0)),
        deducoes.get('percentual', 0),
        nivel=0,
        tipo='deducao',
        valor_anterior=-abs(comparacao.get('deducoes', {}).get('total', 0)) if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('deducoes') if comparacao else None
    )
    for item in deducoes.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            -abs(item.get('valor', 0)),
            0,
            nivel=1,
            tipo='subconta'
        )
    
    # 3. RECEITA LÍQUIDA
    receita_liquida = dre.get('receita_liquida', {})
    adicionar_linha(
        "3. = RECEITA LÍQUIDA",
        receita_liquida.get('total', 0),
        receita_liquida.get('percentual', 0),
        nivel=0,
        tipo='resultado',
        valor_anterior=comparacao.get('receita_liquida', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('receita_liquida') if comparacao else None
    )
    
    # 4. CUSTOS
    custos = dre.get('custos', {})
    adicionar_linha(
        "4. (-) CUSTOS",
        custos.get('total', 0),
        custos.get('percentual', 0),
        nivel=0,
        tipo='deducao',
        valor_anterior=comparacao.get('custos', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('custos') if comparacao else None
    )
    for item in custos.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=1,
            tipo='subconta'
        )
    
    # 5. LUCRO BRUTO
    lucro_bruto = dre.get('lucro_bruto', {})
    adicionar_linha(
        "5. = LUCRO BRUTO",
        lucro_bruto.get('total', 0),
        lucro_bruto.get('percentual', 0),
        nivel=0,
        tipo='resultado',
        valor_anterior=comparacao.get('lucro_bruto', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('lucro_bruto') if comparacao else None
    )
    
    # 6. DESPESAS OPERACIONAIS
    desp_op = dre.get('despesas_operacionais', {})
    adicionar_linha(
        "6. (-) DESPESAS OPERACIONAIS",
        desp_op.get('total', 0),
        desp_op.get('percentual', 0),
        nivel=0,
        tipo='deducao',
        valor_anterior=comparacao.get('despesas_operacionais', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('despesas_operacionais') if comparacao else None
    )
    for item in desp_op.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=1,
            tipo='subconta'
        )
    
    # 7. RESULTADO OPERACIONAL
    res_op = dre.get('resultado_operacional', {})
    adicionar_linha(
        "7. = RESULTADO OPERACIONAL",
        res_op.get('total', 0),
        res_op.get('percentual', 0),
        nivel=0,
        tipo='resultado',
        valor_anterior=comparacao.get('resultado_operacional', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('resultado_operacional') if comparacao else None
    )
    
    # 8. RESULTADO FINANCEIRO
    res_fin = dre.get('resultado_financeiro', {})
    
    # 8.1 Receitas Financeiras
    rec_fin = res_fin.get('receitas', {})
    adicionar_linha(
        "8.1 (+) Receitas Financeiras",
        rec_fin.get('total', 0),
        rec_fin.get('percentual', 0),
        nivel=1,
        tipo='normal'
    )
    for item in rec_fin.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=2,
            tipo='subconta'
        )
    
    # 8.2 Despesas Financeiras
    desp_fin = res_fin.get('despesas', {})
    adicionar_linha(
        "8.2 (-) Despesas Financeiras",
        desp_fin.get('total', 0),
        desp_fin.get('percentual', 0),
        nivel=1,
        tipo='normal'
    )
    for item in desp_fin.get('itens', []):
        adicionar_linha(
            item.get('descricao', ''),
            item.get('valor', 0),
            0,
            nivel=2,
            tipo='subconta'
        )
    
    # 8. Total Resultado Financeiro
    adicionar_linha(
        "8. = RESULTADO FINANCEIRO",
        res_fin.get('total', 0),
        res_fin.get('percentual', 0),
        nivel=0,
        tipo='resultado',
        valor_anterior=comparacao.get('resultado_financeiro', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('resultado_financeiro') if comparacao else None
    )
    
    # 9. LUCRO LÍQUIDO DO EXERCÍCIO
    lucro_liq = dre.get('lucro_liquido', {})
    adicionar_linha(
        "9. = LUCRO LÍQUIDO DO EXERCÍCIO",
        lucro_liq.get('total', 0),
        lucro_liq.get('percentual', 0),
        nivel=0,
        tipo='final',
        valor_anterior=comparacao.get('lucro_liquido', {}).get('total') if comparacao else None,
        variacao=dados_dre.get('variacoes', {}).get('lucro_liquido') if comparacao else None
    )
    
    # ===== RODAPÉ =====
    linha_atual += 2
    ws.merge_cells(f'A{linha_atual}:E{linha_atual}' if comparacao else f'A{linha_atual}:C{linha_atual}')
    cell_rodape = ws[f'A{linha_atual}']
    data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
    cell_rodape.value = f"Relatório gerado em: {data_geracao}"
    cell_rodape.font = Font(name='Arial', size=8, color='7F8C8D', italic=True)
    cell_rodape.alignment = align_right
    
    # ===== SALVAR EM BUFFER =====
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def gerar_dashboard_excel(
    dados_dashboard: Dict,
    nome_empresa: str = "Empresa",
    mes_referencia: str = ""
) -> BytesIO:
    """
    Gera Excel do Dashboard Gerencial
    
    Args:
        dados_dashboard: Dicionário retornado por gerar_dashboard_gerencial()
        nome_empresa: Nome da empresa
        mes_referencia: Mês de referência (ex: "Fevereiro/2026")
    
    Returns:
        BytesIO com o arquivo Excel gerado
    """
    
    # Criar workbook com múltiplas abas
    wb = Workbook()
    
    # Extrair dados
    dashboard = dados_dashboard.get('dashboard', {})
    kpis = dashboard.get('kpis', {})
    evolucao = dashboard.get('evolucao_mensal', [])
    ponto_eq = dashboard.get('ponto_equilibrio', {})
    comparacao = dashboard.get('comparacao', {})
    graficos_adicionais = dashboard.get('graficos_adicionais', {})
    
    # ===== ABA 1: KPIs =====
    ws_kpis = wb.active
    ws_kpis.title = "KPIs"
    
    # Estilos
    font_titulo = Font(name='Arial', size=14, bold=True, color='2C3E50')
    font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    borda_fina = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Cabeçalho
    ws_kpis.merge_cells('A1:D1')
    cell = ws_kpis['A1']
    cell.value = nome_empresa
    cell.font = font_titulo
    cell.alignment = align_center
    
    ws_kpis.merge_cells('A2:D2')
    cell = ws_kpis['A2']
    cell.value = "DASHBOARD GERENCIAL"
    cell.font = font_titulo
    cell.alignment = align_center
    
    if mes_referencia:
        ws_kpis.merge_cells('A3:D3')
        cell = ws_kpis['A3']
        cell.value = f"Mês: {mes_referencia}"
        cell.font = Font(name='Arial', size=11, color='34495E')
        cell.alignment = align_center
    
    # KPIs
    linha = 5
    ws_kpis.merge_cells(f'A{linha}:D{linha}')
    cell = ws_kpis[f'A{linha}']
    cell.value = "📊 INDICADORES PRINCIPAIS"
    cell.font = Font(name='Arial', size=12, bold=True, color='2C3E50')
    linha += 1
    
    # Headers
    headers = ['INDICADOR', 'VALOR', 'VARIAÇÃO', 'STATUS']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_kpis.cell(row=linha, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda_fina
    linha += 1
    
    # Dados KPIs
    kpi_data = [
        ['Receita do Mês', kpis.get('receita_mes', {}).get('valor', 0),
         f"{kpis.get('receita_mes', {}).get('variacao_percentual', 0):+.2f}%",
         kpis.get('receita_mes', {}).get('tendencia', '')],
        ['Despesas do Mês', kpis.get('despesas_mes', {}).get('valor', 0),
         f"{kpis.get('despesas_mes', {}).get('percentual_receita', 0):.2f}% da receita", ''],
        ['Lucro Líquido', kpis.get('lucro_liquido_mes', {}).get('valor', 0),
         f"{kpis.get('lucro_liquido_mes', {}).get('variacao_percentual', 0):+.2f}%",
         kpis.get('lucro_liquido_mes', {}).get('tendencia', '')],
        ['Margem Líquida', kpis.get('margem_liquida', {}).get('percentual', 0),
         '', kpis.get('margem_liquida', {}).get('status', '').upper()]
    ]
    
    for row_data in kpi_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_kpis.cell(row=linha, column=col_idx)
            if col_idx == 2 and isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0.00' if row_data[0] != 'Margem Líquida' else '0.00"%"'
            else:
                cell.value = value
            cell.alignment = align_right if col_idx in [2, 3] else align_left
            cell.border = borda_fina
        linha += 1
    
    # Larguras
    ws_kpis.column_dimensions['A'].width = 25
    ws_kpis.column_dimensions['B'].width = 20
    ws_kpis.column_dimensions['C'].width = 25
    ws_kpis.column_dimensions['D'].width = 15
    
    # ===== ABA 2: EVOLUÇÃO MENSAL =====
    ws_evolucao = wb.create_sheet("Evolução 12 Meses")
    
    # Cabeçalho
    ws_evolucao.merge_cells('A1:E1')
    cell = ws_evolucao['A1']
    cell.value = "📈 EVOLUÇÃO MENSAL (Últimos 12 Meses)"
    cell.font = font_titulo
    cell.alignment = align_center
    
    # Headers
    linha = 3
    headers_ev = ['MÊS', 'RECEITA', 'DESPESAS', 'LUCRO LÍQUIDO', 'MARGEM %']
    for col_idx, header in enumerate(headers_ev, start=1):
        cell = ws_evolucao.cell(row=linha, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda_fina
    linha += 1
    
    # Dados
    for mes_data in evolucao:
        ws_evolucao.cell(row=linha, column=1, value=mes_data.get('mes_nome', '')).alignment = align_left
        ws_evolucao.cell(row=linha, column=2, value=mes_data.get('receita', 0)).number_format = '#,##0.00'
        ws_evolucao.cell(row=linha, column=3, value=mes_data.get('despesas', 0)).number_format = '#,##0.00'
        ws_evolucao.cell(row=linha, column=4, value=mes_data.get('lucro_liquido', 0)).number_format = '#,##0.00'
        ws_evolucao.cell(row=linha, column=5, value=mes_data.get('margem', 0) / 100).number_format = '0.00%'
        
        for col in range(1, 6):
            ws_evolucao.cell(row=linha, column=col).border = borda_fina
            ws_evolucao.cell(row=linha, column=col).alignment = align_right if col > 1 else align_left
        linha += 1
    
    # Larguras
    ws_evolucao.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E']:
        ws_evolucao.column_dimensions[col].width = 18
    
    # ===== ABA 3: DESPESAS POR CATEGORIA =====
    despesas_cat = graficos_adicionais.get('despesas_por_categoria', [])
    if despesas_cat:
        ws_desp = wb.create_sheet("Despesas por Categoria")
        
        ws_desp.merge_cells('A1:C1')
        cell = ws_desp['A1']
        cell.value = "🔴 DESPESAS POR CATEGORIA"
        cell.font = font_titulo
        cell.alignment = align_center
        
        linha = 3
        headers_desp = ['CATEGORIA', 'VALOR', '% DO TOTAL']
        for col_idx, header in enumerate(headers_desp, start=1):
            cell = ws_desp.cell(row=linha, column=col_idx)
            cell.value = header
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = borda_fina
        linha += 1
        
        total_despesas = sum(item['valor'] for item in despesas_cat)
        for item in despesas_cat:
            percentual = (item['valor'] / total_despesas) if total_despesas > 0 else 0
            ws_desp.cell(row=linha, column=1, value=item['categoria']).alignment = align_left
            ws_desp.cell(row=linha, column=2, value=item['valor']).number_format = '#,##0.00'
            ws_desp.cell(row=linha, column=3, value=percentual).number_format = '0.00%'
            
            for col in range(1, 4):
                ws_desp.cell(row=linha, column=col).border = borda_fina
                ws_desp.cell(row=linha, column=col).alignment = align_right if col > 1 else align_left
            linha += 1
        
        ws_desp.column_dimensions['A'].width = 40
        ws_desp.column_dimensions['B'].width = 20
        ws_desp.column_dimensions['C'].width = 15
    
    # ===== ABA 4: RECEITAS POR CATEGORIA =====
    receitas_cat = graficos_adicionais.get('receitas_por_categoria', [])
    if receitas_cat:
        ws_rec = wb.create_sheet("Receitas por Categoria")
        
        ws_rec.merge_cells('A1:C1')
        cell = ws_rec['A1']
        cell.value = "🟢 RECEITAS POR CATEGORIA"
        cell.font = font_titulo
        cell.alignment = align_center
        
        linha = 3
        headers_rec = ['CATEGORIA', 'VALOR', '% DO TOTAL']
        for col_idx, header in enumerate(headers_rec, start=1):
            cell = ws_rec.cell(row=linha, column=col_idx)
            cell.value = header
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = borda_fina
        linha += 1
        
        total_receitas = sum(item['valor'] for item in receitas_cat)
        for item in receitas_cat:
            percentual = (item['valor'] / total_receitas) if total_receitas > 0 else 0
            ws_rec.cell(row=linha, column=1, value=item['categoria']).alignment = align_left
            ws_rec.cell(row=linha, column=2, value=item['valor']).number_format = '#,##0.00'
            ws_rec.cell(row=linha, column=3, value=percentual).number_format = '0.00%'
            
            for col in range(1, 4):
                ws_rec.cell(row=linha, column=col).border = borda_fina
                ws_rec.cell(row=linha, column=col).alignment = align_right if col > 1 else align_left
            linha += 1
        
        ws_rec.column_dimensions['A'].width = 40
        ws_rec.column_dimensions['B'].width = 20
        ws_rec.column_dimensions['C'].width = 15
    
    # Salvar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


# ============================================================================
# EXPORTAÇÃO DE CONTRATOS
# ============================================================================

def gerar_contratos_pdf(contratos: list, nome_empresa: str = "Empresa") -> BytesIO:
    """
    Gera PDF com lista de contratos
    
    Args:
        contratos: Lista de contratos retornada pelo backend
        nome_empresa: Nome da empresa para o cabeçalho
    
    Returns:
        BytesIO com o PDF gerado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(nome_empresa, titulo_style))
    elements.append(Paragraph("RELATÓRIO DE CONTRATOS", titulo_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                             ParagraphStyle('DataStyle', parent=styles['Normal'], fontSize=8, 
                                          textColor=colors.grey, alignment=TA_RIGHT)))
    elements.append(Spacer(1, 10*mm))
    
    # Preparar dados da tabela
    dados_tabela = [[
        'Nº',
        'Cliente',
        'Tipo',
        'Valor',
        'Vigência',
        'Horas',
        'Status'
    ]]
    
    for contrato in contratos:
        # Extrair dados do observacoes JSON
        obs = contrato.get('observacoes_json', {}) or {}
        tipo = obs.get('tipo', 'N/A')
        nome_contrato = obs.get('nome', contrato.get('descricao', 'N/A'))[:30]
        
        # Calcular saldo de horas
        horas_totais = float(contrato.get('horas_totais', 0))
        horas_utilizadas = float(contrato.get('horas_utilizadas', 0))
        horas_restantes = horas_totais - horas_utilizadas
        horas_extras = float(contrato.get('horas_extras', 0))
        
        # Montar string de horas
        if contrato.get('controle_horas_ativo'):
            horas_str = f"{horas_totais:.1f}h\n{horas_restantes:.1f}h rest."
            if horas_extras > 0:
                horas_str += f"\n+{horas_extras:.1f}h extra"
        else:
            horas_str = "N/A"
        
        # Vigência
        data_inicio = contrato.get('data_vigencia_inicio', contrato.get('data_inicio', ''))
        data_fim = contrato.get('data_vigencia_fim', contrato.get('data_fim', ''))
        if data_inicio:
            vigencia = f"{data_inicio[:10]}\naté\n{data_fim[:10] if data_fim else 'Indeterminado'}"
        else:
            vigencia = "N/A"
        
        dados_tabela.append([
            contrato.get('numero', 'N/A'),
            Paragraph(contrato.get('cliente_nome', 'N/A')[:25], styles['Normal']),
            tipo,
            formatar_moeda_pdf(float(contrato.get('valor_contrato', contrato.get('valor', 0)))),
            Paragraph(vigencia, ParagraphStyle('SmallStyle', parent=styles['Normal'], fontSize=7)),
            Paragraph(horas_str, ParagraphStyle('SmallStyle', parent=styles['Normal'], fontSize=7)),
            contrato.get('status_pagamento', contrato.get('status', 'N/A'))[:10]
        ])
    
    # Criar tabela
    tabela = Table(dados_tabela, colWidths=[20*mm, 35*mm, 20*mm, 25*mm, 30*mm, 25*mm, 20*mm])
    
    estilo_tabela = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])
    
    tabela.setStyle(estilo_tabela)
    elements.append(tabela)
    
    # Estatísticas
    elements.append(Spacer(1, 10*mm))
    
    total_contratos = len(contratos)
    valor_total = sum(float(c.get('valor_contrato', c.get('valor', 0))) for c in contratos)
    contratos_com_horas = len([c for c in contratos if c.get('controle_horas_ativo')])
    
    stats_data = [
        ['📊 ESTATÍSTICAS', ''],
        ['Total de Contratos:', str(total_contratos)],
        ['Valor Total:', formatar_moeda_pdf(valor_total)],
        ['Contratos com Controle de Horas:', str(contratos_com_horas)]
    ]
    
    stats_table = Table(stats_data, colWidths=[60*mm, 40*mm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    
    elements.append(stats_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def gerar_contratos_excel(contratos: list, nome_empresa: str = "Empresa") -> BytesIO:
    """
    Gera Excel com lista de contratos
    
    Args:
        contratos: Lista de contratos retornada pelo backend
        nome_empresa: Nome da empresa para o cabeçalho
    
    Returns:
        BytesIO com o Excel gerado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    
    # Estilos
    font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_titulo = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
    fill_header = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = f"{nome_empresa} - RELATÓRIO DE CONTRATOS"
    cell.font = font_titulo
    cell.fill = fill_titulo
    cell.alignment = align_center
    
    # Subtítulo
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = align_center
    
    # Cabeçalhos
    headers = ['Nº', 'Cliente', 'Tipo', 'Nome/Descrição', 'Valor', 'Data Início', 'Data Fim', 
               'Horas Totais', 'Horas Utilizadas', 'Horas Restantes', 'Horas Extras', 'Status']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda
    
    # Dados
    row = 5
    for contrato in contratos:
        obs = contrato.get('observacoes_json', {}) or {}
        
        ws.cell(row=row, column=1, value=contrato.get('numero', 'N/A')).border = borda
        ws.cell(row=row, column=2, value=contrato.get('cliente_nome', 'N/A')).border = borda
        ws.cell(row=row, column=3, value=obs.get('tipo', 'N/A')).border = borda
        ws.cell(row=row, column=4, value=obs.get('nome', contrato.get('descricao', 'N/A'))).border = borda
        
        cell = ws.cell(row=row, column=5, value=float(contrato.get('valor_contrato', contrato.get('valor', 0))))
        cell.number_format = 'R$ #,##0.00'
        cell.border = borda
        cell.alignment = align_right
        
        ws.cell(row=row, column=6, value=contrato.get('data_vigencia_inicio', contrato.get('data_inicio', ''))).border = borda
        ws.cell(row=row, column=7, value=contrato.get('data_vigencia_fim', contrato.get('data_fim', ''))).border = borda
        
        ws.cell(row=row, column=8, value=float(contrato.get('horas_totais', 0))).border = borda
        ws.cell(row=row, column=9, value=float(contrato.get('horas_utilizadas', 0))).border = borda
        
        horas_restantes = float(contrato.get('horas_totais', 0)) - float(contrato.get('horas_utilizadas', 0))
        ws.cell(row=row, column=10, value=horas_restantes).border = borda
        ws.cell(row=row, column=11, value=float(contrato.get('horas_extras', 0))).border = borda
        ws.cell(row=row, column=12, value=contrato.get('status_pagamento', contrato.get('status', 'N/A'))).border = borda
        
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================================
# EXPORTAÇÃO DE SESSÕES
# ============================================================================

def gerar_sessoes_pdf(sessoes: list, nome_empresa: str = "Empresa") -> BytesIO:
    """
    Gera PDF com lista de sessões
    
    Args:
        sessoes: Lista de sessões retornada pelo backend
        nome_empresa: Nome da empresa para o cabeçalho
    
    Returns:
        BytesIO com o PDF gerado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(nome_empresa, titulo_style))
    elements.append(Paragraph("RELATÓRIO DE SESSÕES", titulo_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                             ParagraphStyle('DataStyle', parent=styles['Normal'], fontSize=8, 
                                          textColor=colors.grey, alignment=TA_RIGHT)))
    elements.append(Spacer(1, 10*mm))
    
    # Preparar dados da tabela
    dados_tabela = [[
        'ID',
        'Cliente',
        'Contrato',
        'Data',
        'Horas',
        'Status',
        'Descrição'
    ]]
    
    for sessao in sessoes:
        # Extrair horas trabalhadas
        horas_trab = float(sessao.get('horas_trabalhadas', 0))
        dados_json = sessao.get('dados_json', {}) or {}
        qtd_horas = dados_json.get('quantidade_horas', 0)
        
        horas_str = f"{horas_trab}h" if horas_trab > 0 else (f"{qtd_horas}h" if qtd_horas else "N/A")
        
        # Status com emoji
        status = sessao.get('status', 'N/A')
        status_emoji = {
            'rascunho': '📝',
            'agendada': '📅',
            'em_andamento': '🔄',
            'finalizada': '✅',
            'cancelada': '❌',
            'reaberta': '🔄'
        }.get(status, '❓')
        
        dados_tabela.append([
            str(sessao.get('id', '')),
            Paragraph(sessao.get('cliente_nome', 'N/A')[:20], styles['Normal']),
            sessao.get('contrato_numero', 'N/A')[:10],
            sessao.get('data', '')[:10] if sessao.get('data') else 'N/A',
            horas_str,
            f"{status_emoji} {status}",
            Paragraph(sessao.get('descricao', 'N/A')[:40], 
                     ParagraphStyle('SmallStyle', parent=styles['Normal'], fontSize=7))
        ])
    
    # Criar tabela
    tabela = Table(dados_tabela, colWidths=[12*mm, 30*mm, 20*mm, 20*mm, 15*mm, 25*mm, 50*mm])
    
    estilo_tabela = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])
    
    tabela.setStyle(estilo_tabela)
    elements.append(tabela)
    
    # Estatísticas
    elements.append(Spacer(1, 10*mm))
    
    total_sessoes = len(sessoes)
    sessoes_finalizadas = len([s for s in sessoes if s.get('status') == 'finalizada'])
    total_horas = sum(float(s.get('horas_trabalhadas', 0)) for s in sessoes)
    
    stats_data = [
        ['📊 ESTATÍSTICAS', ''],
        ['Total de Sessões:', str(total_sessoes)],
        ['Sessões Finalizadas:', str(sessoes_finalizadas)],
        ['Total de Horas Trabalhadas:', f"{total_horas:.1f}h"]
    ]
    
    stats_table = Table(stats_data, colWidths=[60*mm, 40*mm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    
    elements.append(stats_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def gerar_sessoes_excel(sessoes: list, nome_empresa: str = "Empresa") -> BytesIO:
    """
    Gera Excel com lista de sessões
    
    Args:
        sessoes: Lista de sessões retornada pelo backend
        nome_empresa: Nome da empresa para o cabeçalho
    
    Returns:
        BytesIO com o Excel gerado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sessões"
    
    # Estilos
    font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_titulo = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
    fill_header = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = f"{nome_empresa} - RELATÓRIO DE SESSÕES"
    cell.font = font_titulo
    cell.fill = fill_titulo
    cell.alignment = align_center
    
    # Subtítulo
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = align_center
    
    # Cabeçalhos
    headers = ['ID', 'Cliente', 'Contrato Nº', 'Contrato Nome', 'Data', 'Horário', 
               'Horas Trabalhadas', 'Status', 'Endereço', 'Descrição']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda
    
    # Dados
    row = 5
    for sessao in sessoes:
        dados_json = sessao.get('dados_json', {}) or {}
        
        ws.cell(row=row, column=1, value=sessao.get('id', '')).border = borda
        ws.cell(row=row, column=2, value=sessao.get('cliente_nome', 'N/A')).border = borda
        ws.cell(row=row, column=3, value=sessao.get('contrato_numero', 'N/A')).border = borda
        ws.cell(row=row, column=4, value=sessao.get('contrato_nome', 'N/A')).border = borda
        ws.cell(row=row, column=5, value=sessao.get('data', '')).border = borda
        ws.cell(row=row, column=6, value=dados_json.get('horario', 'N/A')).border = borda
        
        horas_trab = float(sessao.get('horas_trabalhadas', 0))
        ws.cell(row=row, column=7, value=horas_trab).border = borda
        ws.cell(row=row, column=7).number_format = '0.0'
        
        ws.cell(row=row, column=8, value=sessao.get('status', 'N/A')).border = borda
        ws.cell(row=row, column=9, value=sessao.get('endereco', 'N/A')).border = borda
        ws.cell(row=row, column=10, value=sessao.get('descricao', 'N/A')).border = borda
        
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 40
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================================
# RELATÓRIO DE CONTROLE DE HORAS
# ============================================================================

def gerar_relatorio_controle_horas_pdf(dados: dict, nome_empresa: str = "Empresa") -> BytesIO:
    """
    Gera PDF do Relatório de Controle de Horas
    
    Args:
        dados: Dicionário com dados retornados pelo backend
        nome_empresa: Nome da empresa para o cabeçalho
    
    Returns:
        BytesIO com o PDF gerado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(nome_empresa, titulo_style))
    elements.append(Paragraph("⏱️ RELATÓRIO DE CONTROLE DE HORAS", titulo_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                             ParagraphStyle('DataStyle', parent=styles['Normal'], fontSize=8, 
                                          textColor=colors.grey, alignment=TA_RIGHT)))
    elements.append(Spacer(1, 10*mm))
    
    # Resumo Geral
    resumo = dados.get('resumo', {})
    
    resumo_data = [
        ['📊 RESUMO GERAL', '', '', ''],
        ['Total de Contratos:', str(resumo.get('total_contratos', 0)), 
         'Contratos Ativos:', str(resumo.get('contratos_ativos', 0))],
        ['Contratos com Controle de Horas:', str(resumo.get('contratos_com_controle_horas', 0)),
         'Total de Sessões:', str(resumo.get('total_sessoes', 0))],
        ['Horas Totais Contratadas:', f"{resumo.get('total_horas_contratadas', 0):.1f}h",
         'Horas Utilizadas:', f"{resumo.get('total_horas_utilizadas', 0):.1f}h"],
        ['Horas Restantes:', f"{resumo.get('total_horas_restantes', 0):.1f}h",
         'Horas Extras:', f"{resumo.get('total_horas_extras', 0):.1f}h"]
    ]
    
    resumo_table = Table(resumo_data, colWidths=[50*mm, 40*mm, 50*mm, 40*mm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    
    elements.append(resumo_table)
    elements.append(Spacer(1, 10*mm))
    
    # Detalhamento por Contrato
    elements.append(Paragraph("📋 DETALHAMENTO POR CONTRATO", 
                             ParagraphStyle('SubTitle', parent=styles['Heading2'], 
                                          fontSize=12, textColor=colors.HexColor('#2c3e50'))))
    elements.append(Spacer(1, 5*mm))
    
    contratos = dados.get('contratos', [])
    
    dados_contratos = [[
        'Nº',
        'Cliente',
        'Totais',
        'Utilizadas',
        'Restantes',
        'Extras',
        'Status'
    ]]
    
    for contrato in contratos:
        horas_totais = float(contrato.get('horas_totais', 0))
        horas_utilizadas = float(contrato.get('horas_utilizadas', 0))
        horas_restantes = float(contrato.get('horas_restantes', 0))
        horas_extras = float(contrato.get('horas_extras', 0))
        
        # Determinar status do saldo
        if horas_extras > 0:
            status = "⚠️ Extras"
            cor_status = colors.HexColor('#e74c3c')
        elif horas_restantes <= 5:
            status = "⚡ Baixo"
            cor_status = colors.HexColor('#f39c12')
        elif horas_restantes > 0:
            status = "✅ OK"
            cor_status = colors.HexColor('#27ae60')
        else:
            status = "N/A"
            cor_status = colors.grey
        
        dados_contratos.append([
            contrato.get('numero', 'N/A'),
            Paragraph(contrato.get('cliente_nome', 'N/A')[:20], styles['Normal']),
            f"{horas_totais:.1f}h",
            f"{horas_utilizadas:.1f}h",
            f"{horas_restantes:.1f}h",
            f"{horas_extras:.1f}h" if horas_extras > 0 else "-",
            status
        ])
    
    tabela_contratos = Table(dados_contratos, colWidths=[20*mm, 40*mm, 20*mm, 25*mm, 25*mm, 20*mm, 20*mm])
    
    estilo_contratos = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])
    
    tabela_contratos.setStyle(estilo_contratos)
    elements.append(tabela_contratos)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def gerar_relatorio_controle_horas_excel(dados: dict, nome_empresa: str = "Empresa") -> BytesIO:
    """
    Gera Excel do Relatório de Controle de Horas
    
    Args:
        dados: Dicionário com dados retornados pelo backend
        nome_empresa: Nome da empresa para o cabeçalho
    
    Returns:
        BytesIO com o Excel gerado
    """
    wb = Workbook()
    
    # ABA 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Estilos
    font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_titulo = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
    fill_header = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws_resumo.merge_cells('A1:D1')
    cell = ws_resumo['A1']
    cell.value = f"{nome_empresa} - CONTROLE DE HORAS"
    cell.font = font_titulo
    cell.fill = fill_titulo
    cell.alignment = align_center
    
    ws_resumo.merge_cells('A2:D2')
    cell = ws_resumo['A2']
    cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = align_center
    
    # Resumo
    resumo = dados.get('resumo', {})
    
    row = 4
    dados_resumo = [
        ['INDICADOR', 'VALOR'],
        ['Total de Contratos', resumo.get('total_contratos', 0)],
        ['Contratos Ativos', resumo.get('contratos_ativos', 0)],
        ['Contratos com Controle de Horas', resumo.get('contratos_com_controle_horas', 0)],
        ['Total de Sessões', resumo.get('total_sessoes', 0)],
        ['Horas Totais Contratadas', f"{resumo.get('total_horas_contratadas', 0):.1f}"],
        ['Horas Utilizadas', f"{resumo.get('total_horas_utilizadas', 0):.1f}"],
        ['Horas Restantes', f"{resumo.get('total_horas_restantes', 0):.1f}"],
        ['Horas Extras', f"{resumo.get('total_horas_extras', 0):.1f}"],
    ]
    
    for col_idx, header in enumerate(dados_resumo[0], start=1):
        cell = ws_resumo.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda
    
    row += 1
    for item in dados_resumo[1:]:
        ws_resumo.cell(row=row, column=1, value=item[0]).border = borda
        cell = ws_resumo.cell(row=row, column=2, value=item[1])
        cell.border = borda
        cell.alignment = align_right
        row += 1
    
    ws_resumo.column_dimensions['A'].width = 40
    ws_resumo.column_dimensions['B'].width = 20
    
    # ABA 2: Contratos
    ws_contratos = wb.create_sheet("Contratos")
    
    ws_contratos.merge_cells('A1:I1')
    cell = ws_contratos['A1']
    cell.value = "CONTRATOS - CONTROLE DE HORAS"
    cell.font = font_titulo
    cell.fill = fill_titulo
    cell.alignment = align_center
    
    headers = ['Nº Contrato', 'Cliente', 'Tipo', 'Horas Totais', 'Horas Utilizadas', 
               'Horas Restantes', 'Horas Extras', '% Utilizado', 'Status']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_contratos.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda
    
    contratos = dados.get('contratos', [])
    row = 4
    
    for contrato in contratos:
        ws_contratos.cell(row=row, column=1, value=contrato.get('numero', 'N/A')).border = borda
        ws_contratos.cell(row=row, column=2, value=contrato.get('cliente_nome', 'N/A')).border = borda
        
        obs = contrato.get('observacoes_json', {}) or {}
        ws_contratos.cell(row=row, column=3, value=obs.get('tipo', 'N/A')).border = borda
        
        ws_contratos.cell(row=row, column=4, value=float(contrato.get('horas_totais', 0))).border = borda
        ws_contratos.cell(row=row, column=5, value=float(contrato.get('horas_utilizadas', 0))).border = borda
        ws_contratos.cell(row=row, column=6, value=float(contrato.get('horas_restantes', 0))).border = borda
        ws_contratos.cell(row=row, column=7, value=float(contrato.get('horas_extras', 0))).border = borda
        
        perc_utilizado = float(contrato.get('percentual_utilizado', 0))
        cell = ws_contratos.cell(row=row, column=8, value=perc_utilizado / 100)
        cell.number_format = '0.0%'
        cell.border = borda
        
        # Determinar status
        horas_extras = float(contrato.get('horas_extras', 0))
        horas_restantes = float(contrato.get('horas_restantes', 0))
        
        if horas_extras > 0:
            status = "EXTRAS"
        elif horas_restantes <= 5:
            status = "BAIXO"
        elif horas_restantes > 0:
            status = "OK"
        else:
            status = "N/A"
        
        ws_contratos.cell(row=row, column=9, value=status).border = borda
        row += 1
    
    ws_contratos.column_dimensions['A'].width = 15
    ws_contratos.column_dimensions['B'].width = 30
    ws_contratos.column_dimensions['C'].width = 12
    ws_contratos.column_dimensions['D'].width = 12
    ws_contratos.column_dimensions['E'].width = 15
    ws_contratos.column_dimensions['F'].width = 15
    ws_contratos.column_dimensions['G'].width = 12
    ws_contratos.column_dimensions['H'].width = 12
    ws_contratos.column_dimensions['I'].width = 12
    
    # ABA 3: Sessões por Contrato
    ws_sessoes = wb.create_sheet("Sessões")
    
    ws_sessoes.merge_cells('A1:G1')
    cell = ws_sessoes['A1']
    cell.value = "SESSÕES POR CONTRATO"
    cell.font = font_titulo
    cell.fill = fill_titulo
    cell.alignment = align_center
    
    headers_sessoes = ['Contrato', 'Cliente', 'Data', 'Horas Trabalhadas', 'Status', 'Descrição']
    
    for col_idx, header in enumerate(headers_sessoes, start=1):
        cell = ws_sessoes.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = borda
    
    row = 4
    for contrato in contratos:
        sessoes = contrato.get('sessoes', [])
        if sessoes:
            for sessao in sessoes:
                ws_sessoes.cell(row=row, column=1, value=contrato.get('numero', 'N/A')).border = borda
                ws_sessoes.cell(row=row, column=2, value=contrato.get('cliente_nome', 'N/A')).border = borda
                ws_sessoes.cell(row=row, column=3, value=sessao.get('data', 'N/A')).border = borda
                ws_sessoes.cell(row=row, column=4, value=float(sessao.get('horas_trabalhadas', 0))).border = borda
                ws_sessoes.cell(row=row, column=5, value=sessao.get('status', 'N/A')).border = borda
                ws_sessoes.cell(row=row, column=6, value=sessao.get('descricao', 'N/A')).border = borda
                row += 1
    
    ws_sessoes.column_dimensions['A'].width = 15
    ws_sessoes.column_dimensions['B'].width = 30
    ws_sessoes.column_dimensions['C'].width = 12
    ws_sessoes.column_dimensions['D'].width = 15
    ws_sessoes.column_dimensions['E'].width = 15
    ws_sessoes.column_dimensions['F'].width = 40
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
